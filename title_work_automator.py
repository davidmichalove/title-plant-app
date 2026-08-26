import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from tkinterdnd2 import TkinterDnD, DND_FILES
import threading
import os

# Dynamically resolve the root of the hard drive (works on Mac & Windows)
DRIVE_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import shutil
import time
import base64
import urllib.parse
import traceback
import requests
import re
import PyPDF2
import json
import datetime
import threading
import base64
from bs4 import BeautifulSoup
from send2trash import send2trash

try:
    import geopandas as gpd
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import NoSuchFrameException
except ImportError as e:
    messagebox.showerror("Error", f"Missing dependencies: {e}")
    exit(1)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ZIP_PATH = os.path.join(BASE_DIR, "Polygon_Belmont_County_Web_Parcels_20260501085529 (1).zip")
TEMPLATE_1 = os.path.join(BASE_DIR, "PID OR (DATE)_TEMPLATE (2).xlsx")
TEMPLATE_2 = os.path.join(BASE_DIR, "PID RS (DATE)_TEMPLATE.xlsx")

class AutomatorApp:
    def set_buttons_state(self, state):
        def _set():
            for btn_attr in ['run_btn', 'manual_btn', 'scrape_btn', 'og_btn', 'court_btn', 'gis_btn', 'next_page_btn', 'court_view_btn', 'gis_view_btn']:
                btn = getattr(self, btn_attr, None)
                if btn:
                    try:
                        btn.config(state=state)
                    except Exception:
                        pass
        if hasattr(self, 'root'):
            self.root.after(0, _set)

    def __init__(self, root):
        self.root = root
        self.root.title("Automate Title Work")
        self.root.geometry("700x400")
        self.setup_ui()
        self.gdf = None
        self.viewer_window = None

        self.req_headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8'
        }
        
        self.refresh_missing_logs()

    def setup_ui(self):
        # Create a canvas and a vertical scrollbar for scrolling
        self.canvas = tk.Canvas(self.root)
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.canvas.yview)
        
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        

            
        frame = ttk.Frame(self.canvas, padding=10)
        self.canvas_window = self.canvas.create_window((300, 0), window=frame, anchor="n")
        
        def on_frame_configure(event):
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        
        def on_canvas_configure(event):
            self.canvas.coords(self.canvas_window, event.width / 2, 0)


        frame.bind("<Configure>", on_frame_configure)
        self.canvas.bind("<Configure>", on_canvas_configure)

        # Global mouse wheel scrolling is intentionally disabled here
        # so that scrolling with fingers only scrolls inner widgets 
        # (like the listbox) instead of bouncing the whole window up and down.

        import glob
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        
        def get_dir_mtime(d):
            latest = os.path.getmtime(d)
            try:
                for entry in os.scandir(d):
                    if not entry.name.startswith('.'):
                        try:
                            m = entry.stat().st_mtime
                            if m > latest:
                                latest = m
                        except Exception: pass
            except Exception: pass
            return latest

        pid_folders = []
        for d in glob.glob(os.path.join(base_dir, "PID *")):
            if os.path.isdir(d):
                folder_name = os.path.basename(d)
                raw_name = re.sub(r'^PID[\s_]*', '', folder_name).strip()
                if "TEMPLATE" in raw_name.upper():
                    continue

                u_name = ""
                unit_match = re.search(r'\(([^)]+)\)', raw_name)
                if unit_match:
                    u_name = unit_match.group(1).strip()
                    p_num = raw_name[:unit_match.start()].strip()
                else:
                    p_num = raw_name

                mtime = get_dir_mtime(d)
                pid_folders.append((mtime, p_num, u_name))

        # Sort descending by latest modified time (most recently worked on first)
        pid_folders.sort(key=lambda x: x[0], reverse=True)
        recent_parcels = []
        recent_units = []
        for _, p_num, u_name in pid_folders:
            if p_num not in recent_parcels:
                recent_parcels.append(p_num)
            if u_name and u_name not in recent_units:
                recent_units.append(u_name)

        top_header_wrapper = ttk.Frame(frame)
        top_header_wrapper.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        top_header_frame = ttk.Frame(top_header_wrapper)
        top_header_frame.pack(anchor=tk.CENTER)
        
        ttk.Label(top_header_frame, text="Parcel Number:").pack(side=tk.LEFT)
        self.parcel_entry = ttk.Combobox(top_header_frame, values=recent_parcels, width=15, postcommand=self.refresh_recent_parcels)
        self.parcel_entry.pack(side=tk.LEFT, padx=(5, 20))
        
        ttk.Label(top_header_frame, text="Unit Name:").pack(side=tk.LEFT)
        self.unit_entry = ttk.Combobox(top_header_frame, values=recent_units, width=12, postcommand=self.refresh_recent_parcels)
        self.unit_entry.pack(side=tk.LEFT, padx=5)
        self.unit_entry.set("")
        self.unit_entry.bind('<<ComboboxSelected>>', lambda e: self.on_parcel_change())
        self.unit_entry.bind('<KeyRelease>', lambda e: self.on_parcel_change())

        btn_container_top = ttk.Frame(frame)
        btn_container_top.grid(row=1, column=0, columnspan=2, pady=(10, 5))
        
        self.run_btn = ttk.Button(btn_container_top, text="Run Auto", command=self.start_automation)
        self.run_btn.pack(side=tk.LEFT, padx=3)
        self.complete_btn = ttk.Button(btn_container_top, text="Complete", command=self.complete_assignment)
        self.complete_btn.pack(side=tk.LEFT, padx=3)
        self.checklist_btn = ttk.Button(btn_container_top, text="Checklst", command=self.open_checklist)
        self.checklist_btn.pack(side=tk.LEFT, padx=3)
        self.status_btn = ttk.Button(btn_container_top, text="PID Stat", command=self.open_status_tracker)
        self.status_btn.pack(side=tk.LEFT, padx=3)
        self.open_or_btn = ttk.Button(btn_container_top, text="OR", command=self.open_ownership_report)
        self.open_or_btn.pack(side=tk.LEFT, padx=3)
        self.email_btn = ttk.Button(btn_container_top, text="Email", command=self.open_submission_email)
        self.email_btn.pack(side=tk.LEFT, padx=3)
        self.recorder_btn = ttk.Button(btn_container_top, text="B-Rcrdr", command=self.open_belmont_recorder)
        self.recorder_btn.pack(side=tk.LEFT, padx=3)

        btn_container_bottom = ttk.Frame(frame)
        btn_container_bottom.grid(row=2, column=0, columnspan=2, pady=(0, 10))

        self.chat_btn = ttk.Button(btn_container_bottom, text="Agent", command=self.open_chat)
        self.chat_btn.pack(side=tk.LEFT, padx=3)
        self.rs_editor_btn = ttk.Button(btn_container_bottom, text="Edit RS", command=self.open_rs_editor)
        self.rs_editor_btn.pack(side=tk.LEFT, padx=3)
        self.gemini_rs_btn = ttk.Button(btn_container_bottom, text="Gemini RS", command=self.open_gemini_rs_editor)
        self.gemini_rs_btn.pack(side=tk.LEFT, padx=3)
        self.reformat_rs_btn = ttk.Button(btn_container_bottom, text="Reformat RS", command=self.reformat_active_parcel_runsheet)
        self.reformat_rs_btn.pack(side=tk.LEFT, padx=3)
        self.gis_browser_btn = ttk.Button(btn_container_bottom, text="B-GIS", command=self.open_belmont_gis)
        self.gis_browser_btn.pack(side=tk.LEFT, padx=3)
        self.name_search_btn = ttk.Button(btn_container_bottom, text="Nom Search", command=self.open_name_search)
        self.name_search_btn.pack(side=tk.LEFT, padx=3)
        self.shortcuts_btn = ttk.Button(btn_container_bottom, text="Shrtcs", command=self.show_shortcuts_dialog)
        self.shortcuts_btn.pack(side=tk.LEFT, padx=3)

        self.notebook = ttk.Notebook(frame)
        self.notebook.grid(row=3, column=0, columnspan=2, sticky='nsew', pady=5)
        
        tab_main = ttk.Frame(self.notebook)
        tab_tools = ttk.Frame(self.notebook)
        
        self.notebook.add(tab_main, text="Workflow")
        self.notebook.add(tab_tools, text="Tools")
        
        # --- TOOLS TAB CONTENT ---
        combiner_frame = ttk.Frame(tab_tools)
        combiner_frame.grid(row=97, column=0, columnspan=2, pady=(15, 5), sticky="w", padx=10)
        ttk.Label(combiner_frame, text="PDF Combiner Utility:", font=("Helvetica", 12, "bold")).pack(anchor=tk.W, pady=(0, 2))
        ttk.Label(combiner_frame, text="Merge multiple PDFs into a single file with drag & drop, custom ordering, and auto-naming.", wraplength=450).pack(anchor=tk.W, pady=(0, 5))
        ttk.Button(combiner_frame, text="📑 Open PDF Combiner", command=self.open_pdf_combiner).pack(anchor=tk.W)

        plat_frame = ttk.Frame(tab_tools)
        plat_frame.grid(row=98, column=0, columnspan=2, pady=(15, 5), sticky="w", padx=10)
        
        ttk.Label(plat_frame, text="Belmont Plat Cabinet Searcher:", font=("Helvetica", 12, "bold")).pack(anchor=tk.W, pady=(0, 2))
        ttk.Label(plat_frame, text="Search 2,970+ Plat Cabinets (Cabinets A-F & Plat slides) by subdivision name or slide number.", wraplength=450).pack(anchor=tk.W, pady=(0, 5))
        ttk.Button(plat_frame, text="📑 Open Plat Cabinet Searcher", command=self.open_plat_cabinet_searcher).pack(anchor=tk.W)

        gis_frame = ttk.Frame(tab_tools)
        gis_frame.grid(row=99, column=0, columnspan=2, pady=10, sticky="w", padx=10)
        ttk.Label(gis_frame, text="Manual Utilities:", font=("Helvetica", 12, "bold")).pack(anchor=tk.W, pady=(0, 2))
        
        def run_manual_gis():
            p_num = self.parcel_entry.get().strip()
            if not p_num:
                messagebox.showerror("Error", "Please enter a Parcel Number in the main tab first!")
                return
            
            import subprocess
            import threading
            
            def background_gis():
                base_dir = "/Volumes/davidlls/assignments"
                maps_dir = os.path.join(base_dir, f"PID {p_num}", "MAPS")
                if not os.path.exists(maps_dir):
                    os.makedirs(maps_dir)
                    
                print(f"Triggering background GIS map generation for {p_num}...")
                subprocess.run(["python3", "/Volumes/davidlls/assignments/app/gis_map_generator.py", p_num, maps_dir])
            
            threading.Thread(target=background_gis, daemon=True).start()
            messagebox.showinfo("Started", f"GIS Map generation for {p_num} started in the background!\n\nThe PDF will automatically pop open on your screen in about 60 seconds once it finishes.")

        ttk.Button(gis_frame, text="Generate Professional GIS Map", command=run_manual_gis).pack(anchor=tk.W)
        # -------------------------
        
        tab_notes = ttk.Frame(self.notebook)
        self.notebook.add(tab_notes, text="Notes")
        
        tab_logs = ttk.Frame(self.notebook)
        self.notebook.add(tab_logs, text="Missing")
        
        tab_activity = ttk.Frame(self.notebook)
        self.notebook.add(tab_activity, text="Logs")
        
        self.large_log_text = tk.Text(tab_activity, height=25, state=tk.DISABLED, font=("Helvetica", 11))
        self.large_log_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        notes_split = ttk.PanedWindow(tab_notes, orient=tk.HORIZONTAL)
        notes_split.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Left Panel
        notes_left = ttk.Frame(notes_split)
        notes_split.add(notes_left, weight=1)
        
        self.notes_tree = ttk.Treeview(notes_left, columns=("Date", "Time", "Subject"), show="headings", height=15)
        self.notes_tree.column("Date", width=70)
        self.notes_tree.column("Time", width=70)
        self.notes_tree.column("Subject", width=150)
        self.notes_tree.heading("Date", text="Date")
        self.notes_tree.heading("Time", text="Time")
        self.notes_tree.heading("Subject", text="Subject")
        self.notes_tree.pack(fill=tk.BOTH, expand=True)
        self.notes_tree.bind('<<TreeviewSelect>>', lambda e: self.load_selected_note())
        
        btn_frame = ttk.Frame(notes_left)
        btn_frame.pack(fill=tk.X, pady=5)
        ttk.Button(btn_frame, text="New Note", command=self.new_note).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Delete Note", command=self.delete_selected_note).pack(side=tk.LEFT, padx=5)
        
        # Right Panel
        notes_right = ttk.Frame(notes_split)
        notes_split.add(notes_right, weight=1)
        
        subj_frame = ttk.Frame(notes_right)
        subj_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(subj_frame, text="Subject:").pack(side=tk.LEFT)
        self.note_subject_var = tk.StringVar()
        ttk.Entry(subj_frame, textvariable=self.note_subject_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.note_text = tk.Text(notes_right, height=20, width=40, font=("Helvetica", 14), wrap=tk.WORD)
        self.note_text.pack(fill=tk.BOTH, expand=True)
        
        ttk.Button(notes_right, text="Save Note", command=self.save_note).pack(pady=5)
        
        self.current_note_id = None
        
        # --- TAB 3: Missing Logs ---
        logs_top_frame = ttk.Frame(tab_logs)
        logs_top_frame.pack(fill=tk.X, pady=10, padx=10)
        ttk.Button(logs_top_frame, text="Refresh Logs", command=self.refresh_missing_logs).pack(side=tk.LEFT, padx=5)
        ttk.Button(logs_top_frame, text="Mark as Found", command=self.mark_log_found).pack(side=tk.LEFT, padx=5)
        ttk.Button(logs_top_frame, text="Clear All Logs", command=self.clear_missing_logs).pack(side=tk.LEFT, padx=5)
        
        logs_split = ttk.PanedWindow(tab_logs, orient=tk.HORIZONTAL)
        logs_split.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Left side: Missing Volumes
        vol_frame = ttk.LabelFrame(logs_split, text="Entire Volume Folder Missing")
        logs_split.add(vol_frame, weight=1)
        
        self.missing_vols_tree = ttk.Treeview(vol_frame, columns=("Type", "Vol", "Pg", "PID"), show="headings", height=8)
        self.missing_vols_tree.column("Type", width=60)
        self.missing_vols_tree.column("Vol", width=50)
        self.missing_vols_tree.column("Pg", width=50)
        self.missing_vols_tree.column("PID", width=90)
        self.missing_vols_tree.heading("Type", text="Type")
        self.missing_vols_tree.heading("Vol", text="Volume")
        self.missing_vols_tree.heading("Pg", text="Page Req.")
        self.missing_vols_tree.heading("PID", text="PID")
        self.missing_vols_tree.pack(fill=tk.BOTH, expand=True)
        
        # Right side: Missing Pages
        pg_frame = ttk.LabelFrame(logs_split, text="Page Missing (Volume Exists)")
        logs_split.add(pg_frame, weight=1)
        
        self.missing_pgs_tree = ttk.Treeview(pg_frame, columns=("Type", "Vol", "Pg", "PID"), show="headings", height=8)
        self.missing_pgs_tree.column("Type", width=60)
        self.missing_pgs_tree.column("Vol", width=50)
        self.missing_pgs_tree.column("Pg", width=50)
        self.missing_pgs_tree.column("PID", width=90)
        self.missing_pgs_tree.heading("Type", text="Type")
        self.missing_pgs_tree.heading("Vol", text="Volume")
        self.missing_pgs_tree.heading("Pg", text="Missing Page")
        self.missing_pgs_tree.heading("PID", text="PID")
        self.missing_pgs_tree.pack(fill=tk.BOTH, expand=True)
        
        
        # --- TAB 1 ---
        ttk.Label(tab_main, text="Get Deed / Mortgage").grid(row=0, column=0, columnspan=2, pady=5)
        vol_pg_frame = ttk.Frame(tab_main)
        vol_pg_frame.grid(row=1, column=0, columnspan=2, pady=5)
        
        self.doc_type_combo = ttk.Combobox(vol_pg_frame, values=["Deed", "Mortgage", "ALL"], width=8, state="readonly")
        self.doc_type_combo.set("Deed")
        self.doc_type_combo.grid(row=0, column=0, padx=2)
        
        ttk.Label(vol_pg_frame, text="Volume:").grid(row=0, column=1, padx=2)
        self.vol_entry = ttk.Entry(vol_pg_frame, width=8)
        self.vol_entry.grid(row=0, column=2, padx=2)
        
        ttk.Label(vol_pg_frame, text="Page:").grid(row=0, column=3, padx=2)
        self.pg_entry = ttk.Entry(vol_pg_frame, width=8)
        self.pg_entry.grid(row=0, column=4, padx=2)
        
        self.manual_btn = ttk.Button(vol_pg_frame, text="Get Document", command=self.start_manual_download)
        self.manual_btn.grid(row=0, column=5, padx=5)
        
        self.scrape_btn = ttk.Button(vol_pg_frame, text="Get From Website", command=self.start_manual_scrape)
        self.scrape_btn.grid(row=1, column=5, padx=5, pady=2)
        

        
        self.next_page_btn = ttk.Button(vol_pg_frame, text="Next Page >", command=self.fetch_next_page, state=tk.NORMAL)
        self.next_page_btn.grid(row=0, column=6, padx=2)
        
        viewer_frame = ttk.LabelFrame(tab_main, text="Document Viewer & Docket", padding=5)
        viewer_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        
        top_vframe = ttk.Frame(viewer_frame)
        top_vframe.pack(fill=tk.X)
        
        ttk.Label(top_vframe, text="Folder:").pack(side=tk.LEFT, padx=(0, 5))
        self.viewer_folder_combo = ttk.Combobox(top_vframe, state="readonly", width=15)
        self.viewer_folder_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.viewer_folder_combo.bind('<<ComboboxSelected>>', lambda e: self.refresh_viewer_list())
        
        vbtn_frame = ttk.Frame(top_vframe)
        vbtn_frame.pack(side=tk.RIGHT)
        
        vbtn_row1 = ttk.Frame(vbtn_frame)
        vbtn_row1.pack(fill=tk.X, pady=(0,2))
        ttk.Button(vbtn_row1, text="Combine & Send to Organizer", command=self.combine_to_organizer).pack(side=tk.RIGHT, padx=(5,0))
        ttk.Button(vbtn_row1, text="Open", command=self.open_selected_viewer_file).pack(side=tk.RIGHT, padx=(5,0))
        
        vbtn_row2 = ttk.Frame(vbtn_frame)
        vbtn_row2.pack(fill=tk.X)
        ttk.Button(vbtn_row2, text="Discard", command=self.mark_irrelevant_file).pack(side=tk.RIGHT, padx=(5,0))
        ttk.Button(vbtn_row2, text="Undo Discard", command=self.undo_discard).pack(side=tk.RIGHT, padx=(5,0))
        ttk.Button(vbtn_row2, text="Clean Docket", command=self.clean_docket).pack(side=tk.RIGHT, padx=(5,0))
        
        vbtn_row3 = ttk.Frame(vbtn_frame)
        vbtn_row3.pack(fill=tk.X, pady=(2, 0))
        ttk.Button(vbtn_row3, text="Send Back to Docket", command=self.send_back_to_docket).pack(side=tk.RIGHT, padx=(5,0))
        ttk.Button(vbtn_row3, text="Redact", command=self.redact_document).pack(side=tk.RIGHT, padx=(5,0))
        ttk.Button(vbtn_row3, text="Generate AI Abstract", command=self.generate_ai_abstract_from_docket).pack(side=tk.RIGHT, padx=(5,0))
        
        search_vframe = ttk.Frame(viewer_frame)
        search_vframe.pack(fill=tk.X, pady=(5, 0))
        ttk.Label(search_vframe, text="Search:").pack(side=tk.LEFT, padx=(0, 5))
        self.viewer_search_var = tk.StringVar()
        self.viewer_search_entry = ttk.Entry(search_vframe, textvariable=self.viewer_search_var)
        self.viewer_search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.viewer_search_entry.bind('<KeyRelease>', self.on_viewer_search)
        
        list_vframe = ttk.Frame(viewer_frame)
        list_vframe.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.viewer_listbox = tk.Listbox(list_vframe, selectmode=tk.EXTENDED, height=12)
        scrollbar = ttk.Scrollbar(list_vframe, orient=tk.VERTICAL, command=self.viewer_listbox.yview)
        self.viewer_listbox.configure(yscrollcommand=scrollbar.set)
        self.viewer_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.viewer_listbox.bind('<Double-1>', lambda e: self.open_selected_viewer_file())

        # Remove old standalone label and make org_frame a LabelFrame
        org_frame = ttk.LabelFrame(tab_main, text="File Organizer (Drag & Drop)")
        org_frame.grid(row=4, column=0, columnspan=2, pady=10, sticky=(tk.W, tk.E), padx=10)
        
        # We need a sub-frame to center the contents of the organizer inside the LabelFrame
        org_inner = ttk.Frame(org_frame)
        org_inner.pack(pady=5)
        
        ttk.Label(org_inner, text="Type:").grid(row=0, column=0, padx=2)
        self.type_combo = ttk.Combobox(org_inner, values=["Deed Records", "Lease Records", "Official Records", "Mortgage Records", "Power of Attorney Records", "Will Records", "Probate Records", "Miscellaneous Records"], width=18)
        self.type_combo.set("Deed Records")
        self.type_combo.grid(row=0, column=1, padx=2)
        
        ttk.Label(org_inner, text="Vol:").grid(row=0, column=2, padx=2)
        self.org_vol_entry = ttk.Entry(org_inner, width=5)
        self.org_vol_entry.grid(row=0, column=3, padx=2)
        
        ttk.Label(org_inner, text="Pg:").grid(row=0, column=4, padx=2)
        self.org_pg_entry = ttk.Entry(org_inner, width=5)
        self.org_pg_entry.grid(row=0, column=5, padx=2)

        ttk.Label(org_inner, text="Suffix:").grid(row=0, column=6, padx=2)
        self.suffix_combo = ttk.Combobox(org_inner, values=["", "OUTSALE", "Deed Plot"], width=9)
        self.suffix_combo.grid(row=0, column=7, padx=2)
        
        self.drop_var = tk.StringVar()
        self.drop_var.set("Drag & Drop PDF Here")
        self.drop_label = tk.Label(tab_main, textvariable=self.drop_var, bg="lightgrey", width=50, height=3, relief="sunken")
        self.drop_label.grid(row=5, column=0, columnspan=2, pady=5)
        self.drop_label.drop_target_register(DND_FILES)
        self.drop_label.dnd_bind('<<Drop>>', self.on_drop)
        
        btn_frame = ttk.Frame(tab_main)
        btn_frame.grid(row=6, column=0, columnspan=2, pady=5)
        self.org_btn = ttk.Button(btn_frame, text="Save & Rename", command=self.move_and_rename_file)
        self.org_btn.pack(side=tk.LEFT, padx=5)
        
        self.quick_log_btn = ttk.Button(btn_frame, text="Log Mortgage Release", command=self.open_quick_log_mortgage)
        self.quick_log_btn.pack(side=tk.LEFT, padx=5)
        
        self.save_other_btn = ttk.Button(btn_frame, text="Save Other", command=self.open_save_other_popup)
        self.save_other_btn.pack(side=tk.LEFT, padx=5)
        
        self.cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.cancel_organizer)
        self.cancel_btn.pack(side=tk.LEFT, padx=5)

        # --- TAB 2 ---
        ttk.Label(tab_tools, text="Historical Map Compiler (Drag & Drop)").grid(row=0, column=0, columnspan=2, pady=5)
        map_frame = ttk.Frame(tab_tools)
        map_frame.grid(row=1, column=0, columnspan=2, pady=5)
        
        ttk.Label(map_frame, text="Twp Name:").grid(row=0, column=0, padx=2)
        self.map_twp_entry = ttk.Entry(map_frame, width=12)
        self.map_twp_entry.grid(row=0, column=1, padx=2)
        
        ttk.Label(map_frame, text="S#:").grid(row=0, column=2, padx=2)
        self.map_s_entry = ttk.Entry(map_frame, width=5)
        self.map_s_entry.grid(row=0, column=3, padx=2)
        
        ttk.Label(map_frame, text="T#:").grid(row=0, column=4, padx=2)
        self.map_t_entry = ttk.Entry(map_frame, width=5)
        self.map_t_entry.grid(row=0, column=5, padx=2)
        
        ttk.Label(map_frame, text="R#:").grid(row=0, column=6, padx=2)
        self.map_r_entry = ttk.Entry(map_frame, width=5)
        self.map_r_entry.grid(row=0, column=7, padx=2)
        
        self.map_twp_entry.bind("<FocusOut>", lambda e: self.save_parcel_info())
        self.map_s_entry.bind("<FocusOut>", lambda e: self.save_parcel_info())
        self.map_t_entry.bind("<FocusOut>", lambda e: self.save_parcel_info())
        self.map_r_entry.bind("<FocusOut>", lambda e: self.save_parcel_info())

        self.map_drop_var = tk.StringVar()
        self.map_drop_var.set("Drag & Drop Map + Names Here")
        self.map_drop_label = tk.Label(tab_tools, textvariable=self.map_drop_var, bg="lightblue", width=50, height=2, relief="sunken")
        self.map_drop_label.grid(row=2, column=0, columnspan=2, pady=5)
        self.map_drop_label.drop_target_register(DND_FILES)
        self.map_drop_label.dnd_bind('<<Drop>>', self.on_map_drop)
        
        ttk.Button(tab_tools, text="Auto-Download All Years", command=self.start_auto_download).grid(row=3, column=0, columnspan=2, pady=2)
        
        self.map_dl_status = tk.Text(tab_tools, height=4, width=50, font=("Helvetica", 9))
        self.map_dl_status.grid(row=4, column=0, columnspan=2, pady=2)
        self.map_dl_status.config(state=tk.DISABLED)

        ttk.Separator(tab_tools, orient='horizontal').grid(row=5, column=0, columnspan=2, sticky='ew', pady=10)
        self.court_btn = ttk.Button(tab_tools, text="Court Records Checker (Select CSV)", command=self.start_court_check)
        self.court_btn.grid(row=6, column=0, columnspan=2, pady=5)

        ttk.Separator(tab_tools, orient='horizontal').grid(row=7, column=0, columnspan=2, sticky='ew', pady=10)
        self.og_btn = ttk.Button(tab_tools, text="Oil & Gas Checker", command=self.start_og_check)
        self.og_btn.grid(row=8, column=0, columnspan=2, pady=5)

        # --- SHARED FOOTER ---
        log_frame = ttk.LabelFrame(frame, text="Activity Log")
        log_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), padx=18, pady=10)
        
        self.log_text = tk.Text(log_frame, height=4, state=tk.DISABLED, width=40)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.root.geometry("700x850")
        
        self.parcel_entry.bind('<<ComboboxSelected>>', lambda e: self.on_parcel_change())
        self.parcel_entry.bind('<KeyRelease>', lambda e: self.on_parcel_change())
        
        # Shortcuts (Command+O for Ownership Report, Command+G for Belmont GIS)
        self.root.bind("<Command-o>", self.open_ownership_report)
        self.root.bind("<Control-o>", self.open_ownership_report)
        self.root.bind("<Command-O>", self.open_ownership_report)
        self.root.bind("<Control-O>", self.open_ownership_report)

        self.root.bind("<Command-g>", self.open_belmont_gis)
        self.root.bind("<Control-g>", self.open_belmont_gis)
        self.root.bind("<Command-G>", self.open_belmont_gis)
        self.root.bind("<Control-G>", self.open_belmont_gis)

        self.root.bind("<Command-Shift-E>", self.open_submission_email)
        self.root.bind("<Command-Shift-e>", self.open_submission_email)
        self.root.bind("<Control-Shift-E>", self.open_submission_email)
        self.root.bind("<Control-Shift-e>", self.open_submission_email)

        self.root.bind("<Command-r>", self.open_belmont_recorder)
        self.root.bind("<Control-r>", self.open_belmont_recorder)
        self.root.bind("<Command-R>", self.open_belmont_recorder)
        self.root.bind("<Control-R>", self.open_belmont_recorder)

        self.root.bind("<Command-e>", self.open_rs_editor)
        self.root.bind("<Control-e>", self.open_rs_editor)
        self.root.bind("<Command-E>", self.open_rs_editor)
        self.root.bind("<Control-E>", self.open_rs_editor)

    def refresh_recent_parcels(self):
        import glob
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        
        def get_dir_mtime(d):
            latest = os.path.getmtime(d)
            try:
                for entry in os.scandir(d):
                    if not entry.name.startswith('.'):
                        try:
                            m = entry.stat().st_mtime
                            if m > latest:
                                latest = m
                        except Exception: pass
            except Exception: pass
            return latest

        pid_folders = []
        for d in glob.glob(os.path.join(base_dir, "PID *")):
            if os.path.isdir(d):
                folder_name = os.path.basename(d)
                raw_name = re.sub(r'^PID[\s_]*', '', folder_name).strip()
                if "TEMPLATE" in raw_name.upper():
                    continue

                u_name = ""
                unit_match = re.search(r'\(([^)]+)\)', raw_name)
                if unit_match:
                    u_name = unit_match.group(1).strip()
                    p_num = raw_name[:unit_match.start()].strip()
                else:
                    p_num = raw_name

                mtime = get_dir_mtime(d)
                pid_folders.append((mtime, p_num, u_name))

        pid_folders.sort(key=lambda x: x[0], reverse=True)
        recent_parcels = []
        recent_units = []
        for _, p_num, u_name in pid_folders:
            if p_num not in recent_parcels:
                recent_parcels.append(p_num)
            if u_name and u_name not in recent_units:
                recent_units.append(u_name)
                
        if hasattr(self, 'parcel_entry'):
            self.parcel_entry['values'] = recent_parcels
        if hasattr(self, 'unit_entry'):
            self.unit_entry['values'] = recent_units

    def get_parcel_dir(self, parcel_num):
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        unit_name = self.unit_entry.get().strip() if hasattr(self, 'unit_entry') else ""
        
        # 1. Exact match for PID <parcel_num> (e.g. PID 42-00124.000 TEST or PID 42-00124.000)
        exact_path = os.path.join(base_dir, f"PID {parcel_num}")
        if os.path.isdir(exact_path):
            return exact_path

        # 2. Match with unit name if present
        if unit_name:
            unit_path = os.path.join(base_dir, f"PID {parcel_num} ({unit_name})")
            if os.path.isdir(unit_path):
                return unit_path
            
        # 3. Fallback prefix glob
        import glob
        for d in glob.glob(os.path.join(base_dir, f"PID {parcel_num}*")):
            if os.path.isdir(d):
                return d
                
        return os.path.join(base_dir, f"PID {parcel_num}")

    def log(self, message):
        if hasattr(self, 'log_text'):
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, f"{message}\n")
            self.log_text.see(tk.END)
            self.log_text.config(state=tk.DISABLED)
            
        if hasattr(self, 'large_log_text'):
            self.large_log_text.config(state=tk.NORMAL)
            self.large_log_text.insert(tk.END, f"{message}\n")
            self.large_log_text.see(tk.END)
            self.large_log_text.config(state=tk.DISABLED)
            
        if hasattr(self, 'root'):
            self.root.update()
        print(message)

    def on_parcel_change(self):
        self.update_viewer_folders()
        self.load_parcel_info()
        self.load_notes_for_parcel()

    def save_parcel_info(self):
        parcel_num = self.parcel_entry.get().strip()
        if not parcel_num:
            return
        
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        pid_dir = self.get_parcel_dir(parcel_num)
        os.makedirs(pid_dir, exist_ok=True)
        
        info = {
            "twp": self.map_twp_entry.get().strip(),
            "s": self.map_s_entry.get().strip(),
            "t": self.map_t_entry.get().strip(),
            "r": self.map_r_entry.get().strip()
        }
        
        try:
            with open(os.path.join(pid_dir, "parcel_info.json"), "w") as f:
                json.dump(info, f)
        except Exception as e:
            self.log(f"Warning: Could not save parcel info: {e}")

    def load_parcel_info(self):
        self.map_twp_entry.delete(0, tk.END)
        self.map_s_entry.delete(0, tk.END)
        self.map_t_entry.delete(0, tk.END)
        self.map_r_entry.delete(0, tk.END)
        
        parcel_num = self.parcel_entry.get().strip()
        if not parcel_num:
            return
            
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        info_path = os.path.join(base_dir, f"PID {parcel_num}", "parcel_info.json")
        
        if os.path.exists(info_path):
            try:
                with open(info_path, "r") as f:
                    info = json.load(f)
                    
                self.map_twp_entry.insert(0, info.get("twp", ""))
                self.map_s_entry.insert(0, info.get("s", ""))
                self.map_t_entry.insert(0, info.get("t", ""))
                self.map_r_entry.insert(0, info.get("r", ""))
            except Exception as e:
                self.log(f"Warning: Could not load parcel info: {e}")

    def on_map_drop(self, event):
        try:
            # Handle macOS tkinterdnd2 quirk
            if isinstance(event.data, (list, tuple)):
                filepaths = event.data
            else:
                filepaths = self.root.tk.splitlist(event.data)
                
            if not filepaths: return
            
            parcel_num = self.parcel_entry.get().strip()
            if not parcel_num:
                messagebox.showwarning("Warning", "Please enter a Parcel Number at the top.")
                return
                
            twp = self.map_twp_entry.get().strip()
            s = self.map_s_entry.get().strip()
            t = self.map_t_entry.get().strip()
            r = self.map_r_entry.get().strip()
            
            if not twp or not s or not t or not r:
                messagebox.showwarning("Warning", "Please fill out Twp Name, S#, T#, and R# fields before merging maps.")
                return
                
            if not filepaths: return
            
            # Identify the map file and multiple names files
            map_fp = None
            names_fps = []
            year = ""
            
            for fp in filepaths:
                filename = os.path.basename(fp).lower()
                
                # Extract year if we haven't found one yet
                match = re.search(r'((?:17|18|19|20)\d{2})', filename)
                if match and not year:
                    year = match.group(1)
                    
                if 'name' in filename:
                    names_fps.append(fp)
                else:
                    map_fp = fp
                    
            if not map_fp:
                # Fallback: if no map file was identified, just use the first file
                map_fp = filepaths[0]
                if map_fp in names_fps:
                    names_fps.remove(map_fp)
                    
            if not year:
                messagebox.showwarning("Warning", "Could not find a year (e.g., 1868) in any of the filenames.")
                return
                
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            parcel_dir = self.get_parcel_dir(parcel_num)
            maps_dir = os.path.join(parcel_dir, "MAPS")
            os.makedirs(maps_dir, exist_ok=True)
            
            # Match prompt: S# T# R#
            new_filename = f"Historical Map {twp} Twp - S{s} T{t} R{r} - {year}.pdf"
            dest_path = os.path.join(maps_dir, new_filename)
            
            # Sort the names files numerically (e.g., 1868 names 1.pdf, 1868 names 2.pdf)
            def extract_number(fp):
                fname = os.path.basename(fp).lower()
                if year:
                    fname = fname.replace(year, '')
                m = re.search(r'(\d+)', fname)
                return int(m.group(1)) if m else 999999
                
            names_fps.sort(key=extract_number)
            
            import fitz
            merged_doc = fitz.open()
            
            pdf_doc = fitz.open(map_fp)
            merged_doc.insert_pdf(pdf_doc)
            pdf_doc.close()
            
            for fp in names_fps:
                pdf_doc = fitz.open(fp)
                merged_doc.insert_pdf(pdf_doc)
                pdf_doc.close()
            
            # Append any extra files just in case (that weren't identified as map or name)
            for fp in filepaths:
                if fp != map_fp and fp not in names_fps:
                    pdf_doc = fitz.open(fp)
                    merged_doc.insert_pdf(pdf_doc)
                    pdf_doc.close()
                    
            merged_doc.save(dest_path)
            merged_doc.close()
            
            self.log(f"Merged Maps -> MAPS/{new_filename}")
            
            for fp in filepaths:
                try:
                    self.safe_move_to_irrelevant(fp)
                except Exception as e:
                    self.log(f"Could not trash {fp}: {e}")
            
            # Auto update view
            self.update_viewer_folders()
            self.viewer_folder_combo.set("MAPS")
            self.refresh_viewer_list()
            
        except Exception as e:
            self.log(f"Error during map merge: {e}")
            traceback.print_exc()

    def update_viewer_folders(self, *args):
        parcel_num = self.parcel_entry.get().strip()
        if not parcel_num:
            self.viewer_folder_combo.set('')
            self.viewer_folder_combo['values'] = []
            self.viewer_listbox.delete(0, tk.END)
            return

        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.viewer_pid_dir = self.get_parcel_dir(parcel_num)
            
        folders = ["Root (PID Folder)"]
        docs_exists = False
        
        if os.path.exists(self.viewer_pid_dir):
            for item in os.listdir(self.viewer_pid_dir):
                item_path = os.path.join(self.viewer_pid_dir, item)
                if os.path.isdir(item_path):
                    folders.append(item)
                    if item.upper() == "DOCS":
                        docs_exists = True
                        for root_dir, dirs, files in os.walk(item_path):
                            rel = os.path.relpath(root_dir, self.viewer_pid_dir)
                            if rel != "DOCS" and not any(part.startswith(".") for part in rel.split(os.sep)):
                                if rel not in folders:
                                    folders.append(rel)
                    if item.upper() == "WELL INFO":
                        for sub_item in os.listdir(item_path):
                            if os.path.isdir(os.path.join(item_path, sub_item)):
                                folders.append(f"WELL INFO/{sub_item}")
                        
        self.viewer_folder_combo['values'] = folders
        
        if docs_exists:
            self.viewer_folder_combo.set("DOCS")
        else:
            self.viewer_folder_combo.set("Root (PID Folder)")
            
        self.refresh_viewer_list()
        
    def refresh_viewer_list(self):
        self.viewer_listbox.delete(0, tk.END)
        self.current_viewer_files = []
        
        if not hasattr(self, 'viewer_pid_dir') or not self.viewer_pid_dir:
            return

        folder = self.viewer_folder_combo.get()
        if not folder:
            return

        search_query = getattr(self, 'viewer_search_var', tk.StringVar()).get().lower()

        def get_files_for_folder(fld):
            target_dir = self.viewer_pid_dir if fld == "Root (PID Folder)" else os.path.join(self.viewer_pid_dir, fld)
            files_with_time = []
            if os.path.exists(target_dir):
                for f in os.listdir(target_dir):
                    full_path = os.path.join(target_dir, f)
                    valid_ext = ('.pdf', '.jpg', '.jpeg', '.png', '.tif', '.tiff', '.xlsx', '.xls', '.csv')
                    if os.path.isfile(full_path) and f.lower().endswith(valid_ext):
                        if f.startswith('._'):
                            try:
                                os.remove(full_path)
                            except:
                                pass
                            continue
                        if search_query in f.lower():
                            files_with_time.append((f, full_path, os.path.getmtime(full_path)))
            files_with_time.sort(key=lambda x: x[2], reverse=True)
            return files_with_time

        matches = get_files_for_folder(folder)
        
        # If no matches in current folder but we are searching, check other folders
        if not matches and search_query:
            all_folders = list(self.viewer_folder_combo['values'])
            for other_folder in all_folders:
                if other_folder != folder:
                    other_matches = get_files_for_folder(other_folder)
                    if other_matches:
                        folder = other_folder
                        self.viewer_folder_combo.set(folder)
                        matches = other_matches
                        break
            
        for i, (f, full_path, _) in enumerate(matches):
            self.viewer_listbox.insert(tk.END, f)
            self.current_viewer_files.append(full_path)
            
            # Highlight the most recently added file at the top (only if not searching)
            if i == 0 and not search_query:
                self.viewer_listbox.itemconfig(i, {'fg': 'darkblue', 'bg': '#e6f2ff'})
                
    def open_selected_viewer_file(self):
        selection = self.viewer_listbox.curselection()
        if not selection:
            return
        idx = selection[0]
        filepath = self.current_viewer_files[idx]
        
        try:
            if os.name == 'nt':
                os.startfile(filepath)
            else:
                import subprocess
                subprocess.call(('open', filepath))
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file: {e}")

    def clean_docket(self):
        import os
        import shutil
    def safe_move_to_irrelevant(self, filepath):
        try:
            if not os.path.exists(filepath): return
            curr_dir = os.path.abspath(filepath)
            pid_dir = None
            for _ in range(5):
                curr_dir = os.path.dirname(curr_dir)
                if os.path.basename(curr_dir).startswith("PID "):
                    pid_dir = curr_dir
                    break
            
            if pid_dir:
                irrel_dir = os.path.join(pid_dir, "IRRELEVANT")
                os.makedirs(irrel_dir, exist_ok=True)
                import shutil
                shutil.move(filepath, os.path.join(irrel_dir, os.path.basename(filepath)))
            else:
                import send2trash
                send2trash.send2trash(filepath)
        except Exception as e:
            self.log(f"Safe move failed for {filepath}: {e}")

    def clean_docket(self):
        import os
        import shutil
        from tkinter import messagebox
        parcel_num = self.parcel_entry.get().strip()
        if not parcel_num:
            messagebox.showwarning("Warning", "Please enter a Parcel Number.")
            return

        pid_dir = self.get_parcel_dir(parcel_num)
        docket_dir = os.path.join(pid_dir, "DOCS", "docket")
        
        if os.path.exists(docket_dir):
            try:
                for f in os.listdir(docket_dir):
                    fp = os.path.join(docket_dir, f)
                    if os.path.isfile(fp):
                        self.safe_move_to_irrelevant(fp)
                self.log(f"Cleaned docket for {parcel_num}.")
                self.refresh_viewer_list()
            except Exception as e:
                self.log(f"Error cleaning docket: {e}")

    def combine_to_organizer(self):
        selection = self.viewer_listbox.curselection()
        if not selection:
            from tkinter import messagebox
            messagebox.showwarning("Warning", "Please select files to combine.")
            return
            
        filepaths = [self.current_viewer_files[i] for i in selection]
        
        parcel_num = self.parcel_entry.get().strip()
        if not parcel_num:
            return
            
        import os
        from PIL import Image
        import re
        
        def extract_page_num(fp):
            base = os.path.splitext(os.path.basename(fp))[0]
            parts = base.split("-")
            pg_str = parts[-1] if parts else ""
            nums = re.findall(r'\d+', pg_str)
            return int(nums[0]) if nums else 999999
            
        filepaths.sort(key=extract_page_num)
        
        try:
            pid_dir = self.get_parcel_dir(parcel_num)
            docket_dir = os.path.join(pid_dir, "DOCS", "docket")
            os.makedirs(docket_dir, exist_ok=True)
            temp_pdf_path = os.path.join(docket_dir, "temp_combined.pdf")

            if len(filepaths) == 1 and filepaths[0].lower().endswith('.pdf'):
                import shutil
                shutil.copy2(filepaths[0], temp_pdf_path)
            else:
                import fitz
                merged_doc = fitz.open()
                for fp in filepaths:
                    if fp.lower().endswith('.pdf'):
                        pdf_doc = fitz.open(fp)
                        merged_doc.insert_pdf(pdf_doc)
                        pdf_doc.close()
                    else:
                        try:
                            img_doc = fitz.open(fp)
                            pdf_bytes = img_doc.convert_to_pdf()
                            pdf_doc = fitz.open("pdf", pdf_bytes)
                            merged_doc.insert_pdf(pdf_doc)
                            img_doc.close()
                            pdf_doc.close()
                        except Exception as e:
                            self.log(f"Failed to append {fp}: {e}")
                            
                if len(merged_doc) > 0:
                    merged_doc.save(temp_pdf_path)
                merged_doc.close()
            
            self.dropped_filepaths = [temp_pdf_path]
            self.drop_var.set(temp_pdf_path)
            
            first_filename = os.path.basename(filepaths[0])
            name_no_ext = os.path.splitext(first_filename)[0]
            parts = name_no_ext.split("-")
            
            if len(parts) >= 2:
                vol_part = parts[0].strip()
                if " " in vol_part:
                    prefix_part, real_vol = vol_part.split(" ", 1)
                    
                    type_map_reverse = {
                        "DEED": "Deed Records",
                        "DR": "Deed Records",
                        "LEASE": "Lease Records",
                        "LR": "Lease Records",
                        "OFFICIAL": "Official Records",
                        "OR": "Official Records",
                        "MORTGAGE": "Mortgage Records",
                        "MR": "Mortgage Records",
                        "POWER": "Power of Attorney Records",
                        "PA": "Power of Attorney Records",
                        "WILL": "Will Records",
                        "WR": "Will Records",
                        "PROBATE": "Probate Records",
                        "PR": "Probate Records",
                        "MISC": "Miscellaneous Records"
                    }
                    
                    if prefix_part.upper() in type_map_reverse:
                        self.type_combo.set(type_map_reverse[prefix_part.upper()])
                    
                    self.org_vol_entry.delete(0, 'end')
                    self.org_vol_entry.insert(0, real_vol)
                else:
                    self.org_vol_entry.delete(0, 'end')
                    self.org_vol_entry.insert(0, vol_part)
                    
                self.org_pg_entry.delete(0, 'end')
                self.org_pg_entry.insert(0, parts[1])
            elif len(parts) == 1:
                self.org_pg_entry.delete(0, 'end')
                self.org_pg_entry.insert(0, parts[0])
                
            self.pending_docket_files = filepaths
            self.drop_label.config(bg="lightpink")
            self.update_live_preview()
            self.log(f"Compiled {len(filepaths)} files. Please check File Organizer to Save & Rename.")
            
        except Exception as e:
            self.log(f"Error compiling PDF: {e}")

    def mark_irrelevant_file(self):
        selection = self.viewer_listbox.curselection()
        if not selection:
            return
            
        parcel_num = self.parcel_entry.get().strip()
        if not parcel_num:
            return
            
        import os
        import shutil
        pid_dir = self.get_parcel_dir(parcel_num)
        irrelevant_dir = os.path.join(pid_dir, "DOCS", "Irrelevant")
        os.makedirs(irrelevant_dir, exist_ok=True)
        
        self.last_discarded_files = []
        
        for idx in selection:
            filepath = self.current_viewer_files[idx]
            filename = os.path.basename(filepath)
            dest_path = os.path.join(irrelevant_dir, filename)
            try:
                shutil.move(filepath, dest_path)
                self.log(f"Moved {filename} to Irrelevant folder.")
                self.last_discarded_files.append((filepath, dest_path))
            except Exception as e:
                self.log(f"Failed to move file: {e}")
                
        self.refresh_viewer_list()
        
        try:
            self.close_preview()
        except:
            pass

    def undo_discard(self):
        import shutil
        import os
        if not hasattr(self, 'last_discarded_files') or not self.last_discarded_files:
            self.log("Nothing to undo.")
            return
            
        for source_path, dest_path in self.last_discarded_files:
            if os.path.exists(dest_path):
                try:
                    shutil.move(dest_path, source_path)
                    self.log(f"Restored {os.path.basename(source_path)} to original location.")
                except Exception as e:
                    self.log(f"Error restoring file: {e}")
                    
        self.last_discarded_files = []
        self.refresh_viewer_list()

    def on_drop(self, event):
        try:
            # Handle macOS tkinterdnd2 quirk where event.data might already be a tuple or list
            if isinstance(event.data, (list, tuple)):
                filepaths = event.data
            else:
                filepaths = self.root.tk.splitlist(event.data)
                
            if not filepaths: return
            self.dropped_filepaths = filepaths
            
            if len(filepaths) == 1:
                self.drop_var.set(filepaths[0])
            else:
                self.drop_var.set(f"{os.path.basename(filepaths[0])} and {len(filepaths)-1} more ready to merge")
            
            lowest_pg_val = float('inf')
            best_rtype = None
            best_vol = None
            best_pg = None
            
            for fp in filepaths:
                filename = os.path.basename(fp)
                # Match any combination of prefix, volume, and page, e.g. DR 49-101.pdf or DEED1-000049-0101.pdf
                match = re.search(r'([A-Za-z]+).*?0*([A-Za-z0-9]+)[-_\s]+0*([A-Za-z0-9]+)(?:\.[a-zA-Z0-9]+)?$', filename)
                if match:
                    rtype, vol, pg = match.groups()
                    try:
                        # extract numbers for comparison
                        pg_num = int(re.sub(r'\D', '', pg))
                    except:
                        pg_num = 999999
                        
                    if pg_num < lowest_pg_val:
                        lowest_pg_val = pg_num
                        best_rtype = rtype
                        best_vol = vol
                        best_pg = pg
                        
            if best_rtype:
                self.org_vol_entry.delete(0, tk.END)
                self.org_vol_entry.insert(0, best_vol)
                
                self.org_pg_entry.delete(0, tk.END)
                self.org_pg_entry.insert(0, best_pg)
                
                rtype_upper = best_rtype.upper()
                type_map_reverse = {
                    "DEED": "Deed Records",
                    "DR": "Deed Records",
                    "LEASE": "Lease Records",
                    "LR": "Lease Records",
                    "OFFICIAL": "Official Records",
                    "OR": "Official Records",
                    "MORTGAGE": "Mortgage Records",
                    "MR": "Mortgage Records",
                    "POWER": "Power of Attorney Records",
                    "PA": "Power of Attorney Records",
                    "WILL": "Will Records",
                    "WR": "Will Records",
                    "PROBATE": "Probate Records",
                    "PR": "Probate Records",
                    "MISC": "Miscellaneous Records"
                }
                if rtype_upper in type_map_reverse:
                    self.type_combo.set(type_map_reverse[rtype_upper])
        except Exception as e:
            self.log(f"Error during drag-and-drop: {e}")
            import traceback
            traceback.print_exc()

    def update_live_preview(self, *args):
        if not getattr(self, 'dropped_filepaths', None):
            return
            
        record_type = self.type_combo.get().strip()
        vol = self.org_vol_entry.get().strip()
        pg = self.org_pg_entry.get().strip()
        suffix = self.suffix_combo.get().strip()
        
        type_map = {
            "Deed Records": "DR",
            "Lease Records": "LR",
            "Official Records": "OR",
            "Mortgage Records": "MR"
        }
        prefix = type_map.get(record_type, record_type)
        
        new_filename = f"{prefix} {vol}-{pg}"
        if suffix:
            new_filename += f" {suffix}"
            
        import os
        ext = os.path.splitext(self.dropped_filepaths[0])[1] or ".pdf"
        new_filename += ext
        
        self.drop_var.set(f"Will Save As: {new_filename}")

    def move_and_rename_file(self):
        filepaths = getattr(self, 'dropped_filepaths', None)
        if not filepaths:
            messagebox.showwarning("Warning", "Please drag and drop a valid file first.")
            return
            
        parcel_num = self.parcel_entry.get().strip()
        if not parcel_num:
            messagebox.showwarning("Warning", "Please enter a Parcel Number at the top.")
            return
            
        record_type = self.type_combo.get().strip()
        vol = self.org_vol_entry.get().strip()
        pg = self.org_pg_entry.get().strip()
        suffix = self.suffix_combo.get().strip()
        
        if not vol or not pg:
            messagebox.showwarning("Warning", "Please enter both Volume and Page.")
            return

        type_map = {
            "Deed Records": "DR",
            "Lease Records": "LR",
            "Official Records": "OR",
            "Mortgage Records": "MR",
            "Power of Attorney Records": "PA",
            "Will Records": "WR",
            "Probate Records": "PR",
            "Miscellaneous Records": "MISC"
        }
        abbr = type_map.get(record_type, record_type)

        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        parcel_dir = self.get_parcel_dir(parcel_num)
        docs_dir = os.path.join(parcel_dir, "DOCS")
        
        os.makedirs(docs_dir, exist_ok=True)
        
        ext = os.path.splitext(filepaths[0])[1]
        if not ext:
            ext = ".pdf"
            
        suffix_str = f" {suffix}" if suffix else ""
        new_filename = f"{abbr} {vol}-{pg}{suffix_str}{ext}"
        dest_path = os.path.join(docs_dir, new_filename)
        
        try:
            # Clean up mac hidden files in DOCS
            for root, dirs, files in os.walk(docs_dir):
                for file in files:
                    if file.startswith('._'):
                        try:
                            os.remove(os.path.join(root, file))
                        except:
                            pass
                            
            def launch_auto_open(target):
                try:
                    import sys, subprocess, os
                    if sys.platform == "darwin":
                        subprocess.Popen(["open", target])
                    elif sys.platform == "win32":
                        os.startfile(target)
                    else:
                        subprocess.Popen(["xdg-open", target])
                except Exception as ex:
                    self.log(f"Auto-open warning: {ex}")

            if len(filepaths) == 1:
                shutil.move(filepaths[0], dest_path)
                
                # Strip PDF/A
                if dest_path.lower().endswith('.pdf'):
                    try:
                        import fitz
                        doc = fitz.open(dest_path)
                        
                        # Properly remove PDF/A compliance
                        catalog = doc.pdf_catalog()
                        doc.xref_set_key(catalog, "OutputIntents", "null")
                        doc.xref_set_key(catalog, "Metadata", "null")
                        
                        tmp_path = dest_path + ".tmp"
                        doc.save(tmp_path, incremental=False, deflate=True)
                        doc.close()
                        shutil.move(tmp_path, dest_path)
                    except Exception as e:
                        self.log(f"Warning: Failed to strip PDF/A compliance: {e}")
                
                self.log(f"Moved {os.path.basename(filepaths[0])} -> DOCS/{new_filename}")
                
                # Cleanup if it was temp_combined
                if os.path.basename(filepaths[0]) == "temp_combined.pdf":
                    if hasattr(self, 'pending_docket_files'):
                        for fp in self.pending_docket_files:
                            try:
                                if os.path.exists(fp):
                                    os.remove(fp)
                            except:
                                pass
                        self.pending_docket_files = []
                        self.drop_label.config(bg="lightgrey")
                        self.refresh_viewer_list()
                
                self.close_preview()
                self.root.after(150, lambda p=dest_path: launch_auto_open(p))
            else:
                import fitz
                merged_doc = fitz.open()
                
                def get_page_num(fp):
                    filename = os.path.basename(fp)
                    match = re.search(r'([A-Za-z]+).*?0*([A-Za-z0-9]+)[-_\s]+0*([A-Za-z0-9]+)(?:\.[a-zA-Z0-9]+)?$', filename)
                    if match:
                        _, _, pg_match = match.groups()
                        try:
                            return int(re.sub(r'\D', '', pg_match))
                        except:
                            pass
                    return filename
                    
                sorted_filepaths = sorted(filepaths, key=lambda x: str(get_page_num(x)).zfill(10))
                
                for fp in sorted_filepaths:
                    pdf_doc = fitz.open(fp)
                    merged_doc.insert_pdf(pdf_doc)
                    pdf_doc.close()
                
                merged_doc.save(dest_path)
                merged_doc.close()
                self.log(f"Merged {len(filepaths)} files -> DOCS/{new_filename}")
                
                for fp in filepaths:
                    try:
                        self.safe_move_to_irrelevant(fp)
                    except Exception as e:
                        self.log(f"Could not trash {fp}: {e}")

                self.close_preview()
                self.root.after(150, lambda p=dest_path: launch_auto_open(p))
                        
            self.drop_var.set("Drag & Drop PDF Here")
            self.dropped_filepaths = None
            self.org_vol_entry.delete(0, tk.END)
            self.org_pg_entry.delete(0, tk.END)
            self.suffix_combo.set("")
            
            # Refresh document viewer if it's viewing the same PID
            if getattr(self, 'viewer_pid_dir', None) == parcel_dir:
                # Attempt to switch to DOCS automatically to see the saved file
                try:
                    all_folders = list(self.viewer_folder_combo['values'])
                    if "DOCS" in all_folders:
                        self.viewer_folder_combo.set("DOCS")
                except:
                    pass
                self.refresh_viewer_list()
                    
        except Exception as e:
            self.log(f"Error processing file(s): {e}")
            traceback.print_exc()


    def open_save_other_popup(self):
        filepaths = getattr(self, 'dropped_filepaths', None)
        if not filepaths:
            messagebox.showwarning("Warning", "Please drag and drop a valid file first.")
            return
            
        if len(filepaths) > 1:
            messagebox.showwarning("Warning", "Save Other only works for single files.")
            return
            
        parcel_num = self.parcel_entry.get().strip()
        if not parcel_num:
            messagebox.showwarning("Warning", "Please enter a Parcel Number.")
            return
            
        parcel_dir = self.get_parcel_dir(parcel_num)
        default_docs_dir = os.path.join(parcel_dir, "DOCS")
        
        orig_filename = os.path.basename(filepaths[0])
        
        popup = tk.Toplevel(self.root)
        popup.title("Save Other")
        popup.geometry("500x200")
        
        # Center popup
        popup.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 500) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 200) // 2
        popup.geometry(f"+{x}+{y}")
        
        frame = ttk.Frame(popup, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="File Name:").grid(row=0, column=0, sticky=tk.W, pady=5)
        name_var = tk.StringVar(value=orig_filename)
        ttk.Entry(frame, textvariable=name_var, width=50).grid(row=0, column=1, columnspan=2, sticky=tk.EW, pady=5)
        
        ttk.Label(frame, text="Save In:").grid(row=1, column=0, sticky=tk.W, pady=5)
        dir_var = tk.StringVar(value=default_docs_dir)
        ttk.Entry(frame, textvariable=dir_var, width=40).grid(row=1, column=1, sticky=tk.EW, pady=5)
        
        def browse_dir():
            d = filedialog.askdirectory(initialdir=dir_var.get())
            if d:
                dir_var.set(d)
                
        ttk.Button(frame, text="Browse", command=browse_dir).grid(row=1, column=2, padx=5, pady=5)
        
        def execute_save():
            final_name = name_var.get().strip()
            if not final_name: final_name = orig_filename
            final_dir = dir_var.get().strip()
            
            if not os.path.exists(final_dir):
                os.makedirs(final_dir)
                
            dest_path = os.path.join(final_dir, final_name)
            try:
                shutil.move(filepaths[0], dest_path)
                self.log(f"Moved {orig_filename} -> {dest_path}")
                self.drop_var.set("Drag & Drop PDF Here")
                self.dropped_filepaths = None
                
                # Cleanup docket if it was temp_combined
                if orig_filename == "temp_combined.pdf" and hasattr(self, 'pending_docket_files'):
                    for fp in self.pending_docket_files:
                        try:
                            if os.path.exists(fp):
                                os.remove(fp)
                        except:
                            pass
                self.close_preview()
                self.refresh_viewer_list()
                popup.destroy()

                def auto_open_saved(p):
                    try:
                        import sys, subprocess, os
                        if sys.platform == "darwin":
                            subprocess.Popen(["open", p])
                        elif sys.platform == "win32":
                            os.startfile(p)
                        else:
                            subprocess.Popen(["xdg-open", p])
                    except Exception as ex:
                        self.log(f"Auto-open error: {ex}")
                self.root.after(150, lambda p=dest_path: auto_open_saved(p))
            except Exception as e:
                messagebox.showerror("Error", f"Could not save file: {e}")
                
        ttk.Button(frame, text="Save", command=execute_save).grid(row=2, column=1, pady=20)
        ttk.Button(frame, text="Cancel", command=popup.destroy).grid(row=2, column=2, pady=20)

    def start_auto_download(self):
        parcel_num = self.parcel_entry.get().strip()
        if not parcel_num:
            messagebox.showwarning("Warning", "Please enter a Parcel Number at the top.")
            return
            
        twp = self.map_twp_entry.get().strip()
        s = self.map_s_entry.get().strip()
        t = self.map_t_entry.get().strip()
        r = self.map_r_entry.get().strip()
        
        if not all([twp, s, t, r]):
            messagebox.showwarning("Warning", "Please fill in Twp Name, S#, T#, and R#.")
            return

        self.map_dl_status.config(state=tk.NORMAL)
        self.map_dl_status.delete('1.0', tk.END)
        self.map_dl_status.insert(tk.END, "Starting auto-download...\n")
        self.map_dl_status.config(state=tk.DISABLED)
        
        threading.Thread(target=self._run_auto_download, args=(parcel_num, twp, s, t, r), daemon=True).start()

    def _update_dl_status(self, text, needs_attention=False):
        def update():
            self.map_dl_status.config(state=tk.NORMAL)
            self.map_dl_status.insert(tk.END, text + "\n")
            self.map_dl_status.see(tk.END)
            if needs_attention:
                self.attention_list.append(text)
            self.map_dl_status.config(state=tk.DISABLED)
        self.root.after(0, update)

    def _run_auto_download(self, parcel_num, twp, s, t, r):
        self.attention_list = []
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        parcel_dir = self.get_parcel_dir(parcel_num)
        maps_dir = os.path.join(parcel_dir, "MAPS")
        os.makedirs(maps_dir, exist_ok=True)
        
        years = [
            "2026", "2025", "2024", "2023", "2022", "2021", "2020", "2019", "2018", "2017",
            "2016", "2015", "2014", "2013", "2012", "2011", "2010", "2009", "2008", "2007",
            "2006", "2005", "2004", "2003", "1914", "1910", "1905", "1900", "1890", "1880", "1870", "1868"
        ]
        
        req_area = twp[:3].capitalize()
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7'
        }
        
        s_val = str(s).lstrip('0')
        t_val = str(t).lstrip('0')
        r_val = str(r).lstrip('0')
        
        # Match modern or historical
        pattern_str = rf"(Section\s+0?{s_val}\s+T-0?{t_val}\s+R-0?{r_val}|Sec\s+0?{s_val}-0?{t_val}-0?{r_val})"
        pattern = re.compile(pattern_str, re.IGNORECASE)
        
        for year in years:
            self._update_dl_status(f"Checking {year}...")
            url = "https://belcogis.com/php/taxmaplist.php"
            params = {
                'YearSelection': year,
                'TaxMapSelection': 'section',
                'reqArea': req_area
            }
            try:
                response = requests.get(url, headers=headers, params=params, timeout=10)
                soup = BeautifulSoup(response.text, 'html.parser')
                
                table = soup.find('table')
                found_links = []
                if table:
                    for row in table.find_all('tr'):
                        cols = row.find_all('td')
                        if not cols: continue
                        a = cols[0].find('a')
                        if a and pattern.search(a.text):
                            found_links.append((a.text.strip(), a.get('href')))
                
                if not found_links:
                    self._update_dl_status(f"-> {year}: Nothing found.", needs_attention=True)
                    continue
                    
                def sort_key(item):
                    text = item[0].lower()
                    if 'map' in text: return (0, text)
                    if 'name' in text: return (1, text)
                    return (2, text)
                
                found_links.sort(key=sort_key)
                
                map_count = sum(1 for item in found_links if 'map' in item[0].lower())
                if map_count > 1:
                    self._update_dl_status(f"-> {year}: Attention! Multiple maps found.", needs_attention=True)
                
                self._update_dl_status(f"-> {year}: Found {len(found_links)} files. Downloading...")
                
                temp_files = []
                for idx, (text, href) in enumerate(found_links):
                    pdf_url = "https://belcogis.com" + href
                    pdf_resp = requests.get(pdf_url, headers=headers, timeout=30)
                    if pdf_resp.status_code == 200:
                        pdf_data = None
                        if pdf_resp.content.startswith(b'%PDF'):
                            pdf_data = pdf_resp.content
                        else:
                            # Parse HTML for base64 object
                            page_soup = BeautifulSoup(pdf_resp.text, 'html.parser')
                            obj = page_soup.find('object')
                            if obj and obj.get('data') and obj['data'].startswith('data:application/pdf;base64,'):
                                b64_str = obj['data'].split('base64,')[1]
                                try:
                                    pdf_data = base64.b64decode(b64_str)
                                except Exception:
                                    pass
                                    
                        if pdf_data and pdf_data.startswith(b'%PDF'):
                            temp_fp = os.path.join(maps_dir, f"temp_{year}_{idx}.pdf")
                            with open(temp_fp, 'wb') as f:
                                f.write(pdf_data)
                            temp_files.append(temp_fp)
                        else:
                            self._update_dl_status(f"-> {year}: Failed to download '{text}' (No PDF data found)", needs_attention=True)
                    else:
                        self._update_dl_status(f"-> {year}: Failed to download '{text}' (Status {pdf_resp.status_code})", needs_attention=True)
                
                if temp_files:
                    new_filename = f"Historical Map {twp} Twp - S{s} T{t} R{r} - {year}.pdf"
                    dest_path = os.path.join(maps_dir, new_filename)
                    if len(temp_files) == 1:
                        import shutil
                        shutil.move(temp_files[0], dest_path)
                    else:
                        import fitz
                        merged_doc = fitz.open()
                        for fp in temp_files:
                            pdf_doc = fitz.open(fp)
                            merged_doc.insert_pdf(pdf_doc)
                            pdf_doc.close()
                        merged_doc.save(dest_path)
                        merged_doc.close()
                        for fp in temp_files:
                            try: os.remove(fp)
                            except: pass
                    self._update_dl_status(f"-> {year}: Saved successfully!")
                    
            except Exception as e:
                self._update_dl_status(f"-> {year}: Error - {str(e)}", needs_attention=True)
        
        if self.attention_list:
            self._update_dl_status("\n--- SUMMARY OF ATTENTION ITEMS ---")
            for item in self.attention_list:
                self._update_dl_status(item)
        else:
            self._update_dl_status("\n--- FINISHED SUCCESSFULLY ---")

    def open_checklist(self):
        parcel_num = self.parcel_entry.get().strip()
        if not parcel_num:
            messagebox.showwarning("Warning", "Please enter a Parcel Number to view the checklist.")
            return
            
        pid_dir = os.path.join(DRIVE_ROOT, "assignments", f"PID {parcel_num}")
        if not os.path.exists(pid_dir):
            messagebox.showwarning("Warning", f"Directory for PID {parcel_num} does not exist yet. Run automation first.")
            return
            
        checklist_file = os.path.join(pid_dir, "harbinger_checklist.json")
        
        import json
        from datetime import datetime
        
        state = {}
        if os.path.exists(checklist_file):
            try:
                with open(checklist_file, "r") as f:
                    state = json.load(f)
            except: pass
            
        items = [
            "Surface Chain: Breached 1870 with Warranty Deed or Plat Dedication.",
            "Mortgages: 30-Year Lookback (Included active/releases within 30 yrs).",
            "Liens & Civil Actions: 10-Year Lookback (Excluding NOCs).",
            "OGLs: 40-Year Lookback (Pulled all unreleased OGLs + assignments/amendments).",
            "Mineral Estate: Traced O&G forward (No coal chaining).",
            "Easements/ROWs: Searched back to source deed.",
            "File Naming: PID XX-XXXXX.XXX RS (MM_DD_YYYY) used.",
            "Base Chain Note: Correctly identifies base chain PID at top.",
            "Ordering: Strict chronological order by effective date/Vol-Pg.",
            "Acreage: Only input if explicitly stated in deed body text.",
            "Comments: BOLD ONLY Mineral, Oil & Gas reservations.",
            "DOCS Folder: All PDFs. Subject lands highlighted on assignment exhibits.",
            "MAPS Folder: Tax, GIS, Historical maps present and highlighted.",
            "TAX Folder: Tax Card, Property Card, Current Tax, Payment History included.",
            "WELL INFO Folder: ODNR Interactive Map & Data Link reports included.",
            "Dower Notes: Stated Dower released or No dower mentioned (if applicable).",
            "Prior References: Included only when written on instrument face.",
            "Parse Import Audit: Fixed commas/dates, verified OR vs DR types.",
            "OR Template: Changed RED text to BLACK. Names in ALL CAPS.",
            "OR Legal Description: Prepend Lot Number, acreage to 6 decimal places.",
            "Records Examined Date: Matches starting date on Runsheet.",
            "Leasehold Tab: Completed Schedule A. Bounded-Bys mapped via GIS.",
            "Final Check: Spell-check, page breaks, cell alignment verified.",
            "Submission: Email sent to designated recipients with all required info."
        ]
        
        top = tk.Toplevel(self.root)
        top.title(f"Pre-Flight Checklist - PID {parcel_num}")
        top.geometry("800x600")
        
        # Scrollable frame setup
        canvas = tk.Canvas(top)
        scrollbar = ttk.Scrollbar(top, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        header = ttk.Label(scrollable_frame, text="Harbinger Land Pre-Flight QC Checklist", font=("Helvetica", 16, "bold"))
        header.grid(row=0, column=0, columnspan=3, pady=10, sticky='w', padx=10)
        
        vars_dict = {}
        labels_dict = {}
        
        def on_check(idx):
            var = vars_dict[idx]
            if var.get():
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                state[str(idx)] = now_str
                labels_dict[idx].config(text=now_str)
                # If ttk label:
                # labels_dict[idx].config(foreground="green")
            else:
                if str(idx) in state:
                    del state[str(idx)]
                labels_dict[idx].config(text="")
                
            with open(checklist_file, "w") as f:
                json.dump(state, f)
        
        for i, item_text in enumerate(items):
            var = tk.BooleanVar(value=str(i) in state)
            vars_dict[i] = var
            
            cb = tk.Checkbutton(scrollable_frame, text=item_text, variable=var, 
                                command=lambda idx=i: on_check(idx),
                                justify=tk.LEFT, anchor='w', wraplength=600)
            cb.grid(row=i+1, column=0, sticky='w', padx=10, pady=5)
            
            date_text = state.get(str(i), "")
            date_lbl = tk.Label(scrollable_frame, text=date_text, fg="green")
            date_lbl.grid(row=i+1, column=1, sticky='w', padx=10)
            labels_dict[i] = date_lbl

    def complete_assignment(self):
        parcel_num = self.parcel_entry.get().strip()
        unit = self.suffix_combo.get().strip() if hasattr(self, 'suffix_combo') else ""
        
        from tkinter import messagebox, simpledialog
        if not parcel_num:
            messagebox.showwarning("Warning", "Please enter a Parcel Number to complete.")
            return
            
        if not unit:
            unit = simpledialog.askstring("Suffix Required", f"No suffix provided for PID {parcel_num}.\nPlease enter the suffix (e.g., Norma North II):")
            if not unit or not unit.strip():
                messagebox.showwarning("Warning", "A suffix is required to complete the assignment.")
                return
            unit = unit.strip()
            if hasattr(self, 'suffix_combo'):
                self.suffix_combo.set(unit)
                
        import shutil
        import datetime
        import glob
        import openpyxl
        
        parcel_dir = self.get_parcel_dir(parcel_num)
        if not os.path.exists(parcel_dir):
            messagebox.showwarning("Warning", "The working folder for this parcel doesn't exist.")
            return
            
        # Validation checks
        maps_dir = os.path.join(parcel_dir, "MAPS")
        well_dir = os.path.join(parcel_dir, "WELL INFO")
        docs_dir = os.path.join(parcel_dir, "DOCS")
        
        # Check well interactive map
        well_map_found = False
        if os.path.exists(well_dir):
            for f in os.listdir(well_dir):
                if f"PID {parcel_num} Well Interactive Map" in f:
                    well_map_found = True
                    break
        
        if not well_map_found:
            messagebox.showwarning("Missing Map", f"Cannot complete!\nMissing: 'PID {parcel_num} Well Interactive Map'\nin WELL INFO directory.")
            return
            
        # Check GIS 2026 Map
        gis_map_found = False
        if os.path.exists(maps_dir):
            for f in os.listdir(maps_dir):
                if f"PID {parcel_num} GIS 2026 Map" in f:
                    gis_map_found = True
                    break
                    
        if not gis_map_found:
            messagebox.showwarning("Missing Map", f"Cannot complete!\nMissing: 'PID {parcel_num} GIS 2026 Map'\nin MAPS directory.")
            return
            
        # Parse RS Spreadsheet for DOCS to copy
        import glob
        rs_matches = glob.glob(os.path.join(parcel_dir, f"PID {parcel_num} RS*.xlsx"))
        rs_spreadsheet = rs_matches[0] if rs_matches else None
        valid_doc_prefixes = []
        
        if os.path.exists(rs_spreadsheet):
            try:
                wb = openpyxl.load_workbook(rs_spreadsheet, data_only=True)
                ws = wb.active
                # Start at row 4 as per screenshot
                for row in range(4, ws.max_row + 1):
                    book_type = ws.cell(row=row, column=2).value # Col B
                    vol = ws.cell(row=row, column=3).value # Col C
                    page = ws.cell(row=row, column=4).value # Col D
                    
                    if book_type and vol and page:
                        b_type_str = str(book_type).strip()
                        if b_type_str.upper() == "DEED":
                            b_type_str = "DR"
                            
                        vol_str = str(vol).strip()
                        page_str = str(page).strip()
                        # Form standard prefix, e.g. "DR 571-913"
                        prefix = f"{b_type_str} {vol_str}-{page_str}"
                        valid_doc_prefixes.append(prefix)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to parse RS spreadsheet:\n{e}")
                return
        else:
            messagebox.showwarning("Warning", f"Could not find RS Spreadsheet:\n{rs_spreadsheet}\nMake sure it is named exactly like this in the root PID folder!")
            return
            
        # Execute copy
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        assignments_dir = os.path.join(base_dir, "assignments")
        if not os.path.exists(assignments_dir):
            assignments_dir = base_dir
            
        folder_name = f"PID {parcel_num} ({unit})"
            
        dest_dir = os.path.join(parcel_dir, folder_name)
        if dest_dir == parcel_dir:
            messagebox.showwarning("Warning", "The destination folder is exactly the same as the working folder. Cannot copy to itself.")
            return
            
        if not os.path.exists(dest_dir):
            os.makedirs(dest_dir)
            
        def copy_without_subdirs(src, dst, is_well_info=False):
            if not os.path.exists(src): return
            if not os.path.exists(dst): os.makedirs(dst)
            for item in os.listdir(src):
                s = os.path.join(src, item)
                d = os.path.join(dst, item)
                if os.path.isdir(s):
                    if not is_well_info: continue
                    elif "buffer" in item.lower(): continue
                    else:
                        try: shutil.copytree(s, d)
                        except: pass
                else:
                    try: shutil.copy2(s, d)
                    except: pass
                    
        # 1. Copy DOCS based on valid_doc_prefixes
        dest_docs_dir = os.path.join(dest_dir, "DOCS")
        if not os.path.exists(dest_docs_dir):
            os.makedirs(dest_docs_dir)
            
        docs_copied = 0
        prefixes_found = set()
        if os.path.exists(docs_dir):
            for item in os.listdir(docs_dir):
                if os.path.isdir(os.path.join(docs_dir, item)):
                    continue # Skip subdirectories
                    
                # Safe prefix checking
                for prefix in valid_doc_prefixes:
                    if item.startswith(prefix):
                        # Ensure we don't match DR 34-343 with DR 34-3430
                        next_char = item[len(prefix):len(prefix)+1]
                        if not next_char.isdigit():
                            try:
                                shutil.copy2(os.path.join(docs_dir, item), os.path.join(dest_docs_dir, item))
                                docs_copied += 1
                                prefixes_found.add(prefix)
                            except: pass
                            break
                            
        missing_docs = [p for p in valid_doc_prefixes if p not in prefixes_found]
        missing_msg = ""
        if missing_docs:
            missing_msg = "\n\nWARNING: The following logged documents were NOT found in DOCS:\n" + "\n".join(missing_docs)
        
        # 2. Copy MAPS (no subdirs)
        copy_without_subdirs(os.path.join(parcel_dir, "MAPS"), os.path.join(dest_dir, "MAPS"), is_well_info=False)
        
        # 3. Copy TAX (full copy)
        src_tax = os.path.join(parcel_dir, "TAX")
        dst_tax = os.path.join(dest_dir, "TAX")
        if os.path.exists(src_tax):
            try: shutil.copytree(src_tax, dst_tax, dirs_exist_ok=True)
            except: pass
            
        # 4. Copy WELL INFO (skip buffer subdirs)
        copy_without_subdirs(os.path.join(parcel_dir, "WELL INFO"), os.path.join(dest_dir, "WELL INFO"), is_well_info=True)
        
        # 5. Templates renaming
        now = datetime.datetime.now()
        today_str = now.strftime("(%m-%d-%Y)")
        
        # RS Template copy
        rs_matches = glob.glob(os.path.join(parcel_dir, f"PID {parcel_num} RS*.xlsx"))
        if rs_matches:
            rs_spreadsheet = rs_matches[0]
            new_rs = os.path.join(dest_dir, f"PID {parcel_num} RS {today_str}.xlsx")
            try: shutil.copy2(rs_spreadsheet, new_rs)
            except: pass
            
        # OR Template copy
        or_matches = glob.glob(os.path.join(parcel_dir, f"PID {parcel_num} OR*.xlsx"))
        if or_matches:
            or_spreadsheet = or_matches[0]
            new_or = os.path.join(dest_dir, f"PID {parcel_num} OR {today_str}.xlsx")
            try: shutil.copy2(or_spreadsheet, new_or)
            except: pass
            
        # Clean macOS junk files from the completed directory
        self._clean_directory(dest_dir)
            
        messagebox.showinfo("Success", f"Completed!\nCopied {docs_copied} logged documents.\nFiles successfully saved to:\n{dest_dir}{missing_msg}")

    def _clean_directory(self, target_dir):
        for root, dirs, files in os.walk(target_dir, topdown=False):
            # 1. Clean files starting with ._ or named .DS_Store
            for file in files:
                if file.startswith("._") or file.lower() == ".ds_store":
                    file_path = os.path.join(root, file)
                    try:
                        os.remove(file_path)
                    except Exception:
                        pass

            # 2. Clean hidden macOS folders
            for d in dirs:
                if d in [".Trashes", ".Spotlight-V100", ".fseventsd"]:
                    dir_path = os.path.join(root, d)
                    try:
                        for sub_root, sub_dirs, sub_files in os.walk(dir_path, topdown=False):
                            for f in sub_files:
                                os.remove(os.path.join(sub_root, f))
                            for sd in sub_dirs:
                                os.rmdir(os.path.join(sub_root, sd))
                        os.rmdir(dir_path)
                    except Exception:
                        pass

    def start_automation(self):
        parcel_num = self.parcel_entry.get().strip()
        if not parcel_num:
            messagebox.showwarning("Input Error", "Please enter a Parcel Number")
            return
            
        parcel_dir = self.get_parcel_dir(parcel_num)
        if os.path.exists(parcel_dir):
            if not messagebox.askyesno("Already Created", "WARNING this has already been created.\n\nDo you want to continue automation anyway? (Files may be overwritten)"):
                return
        
        self.set_buttons_state(tk.DISABLED)
        threading.Thread(target=self.run_process, args=(parcel_num,), daemon=True).start()

    def start_manual_download(self):
        parcel_num = self.parcel_entry.get().strip()
        vol = self.vol_entry.get().strip()
        pg = self.pg_entry.get().strip()
        doc_type = getattr(self, 'doc_type_combo', None)
        doc_type_val = doc_type.get() if doc_type else "Deed"
        
        if not vol or not pg:
            messagebox.showwarning("Input Error", "Please enter both Volume and Page.")
            return
            
        if not parcel_num:
            messagebox.showwarning("Input Error", "Please enter a Parcel Number to set the save folder.")
            return

        self.set_buttons_state(tk.DISABLED)
        import threading
        threading.Thread(target=self.run_manual_process, args=(parcel_num, vol, pg, doc_type_val), daemon=True).start()

    def start_manual_scrape(self):
        parcel_num = self.parcel_entry.get().strip()
        vol = self.vol_entry.get().strip()
        pg = self.pg_entry.get().strip()
        doc_type = getattr(self, 'doc_type_combo', None)
        doc_type_val = doc_type.get() if doc_type else "ALL"
        
        if not vol or not pg:
            messagebox.showwarning("Input Error", "Please enter both Volume and Page.")
            return
            
        if not parcel_num:
            messagebox.showwarning("Input Error", "Please enter a Parcel Number to set the save folder.")
            return

        self.set_buttons_state(tk.DISABLED)
        import threading
        threading.Thread(target=self.run_manual_scrape, args=(parcel_num, vol, pg, doc_type_val), daemon=True).start()

    def run_manual_scrape(self, parcel_num, vol, pg, doc_type="ALL"):
        try:
            self.log(f"Starting manual web scrape for Vol: {vol}, Pg: {pg}")
            self._fetch_kofile_deed_background(vol, pg, parcel_num, doc_type, auto_open=True)
            
            # Auto-update the viewer to show the docket folder
            def switch_to_docket():
                self.update_viewer_folders()
                self.viewer_folder_combo.set("DOCS/docket")
                self.refresh_viewer_list()
            self.root.after(500, switch_to_docket)
            
        except Exception as e:
            self.log(f"Manual web scrape failed: {e}")
        finally:
            self.set_buttons_state(tk.NORMAL)

    def start_og_check(self):
        parcel_num = self.parcel_entry.get().strip()
        if not parcel_num:
            messagebox.showwarning("Input Error", "Please enter a Parcel Number")
            return
            
        self.set_buttons_state(tk.DISABLED)
        threading.Thread(target=self.run_og_process, args=(parcel_num,), daemon=True).start()

    def run_og_process(self, parcel_num):
        try:
            import og_checker
            og_checker.check_parcel_og_activity(parcel_num, log_callback=self.log)
        except Exception as e:
            self.log(f"Error running O&G Checker: {e}")
        finally:
            self.set_buttons_state(tk.NORMAL)

    def start_court_check(self):
        csv_path = filedialog.askopenfilename(
            title="Select CSV for Court Records Checker",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
        )
        if not csv_path:
            return
            
        self.set_buttons_state(tk.DISABLED)
        
        self.log(f"Starting Court Records Checker with {csv_path}")
        threading.Thread(target=self.run_court_check, args=(csv_path,), daemon=True).start()
        
    def run_court_check(self, csv_path):
        try:
            import court_checker
            court_checker.process_court_records(csv_path, update_status_callback=self.log)
        except Exception as e:
            self.log(f"Error running Court Records Checker: {e}")
        finally:
            self.set_buttons_state(tk.NORMAL)

    def run_manual_process(self, parcel_num, vol, pg, doc_type="Deed"):
        try:
            self.log(f"Starting manual {doc_type} copy for Vol: {vol}, Pg: {pg}")
            self.copy_local_deed(parcel_num, vol, pg, doc_type=doc_type)
            self.log(f"Manual copy complete. Check the docket folder for PID {parcel_num}.")
            
            # Auto-update the viewer to show the docket folder
            def switch_to_docket():
                self.update_viewer_folders()
                self.viewer_folder_combo.set("DOCS/docket")
                self.refresh_viewer_list()
            self.root.after(500, switch_to_docket)
            
        except Exception as e:
            self.log(f"Manual download failed: {e}")
        finally:
            self.set_buttons_state(tk.NORMAL)

    def load_shapefile(self):
        if self.gdf is None:
            self.log("Loading shapefile (this takes a few seconds on first run)...")
            self.gdf = gpd.read_file(f"zip://{ZIP_PATH}")
            self.log("Shapefile loaded.")

    def fetch_base64_pdf(self, url, output_path):
        try:
            r = requests.get(url, headers=self.req_headers, timeout=20)
            match = re.search(r'data:application/pdf;base64,([^"\']+)', r.text)
            if match:
                pdf_data = base64.b64decode(match.group(1))
                with open(output_path, "wb") as f:
                    f.write(pdf_data)
                return True
            return False
        except Exception as e:
            self.log(f"Error fetching PDF from {url}: {e}")
            return False

    def copy_local_deed(self, parcel_num, vol, pg, is_next_page=False, doc_type="Deed", auto_open=True, custom_dest_dir=None):
        try:
            pid_dir = self.get_parcel_dir(parcel_num)
            docket_dir = custom_dest_dir if custom_dest_dir else os.path.join(pid_dir, "DOCS", "docket")
            os.makedirs(docket_dir, exist_ok=True)
            
            vol_str = str(vol).strip()
            if vol_str.isdigit():
                vol_pad = vol_str.zfill(3)
            else:
                vol_pad = vol_str.rjust(3, '0')
                
            pg_str = str(pg).strip()
            try:
                pg_int = int(float(pg_str))
                pg_pad = str(pg_int).zfill(4)
            except:
                pg_pad = pg_str.zfill(4)
                
            clean_vol = vol_str.lstrip("0") or "0"
            clean_pg = str(pg).strip().lstrip("0") or "0"
            search_str = f"{clean_vol}-{clean_pg}"
            
            # Check if already saved
            docs_dir = os.path.join(pid_dir, "DOCS")
            found_saved = []
            if os.path.exists(docs_dir):
                for root, dirs, files in os.walk(docs_dir):
                    for f in files:
                        if search_str in f and not f.startswith("._"):
                            found_saved.append(os.path.join(root, f))
            if found_saved:
                import subprocess
                
                import threading
                user_choice = tk.StringVar(value="cancel")
                done_event = threading.Event()
                
                def show_dialog():
                    dialog = tk.Toplevel(self.root)
                    dialog.title("Already Saved")
                    dialog.update_idletasks()
                    x = self.root.winfo_x() + (self.root.winfo_width() - 400) // 2
                    y = self.root.winfo_y() + (self.root.winfo_height() - 150) // 2
                    dialog.geometry(f"+{x}+{y}")
                    dialog.transient(self.root)
                    dialog.grab_set()
                    
                    ttk.Label(dialog, text=f"A document matching Volume {clean_vol} Page {clean_pg} is already saved!\n\nWhat would you like to do?", justify=tk.CENTER).pack(pady=20, padx=20)
                    
                    btn_frame = ttk.Frame(dialog)
                    btn_frame.pack(pady=10)
                    
                    def on_open():
                        user_choice.set("open")
                        dialog.destroy()
                        done_event.set()
                        
                    def on_override():
                        user_choice.set("override")
                        dialog.destroy()
                        done_event.set()
                        
                    def on_cancel():
                        user_choice.set("cancel")
                        dialog.destroy()
                        done_event.set()
                        
                    dialog.protocol("WM_DELETE_WINDOW", on_cancel)
                    
                    ttk.Button(btn_frame, text="Open Saved File", command=on_open).pack(side=tk.LEFT, padx=5)
                    ttk.Button(btn_frame, text="Download Anyway", command=on_override).pack(side=tk.LEFT, padx=5)
                    ttk.Button(btn_frame, text="Cancel", command=on_cancel).pack(side=tk.LEFT, padx=5)

                self.root.after(0, show_dialog)
                done_event.wait()
                choice = user_choice.get()
                
                if choice == "open":
                    try:
                        if os.name == 'nt':
                            os.startfile(found_saved[0])
                        else:
                            subprocess.call(('open', found_saved[0]))
                    except:
                        pass
                    return False
                elif choice == "override":
                    pass # Let execution fall through to download the file
                else:
                    return False
                
            vol_pad_4 = str(vol).zfill(4) if str(vol).isdigit() else str(vol)
            ext_deed_dir = f"/Volumes/davidlls/Belmont_Drive_External/Belmont County Court House/2. Belmont Deeds/DEED {vol_pad}"
            ext_lease_dir = f"/Volumes/davidlls/Belmont_Drive_External/Belmont County Court House/3. Belmont Leases/{vol_pad_4}"
            
            if doc_type == "Mortgage":
                primary_dir = os.path.join(DRIVE_ROOT, "drive", "MTGS", f"MTG {vol_pad}")
                backup_dir = os.path.join(DRIVE_ROOT, "drive", "extracted", "MTGS", f"MTG {vol_pad}")
            else:
                primary_dir = os.path.join(DRIVE_ROOT, "drive", "DEEDS", f"DEED {vol_pad}")
                backup_dir = os.path.join(DRIVE_ROOT, "drive", "extracted", "DEEDS", f"DEED {vol_pad}")
                
            source_file = None
            found = False
            
            for check_dir in [backup_dir, primary_dir, ext_deed_dir, ext_lease_dir]:
                test_file = os.path.join(check_dir, f"{pg_pad}.jpg")
                if os.path.exists(test_file):
                    source_file = test_file
                    found = True
                    break
                for ext in ['.pdf', '.tif', '.png', '.jpeg']:
                    if os.path.exists(os.path.join(check_dir, f"{pg_pad}{ext}")):
                        source_file = os.path.join(check_dir, f"{pg_pad}{ext}")
                        found = True
                        break
                if found:
                    break
                    
            if not found:
                missing_type = "Volume" if not os.path.exists(primary_dir) and not os.path.exists(backup_dir) else "Page"
                self.log(f"ℹ️ {doc_type} Vol {vol} Pg {pg} not found in local drive ({missing_type} missing). Logged to tracker.")
                self.log_missing_document(parcel_num, doc_type, vol, pg, missing_type)
                return False
                    
            import shutil
            ext = os.path.splitext(source_file)[1] or ".jpg"
            clean_vol = vol_str.lstrip("0") or "0"
            clean_pg = str(pg).strip().lstrip("0") or "0"
            filename = f"{clean_vol}-{clean_pg}{ext}"
            dest_file = os.path.join(docket_dir, filename)
            shutil.copy2(source_file, dest_file)
            
            # Auto-update the viewer to show the docket folder
            def switch_to_docket():
                if hasattr(self, 'update_viewer_folders'):
                    self.update_viewer_folders()
                    self.viewer_folder_combo.set("DOCS/docket")
                    self.refresh_viewer_list()
            self.root.after(500, switch_to_docket)
            
            # Clean up mac hidden files in docket
            for file in os.listdir(docket_dir):
                if file.startswith('._'):
                    try:
                        os.remove(os.path.join(docket_dir, file))
                    except:
                        pass
            self.log(f"Copied deed page to {dest_file}")
            
            self.last_deed_download = {
                'parcel_num': parcel_num,
                'vol': vol_str,
                'pg': pg_pad,
                'doc_type': doc_type
            }
            try:
                self.last_deed_download['pg_int'] = int(float(pg_str))
            except:
                self.last_deed_download['pg_int'] = None
                
            if hasattr(self, 'next_page_btn'):
                self.next_page_btn.config(state=tk.NORMAL)
            
            if auto_open:
                import subprocess, sys
                try:
                    if sys.platform == "darwin":
                        subprocess.Popen(["open", dest_file])
                    elif sys.platform == "win32":
                        os.startfile(dest_file)
                    else:
                        subprocess.Popen(["xdg-open", dest_file])
                except Exception as ex:
                    self.log(f"Error opening photo: {ex}")
            
            return True
            
        except Exception as e:
            self.log(f"Error copying local deed: {e}")
            return False
            
    def fetch_next_page(self):
        parcel = self.parcel_entry.get().strip()
        vol = ""
        pg_int = None
        doc_type = "Deed"

        # 1. Check entry boxes first
        entry_vol = self.vol_entry.get().strip()
        entry_pg = self.pg_entry.get().strip()
        if entry_vol and entry_pg:
            vol = entry_vol
            try:
                pg_int = int(float(entry_pg))
            except:
                pass
            if hasattr(self, 'doc_type_combo'):
                doc_type = self.doc_type_combo.get()

        # 2. Check last_deed_download metadata
        if (not vol or pg_int is None) and getattr(self, 'last_deed_download', None):
            parcel = parcel or self.last_deed_download.get('parcel_num')
            vol = self.last_deed_download.get('vol')
            pg_int = self.last_deed_download.get('pg_int')
            doc_type = self.last_deed_download.get('doc_type', doc_type)

        # 3. Check selected item in Document Viewer
        if (not vol or pg_int is None):
            try:
                selected = self.viewer_listbox.curselection()
                if selected:
                    fname = self.viewer_listbox.get(selected[0])
                    m = re.search(r'(\d+)[-_](\d+)', fname)
                    if m:
                        vol = m.group(1)
                        pg_int = int(m.group(2))
            except:
                pass

        if not parcel:
            self.log("Cannot fetch next page: Please enter a Parcel Number.")
            return

        if not vol or pg_int is None:
            self.log("Cannot determine current Volume and Page for Next Page. Please enter Volume and Page in the fields above.")
            return

        next_pg = pg_int + 1
        
        # Advance the page entry box in the UI
        self.vol_entry.delete(0, tk.END)
        self.vol_entry.insert(0, str(vol))
        self.pg_entry.delete(0, tk.END)
        self.pg_entry.insert(0, str(next_pg))

        self.last_deed_download = {
            'parcel_num': parcel,
            'vol': str(vol),
            'pg': str(next_pg),
            'pg_int': next_pg,
            'doc_type': doc_type
        }

        self.log(f"Fetching next page ({next_pg}) for Vol {vol} ({doc_type})...")
        
        def do_fetch():
            res = self.copy_local_deed(parcel, vol, next_pg, True, doc_type)
            # If not found locally, try scraping from website as fallback
            if not res:
                self.log(f"Next page {vol}-{next_pg} not found in local archives. Attempting website fetch...")
                self._fetch_kofile_deed_background(vol, next_pg, parcel, doc_type, auto_open=True)

        import threading
        threading.Thread(target=do_fetch, daemon=True).start()



    def set_api_key(self):
        from tkinter import simpledialog, messagebox
        import json
        import os
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
        current_key = ""
        if os.path.exists(config_path):
            with open(config_path, 'r') as f:
                data = json.load(f)
                current_key = data.get("GEMINI_API_KEY", "")
                
        new_key = simpledialog.askstring("Set API Key", "Enter your Google Gemini API Key:", initialvalue=current_key)
        if new_key is not None:
            new_key = new_key.strip()
            
            with open(config_path, 'w') as f:
                json.dump({"GEMINI_API_KEY": new_key}, f)
            messagebox.showinfo("Success", "API Key saved successfully!")

    def generate_ai_abstract_from_docket(self):
        parcel_num = self.parcel_entry.get().strip()
        if not parcel_num:
            from tkinter import messagebox
            messagebox.showwarning("Warning", "Please enter a Parcel Number first.")
            return

        selected = self.viewer_listbox.curselection()
        folder = self.viewer_folder_combo.get()
        
        if not folder:
            from tkinter import messagebox
            messagebox.showwarning("Warning", "Please select a folder.")
            return
            
        if folder == "Root (PID Folder)":
            current_dir = self.viewer_pid_dir
        else:
            current_dir = os.path.join(self.viewer_pid_dir, folder)
            
        filepaths = []
        if selected:
            for idx in selected:
                filename = self.viewer_listbox.get(idx)
                filepaths.append(os.path.join(current_dir, filename))
        else:
            from tkinter import messagebox
            if messagebox.askyesno("Confirm", f"No specific files selected. Summarize ALL PDFs in {folder}?"):
                for f in os.listdir(current_dir):
                    if f.lower().endswith('.pdf'):
                        filepaths.append(os.path.join(current_dir, f))
            else:
                return
                    
        if not filepaths:
            from tkinter import messagebox
            messagebox.showwarning("Warning", "No PDF files found to summarize.")
            return

        import json
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
        api_key = None
        if os.path.exists(config_path):
            with open(config_path, 'r') as cf:
                config = json.load(cf)
                api_key = config.get('GEMINI_API_KEY')
                
        if not api_key:
            from tkinter import messagebox
            messagebox.showwarning("Warning", "No GEMINI_API_KEY found in config.json. Go to File -> Set Gemini API Key.")
            return

        def _run_abstract():
            self.log(f"Uploading {len(filepaths)} files to Gemini for Abstracting Summary...")
            try:
                import ai_parser
                summary = ai_parser.generate_abstracting_summary_multiple(api_key, filepaths)
                if summary and not summary.startswith("Error"):
                    summary_path = os.path.join(self.viewer_pid_dir, f"PID {parcel_num} Chain of Title Summary.txt")
                    with open(summary_path, 'w', encoding='utf-8') as f:
                        f.write("GEMINI ABSTRACTING SUMMARY\n")
                        f.write("==========================\n\n")
                        f.write(summary)
                    self.log(f"Saved Gemini Chain of Title Abstracting Summary to workspace.")
                else:
                    self.log(f"Gemini Summary Error: {summary}")
            except Exception as e:
                self.log(f"Gemini Abstract Error: {e}")

        import threading
        threading.Thread(target=_run_abstract, daemon=True).start()

    def redact_document(self):
        selection = self.viewer_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a file to redact.")
            return
            
        filename = self.viewer_listbox.get(selection[0])
        current_folder = self.viewer_folder_combo.get()
        source_path = os.path.join(self.viewer_pid_dir, current_folder, filename)
        
        if not os.path.exists(source_path):
            return
            
        redacted_dir = os.path.join(self.viewer_pid_dir, "REDACTED")
        if not os.path.exists(redacted_dir):
            os.makedirs(redacted_dir)
            
        dest_path = os.path.join(redacted_dir, filename)
        
        try:
            import shutil
            import subprocess
            shutil.copy2(source_path, dest_path)
            self.log(f"Copied for redaction: {filename} into REDACTED folder.")
            
            if os.name == 'nt':
                os.startfile(dest_path)
            else:
                # Force open with Apple Preview on macOS
                subprocess.call(['open', '-a', 'Preview', dest_path])
        except Exception as e:
            messagebox.showerror("Error", f"Failed to setup redaction copy: {e}")

    def send_back_to_docket(self):
        selection = self.viewer_listbox.curselection()
        if not selection:
            from tkinter import messagebox
            messagebox.showwarning("No Selection", "Please select a file to send back to the docket.")
            return
            
        filename = self.viewer_listbox.get(selection[0])
        current_folder = self.viewer_folder_combo.get()
        source_path = os.path.join(self.viewer_pid_dir, current_folder, filename)
        
        if not os.path.exists(source_path):
            return
            
        docket_dir = os.path.join(self.viewer_pid_dir, "DOCS", "docket")
        os.makedirs(docket_dir, exist_ok=True)
        
        # Strip prefixes like "Deed Records "
        import re
        import shutil
        clean_name = re.sub(r'^[A-Za-z\s]+ Records ', '', filename)
        
        # Convert dashes to underscores for volume-page separation if present
        clean_name = clean_name.replace('-', '_')
        
        # Strip extension to append temp_combined
        name_no_ext, ext = os.path.splitext(clean_name)
        new_filename = f"{name_no_ext}_temp_combined{ext}"
        
        dest_path = os.path.join(docket_dir, new_filename)
        shutil.move(source_path, dest_path)
        
        self.log(f"Moved {filename} back to docket as {new_filename}")
        self.refresh_viewer_list()
        
    def auto_fetch_gemini_deed(self, parcel_num, vol, pg, date):
        try:
            pid_dir = self.get_parcel_dir(parcel_num)
            docket_dir = os.path.join(pid_dir, "DOCS", "docket")
            os.makedirs(docket_dir, exist_ok=True)
            
            vol_str = str(vol).strip()
            if vol_str.isdigit():
                vol_pad = vol_str.zfill(3)
            else:
                vol_pad = vol_str.rjust(3, '0')
                
            pg_str = str(pg).strip()
            try:
                pg_int = int(float(pg_str))
                pg_pad = str(pg_int).zfill(4)
            except:
                pg_pad = pg_str.zfill(4)
                
            clean_vol = vol_str.lstrip("0") or "0"
            clean_pg = str(pg).strip().lstrip("0") or "0"
            search_str = f"{clean_vol}-{clean_pg}"
            
            # Check if already saved silently for Gemini
            docs_dir = os.path.join(pid_dir, "DOCS")
            if os.path.exists(docs_dir):
                for root, dirs, files in os.walk(docs_dir):
                    for f in files:
                        if search_str in f and not f.startswith("._"):
                            self.log(f"Gemini auto-fetch: Skipped {clean_vol}-{clean_pg} (already saved)")
                            return f"- Vol {vol}, Pg {pg} (Already saved)" 
                
            primary_dir = os.path.join(DRIVE_ROOT, "drive", "DEEDS", f"DEED {vol_pad}")
            backup_dir = os.path.join(DRIVE_ROOT, "drive", "extracted", "DEEDS", f"DEED {vol_pad}")
            
            vol_pad_4 = str(vol).zfill(4) if str(vol).isdigit() else str(vol)
            ext_deed_dir = f"/Volumes/davidlls/Belmont_Drive_External/Belmont County Court House/2. Belmont Deeds/DEED {vol_pad}"
            ext_lease_dir = f"/Volumes/davidlls/Belmont_Drive_External/Belmont County Court House/3. Belmont Leases/{vol_pad_4}"
            
            source_file = None
            found = False
            
            for check_dir in [backup_dir, primary_dir, ext_deed_dir, ext_lease_dir]:
                test_file = os.path.join(check_dir, f"{pg_pad}.jpg")
                if os.path.exists(test_file):
                    source_file = test_file
                    found = True
                    break
                for ext in ['.pdf', '.tif', '.png', '.jpeg']:
                    if os.path.exists(os.path.join(check_dir, f"{pg_pad}{ext}")):
                        source_file = os.path.join(check_dir, f"{pg_pad}{ext}")
                        found = True
                        break
                if found:
                    break
                    
            if not found:
                self.log(f"Gemini auto-fetch: Deed not found locally {vol}-{pg}")
                missing_t = "Volume" if not os.path.exists(primary_dir) and not os.path.exists(backup_dir) else "Page"
                self.log_missing_document(parcel_num, "Deed", vol, pg, missing_t)
                
                if missing_t == "Volume":
                    return f"- Vol {vol}, Pg {pg} (Entire Volume missing)"
                else:
                    return f"- Vol {vol}, Pg {pg} (Page missing)" 
                    
            import shutil
            ext = os.path.splitext(source_file)[1] or ".jpg"
            clean_vol = vol_str.lstrip("0") or "0"
            clean_pg = str(pg).strip().lstrip("0") or "0"
            
            safe_date = date.replace('/', '-').strip()
            
            filename = f"{clean_vol}_{clean_pg}_{safe_date}_temp_combined{ext}"
            dest_file = os.path.join(docket_dir, filename)
            shutil.copy2(source_file, dest_file)
            self.log(f"Auto-fetched Gemini deed: {filename}")
            return True
        except Exception as e:
            self.log(f"Error auto-fetching gemini deed {vol}-{pg}: {e}")
            return False

    def run_process(self, parcel_num):
        try:
            self.log(f"Starting automation for Parcel: {parcel_num}")
            self.load_shapefile()

            row = self.gdf[self.gdf['parcel_no'] == parcel_num]
            if row.empty:
                self.log(f"ERROR: Parcel {parcel_num} not found in shapefile.")
                self.set_buttons_state(tk.NORMAL)
                return

            parcel_data = row.iloc[0]
            pin = parcel_data.get('pin', '')
            hyperlink = parcel_data.get('hyperlink', '')
            sec = str(parcel_data.get('sec', ''))
            twp = str(parcel_data.get('twp', ''))
            vol_val = str(parcel_data.get('vol', '')).strip()
            gis_ac = str(parcel_data.get('ac', 'Unknown'))
            acres_placeholder = f"<LOT_PLACEHOLDER>{gis_ac} Acres, more or less"
            pg_val = str(parcel_data.get('pg', '')).strip()
            
            self.log(f"Found PIN: {pin}, VOL: {vol_val}, PG: {pg_val}")

            pid_dir = self.get_parcel_dir(parcel_num)
            docs_dir = os.path.join(pid_dir, "DOCS")
            maps_dir = os.path.join(pid_dir, "MAPS")
            tax_dir = os.path.join(pid_dir, "TAX")
            well_dir = os.path.join(pid_dir, "WELL INFO")

            for d in [pid_dir, docs_dir, maps_dir, tax_dir, well_dir]:
                os.makedirs(d, exist_ok=True)
            self.log("Created directory structure.")

            today_str = datetime.datetime.today().strftime("(%m-%d-%Y)")
            target_template_1 = os.path.join(pid_dir, f"PID {parcel_num} OR {today_str}.xlsx")
            target_template_2 = os.path.join(pid_dir, f"PID {parcel_num} RS {today_str}.xlsx")

            if os.path.exists(TEMPLATE_1):
                shutil.copy(TEMPLATE_1, target_template_1)
            if os.path.exists(TEMPLATE_2):
                shutil.copy(TEMPLATE_2, target_template_2)
            self.log("Copied Excel templates.")
            
            # Calculate Quarter Section
            qtr_val = ""
            try:
                import re
                desc_str = str(parcel_data.get('desc_', '') or '')
                m_q = re.search(r'\b(NW|NE|SW|SE)\b', desc_str, re.IGNORECASE)
                if m_q:
                    qtr_val = f"{m_q.group(1).upper()}4"
                else:
                    p_geom = parcel_data.geometry if hasattr(parcel_data, 'geometry') else None
                    if p_geom is not None:
                        plss_shp = "/Volumes/davidlls/various_GIS_shapefiles/OH-CADNSDI-v2_SPSNAD83/PLSSFirstDivision.shp"
                        if os.path.exists(plss_shp):
                            import geopandas as gpd
                            from shapely.geometry import box
                            plss_gdf = gpd.read_file(plss_shp)
                            m_secs = plss_gdf[plss_gdf.intersects(p_geom)]
                            for _, sec_row in m_secs.iterrows():
                                minx, miny, maxx, maxy = sec_row.geometry.bounds
                                midx = (minx + maxx) / 2.0
                                midy = (miny + maxy) / 2.0
                                q_boxes = {
                                    "NW": box(minx, midy, midx, maxy),
                                    "NE": box(midx, midy, maxx, maxy),
                                    "SW": box(minx, miny, midx, midy),
                                    "SE": box(midx, miny, maxx, midy),
                                }
                                overlaps = {}
                                for q_code, q_box in q_boxes.items():
                                    if p_geom.intersects(q_box):
                                        overlaps[q_code] = p_geom.intersection(q_box).area
                                if overlaps:
                                    best_q = max(overlaps, key=overlaps.get)
                                    qtr_val = f"{best_q}4"
                                    break
                if qtr_val:
                    self.log(f"Calculated Quarter Section: {qtr_val}")
            except Exception as e:
                self.log(f"Note: Could not calculate quarter section: {e}")

            # Autofill Excel templates
            try:
                import openpyxl
                self.log("Autofilling Excel templates...")
                replacements = {
                    "<PARCEL>": str(parcel_num),
                    "<PIN>": str(pin),
                    "<SEC>": str(sec),
                    "<TWP>": str(twp),
                    "<QTR>": str(qtr_val),
                    "<QUARTER>": str(qtr_val),
                    "<QTR_CALL>": str(qtr_val),
                    "QUARTER CALL": str(qtr_val),
                    "<VOL>": str(vol_val),
                    "<PG>": str(pg_val),
                    "<ACRES_IN2>": str(gis_ac),
                    "<ACRES_IN>": f"<LOT_PLACEHOLDER>{gis_ac}",
                    "<ACRES>": str(acres_placeholder)
                }
                
                for excel_file in [target_template_1, target_template_2]:
                    if os.path.exists(excel_file):
                        wb = openpyxl.load_workbook(excel_file)
                        ws = wb.active
                        
                        # Replace in cells
                        from openpyxl.cell.rich_text import CellRichText
                        for row in ws.iter_rows():
                            for cell in row:
                                if cell.value:
                                    if isinstance(cell.value, str):
                                        for key, val in replacements.items():
                                            if key in cell.value:
                                                cell.value = cell.value.replace(key, val)
                        # Replace in headers and footers
                        sections = [ws.oddHeader, ws.evenHeader, ws.firstHeader, 
                                    ws.oddFooter, ws.evenFooter, ws.firstFooter]
                        for section in sections:
                            for part in [section.left, section.center, section.right]:
                                if part.text:
                                    for key, val in replacements.items():
                                        if key in part.text:
                                            part.text = part.text.replace(key, val)
                                            
                        wb.save(excel_file)
            except Exception as e:
                self.log(f"Warning: Failed to autofill excel templates: {e}")

            try:
                import openpyxl
                self.log("Updating Excel template data...")
                
                TWP_NAMES = {
                    'YOR': 'York', 'WAS': 'Washington', 'SOM': 'Somerset', 'MEA': 'Mead', 'WAY': 'Wayne',
                    'WAR': 'Warren', 'SMI': 'Smith', 'PUL': 'Pultney', 'GOS': 'Goshen', 'RIC': 'Richland',
                    'KIR': 'Kirkwood', 'UNI': 'Union', 'PEA': 'Pease', 'COL': 'Colerain', 'WHE': 'Wheeling',
                    'FLU': 'Flushing'
                }
                
                full_twp_name = TWP_NAMES.get(twp.upper()[:3] if twp else '', twp)
                t = str(parcel_data.get('t', '') or '').strip()
                r = str(parcel_data.get('r', '') or '').strip()
                sec_val = str(parcel_data.get('sec', '') or '').strip()
                acreage_val = str(parcel_data.get('ac', 0.0) or parcel_data.get('acreage', 0.0)).strip()
                
                t_pad = str(int(t)).zfill(2) if t.isdigit() else t
                r_pad = str(int(r)).zfill(2) if r.isdigit() else r

                # Calculate Bounding Owners (North, South, East, West)
                north_owner = "Unknown"
                south_owner = "Unknown"
                east_owner = "Unknown"
                west_owner = "Unknown"
                
                try:
                    if hasattr(self, 'gdf') and self.gdf is not None:
                        # Find the target geometry
                        target_matches = self.gdf[self.gdf['pin'].astype(str).str.contains(str(parcel_num), na=False)]
                        if not target_matches.empty:
                            target_geom = target_matches.iloc[0].geometry
                            target_centroid = target_geom.centroid
                            
                            # Find all neighbors that touch or intersect
                            neighbors = self.gdf[self.gdf.geometry.touches(target_geom) | self.gdf.geometry.intersects(target_geom)]
                            
                            def format_name(row):
                                fname = str(row.get('fname', '')).strip()
                                lname = str(row.get('lname', '')).strip()
                                if fname.lower() == 'nan': fname = ''
                                if lname.lower() == 'nan': lname = ''
                                
                                if fname and lname:
                                    name = f"{fname} {lname}"
                                elif lname:
                                    name = lname
                                else:
                                    coname = str(row.get('coname', '')).strip()
                                    name = coname if coname.lower() != 'nan' else ''
                                    
                                return ' '.join([w.capitalize() for w in name.split()])

                            import math
                            for _, neighbor in neighbors.iterrows():
                                n_pin = str(neighbor.get('pin', ''))
                                if str(parcel_num) in n_pin:
                                    continue # Skip self
                                    
                                n_geom = neighbor.geometry
                                n_centroid = n_geom.centroid
                                
                                # Angle between centroids
                                dx = n_centroid.x - target_centroid.x
                                dy = n_centroid.y - target_centroid.y
                                angle = math.degrees(math.atan2(dy, dx))
                                
                                formatted_name = format_name(neighbor)
                                if not formatted_name:
                                    continue
                                    
                                # Classify direction based on angle
                                # East: -45 to 45
                                # North: 45 to 135
                                # West: 135 to 180 and -180 to -135
                                # South: -135 to -45
                                if -45 <= angle <= 45:
                                    if east_owner == "Unknown": east_owner = formatted_name
                                elif 45 < angle <= 135:
                                    if north_owner == "Unknown": north_owner = formatted_name
                                elif angle > 135 or angle <= -135:
                                    if west_owner == "Unknown": west_owner = formatted_name
                                elif -135 < angle < -45:
                                    if south_owner == "Unknown": south_owner = formatted_name
                except Exception as e:
                    self.log(f"Warning: Failed to calculate adjacent neighbors: {e}")

                replacements = {
                    "XX-XXXXXXX.XXX": parcel_num,
                    "0.000000": acreage_val,
                    "QUARTER CALL": str(qtr_val or "SE4"),
                    "<QTR>": str(qtr_val or "SE4"),
                    "<QUARTER>": str(qtr_val or "SE4"),
                    "AGENT NAME": "DAVID MICHALOVE",
                    "<AGENT>": "DAVID MICHALOVE",
                    "Section XX": f"Section {sec_val}",
                    "SECTION 10": f"SECTION {sec_val}",
                    "Township XN": f"Township {t}N",
                    "Range XW": f"Range {r}W",
                    "Township of Mead": f"Township of {full_twp_name}",
                    "Noble County, OH": "Belmont County, OH",
                    "Noble County": "Belmont County",
                    "TOWNSHIP 2N": f"TOWNSHIP {t}N",
                    "RANGE 2W": f"RANGE {r}W",
                    "02 NORTH": f"{t_pad} NORTH",
                    "02 WEST": f"{r_pad} WEST",
                    "<NORTH>": north_owner,
                    "<SOUTH>": south_owner,
                    "<EAST>": east_owner,
                    "<WEST>": west_owner
                }

                wb = openpyxl.load_workbook(target_template_1)
                from openpyxl.cell.rich_text import CellRichText
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value:
                                if isinstance(cell.value, str):
                                    new_val = cell.value
                                    for old_str, new_str in replacements.items():
                                        new_val = new_val.replace(old_str, new_str)
                                    cell.value = new_val
                                elif type(cell.value).__name__ == 'CellRichText':
                                    new_crt = CellRichText()
                                    changed = False
                                    for part in cell.value:
                                        if isinstance(part, str):
                                            new_text = part
                                            for old_str, new_str in replacements.items():
                                                if old_str in new_text:
                                                    new_text = new_text.replace(old_str, new_str)
                                                    changed = True
                                            new_crt.append(new_text)
                                        else:
                                            new_text = part.text
                                            for old_str, new_str in replacements.items():
                                                if old_str in new_text:
                                                    new_text = new_text.replace(old_str, new_str)
                                                    changed = True
                                            part.text = new_text
                                            new_crt.append(part)
                                    if changed:
                                        cell.value = new_crt
                wb.save(target_template_1)
                self.log("Updated Excel templates successfully.")
            except Exception as e:
                self.log(f"Warning: Could not update Excel templates: {e}")

            # Download actual Transfer Card PDF
            self.log(f"Fetching Transfer Card...")
            if hyperlink:
                transfer_pdf_path = os.path.join(tax_dir, f"PID {parcel_num} Transfer Card.pdf")
                if self.fetch_base64_pdf(hyperlink, transfer_pdf_path):
                    self.log("Saved Transfer Card (actual PDF).")
                    
                    # Gemini Integration
                    self.log("Sending Transfer Card to Gemini AI for analysis...")
                    import sys
                    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
                    try:
                        import ai_parser
                        import json
                        
                        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
                        api_key = None
                        if os.path.exists(config_path):
                            with open(config_path, 'r') as cf:
                                config = json.load(cf)
                                api_key = config.get('GEMINI_API_KEY')
                                
                        if api_key:
                            # Launch the GIS map generator in the background exactly when Gemini starts thinking!
                            import subprocess, threading
                            def trigger_gis():
                                self.log(f"Spawning background process for Belmont County GIS Map...")
                                subprocess.run(["python3", "/Volumes/davidlls/assignments/app/gis_map_generator.py", parcel_num, maps_dir])
                                self.log(f"Belmont GIS complete. Spawning ODNR Map generator...")
                                subprocess.run(["python3", "/Volumes/davidlls/assignments/app/odnr_map_generator.py", parcel_num, well_dir])
                                self.log(f"All background map generation complete!")
                            threading.Thread(target=trigger_gis, daemon=True).start()
                            
                            ai_res, err = ai_parser.extract_deeds_from_transfer_sheet(api_key, transfer_pdf_path)
                            if err:
                                self.log(f"Gemini API Error: {err}")
                            elif isinstance(ai_res, dict):
                                deeds = ai_res.get('deeds', [])
                                lot = ai_res.get('lot', '')
                                if not lot or lot.upper() == 'NONE':
                                    lot = ''
                                else:
                                    lot = lot.title() + ', '
                                
                                # Replace the lot placeholder in the Excel files!
                                import openpyxl
                                for excel_file in [target_template_1, target_template_2]:
                                    if os.path.exists(excel_file):
                                        try:
                                            wb = openpyxl.load_workbook(excel_file)
                                            ws = wb.active
                                            for row_iter in ws.iter_rows():
                                                for cell in row_iter:
                                                    if cell.value and isinstance(cell.value, str):
                                                        if "<LOT_PLACEHOLDER>" in cell.value:
                                                            cell.value = cell.value.replace("<LOT_PLACEHOLDER>", str(lot))
                                            sections = [ws.oddHeader, ws.evenHeader, ws.firstHeader, ws.oddFooter, ws.evenFooter, ws.firstFooter]
                                            for section in sections:
                                                for part in [section.left, section.center, section.right]:
                                                    if part.text and "<LOT_PLACEHOLDER>" in part.text:
                                                        part.text = part.text.replace("<LOT_PLACEHOLDER>", str(lot))
                                            wb.save(excel_file)
                                        except Exception as e:
                                            self.log(f"Warning: Failed to inject Lot number into template: {e}")
                                            
                                self.log(f"Gemini extracted Lot: {lot}, and found {len(deeds)} historical deeds!")
                                missing_gemini_deeds = []
                                for d in deeds:
                                    vol = d.get('vol')
                                    pg = d.get('pg')
                                    date = d.get('date') or 'Unknown'
                                    if vol and pg:
                                        self.log(f"Gemini Auto-Fetch: Vol {vol}, Pg {pg}")
                                        res = self.auto_fetch_gemini_deed(parcel_num, vol, pg, str(date))
                                        if isinstance(res, str):
                                            missing_gemini_deeds.append(res)
                                        # Also run the web scraper for this historical deed
                                        self._fetch_kofile_deed_background(vol, pg, parcel_num)
                                
                                if missing_gemini_deeds:
                                    self.log(f"ℹ️ Gemini Auto-Fetch missing local deeds: {', '.join(missing_gemini_deeds)}")
                                        
                            else:
                                self.log("Gemini successfully analyzed the document but did not find any deeds.")
                        else:
                            self.log("No GEMINI_API_KEY found in config.json. Skipping AI parse.")
                    except Exception as ai_e:
                        self.log(f"Gemini Integration Failed: {ai_e}")
                        
                else:
                    self.log("Failed to extract actual Transfer Card PDF.")

            # Download actual Tax Map PDF
            self.log("Fetching Tax Map...")
            sec_str = str(sec).zfill(2) if sec else ""
            tax_map_url = None
            if twp.upper() == 'WAR':
                tax_map_url = f"https://belcogis.com/php/taxmapview.php?TableAndName=taxmaps2026:{sec_str}0806:War"
            elif twp.upper() == 'SOM':
                tax_map_url = f"https://belcogis.com/php/taxmapview.php?TableAndName=taxmaps2026:{sec_str}0706:Som"
            elif twp.upper() == 'BARNESVILLE' or twp.upper() == 'BAR':
                tax_map_url = f"https://belcogis.com/php/taxmapview.php?TableAndName=taxmaps2026:sh0{sec_str}:Barnesville"

            if tax_map_url:
                tax_map_pdf_path = os.path.join(maps_dir, f"PID {parcel_num} Tax Map.pdf")
                if self.fetch_base64_pdf(tax_map_url, tax_map_pdf_path):
                    self.log("Saved Tax Map (actual PDF).")
                else:
                    self.log("Failed to extract actual Tax Map PDF.")
            else:
                self.log(f"Tax Map URL generation not yet configured for township: {twp}")

            self.log("Setting up headless browser for Auditor Site...")
            options = Options()
            options.add_argument('--headless')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
            driver = webdriver.Chrome(options=options)

            try:
                self.log("Searching Auditor Site...")
                driver.get("https://belmontcountyauditor.org/Disclaimer")
                time.sleep(2)
                try:
                    agree_btn = driver.find_element(By.XPATH, "//button[contains(text(), 'I AGREE')]")
                    driver.execute_script("arguments[0].click();", agree_btn)
                    time.sleep(2)
                except Exception as e:
                    pass
                
                driver.get("https://belmontcountyauditor.org/RealEstate/Search")
                time.sleep(2)
                
                search_box = driver.find_element(By.ID, "quickSearch")
                search_box.send_keys(parcel_num)
                search_box.send_keys(Keys.RETURN)
                time.sleep(4)
                
                property_id = None
                
                # Check current URL first
                url = driver.current_url
                if "property_Id=" in url:
                    property_id = url.split("property_Id=")[1].split("&")[0]
                elif "property_id=" in url.lower():
                    # Fallback for case variations
                    lower_url = url.lower()
                    property_id = lower_url.split("property_id=")[1].split("&")[0]
                    
                # If not in URL, we might be on the Results page still
                if not property_id:
                    links = driver.find_elements(By.XPATH, "//a[contains(@href, 'property_Id=')]")
                    if links:
                        href = links[0].get_attribute("href")
                        property_id = href.split("property_Id=")[1].split("&")[0]

                if not property_id:
                    self.log(f"Could not find Property ID from search.")
                else:
                    self.log(f"Found Property ID: {property_id}")
                    
                    # Print Current Tax, Payment History, Property Card to PDF using driver
                    self.log("Printing Current Tax to PDF...")
                    driver.get(f"https://belmontcountyauditor.org/RealEstate/Tax?property_Id={property_id}&rowNumber=0")
                    time.sleep(2)
                    with open(os.path.join(tax_dir, f"PID {parcel_num} Current Tax.pdf"), "wb") as f:
                        f.write(base64.b64decode(driver.print_page()))
                    
                    self.log("Printing Payment History to PDF...")
                    driver.get(f"https://belmontcountyauditor.org/RealEstate/Payment?property_Id={property_id}&rowNumber=0")
                    time.sleep(2)
                    with open(os.path.join(tax_dir, f"PID {parcel_num} Payment History Card.pdf"), "wb") as f:
                        f.write(base64.b64decode(driver.print_page()))
                    
                    self.log("Printing Property Card to PDF...")
                    driver.get(f"https://belmontcountyauditor.org/RealEstate/Summary?property_Id={property_id}&rowNumber=0")
                    time.sleep(2)
                    with open(os.path.join(tax_dir, f"PID {parcel_num} Property Card.pdf"), "wb") as f:
                        f.write(base64.b64decode(driver.print_page()))
                        
                    self.log("Printing Interactive Map to PDF...")
                    driver.get(f"https://belmontcountyauditor.org/RealEstate/Map?property_Id={property_id}&rowNumber=0")
                    time.sleep(5) # wait for map to load
                    with open(os.path.join(tax_dir, f"PID {parcel_num} Tax Map.pdf"), "wb") as f:
                        f.write(base64.b64decode(driver.print_page()))
                    
                    # Download ACTUAL Tax Card PDF
                    self.log("Downloading actual Tax Card PDF...")
                    tax_card_url = f"https://belmontcountyauditor.org/RealEstate/Default/TaxCard?Property_ID={property_id}&Tax_Year=2025"
                    
                    # Get cookies from selenium to use in requests
                    s = requests.Session()
                    for cookie in driver.get_cookies():
                        s.cookies.set(cookie['name'], cookie['value'])
                    
                    r = s.get(tax_card_url, headers=self.req_headers)
                    if r.status_code == 200 and 'pdf' in r.headers.get('content-type', '').lower():
                        with open(os.path.join(tax_dir, f"PID {parcel_num} Tax Card.pdf"), "wb") as f:
                            f.write(r.content)
                        self.log("Saved Tax Card (actual PDF).")
                    else:
                        self.log(f"Failed to download Tax Card PDF. Status: {r.status_code}")

            finally:
                driver.quit()

            # Execute Kofile scraping if vol and pg exist
            if vol_val and pg_val and vol_val != '0' and pg_val != '0':
                self.log(f"Volume and Page found ({vol_val}/{pg_val}). Copying local deed to docket...")
                self.copy_local_deed(parcel_num, vol_val, pg_val)
            else:
                self.log("Skipping local deed copy and county fetch: Volume/Page empty or 0 in shapefile.")

            # Automatically run Oil & Gas Checker
            self.log("Automatically running Oil & Gas Checker as secondary step...")
            try:
                import og_checker
                og_checker.check_parcel_og_activity(parcel_num, log_callback=self.log, out_dir=well_dir)
            except Exception as e:
                self.log(f"Error in automatic O&G check: {e}")

            # Auto-generate AI Abstracting Summary from pulled deeds (Add-on at the end)
            try:
                self.log("All main tasks completed. Starting automated AI Abstracting Summary as an add-on...")
                api_key = None
                config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
                if os.path.exists(config_path):
                    import json
                    with open(config_path, 'r') as f:
                        config = json.load(f)
                        api_key = config.get('GEMINI_API_KEY')
                if api_key:
                    import ai_parser
                    if os.path.exists(docs_dir):
                        pulled_pdfs = []
                        import glob
                        all_docs = glob.glob(os.path.join(docs_dir, "**", "*.pdf"), recursive=True)
                        for pdf_path in all_docs:
                            f = os.path.basename(pdf_path)
                            if not f.startswith("._") and (f.startswith("DR") or f.startswith("OR")):
                                pulled_pdfs.append(pdf_path)
                                
                        if pulled_pdfs:
                            self.log(f"Uploading {len(pulled_pdfs)} downloaded deed(s) to Gemini for Abstracting Summary...")
                            summary = ai_parser.generate_abstracting_summary_multiple(api_key, pulled_pdfs)
                            if summary and not summary.startswith("Error"):
                                summary_path = os.path.join(pid_dir, f"PID {parcel_num} Chain of Title Summary.txt")
                                with open(summary_path, 'w', encoding='utf-8') as f:
                                    f.write("GEMINI ABSTRACTING SUMMARY\n")
                                    f.write("==========================\n\n")
                                    f.write(summary)
                                self.log("Saved Gemini Chain of Title Abstracting Summary to workspace.")
                                
                                # Kofile Name Search Add-on
                                self.log("Extracting structured name/date search parameters from deeds for Kofile Name Search...")
                                search_params = ai_parser.generate_kofile_search_params(api_key, pulled_pdfs)
                                if search_params:
                                    for sp in search_params:
                                        s_name = sp.get("name", "")
                                        s_acq = sp.get("acquisition_date", "")
                                        s_disp = sp.get("disposal_date", "")
                                        s_lot = sp.get("target_lot", "")
                                        s_pid = sp.get("target_parcel", parcel_num)
                                        if s_name:
                                            self._run_kofile_streaming_pipeline(s_name, s_acq, s_disp, s_lot, s_pid, pid_dir)
                                    self._fetch_court_name_search(search_params, pid_dir)
                                else:
                                    self.log("No valid name search parameters extracted.")
                            else:
                                self.log(f"Gemini Summary Error: {summary}")
                        else:
                            self.log("No downloaded deed PDFs found in docket to summarize.")
                    else:
                        self.log("No docket directory found to summarize.")
                else:
                    self.log("No Gemini API key found. Skipping automated summary.")
            except Exception as summary_e:
                self.log(f"Automated AI Abstract add-on failed (this is non-critical): {summary_e}")

            self.log("Automation completed successfully.")
            self.log(f"Automation Complete for {parcel_num}")
            
            # Ensure Document Viewer folders are refreshed so DOCS, MAPS, etc. show up immediately
            if hasattr(self, 'root'):
                self.root.after(0, self.update_viewer_folders)
            
        except Exception as e:
            self.log(f"Error during automation: {e}")
            import traceback
            traceback.print_exc()
        finally:
            self.set_buttons_state(tk.NORMAL)

    def cancel_organizer(self):
        import os
        import tkinter as tk
        if hasattr(self, 'dropped_filepaths') and self.dropped_filepaths:
            for fp in self.dropped_filepaths:
                if fp and "temp_combined.pdf" in os.path.basename(fp):
                    try:
                        os.remove(fp)
                    except:
                        pass
                        
        self.dropped_filepaths = None
        self.drop_var.set("Drop file here or Combine")
        self.drop_label.config(bg="lightgrey")
        self.org_vol_entry.delete(0, tk.END)
        self.org_pg_entry.delete(0, tk.END)
        self.suffix_combo.set("")
        
        if hasattr(self, 'pending_docket_files'):
            self.pending_docket_files = []
            
        self.log("Action cancelled. Staging cleared.")

    def close_preview(self):
        try:
            import os
            if os.name != 'nt':
                import subprocess
                # Force close all windows without prompting to save
                script = '''tell application "Preview"
    close every window saving no
end tell'''
                subprocess.call(['osascript', '-e', script])
        except Exception as e:
            self.log(f"Error closing preview: {e}")

    def log_missing_document(self, pid, doc_type, vol, pg, missing_type):
        import csv
        import os
        log_file = os.path.join(DRIVE_ROOT, "assignments", "missing_documents.csv")
        file_exists = os.path.exists(log_file)
        
        target_row = [str(pid), str(doc_type), str(vol), str(pg), str(missing_type)]
        
        try:
            # Check for duplicates first
            if file_exists:
                with open(log_file, 'r', newline='') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        # Compare elements as strings
                        if [str(x) for x in row] == target_row:
                            return  # Already exists, do not log again
                            
            with open(log_file, 'a', newline='') as f:
                writer = csv.writer(f)
                if not file_exists:
                    writer.writerow(["PID", "Type", "Volume", "Page", "Missing_Type"])
                writer.writerow(target_row)
            self.refresh_missing_logs()
        except Exception as e:
            self.log(f"Error writing to missing log: {e}")

    def refresh_missing_logs(self):
        import csv
        import os
        
        # clear trees
        for item in self.missing_vols_tree.get_children():
            self.missing_vols_tree.delete(item)
        for item in self.missing_pgs_tree.get_children():
            self.missing_pgs_tree.delete(item)
            
        log_file = os.path.join(DRIVE_ROOT, "assignments", "missing_documents.csv")
        if not os.path.exists(log_file):
            return
            
        try:
            with open(log_file, 'r') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    vals = (row.get("Type"), row.get("Volume"), row.get("Page"), row.get("PID"))
                    if row.get("Missing_Type") == "Volume":
                        self.missing_vols_tree.insert('', 'end', values=vals)
                    else:
                        self.missing_pgs_tree.insert('', 'end', values=vals)
        except Exception as e:
            self.log(f"Error reading missing log: {e}")

    def mark_log_found(self):
        import csv
        import os
        from tkinter import messagebox
        
        selection = self.missing_vols_tree.selection()
        tree = self.missing_vols_tree
        if not selection:
            selection = self.missing_pgs_tree.selection()
            tree = self.missing_pgs_tree
            
        if not selection:
            messagebox.showwarning("Warning", "Please select a log entry first.")
            return
            
        item = selection[0]
        vals = tree.item(item, "values")
        t_type, t_vol, t_pg, t_pid = vals
        
        log_file = os.path.join(DRIVE_ROOT, "assignments", "missing_documents.csv")
        if not os.path.exists(log_file): return
        
        rows_to_keep = []
        try:
            with open(log_file, 'r') as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames
                for row in reader:
                    if str(row.get("PID")) == str(t_pid) and str(row.get("Type")) == str(t_type) and str(row.get("Volume")) == str(t_vol) and str(row.get("Page")) == str(t_pg):
                        continue
                    rows_to_keep.append(row)
                    
            with open(log_file, 'w', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows_to_keep)
                
            tree.delete(item)
            self.log(f"Marked Vol {t_vol} Pg {t_pg} as Found.")
        except Exception as e:
            self.log(f"Error marking log found: {e}")

    def clear_missing_logs(self):
        import os
        from tkinter import messagebox
        if messagebox.askyesno("Confirm Clear", "Are you sure you want to delete all missing logs?"):
            log_file = os.path.join(DRIVE_ROOT, "assignments", "missing_documents.csv")
            if os.path.exists(log_file):
                try:
                    os.remove(log_file)
                    self.refresh_missing_logs()
                    self.log("Missing logs cleared.")
                except Exception as e:
                    self.log(f"Error clearing logs: {e}")


    def on_viewer_search(self, event):
        query = getattr(self, 'viewer_search_var', tk.StringVar()).get().strip()
        if not query:
            if "DOCS" in self.viewer_folder_combo['values']:
                self.viewer_folder_combo.set("DOCS")
        self.refresh_viewer_list()

    def get_notes_file(self):
        parcel_num = self.parcel_entry.get().strip()
        if not parcel_num:
            return None
        pid_dir = self.get_parcel_dir(parcel_num)
        notes_dir = os.path.join(pid_dir, "NOTES")
        return os.path.join(notes_dir, "notes.json")

    def load_notes_for_parcel(self):
        self.notes_tree.delete(*self.notes_tree.get_children())
        self.new_note()
        
        notes_file = self.get_notes_file()
        if not notes_file or not os.path.exists(notes_file):
            return
            
        import json
        try:
            with open(notes_file, 'r') as f:
                notes = json.load(f)
                
            for i, note in enumerate(notes):
                # note is a dict with date, time, subject, content
                self.notes_tree.insert("", "end", iid=str(i), values=(note.get("date", ""), note.get("time", ""), note.get("subject", "")))
        except Exception as e:
            self.log(f"Error loading notes: {e}")

    def load_selected_note(self):
        selection = self.notes_tree.selection()
        if not selection:
            return
            
        note_id = selection[0]
        notes_file = self.get_notes_file()
        if not notes_file or not os.path.exists(notes_file):
            return
            
        import json
        try:
            with open(notes_file, 'r') as f:
                notes = json.load(f)
            note = notes[int(note_id)]
            
            self.note_subject_var.set(note.get("subject", ""))
            self.note_text.delete("1.0", tk.END)
            
            content = note.get("content", "")
            
            # Setup hyperlink styling
            self.note_text.tag_configure("hyperlink", foreground="blue", underline=True)
            self.note_text.tag_bind("hyperlink", "<Enter>", lambda e: self.note_text.config(cursor="hand2"))
            self.note_text.tag_bind("hyperlink", "<Leave>", lambda e: self.note_text.config(cursor=""))
            
            # Custom click handler for viewing the deed
            def on_link_click(event):
                try:
                    index = self.note_text.index(f"@{event.x},{event.y}")
                    tags = self.note_text.tag_names(index)
                    for t in tags:
                        if t.startswith("link_val_"):
                            vol_pg = t.split("link_val_")[1]
                            vol, pg = vol_pg.split('-')
                            clean_v = str(vol).strip().lstrip('0') or '0'
                            clean_p = str(pg).strip().lstrip('0') or '0'
                            vol_pad = vol.zfill(3) if vol.isdigit() else vol
                            vol_pad_4 = vol.zfill(4) if vol.isdigit() else vol
                            
                            import glob, subprocess
                            pid_dir = self.get_parcel_dir(self.parcel_entry.get().strip())
                            
                            # 1. Search DOCS and parcel folder first
                            local_docs = []
                            for ext in ("*.pdf", "*.txt", "*.doc", "*.docx", "*.rtf", "*.png", "*.jpg", "*.tif", "*.tiff"):
                                local_docs.extend(glob.glob(os.path.join(pid_dir, "**", ext), recursive=True))
                                
                            for doc in local_docs:
                                fname = os.path.basename(doc)
                                if fname.startswith("._"): continue
                                if re.search(rf'(?<!\d){clean_v}[-_/ ]+{clean_p}(?!\d)', fname):
                                    try:
                                        if os.name == 'nt': os.startfile(doc)
                                        else: subprocess.Popen(['open', doc])
                                        self.log(f"Opened {fname} from Notes link!")
                                        return
                                    except: pass
                                    
                            # 2. If not found locally, search archives instantly using targeted paths
                            archives = [
                                f"/Volumes/davidlls/drive/DEEDS/DEED {vol_pad}",
                                f"/Volumes/davidlls/drive/MTGS/MTG {vol_pad}",
                                f"/Volumes/davidlls/drive/extracted/DEEDS/DEED {vol_pad}",
                                f"/Volumes/davidlls/drive/extracted/MTGS/MTG {vol_pad}",
                                f"/Volumes/davidlls/Belmont_Drive_External/Belmont County Court House/2. Belmont Deeds/DEED {vol_pad}",
                                f"/Volumes/davidlls/Belmont_Drive_External/Belmont County Court House/3. Belmont Leases/{vol_pad_4}"
                            ]
                            
                            def handle_archive_match(doc_path):
                                import shutil
                                docket_dir = os.path.join(pid_dir, "DOCS", "docket")
                                os.makedirs(docket_dir, exist_ok=True)
                                dest_path = os.path.join(docket_dir, os.path.basename(doc_path))
                                if not os.path.exists(dest_path):
                                    shutil.copy2(doc_path, dest_path)
                                try:
                                    if os.name == 'nt': os.startfile(dest_path)
                                    else: subprocess.Popen(['open', dest_path])
                                    self.log(f"Copied {os.path.basename(doc_path)} to docket and opened it!")
                                except: pass
                                return True
                                
                            for archive_dir in archives:
                                if os.path.exists(archive_dir):
                                    for ext in ("*.pdf", "*.tif", "*.jpg", "*.png"):
                                        for doc in glob.glob(os.path.join(archive_dir, f"*{clean_v}-{clean_p}*{ext}")):
                                            if handle_archive_match(doc): return
                                        for doc in glob.glob(os.path.join(archive_dir, f"*{clean_v}_{clean_p}*{ext}")):
                                            if handle_archive_match(doc): return
                                        for doc in glob.glob(os.path.join(archive_dir, f"*{clean_v} {clean_p}*{ext}")):
                                            if handle_archive_match(doc): return
                except Exception as ex:
                    print(ex)
                    
            self.note_text.tag_bind("hyperlink", "<Button-1>", on_link_click)
            
            import re
            # Convert any existing <link:X-Y> to clean X/Y text
            clean_content = re.sub(r"<link:([0-9a-zA-Z\-]+)>", lambda m: m.group(1).replace('-', '/'), content)
            self.note_text.insert("1.0", clean_content)
            
            # Tag all volume/page occurrences as clickable hyperlinks
            patterns = [
                r'\b(?:DR|OR|MR|LR|PR|PA|WR|MISC|DB|MB|Book|Record)\s*(\d{1,4})\s*[-/,\s]+(?:(?:Page|Pg|p)\.?\s*)?(\d{1,4})\b',
                r'\bVol(?:ume|\.)?\s*(\d{1,4})\s*[-/,\s]+(?:(?:Page|Pg|p)\.?\s*)?(\d{1,4})\b',
                r'(?<![\d/])(\d{1,4})\s*[-/]\s*(\d{1,4})(?![\d/])'
            ]
            seen_spans = []
            for pat in patterns:
                for m in re.finditer(pat, clean_content, re.IGNORECASE):
                    start, end = m.start(), m.end()
                    if any(s <= start < e or s < end <= e for s, e in seen_spans):
                        continue
                    seen_spans.append((start, end))
                    v, p = m.group(1), m.group(2)
                    vpg = f"{v}-{p}"
                    start_idx = f"1.0+{start}c"
                    end_idx = f"1.0+{end}c"
                    self.note_text.tag_add("hyperlink", start_idx, end_idx)
                    self.note_text.tag_add(f"link_val_{vpg}", start_idx, end_idx)
                    
            self.current_note_id = int(note_id)
        except Exception as e:
            self.log(f"Error loading note: {e}")

    def new_note(self):
        self.current_note_id = None
        self.note_subject_var.set("")
        self.note_text.delete("1.0", tk.END)
        self.notes_tree.selection_remove(self.notes_tree.selection())

    def save_note(self):
        notes_file = self.get_notes_file()
        if not notes_file:
            from tkinter import messagebox
            messagebox.showwarning("Warning", "Select a parcel first!")
            return
            
        import json
        from datetime import datetime
        
        notes = []
        if os.path.exists(notes_file):
            try:
                with open(notes_file, 'r') as f:
                    notes = json.load(f)
            except:
                pass
                
        subject = self.note_subject_var.get().strip()
        content = self.note_text.get("1.0", tk.END).strip()
        
        if not subject and not content:
            return
            
        now = datetime.now()
        
        if self.current_note_id is not None and self.current_note_id < len(notes):
            # Update existing
            notes[self.current_note_id]["subject"] = subject
            notes[self.current_note_id]["content"] = content
            notes[self.current_note_id]["date"] = now.strftime("%Y-%m-%d")
            notes[self.current_note_id]["time"] = now.strftime("%I:%M %p")
        else:
            # Create new
            notes.append({
                "subject": subject,
                "content": content,
                "date": now.strftime("%Y-%m-%d"),
                "time": now.strftime("%I:%M %p")
            })
            
        try:
            os.makedirs(os.path.dirname(notes_file), exist_ok=True)
            with open(notes_file, 'w') as f:
                json.dump(notes, f, indent=4)
            self.log("Note saved successfully.")
            self.load_notes_for_parcel()
        except Exception as e:
            self.log(f"Error saving note: {e}")

    def delete_selected_note(self):
        selection = self.notes_tree.selection()
        if not selection:
            return
            
        from tkinter import messagebox
        if not messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this note?"):
            return
            
        note_id = int(selection[0])
        notes_file = self.get_notes_file()
        if not notes_file or not os.path.exists(notes_file):
            return
            
        import json
        try:
            with open(notes_file, 'r') as f:
                notes = json.load(f)
                
            if note_id < len(notes):
                del notes[note_id]
                
            with open(notes_file, 'w') as f:
                json.dump(notes, f, indent=4)
                
            self.log("Note deleted.")
            self.load_notes_for_parcel()
        except Exception as e:
            self.log(f"Error deleting note: {e}")

    def _run_kofile_streaming_pipeline(self, name, start_date, end_date, target_lot, target_parcel, pid_dir, progress_win=None):
        import time, json, os, re, shutil, queue, threading
        from playwright.sync_api import sync_playwright
        from google import genai
        from google.genai import types

        t_start = time.time()
        clean_folder_name = "".join(c for c in name if c.isalnum() or c in " _-").strip() or "Name_Search_Docs"
        target_dir = os.path.join(pid_dir, "DOCS", clean_folder_name)
        hits_dir = os.path.join(target_dir, "Hits")
        os.makedirs(target_dir, exist_ok=True)
        os.makedirs(hits_dir, exist_ok=True)

        # Extract last name
        last_name_cand = name.split()[0] if name else ""
        if len(name.split()) > 1 and name.split()[1] and not name.split()[0].endswith(","):
            last_name_cand = name.split()[-1] if not any(c in name for c in [",", "/"]) else name.split()[0]
        last_name = last_name_cand.replace(",", "").strip().upper()

        def log_msg(msg):
            self.log(msg)
            if progress_win and hasattr(progress_win, 'log_text'):
                try:
                    progress_win.after(0, lambda m=msg: [
                        progress_win.log_text.insert(tk.END, f"{m}\n"),
                        progress_win.log_text.see(tk.END)
                    ])
                except Exception: pass

        def update_status(text, progress_val=None, max_val=None):
            if progress_win and hasattr(progress_win, 'status_lbl'):
                try:
                    def _update():
                        progress_win.status_lbl.config(text=text)
                        if max_val is not None:
                            progress_win.progress_bar.config(mode="determinate", maximum=max_val)
                        if progress_val is not None:
                            progress_win.progress_bar.config(value=progress_val)
                    progress_win.after(0, _update)
                except Exception: pass

        log_msg(f"🚀 Starting Automated Title AI Streaming Engine for '{name}'...")
        log_msg(f"🎯 Target Tract: Lots: '{target_lot}' | Parcel ID: '{target_parcel}' | Last Name: '{last_name}'")
        log_msg(f"📁 Destination Directory: {target_dir}")
        update_status(f"Step 1/3: Searching Belmont County records for '{name}'...")

        # -------------------------------------------------------------
        # STEP 1: Search Public Records on Kofile
        # -------------------------------------------------------------
        all_parsed_rows = []
        seen = set()
        out_file = os.path.join(target_dir, "Kofile_Name_Search_Results.txt")
        manifest_file = os.path.join(target_dir, "Search_Results_Manifest.txt")

        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                context = browser.new_context()
                page = context.new_page()
                page.goto("https://countyfusion13.kofiletech.us/countyweb/loginDisplay.action?countyname=BelmontOH", timeout=45000)
                page.locator("input[value='Login as Guest']").click(no_wait_after=True)
                page.wait_for_load_state('domcontentloaded')
                page.wait_for_timeout(1000)

                try:
                    accept_btn = page.frame_locator("iframe[name='bodyframe']").locator("input#accept")
                    accept_btn.wait_for(state="visible", timeout=6000)
                    accept_btn.click()
                    page.wait_for_load_state('domcontentloaded')
                    page.wait_for_timeout(600)
                except Exception: pass

                try:
                    page.frame_locator("iframe[name='bodyframe']").locator("text='Search Public Records'").first.click(timeout=6000)
                except Exception: pass
                page.wait_for_timeout(1000)

                page.frame_locator("iframe[name='bodyframe']").frame_locator("iframe[name='dynSearchFrame']").get_by_role("tab", name="Name").click()
                page.wait_for_timeout(1000)

                criteria_frame = page.frame_locator("iframe[name='bodyframe']").frame_locator("iframe[name='dynSearchFrame']").frame_locator("iframe[name='criteriaframe']")
                try:
                    criteria_frame.locator("img#clearIcon").click(timeout=1000)
                except Exception: pass

                criteria_frame.get_by_label("Name", exact=True).fill(name)
                if start_date:
                    try:
                        criteria_frame.locator("input[aria-label='Recorded Date From'].textbox-text, input[name*='From']").first.fill(start_date)
                    except Exception: pass
                if end_date:
                    try:
                        criteria_frame.locator("input[aria-label='Recorded Date To'].textbox-text, input[name*='To']").first.fill(end_date)
                    except Exception: pass

                page.frame_locator("iframe[name='bodyframe']").frame_locator("iframe[name='dynSearchFrame']").locator("img#imgSearch").click()
                page.wait_for_timeout(3500)

                reslist = page.frame_locator("iframe[name='bodyframe']").frame_locator("iframe[name='resultFrame']").frame_locator("iframe[name='resultListFrame']")
                reslist.locator("body").wait_for(state="visible", timeout=15000)

                # Hydrate virtual DOM
                for _ in range(3):
                    reslist.locator("body").evaluate("() => window.scrollTo(0, document.body.scrollHeight)")
                    page.wait_for_timeout(800)

                table_data_json = reslist.locator("body").evaluate("""
                    () => {
                        let rows = Array.from(document.querySelectorAll('table tr'));
                        let res = [];
                        for (let r of rows) {
                            let cells = Array.from(r.querySelectorAll('th, td'));
                            let rowVals = cells.map(c => c.innerText.split('\\n').join(' ').trim());
                            if (rowVals.length >= 6) {
                                res.push(rowVals);
                            }
                        }
                        return res;
                    }
                """)

                for row in table_data_json:
                    if any("Instrument" in c or "Document Type" in c for c in row):
                        continue
                    date_idx = -1
                    for i, val in enumerate(row):
                        if re.match(r"^\d{2}/\d{2}/\d{4}$", val.strip()):
                            date_idx = i
                            break
                    if date_idx == -1:
                        continue

                    inst = row[date_idx - 3].strip() if date_idx >= 3 else ""
                    vol = row[date_idx - 2].strip() if date_idx >= 2 else ""
                    pg = row[date_idx - 1].strip() if date_idx >= 1 else ""
                    date = row[date_idx].strip()
                    dtype = row[date_idx + 1].strip() if len(row) > date_idx + 1 else ""
                    name_role = row[date_idx + 2].strip() if len(row) > date_idx + 2 else ""
                    name_val = row[date_idx + 3].strip() if len(row) > date_idx + 3 else ""
                    other_role = row[date_idx + 4].strip() if len(row) > date_idx + 4 else ""
                    other_name = row[date_idx + 5].strip() if len(row) > date_idx + 5 else ""
                    legal = row[date_idx + 6].strip() if len(row) > date_idx + 6 else ""

                    grantor = ""
                    grantee = ""
                    if name_role == "R": grantor = name_val
                    elif name_role in ("E", "D"): grantee = name_val
                    if other_role == "R": grantor = other_name if not grantor else f"{grantor} & {other_name}"
                    elif other_role in ("E", "D"): grantee = other_name if not grantee else f"{grantee} & {other_name}"

                    volpg = f"{vol}/{pg}" if vol and pg else vol or pg
                    key = (inst, volpg, date, dtype, grantor, grantee)
                    if key not in seen:
                        seen.add(key)
                        all_parsed_rows.append((inst, dtype, volpg, date, grantor, grantee, legal))

                browser.close()
        except Exception as e:
            log_msg(f"❌ Kofile search error: {e}")

        # Save manifests
        try:
            with open(out_file, "w") as f:
                f.write(f"--- Results for: {name} ({start_date or 'All'} to {end_date or 'Present'}) ---\n")
                if all_parsed_rows:
                    f.write(f"Found {len(all_parsed_rows)} documents:\n")
                    for r in all_parsed_rows:
                        f.write(f"  [{r[1]}] Vol/Pg: {r[2]} | Date: {r[3]} | Grantor: {r[4]} | Grantee: {r[5]} | Legal: {r[6]}\n")
                else:
                    f.write("  No documents found.\n")

            with open(manifest_file, "w") as f:
                for r in all_parsed_rows:
                    f.write(f"[{r[1]}] Vol/Pg: {r[2]} | Date: {r[3]} | Grantor: {r[4]} | Grantee: {r[5]} | Legal: {r[6]}\n")
        except Exception: pass

        if not all_parsed_rows:
            log_msg(f"⚠️ No documents found on Belmont County Recorder for '{name}'.")
            update_status(f"Search complete: 0 records found for '{name}'.", progress_val=0, max_val=1)
            if hasattr(self, 'root'):
                self.root.after(0, lambda: messagebox.showinfo("No Results", f"No documents found on Kofile for '{name}'.", parent=self.root))
            return

        log_msg(f"📦 Total Search Results: {len(all_parsed_rows)} records found.")

        # -------------------------------------------------------------
        # STEP 2: Deduplicate Book/Pages & Setup Queues
        # -------------------------------------------------------------
        items_to_download = []
        seen_bp = set()
        for r in all_parsed_rows:
            inst, dtype, volpg, date, grantor, grantee, legal = r
            m = re.search(r'(\d+)\s*[-/]\s*(\d+)', volpg)
            if m:
                v, p = m.group(1), m.group(2)
                if (v, p) not in seen_bp:
                    seen_bp.add((v, p))
                    items_to_download.append((v, p, dtype, r))
            elif volpg.isdigit():
                if (volpg, "1") not in seen_bp:
                    seen_bp.add((volpg, "1"))
                    items_to_download.append((volpg, "1", dtype, r))

        total_docs = len(items_to_download)
        log_msg(f"⚡ Unique Documents to Download & Screen: {total_docs}")
        update_status(f"Step 2/3: Streaming Downloads & AI Screening (0 / {total_docs})...", progress_val=0, max_val=total_docs)

        download_queue = queue.Queue()
        for item in items_to_download:
            download_queue.put(item)

        ai_queue = queue.Queue()
        all_ai_results = []
        results_lock = threading.Lock()
        download_complete_event = threading.Event()
        completed_downloads = [0]

        # Init Gemini
        client = None
        try:
            cfg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
            with open(cfg_path) as f:
                api_key = json.load(f)["GEMINI_API_KEY"]
            client = genai.Client(api_key=api_key)
        except Exception as e:
            log_msg(f"⚠️ Gemini Client init warning: {e}")

        # -------------------------------------------------------------
        # STEP 3: AI Screening Worker (Direct Byte Inlining)
        # -------------------------------------------------------------
        def screen_doc(pdf_path, doc_type):
            if not client:
                return {"filename": os.path.basename(pdf_path), "filepath": pdf_path, "is_direct_hit": False, "referenced_prior_vol_pgs": []}
            fn = os.path.basename(pdf_path)
            try:
                t0_ai = time.time()
                with open(pdf_path, "rb") as f:
                    pdf_bytes = f.read()

                prompt = f"""
                You are an expert real estate, oil & gas title attorney examining title records from Belmont County, Ohio.
                Examine this entire document carefully, including Page 1, Page 2, Page 3, and any attached EXHIBIT "A", EXHIBIT "B", Legal Schedules, or Parcel Tables.

                TARGET SEARCH OBJECTIVES:
                1. Check if this document conveys, leases, encumbers, ratifies, assigns, or references:
                   - Target Lot(s): "{target_lot}" (including In-Lot, Out-Lot, or Lot numbers)
                   - Target Parcel ID(s): "{target_parcel}"
                2. OIL & GAS LEASE SPECIAL INSTRUCTION:
                   - Carefully read all tabular columns in Exhibit "A" / Exhibit "B" schedules. Look for matching Lots, Parcels, or Prior Deed references.
                3. MORTGAGES & LIENS:
                   - Does this mortgage encumber the target Lot or Parcel?
                4. RELEASES / SATISFACTIONS / ASSIGNMENTS:
                   - What original Mortgage or Lease Book/Volume and Page numbers does this document release, assign, modify, or satisfy?

                Return a strict JSON object:
                {{
                  "is_direct_hit": true or false,
                  "hit_reasons": ["List matching reasons"],
                  "lots_found": ["List all lot numbers"],
                  "parcels_found": ["List all parcel IDs"],
                  "document_type": "{doc_type}",
                  "recorded_date": "MM/DD/YYYY if visible",
                  "grantor": "Grantor / Direct Party",
                  "grantee": "Grantee / Reverse Party",
                  "is_mortgage": true or false,
                  "referenced_prior_vol_pgs": ["List referenced prior deed, mortgage, or lease Vol/Pg numbers (e.g. '865/633', '618/161')"],
                  "legal_summary": "1-2 sentence concise summary of the land or tract description",
                  "exact_excerpt": "Exact text or table line from document"
                }}
                """
                res = client.models.generate_content(
                    model='gemini-3.6-flash',
                    contents=[types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf"), prompt],
                    config={"response_mime_type": "application/json"}
                )
                data = json.loads(res.text.strip())
                data["filename"] = fn
                data["filepath"] = pdf_path
                data["ai_time"] = round(time.time() - t0_ai, 1)
                return data
            except Exception as e:
                return {"filename": fn, "filepath": pdf_path, "is_direct_hit": False, "hit_reasons": [str(e)], "referenced_prior_vol_pgs": []}

        def extract_volpg_regex(pdf_path):
            import fitz
            refs = set()
            try:
                doc = fitz.open(pdf_path)
                text = " ".join([page.get_text() for page in doc])
                m = re.findall(r'(?:VOL(?:UME)?\.?|BOOK)\s*#?\s*(\d{1,4})\s*(?:AT\s*)?(?:PAGE|PG\.?)\s*#?\s*(\d{1,4})', text, re.IGNORECASE)
                for v, p in m:
                    refs.add(f"{int(v)}/{int(p)}")
                    refs.add(f"{v}/{p}")
            except: pass
            return list(refs)

        def ai_consumer(worker_id):
            while not download_complete_event.is_set() or not ai_queue.empty():
                try:
                    pdf_path, doc_type = ai_queue.get(timeout=1.0)
                except queue.Empty:
                    continue

                data = screen_doc(pdf_path, doc_type)
                regex_refs = extract_volpg_regex(pdf_path)
                existing_refs = data.get("referenced_prior_vol_pgs", [])
                for rf in regex_refs:
                    if rf not in existing_refs:
                        existing_refs.append(rf)
                data["referenced_prior_vol_pgs"] = existing_refs

                is_hit = data.get("is_direct_hit", False)
                fn = os.path.basename(pdf_path)
                t_ai = data.get("ai_time", 0)
                if is_hit:
                    log_msg(f"⭐ [AI-{worker_id}] CONFIRMED HIT in {t_ai}s: {fn}")
                else:
                    log_msg(f"⏭️ [AI-{worker_id}] Non-Target Tract ({t_ai}s): {fn}")

                with results_lock:
                    all_ai_results.append(data)
                ai_queue.task_done()

        ai_threads = []
        for i in range(4):
            t = threading.Thread(target=ai_consumer, args=(i+1,), daemon=True)
            t.start()
            ai_threads.append(t)

        # -------------------------------------------------------------
        # STEP 4: Parallel Scraper Producers (5 Workers)
        # -------------------------------------------------------------
        def scraper_producer(worker_id):
            time.sleep(worker_id * 0.4)
            while True:
                try:
                    vol, pg, dtype, r = download_queue.get_nowait()
                except queue.Empty:
                    break

                clean_type = "".join(c for c in dtype if c.isalnum() or c in " _-").strip() or "DOC"
                expected_pdf = os.path.join(target_dir, f"{vol}-{pg} {clean_type}.pdf")

                # Always grab directly from Kofile website matching party last name
                if os.path.exists(expected_pdf) and os.path.getsize(expected_pdf) > 1000:
                    res = True
                else:
                    log_msg(f"⬇️ [Worker-{worker_id}] Downloading Kofile: {dtype} Vol {vol} Pg {pg}...")
                    res = self._fetch_kofile_single_doc_worker(vol, pg, dtype, target_dir, last_name=last_name)

                if res and os.path.exists(expected_pdf):
                    ai_queue.put((expected_pdf, dtype))
                    log_msg(f"✅ [Worker-{worker_id}] Downloaded: {os.path.basename(expected_pdf)}")
                else:
                    log_msg(f"⚠️ [Worker-{worker_id}] Could not download: {dtype} Vol {vol} Pg {pg}")

                completed_downloads[0] += 1
                cur = completed_downloads[0]
                update_status(f"Step 2/3: Streaming Downloads & AI Screening ({cur} / {total_docs})...", progress_val=cur, max_val=total_docs)
                download_queue.task_done()

        scraper_threads = []
        for i in range(5):
            t = threading.Thread(target=scraper_producer, args=(i+1,), daemon=True)
            t.start()
            scraper_threads.append(t)

        for t in scraper_threads:
            t.join()

        download_complete_event.set()
        for t in ai_threads:
            t.join()

        # -------------------------------------------------------------
        # STEP 5: Relational Linking & Hits Isolation
        # -------------------------------------------------------------
        update_status("Step 3/3: Evaluating relational links and isolating confirmed hits...")
        log_msg("🔗 Resolving relational cross-references (Releases, Satisfactions, Amendments)...")

        confirmed_hits = []
        confirmed_vol_pgs = set()

        for res in all_ai_results:
            fn = res.get("filename", "")
            m = re.match(r'(\d+)-(\d+)', fn)
            if m:
                vp = f"{m.group(1)}/{m.group(2)}"
                if res.get("is_direct_hit"):
                    confirmed_hits.append(res)
                    confirmed_vol_pgs.add(vp)

        # Relational pass
        for res in all_ai_results:
            fn = res.get("filename", "")
            if res.get("is_direct_hit"):
                continue
            refs = res.get("referenced_prior_vol_pgs", [])
            for rf in refs:
                if rf in confirmed_vol_pgs:
                    res["is_relational_hit"] = True
                    res["relational_reason"] = f"References confirmed hit {rf}"
                    confirmed_hits.append(res)
                    m = re.match(r'(\d+)-(\d+)', fn)
                    if m: confirmed_vol_pgs.add(f"{m.group(1)}/{m.group(2)}")
                    log_msg(f"🔗 Relational Link Confirmed: {fn} (Linked via {rf})")
                    break

        # Copy Hits to Hits/
        for hit in confirmed_hits:
            src = hit.get("filepath", "")
            if src and os.path.exists(src):
                dst = os.path.join(hits_dir, os.path.basename(src))
                shutil.copy2(src, dst)

        # Generate Markdown Report
        report_path = os.path.join(target_dir, "Title_AI_Hits_Report.md")
        try:
            with open(report_path, "w") as f:
                f.write(f"# 📋 Belmont County Title AI Screening Report\n\n")
                f.write(f"- **Party Name**: {name}\n")
                f.write(f"- **Target Lot(s)**: {target_lot or 'All'}\n")
                f.write(f"- **Target Parcel**: {target_parcel}\n")
                f.write(f"- **Total Documents Examined**: {total_docs}\n")
                f.write(f"- **Confirmed Title Hits**: {len(confirmed_hits)}\n")
                f.write(f"- **Execution Time**: {round(time.time() - t_start, 2)} seconds\n\n")
                f.write(f"## 🎯 Confirmed Hits Summary\n\n")
                for hit in confirmed_hits:
                    fn = hit.get("filename", "")
                    dt = hit.get("document_type", "DOC")
                    reasons = ", ".join(hit.get("hit_reasons", [])) or hit.get("relational_reason", "Direct Hit")
                    f.write(f"### `{fn}` — {dt}\n")
                    f.write(f"- **Match Basis**: {reasons}\n")
                    f.write(f"- **Legal Excerpt**: {hit.get('legal_summary', hit.get('exact_excerpt', 'N/A'))}\n\n")
        except Exception: pass

        duration = round(time.time() - t_start, 1)
        log_msg("=" * 70)
        log_msg(f"🎉 TITLE AI STREAMING COMPLETE in {duration}s!")
        log_msg(f"📦 Total Examined: {total_docs} | ⭐ Confirmed Hits: {len(confirmed_hits)}")
        log_msg(f"📁 Hits Saved to: DOCS/{clean_folder_name}/Hits/")
        log_msg("=" * 70)

        update_status(f"🎉 Complete in {duration}s! Confirmed Hits: {len(confirmed_hits)} / {total_docs}", progress_val=total_docs, max_val=total_docs)

        # Refresh Document Viewer & select Hits folder
        if hasattr(self, 'root'):
            def update_ui_viewer():
                self.update_viewer_folders()
                hits_fld = f"DOCS/{clean_folder_name}/Hits"
                if hits_fld in self.viewer_folder_combo['values']:
                    self.viewer_folder_combo.set(hits_fld)
                else:
                    self.viewer_folder_combo.set(f"DOCS/{clean_folder_name}")
                self.refresh_viewer_list()

                from tkinter import messagebox
                msg = (
                    f"🎉 Title Examination & AI Triage Complete in {duration}s!\n\n"
                    f"• Party: {name}\n"
                    f"• Examined: {total_docs} documents\n"
                    f"• Confirmed Hits: {len(confirmed_hits)} documents\n\n"
                    f"Hits are isolated in:\nDOCS/{clean_folder_name}/Hits/\n\n"
                    "Would you like to open the Hits folder in Finder?"
                )
                if messagebox.askyesno("AI Triage Complete", msg, parent=self.root):
                    import subprocess, sys
                    try:
                        if sys.platform == "darwin":
                            subprocess.Popen(["open", hits_dir])
                        elif sys.platform == "win32":
                            os.startfile(hits_dir)
                        else:
                            subprocess.Popen(["xdg-open", hits_dir])
                    except Exception: pass

            self.root.after(0, update_ui_viewer)

    def _fetch_court_name_search(self, search_params, pid_dir):
        if not search_params:
            return
            
        self.log(f"Starting CourtView Name Search for {len(search_params)} owners...")
        court_dir = os.path.join(pid_dir, "COURT")
        os.makedirs(court_dir, exist_ok=True)
        
        try:
            from playwright.sync_api import sync_playwright
            import tkinter as tk
            from tkinter import messagebox
            import time
            
            with sync_playwright() as p:
                self.log("Launching visible browser for CourtView (reCAPTCHA solving required)...")
                # Must be headless=False so user can solve the captcha
                browser = p.chromium.launch(headless=False)
                context = browser.new_context()
                page = context.new_page()
                
                self.log("Navigating to Belmont County CourtView...")
                page.goto("https://eservices.belmontcountycourts.com/eservices/home.page.4")
                
                self.log("WAITING FOR HUMAN: Please solve the reCAPTCHA in the browser and click 'Click Here'.")
                self.root.after(0, self.root.bell) # Play an alert sound
                
                # Wait for the user to solve the captcha and for the URL to change to the search page
                try:
                    page.wait_for_url("**/search.page.3**", timeout=120000) # 2 minutes to solve
                    self.log("reCAPTCHA bypassed successfully! Taking over automation...")
                except Exception as e:
                    self.log(f"Timeout waiting for reCAPTCHA bypass or URL change: {e}")
                    browser.close()
                    return
                
                for owner in reversed(search_params):
                    first_name = owner.get("first_name", "").strip()
                    last_name = owner.get("last_name", "").strip()
                    
                    if not last_name and not first_name:
                        name = owner.get("name", "").strip()
                        if not name: continue
                        
                        if "," in name:
                            parts = [p.strip() for p in name.split(",", 1)]
                            last_name = parts[0]
                            first_name = parts[1] if len(parts) > 1 else ""
                        else:
                            parts = name.split()
                            if len(parts) >= 2:
                                # Title Work / AI abstracting formats names as "Last First" (e.g. "Vanfossen Timothy")
                                last_name = parts[0]
                                first_name = " ".join(parts[1:])
                            else:
                                last_name = name
                                first_name = ""
                    
                    self.log(f"Searching CourtView for: Last='{last_name}', First='{first_name}'")
                    
                    # Navigate back to search page just in case we are on a results page
                    if "search.page.3" not in page.url:
                        page.goto("https://eservices.belmontcountycourts.com/eservices/search.page.3")
                    
                    page.wait_for_selector("input[name='lastName']", timeout=10000)
                    page.fill("input[name='lastName']", last_name)
                    if first_name:
                        page.fill("input[name='firstName']", first_name)
                    else:
                        page.fill("input[name='firstName']", "")
                        
                    # Click search
                    page.click("input[type='submit'][value='Search']")
                    page.wait_for_load_state("domcontentloaded")
                    page.wait_for_timeout(3000) # Give results time to render
                    
                    clean_l = "".join(c for c in last_name if c.isalnum() or c in " _-")
                    clean_f = "".join(c for c in first_name if c.isalnum() or c in " _-")
                    pdf_path = os.path.join(court_dir, f"CourtView_{clean_l}_{clean_f}.pdf")
                    self.log(f"Saving PDF to {pdf_path}...")
                    page.pdf(path=pdf_path, format="A4")
                    
                browser.close()
                self.log("CourtView Name Search completed.")
        except Exception as e:
            self.log(f"CourtView Name Search failed: {e}")

    def _fetch_kofile_single_doc_worker(self, vol_val, pg_val, doc_type, target_dir, stagger_sec=0.0, last_name=""):
        import time
        if stagger_sec > 0:
            time.sleep(stagger_sec)
            
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(accept_downloads=True)
            page = context.new_page()
            try:
                page.goto("https://countyfusion13.kofiletech.us/countyweb/loginDisplay.action?countyname=BelmontOH", timeout=45000)
                page.locator("input[value='Login as Guest']").click(no_wait_after=True)
                page.wait_for_load_state('domcontentloaded')
                page.wait_for_timeout(800)

                try:
                    accept_btn = page.frame_locator("iframe[name='bodyframe']").locator("input#accept")
                    accept_btn.wait_for(state="visible", timeout=8000)
                    accept_btn.click()
                    page.wait_for_load_state('domcontentloaded')
                    page.wait_for_timeout(800)
                except Exception:
                    pass

                search_pub = page.frame_locator("iframe[name='bodyframe']").locator("text='Search Public Records'").first
                search_pub.wait_for(state="visible", timeout=10000)
                search_pub.click()
                page.wait_for_timeout(800)

                bp_tab = page.frame_locator("iframe[name='bodyframe']").frame_locator("iframe[name='dynSearchFrame']").get_by_role("tab", name="Book / Page")
                bp_tab.wait_for(state="visible", timeout=10000)
                bp_tab.click()
                page.wait_for_timeout(400)

                crit = page.frame_locator("iframe[name='bodyframe']").frame_locator("iframe[name='dynSearchFrame']").frame_locator("iframe[name='criteriaframe']")
                book_input = crit.get_by_role("textbox", name="Book")
                book_input.wait_for(state="visible", timeout=10000)
                book_input.fill(str(vol_val))
                crit.get_by_role("textbox", name="Page").fill(str(pg_val))

                page.frame_locator("iframe[name='bodyframe']").frame_locator("iframe[name='dynSearchFrame']").locator("img#imgSearch").click()
                page.wait_for_timeout(1200)

                reslist = page.frame_locator("iframe[name='bodyframe']").frame_locator("iframe[name='resultFrame']").frame_locator("iframe[name='resultListFrame']")
                reslist.locator("tr").first.wait_for(state="visible", timeout=25000)

                rows = reslist.locator("tr")
                target_row = None
                for i in range(rows.count()):
                    txt = rows.nth(i).inner_text().strip()
                    if not txt or "Instrument" in txt or "Book/Page" in txt:
                        continue
                    if last_name and last_name.upper() in txt.upper():
                        target_row = rows.nth(i)
                        break
                    elif not target_row:
                        target_row = rows.nth(i)

                if not target_row:
                    browser.close()
                    return False

                target_row.dblclick()
                page.on("dialog", lambda d: d.accept())

                page.frame(name='bodyframe').wait_for_function("""
                    () => {
                        try {
                            var docFrame = document.getElementById("documentFrame");
                            return docFrame && docFrame.contentWindow && typeof docFrame.contentWindow.getNumPages === 'function' && docFrame.contentWindow.getNumPages() > 0;
                        } catch (e) { return false; }
                    }
                """, timeout=45000)

                clean_type = "".join(c for c in doc_type if c.isalnum() or c in " _-").strip() or "DOC"
                with page.expect_download(timeout=60000) as download_info:
                    page.frame(name='bodyframe').evaluate("""
                        var instrId = document.getElementById("documentFrame").contentWindow.getInstrumentId();
                        var numPages = document.getElementById("documentFrame").contentWindow.getNumPages();
                        continueDownloadDocImage(instrId, true, numPages, "printall", false);
                    """)

                download = download_info.value
                out_file = os.path.join(target_dir, f"{vol_val}-{pg_val} {clean_type}.pdf")
                download.save_as(out_file)

                if out_file.lower().endswith('.pdf'):
                    try:
                        import fitz
                        doc = fitz.open(out_file)
                        catalog = doc.pdf_catalog()
                        doc.xref_set_key(catalog, "OutputIntents", "null")
                        doc.xref_set_key(catalog, "Metadata", "null")
                        tmp_path = out_file + ".tmp"
                        doc.save(tmp_path, incremental=False, deflate=True)
                        doc.close()
                        import shutil
                        shutil.move(tmp_path, out_file)
                    except Exception: pass

                browser.close()
                return True
            except Exception as e:
                try: browser.close()
                except: pass
                return False

    def _fetch_kofile_deed_background(self, vol_val, pg_val, parcel_num, doc_type="ALL", custom_docs_dir=None, auto_open=True):
        try:
            self.log(f"⚡ Starting Kofile direct fetch for Volume {vol_val}, Page {pg_val} ({doc_type})...")
            
            from playwright.sync_api import sync_playwright
            import time
            import os
            import re
            
            pid_dir = self.get_parcel_dir(parcel_num)
            docs_dir = custom_docs_dir if custom_docs_dir else os.path.join(pid_dir, "DOCS")
            docket_dir = docs_dir if custom_docs_dir else os.path.join(docs_dir, "docket")
            os.makedirs(docs_dir, exist_ok=True)
            os.makedirs(docket_dir, exist_ok=True)
            
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                context = browser.new_context(accept_downloads=True)
                page = context.new_page()

                self.log("Navigating to Belmont County Recorder...")
                page.goto("https://countyfusion13.kofiletech.us/countyweb/loginDisplay.action?countyname=BelmontOH", timeout=45000)
                
                self.log("Logging in as guest...")
                page.locator("input[value='Login as Guest']").click(no_wait_after=True)
                page.wait_for_load_state('domcontentloaded')
                page.wait_for_timeout(800)
                
                try:
                    accept_btn = page.frame_locator("iframe[name='bodyframe']").locator("input#accept")
                    accept_btn.wait_for(state="visible", timeout=8000)
                    accept_btn.click()
                    page.wait_for_load_state('domcontentloaded')
                    page.wait_for_timeout(800)
                except Exception:
                    pass
                
                self.log("Selecting Book/Page search...")
                search_pub = page.frame_locator("iframe[name='bodyframe']").locator("text='Search Public Records'").first
                search_pub.wait_for(state="visible", timeout=10000)
                search_pub.click()
                page.wait_for_timeout(800)
                
                bp_tab = page.frame_locator("iframe[name='bodyframe']").frame_locator("iframe[name='dynSearchFrame']").get_by_role("tab", name="Book / Page")
                bp_tab.wait_for(state="visible", timeout=10000)
                bp_tab.click()
                page.wait_for_timeout(400)

                self.log(f"Searching Volume {vol_val}, Page {pg_val}...")
                crit = page.frame_locator("iframe[name='bodyframe']").frame_locator("iframe[name='dynSearchFrame']").frame_locator("iframe[name='criteriaframe']")
                book_input = crit.get_by_role("textbox", name="Book")
                book_input.wait_for(state="visible", timeout=10000)
                book_input.fill(str(vol_val))
                crit.get_by_role("textbox", name="Page").fill(str(pg_val))
                
                page.frame_locator("iframe[name='bodyframe']").frame_locator("iframe[name='dynSearchFrame']").locator("img#imgSearch").click()
                page.wait_for_timeout(1500)

                reslist = page.frame_locator("iframe[name='bodyframe']").frame_locator("iframe[name='resultFrame']").frame_locator("iframe[name='resultListFrame']")
                try:
                    reslist.locator("tr").first.wait_for(state="visible", timeout=25000)
                except Exception:
                    self.log(f"❌ No records found for Volume {vol_val}, Page {pg_val} on county records.")
                    browser.close()
                    return False

                rows = reslist.locator("tr")
                target_row = None
                detected_type = doc_type if doc_type and doc_type != "ALL" else "DOC"
                
                for i in range(rows.count()):
                    txt = rows.nth(i).inner_text().strip()
                    if not txt or "Instrument" in txt or "Book/Page" in txt:
                        continue
                    if doc_type and doc_type.upper() == "DEED" and ("DEED" in txt.upper() or "DR" in txt.upper()):
                        target_row = rows.nth(i)
                        detected_type = "Deed"
                        break
                    elif doc_type and doc_type.upper() == "MORTGAGE" and ("MORT" in txt.upper() or "MTG" in txt.upper()):
                        target_row = rows.nth(i)
                        detected_type = "Mortgage"
                        break
                    elif not target_row:
                        target_row = rows.nth(i)
                        if "DEED" in txt.upper(): detected_type = "Deed"
                        elif "MORT" in txt.upper() or "MTG" in txt.upper(): detected_type = "Mortgage"
                        elif "LEASE" in txt.upper() or "OGL" in txt.upper(): detected_type = "Lease"
                        elif "ASSIGN" in txt.upper(): detected_type = "Assignment"
                        elif "RELEASE" in txt.upper() or "SATIS" in txt.upper(): detected_type = "Release"

                if not target_row:
                    self.log(f"❌ No valid result rows found for Volume {vol_val}, Page {pg_val}.")
                    browser.close()
                    return False

                self.log("Opening document in county viewer...")
                target_row.dblclick()
                page.on("dialog", lambda d: d.accept())

                self.log("Waiting for document pages to load...")
                page.frame(name='bodyframe').wait_for_function("""
                    () => {
                        try {
                            var docFrame = document.getElementById("documentFrame");
                            return docFrame && docFrame.contentWindow && typeof docFrame.contentWindow.getNumPages === 'function' && docFrame.contentWindow.getNumPages() > 0;
                        } catch (e) { return false; }
                    }
                """, timeout=50000)

                clean_type = "".join(c for c in detected_type if c.isalnum() or c in " _-").strip() or "DOC"
                self.log(f"Downloading high-resolution document ({clean_type})...")
                
                with page.expect_download(timeout=60000) as download_info:
                    page.frame(name='bodyframe').evaluate("""
                        var instrId = document.getElementById("documentFrame").contentWindow.getInstrumentId();
                        var numPages = document.getElementById("documentFrame").contentWindow.getNumPages();
                        continueDownloadDocImage(instrId, true, numPages, "printall", false);
                    """)

                download = download_info.value
                out_file = os.path.join(docket_dir, f"{vol_val}-{pg_val} {clean_type}.pdf")
                download.save_as(out_file)

                if out_file.lower().endswith('.pdf'):
                    try:
                        import fitz
                        doc = fitz.open(out_file)
                        catalog = doc.pdf_catalog()
                        doc.xref_set_key(catalog, "OutputIntents", "null")
                        doc.xref_set_key(catalog, "Metadata", "null")
                        tmp_path = out_file + ".tmp"
                        doc.save(tmp_path, incremental=False, deflate=True)
                        doc.close()
                        import shutil
                        shutil.move(tmp_path, out_file)
                    except Exception as e:
                        self.log(f"PDF/A sanitization notice: {e}")

                self.log(f"✅ Successfully downloaded {os.path.basename(out_file)} to docket!")
                browser.close()

                # Refresh UI Document Viewer
                if hasattr(self, 'root'):
                    def update_viewer():
                        self.update_viewer_folders()
                        docket_val = "DOCS/docket" if not custom_docs_dir else os.path.relpath(docket_dir, pid_dir)
                        if docket_val in self.viewer_folder_combo['values']:
                            self.viewer_folder_combo.set(docket_val)
                        self.refresh_viewer_list()
                    self.root.after(0, update_viewer)

                if auto_open:
                    import subprocess, sys
                    try:
                        if sys.platform == "darwin":
                            subprocess.Popen(["open", out_file])
                        elif sys.platform == "win32":
                            os.startfile(out_file)
                        else:
                            subprocess.Popen(["xdg-open", out_file])
                    except Exception as ex:
                        self.log(f"Error opening downloaded document: {ex}")

                return True
        except Exception as e:
            self.log(f"❌ Kofile scraper error for Vol {vol_val} Pg {pg_val}: {e}")
            return False

    def open_name_search(self):
        pid = self.parcel_entry.get().strip()
        if not pid:
            from tkinter import messagebox
            messagebox.showerror("Error", "Please enter a Parcel Number (PID) first.", parent=self.root)
            return
            
        pid_dir = self.get_parcel_dir(pid)
        docs_dir = os.path.join(pid_dir, "DOCS")
        os.makedirs(docs_dir, exist_ok=True)
        out_file = os.path.join(docs_dir, "Kofile_Name_Search_Results.txt")
        
        history_names = []
        if os.path.exists(out_file):
            import re
            try:
                with open(out_file, "r") as f:
                    content = f.read()
                    matches = re.findall(r'--- Results for: (.*?)\s*\(', content)
                    history_names = list(set(matches))
            except: pass
            
        dialog = tk.Toplevel(self.root)
        dialog.title("⚡ Kofile Name Search & Automated AI Triage")
        dialog.geometry("520x400")
        dialog.transient(self.root)
        
        active_pid = self.parcel_entry.get().strip() if hasattr(self, 'parcel_entry') else ""
        
        top_banner = ttk.Frame(dialog, padding=(16, 12))
        top_banner.pack(fill=tk.X)
        ttk.Label(top_banner, text="⚡ Belmont County Title Streamer & AI Triage", font=("Helvetica", 13, "bold")).pack(anchor=tk.W)
        ttk.Label(top_banner, text="Automatically searches public records, downloads PDFs with parallel workers,\nand runs Gemini 3.6 Flash multimodal AI screening to isolate confirmed tract hits.", foreground="#4a5568").pack(anchor=tk.W, pady=(2, 0))
        
        form_frame = ttk.Frame(dialog, padding=(16, 4))
        form_frame.pack(fill=tk.X)
        
        ttk.Label(form_frame, text="Party Name (e.g. McLaughlin Douglas):", font=("Helvetica", 11, "bold")).pack(anchor=tk.W, pady=(4, 2))
        name_var = tk.StringVar()
        name_cb = ttk.Combobox(form_frame, textvariable=name_var, values=history_names, font=("Helvetica", 11))
        name_cb.pack(fill=tk.X, pady=(0, 8))
        if history_names:
            name_cb.set(history_names[0])
            
        date_frame = ttk.Frame(form_frame)
        date_frame.pack(fill=tk.X, pady=4)
        
        ttk.Label(date_frame, text="Start Date:").grid(row=0, column=0, sticky=tk.W, pady=3)
        start_date_var = tk.StringVar(value="01/01/1980")
        ttk.Entry(date_frame, textvariable=start_date_var, width=13, font=("Helvetica", 10)).grid(row=0, column=1, sticky=tk.W, padx=(4, 15), pady=3)
        
        ttk.Label(date_frame, text="End Date:").grid(row=0, column=2, sticky=tk.W, pady=3)
        from datetime import datetime
        today_str = datetime.now().strftime("%m/%d/%Y")
        end_date_var = tk.StringVar(value=today_str)
        ttk.Entry(date_frame, textvariable=end_date_var, width=13, font=("Helvetica", 10)).grid(row=0, column=3, sticky=tk.W, padx=4, pady=3)
        
        # Target Tract Criteria Frame
        t_frame = ttk.LabelFrame(form_frame, text="🎯 Target Tract Criteria (AI Screening Filter)", padding=10)
        t_frame.pack(fill=tk.X, pady=8)
        
        ttk.Label(t_frame, text="Target Lot(s):").grid(row=0, column=0, sticky=tk.W, pady=3)
        target_lot_var = tk.StringVar(value="")
        ttk.Entry(t_frame, textvariable=target_lot_var, width=13, font=("Helvetica", 10)).grid(row=0, column=1, sticky=tk.W, padx=(4, 15), pady=3)
        
        ttk.Label(t_frame, text="Parcel ID:").grid(row=0, column=2, sticky=tk.W, pady=3)
        target_parcel_var = tk.StringVar(value=active_pid)
        ttk.Entry(t_frame, textvariable=target_parcel_var, width=15, font=("Helvetica", 10)).grid(row=0, column=3, sticky=tk.W, padx=4, pady=3)
        
        def on_search():
            name = name_var.get().strip()
            if not name:
                from tkinter import messagebox
                messagebox.showwarning("Missing Name", "Please enter a Party Name to search.", parent=dialog)
                return
            start = start_date_var.get().strip()
            end = end_date_var.get().strip()
            lot_crit = target_lot_var.get().strip()
            parcel_crit = target_parcel_var.get().strip()
            dialog.destroy()
            
            # Launch Live Progress Window
            prog_win = KofileStreamingProgressWindow(self.root, name)
            
            import threading
            threading.Thread(
                target=self._run_kofile_streaming_pipeline,
                args=(name, start, end, lot_crit, parcel_crit, pid_dir, prog_win),
                daemon=True
            ).start()
            
        btn_box = ttk.Frame(dialog, padding=(16, 12))
        btn_box.pack(fill=tk.X)
        ttk.Button(btn_box, text="⚡ Start Search & AI Triage", command=on_search, style="Accent.TButton").pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(btn_box, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT)

    def open_quick_log_mortgage(self):
        vol = self.org_vol_entry.get().strip()
        pg = self.org_pg_entry.get().strip()
        pid = self.parcel_entry.get().strip()
        
        if not vol or not pg:
            vol = self.vol_entry.get().strip()
            pg = self.pg_entry.get().strip()
            
        if not vol or not pg:
            selection = getattr(self, 'viewer_listbox', None) and self.viewer_listbox.curselection()
            if selection:
                fname = self.viewer_listbox.get(selection[0])
                import re
                m = re.search(r'(\d+)[-_](\d+)', fname)
                if m:
                    vol = m.group(1)
                    pg = m.group(2)
        
        if not vol or not pg or not pid:
            from tkinter import messagebox
            messagebox.showwarning("Missing Info", "Please enter Parcel ID, Volume, and Page first.")
            return
            
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Log Mortgage - Vol {vol} Pg {pg}")
        dialog.geometry("350x250")
        
        ttk.Label(dialog, text="Grantor:").grid(row=0, column=0, padx=10, pady=10, sticky='e')
        grantor_entry = ttk.Entry(dialog)
        grantor_entry.grid(row=0, column=1, padx=10, pady=10, sticky='ew')
        
        ttk.Label(dialog, text="Grantee:").grid(row=1, column=0, padx=10, pady=10, sticky='e')
        grantee_entry = ttk.Entry(dialog)
        grantee_entry.grid(row=1, column=1, padx=10, pady=10, sticky='ew')
        
        released_var = tk.BooleanVar()
        chk = ttk.Checkbutton(dialog, text="Mortgage Released?", variable=released_var)
        chk.grid(row=2, column=0, columnspan=2, pady=5)
        
        ttk.Label(dialog, text="Release Vol/Pg:").grid(row=3, column=0, padx=10, pady=10, sticky='e')
        release_entry = ttk.Entry(dialog)
        release_entry.grid(row=3, column=1, padx=10, pady=10, sticky='ew')
        
        def on_save():
            import json, datetime, os
            pid_dir = self.get_parcel_dir(pid)
            notes_dir = os.path.join(pid_dir, "NOTES")
            os.makedirs(notes_dir, exist_ok=True)
            notes_file = os.path.join(notes_dir, "notes.json")
            
            notes = []
            if os.path.exists(notes_file):
                try:
                    with open(notes_file, 'r') as f:
                        notes = json.load(f)
                except: pass
                
            mortgage_note_idx = None
            for i, note in enumerate(notes):
                if note.get("subject") == "Mortgage Tracking":
                    mortgage_note_idx = i
                    break
                    
            if mortgage_note_idx is None:
                notes.append({
                    "subject": "Mortgage Tracking",
                    "content": "List of Mortgages and their Release statuses:\n\n",
                    "date": datetime.datetime.now().strftime("%Y-%m-%d"),
                    "time": datetime.datetime.now().strftime("%I:%M %p")
                })
                mortgage_note_idx = len(notes) - 1
                
            content = notes[mortgage_note_idx].get("content", "")
            lines = content.split('\n')
            
            mortgage_link = f"<link:{vol}-{pg}>"
            release_val = release_entry.get().strip()
            release_link = f"<link:{release_val.replace(' ', '').replace('/', '-')}>" if release_val else ""
            
            grantor = grantor_entry.get().strip() or "Unknown"
            grantee = grantee_entry.get().strip() or "Unknown"
            is_released = released_var.get()
            
            if is_released:
                new_status = f"[RELEASED by {release_link}]" if release_val else "[RELEASED]"
            else:
                new_status = "[UNRELEASED]"
                
            entry_prefix = f"• Mortgage {mortgage_link} "
            new_line = f"• Mortgage {mortgage_link} - {grantor} to {grantee} -> {new_status}"
            
            found = False
            for i, line in enumerate(lines):
                if line.startswith(entry_prefix):
                    lines[i] = new_line
                    found = True
                    break
                    
            if not found:
                lines.append(new_line)
                
            notes[mortgage_note_idx]["content"] = '\n'.join(lines)
            notes[mortgage_note_idx]["date"] = datetime.datetime.now().strftime("%Y-%m-%d")
            notes[mortgage_note_idx]["time"] = datetime.datetime.now().strftime("%I:%M %p")
            
            with open(notes_file, 'w') as f:
                json.dump(notes, f, indent=4)
                
            self.log(f"Logged Mortgage {vol}/{pg} tracking note!")
            self.load_notes_for_parcel()
            dialog.destroy()
            
        ttk.Button(dialog, text="Save to Notes", command=on_save).grid(row=4, column=0, columnspan=2, pady=15)
        
    def open_submission_email(self, event=None):
        pid = self.parcel_entry.get().strip()
        pid_dir = self.get_parcel_dir(pid) if pid and os.path.exists(self.get_parcel_dir(pid)) else None
        import submission_email_dialog
        submission_email_dialog.SubmissionEmailDialog(self.root, pid_dir, parcel_num=pid)
        return "break"

    def open_pdf_combiner(self):
        pid = self.parcel_entry.get().strip()
        pid_dir = self.get_parcel_dir(pid) if pid and os.path.exists(self.get_parcel_dir(pid)) else None
        
        import pdf_combiner
        pdf_combiner.PDFCombinerWindow(self.root, parcel_dir=pid_dir)

    def open_plat_cabinet_searcher(self):
        pid = self.parcel_entry.get().strip()
        pid_dir = self.get_parcel_dir(pid) if pid and os.path.exists(self.get_parcel_dir(pid)) else None
        
        import plat_cabinet_searcher
        plat_cabinet_searcher.PlatCabinetSearchWindow(self.root, parcel_dir=pid_dir)

    def show_shortcuts_dialog(self):
        popup = tk.Toplevel(self.root)
        popup.title("Keyboard Shortcuts Cheat Sheet")
        popup.geometry("640x520")
        popup.attributes("-topmost", True)
        popup.transient(self.root)

        try:
            x = self.root.winfo_rootx() + (self.root.winfo_width() // 2) - 320
            y = self.root.winfo_rooty() + (self.root.winfo_height() // 2) - 260
            popup.geometry(f"+{x}+{y}")
        except Exception: pass

        main_frame = ttk.Frame(popup, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="⌨️ Title Work Shortcuts Cheat Sheet", font=("Helvetica", 16, "bold")).pack(anchor=tk.W, pady=(0, 10))

        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        cols = ("shortcut", "area", "description")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=15)
        tree.heading("shortcut", text="Shortcut Key")
        tree.heading("area", text="Area")
        tree.heading("description", text="Action / Description")

        tree.column("shortcut", width=180, anchor=tk.W)
        tree.column("area", width=100, anchor=tk.W)
        tree.column("description", width=320, anchor=tk.W)

        shortcuts = [
            # Main Portal
            ("Cmd + E / Ctrl + E", "Main Portal", "Open Runsheet Form Editor (Edit RS)"),
            ("Cmd + O / Ctrl + O", "Main Portal", "Open active parcel's Ownership Report (*OR*.xlsx)"),
            ("Cmd + G / Ctrl + G", "Main Portal", "Open Belmont GIS zoomed to active parcel"),
            ("Cmd + R / Ctrl + R", "Main Portal", "Open Belmont Recorder website (Kofile)"),
            ("Cmd+Shift+E / Ctrl+Shift+E", "Main Portal", "✉️ Open Completion & Submission Email Generator"),
            
            # Editor
            ("Cmd+Shift+O / Ctrl+Shift+O", "RS / Gemini", "📊 Sync Runsheet to Ownership Report (*OR*.xlsx)"),
            ("Cmd + S / Ctrl + S", "RS Editor", "Save current row & Excel workbook"),
            ("Cmd + G / Ctrl + G", "Gemini Editor", "✨ Draft active row with Gemini AI"),
            ("Cmd + A / Ctrl + A", "RS / Gemini", "View Gemini Source Provenance, Quotes, & Warnings"),
            ("Cmd+Shift+D / Opt+D", "RS / Gemini", "Strip '--- Gemini Draft ---' / '--- Original ---' blocks"),
            ("Cmd + P / Ctrl + P", "RS Editor", "Set row status to 'In Progress'"),
            ("Cmd + F / Ctrl + F", "RS Editor", "Set row status to 'Completed'"),
            ("Cmd + D / Ctrl + D", "RS Editor", "Toggle 'Dower Reviewed' checkbox"),
            ("Cmd + O / Ctrl + O", "RS Editor", "Open PDF Document for current row"),
            ("Cmd + N / Ctrl + N", "RS Editor", "Convert selection or field to Title Case"),
            ("Cmd + L / Ctrl + L", "RS Editor", "Open Phrase Library window"),
            ("Ctrl + 1 .. 9, 0", "RS Editor", "Insert Phrase #1 through #10 at cursor"),
            ("Ctrl + ↑ / ↓ (Alt + ↑/↓)", "RS Editor", "Save & jump to Previous / Next Row"),
            ("Click 🔍 Blue Label", "RS Editor", "Auto-apply AI suggested field value"),
            ("Left-Click Link", "RS Editor", "Jump to referenced runsheet row"),
            ("Right-Click Link", "RS Editor", "Open referenced PDF Document")
        ]

        for sc, area, desc in shortcuts:
            tree.insert("", tk.END, values=(sc, area, desc))

        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        btn_close = ttk.Button(main_frame, text="Close (Esc)", command=popup.destroy)
        btn_close.pack(pady=(10, 0), anchor=tk.E)
        btn_close.focus_set()

        popup.bind("<Escape>", lambda e: popup.destroy())
        popup.bind("<Return>", lambda e: popup.destroy())

    def open_belmont_recorder(self, event=None):
        import subprocess, sys
        script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "launch_kofile.py")
        try:
            subprocess.Popen([sys.executable, script_path])
            self.log("🏛️ Launching Belmont County Recorder and navigating directly to Search Public Records...")
        except Exception as e:
            import webbrowser
            recorder_url = "https://countyfusion13.kofiletech.us/countyweb/loginDisplay.action?countyname=BelmontOH"
            webbrowser.open_new_tab(recorder_url)
            self.log(f"🏛️ Opened Belmont County Recorder in browser fallback: {e}")
        return "break"

    def open_belmont_gis(self, event=None):
        raw_p_num = self.parcel_entry.get().strip()
        if not raw_p_num:
            from tkinter import messagebox
            messagebox.showerror("Error", "Please enter or select a Parcel Number (PID) first.")
            return "break"

        import re, webbrowser
        m = re.search(r'\d{2}-\d{5}\.\d{3}', raw_p_num)
        if m:
            clean_parcel = m.group(0)
        else:
            clean_parcel = raw_p_num.split()[0].replace("PID", "").strip()

        gis_url = f"https://gis.belcogis.com/ParcelMap/#widget_48=text:{clean_parcel}&zoom_to_selection=true"

        try:
            webbrowser.open_new_tab(gis_url)
            self.log(f"🗺️ Opened Belmont County GIS for Parcel: {clean_parcel}")
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Error", f"Could not open Belmont GIS URL:\n{e}")

        return "break"

    def open_ownership_report(self, event=None):
        pid = self.parcel_entry.get().strip()
        if not pid:
            from tkinter import messagebox
            messagebox.showerror("Error", "Please enter or select a Parcel Number (PID) first.")
            return "break"

        pid_dir = self.get_parcel_dir(pid)
        if not os.path.exists(pid_dir):
            from tkinter import messagebox
            messagebox.showerror("Error", f"Parcel folder does not exist:\n{pid_dir}")
            return "break"

        import glob
        matches = glob.glob(os.path.join(pid_dir, "*OR*.xlsx")) + glob.glob(os.path.join(pid_dir, "*Ownership*.xlsx"))
        matches = [
            m for m in matches 
            if not os.path.basename(m).startswith("~$") 
            and not os.path.basename(m).startswith("._") 
            and "_Backup" not in m
        ]

        if not matches:
            from tkinter import messagebox
            messagebox.showerror("Error", f"Could not find any Ownership Report (*OR*.xlsx) in:\n{pid_dir}")
            return "break"

        target_file = matches[0]
        for m in matches:
            b_name = os.path.basename(m).upper()
            if "TEMPLATE" not in b_name and "BLANK" not in b_name:
                target_file = m
                break

        try:
            if os.name == 'nt':
                os.startfile(target_file)
            else:
                import subprocess
                subprocess.call(('open', target_file))
            self.log(f"📊 Opened Ownership Report: {os.path.basename(target_file)}")
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Error", f"Could not open Ownership Report Excel:\n{e}")

        return "break"

    def reformat_active_parcel_runsheet(self):
        pid = self.parcel_entry.get().strip()
        if not pid:
            from tkinter import messagebox
            messagebox.showerror("Error", "Please enter a Parcel Number (PID) first.", parent=self.root)
            return

        pid_dir = self.get_parcel_dir(pid)
        if not os.path.exists(pid_dir):
            from tkinter import messagebox
            messagebox.showerror("Error", f"Parcel folder does not exist:\n{pid_dir}", parent=self.root)
            return

        import glob
        matches = glob.glob(os.path.join(pid_dir, "*RS*.xlsx")) + glob.glob(os.path.join(pid_dir, "*Runsheet*.xlsx"))
        matches = [
            m for m in matches 
            if not os.path.basename(m).startswith("~$") 
            and not os.path.basename(m).startswith("._") 
            and "_Backup" not in m
        ]

        if not matches:
            from tkinter import messagebox
            messagebox.showerror("Error", f"Could not find any Runsheet (*RS*.xlsx) in:\n{pid_dir}", parent=self.root)
            return

        target_file = matches[0]
        for m in matches:
            b_name = os.path.basename(m).upper()
            if "TEMPLATE" not in b_name and "BLANK" not in b_name:
                target_file = m
                break

        # Ask user for confirmation
        from tkinter import messagebox
        confirm = messagebox.askyesno(
            "Reformat Runsheet Excel",
            f"Do you want to reformat the Runsheet Excel workbook for PID {pid}?\n\n"
            f"File: {os.path.basename(target_file)}\n\n"
            "This will:\n"
            "• Create a timestamped backup in BACKUPS/\n"
            "• Clean messy date formats and remove quote prefixes\n"
            "• Parse ARTI, Dower, Amounts, Maturities, and Prior References\n"
            "• Apply bold formatted section headers and standardized styling\n\n"
            "Proceed?",
            parent=self.root
        )
        if not confirm:
            return

        # Perform the complete reformatting pipeline
        try:
            import datetime, shutil, json, re, openpyxl
            from openpyxl.cell.rich_text import TextBlock, CellRichText

            # 1. Backup
            backups_dir = os.path.join(pid_dir, "BACKUPS")
            os.makedirs(backups_dir, exist_ok=True)
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = os.path.splitext(os.path.basename(target_file))[0]
            backup_file = os.path.join(backups_dir, f"{base_name}_backup_{ts}.xlsx")
            shutil.copy2(target_file, backup_file)

            wb = openpyxl.load_workbook(target_file, rich_text=True)
            ws = wb.active

            headers = [str(cell.value or "").strip() for cell in ws[2]]
            date_cols = [i for i, h in enumerate(headers) if "date" in h.lower()]
            comments_col = None
            for i, h in enumerate(headers):
                if "comment" in h.lower() or "note" in h.lower():
                    comments_col = i + 1
                    break

            # 2. Clean dates
            for row_idx in range(3, ws.max_row + 1):
                row = ws[row_idx]
                for col_idx in date_cols:
                    if col_idx < len(row):
                        cell = row[col_idx]
                        val = str(cell.value or "").strip()
                        if getattr(cell, 'quotePrefix', False):
                            cell.quotePrefix = False
                        if len(val) > 20 and "GMT" in val:
                            parts = val.split()
                            if len(parts) >= 4 and parts[0] in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]:
                                try:
                                    dt = datetime.datetime.strptime(" ".join(parts[:4]), "%a %b %d %Y")
                                    cell.value = dt.strftime("%m/%d/%Y")
                                    cell.number_format = '@'
                                    cell.data_type = 's'
                                except Exception: pass
                        elif val.startswith("'"):
                            cell.value = val[1:]
                            cell.number_format = '@'
                            cell.data_type = 's'

            # 3. Format Comments
            changed_count = 0
            if comments_col:
                import runsheet_editor
                editor_dummy = runsheet_editor.RunsheetEditorWindow.__new__(runsheet_editor.RunsheetEditorWindow)
                editor_dummy.headers = headers
                editor_dummy.pid_dir = pid_dir
                editor_dummy.base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                
                for row_idx in range(3, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=comments_col)
                    val = cell.value
                    if not val:
                        continue
                    if type(val).__name__ == 'CellRichText':
                        txt = "".join(str(p) for p in val)
                    else:
                        txt = str(val)

                    if "--- Original ---" in txt:
                        raw_text = txt.split("--- Original ---")[1].strip()
                    else:
                        raw_text = txt.replace("\u200B", "").strip()

                    inst_type = str(ws.cell(row=row_idx, column=1).value or "").lower()
                    formatted_txt = runsheet_editor.RunsheetEditorWindow.apply_initial_formatting_pipeline(editor_dummy, raw_text, inst_type)
                    parts = runsheet_editor.RunsheetEditorWindow._parse_bold_tokens(editor_dummy, formatted_txt)

                    if not any(isinstance(p, TextBlock) for p in parts):
                        cell.value = "".join(parts).strip()
                    else:
                        cell.value = CellRichText(*parts)
                    changed_count += 1

            wb.save(target_file)

            # Update initial_formatting_done.json
            fmt_state = os.path.join(pid_dir, "initial_formatting_done.json")
            with open(fmt_state, "w") as f:
                json.dump({"initial_formatting_completed": True}, f)

            self.log(f"⚡ Reformat complete: {changed_count} rows formatted in {os.path.basename(target_file)}")

            # Notify user
            open_confirm = messagebox.askyesno(
                "Reformat Complete",
                f"🎉 Runsheet reformatted successfully!\n\n"
                f"• Reformatted {changed_count} rows with clean dates and bold formatting.\n"
                f"• Backup saved to: BACKUPS/{os.path.basename(backup_file)}\n\n"
                "Would you like to open the Excel spreadsheet now?",
                parent=self.root
            )
            if open_confirm:
                import subprocess
                if os.name == 'nt':
                    os.startfile(target_file)
                else:
                    subprocess.call(('open', target_file))

        except Exception as e:
            self.log(f"Reformat failed: {e}")
            messagebox.showerror("Error", f"Failed to reformat Runsheet Excel:\n{e}", parent=self.root)

    def open_gemini_rs_editor(self):
        pid = self.parcel_entry.get().strip()
        if not pid:
            from tkinter import messagebox
            messagebox.showerror("Error", "Please enter a Parcel Number (PID) first.")
            return

        import gemini_runsheet_editor
        gemini_runsheet_editor.GeminiRunsheetEditorWindow(self.root, pid, os.path.dirname(os.path.abspath(__file__)))

    def open_rs_editor(self, event=None):
        pid = self.parcel_entry.get().strip()
        if not pid:
            from tkinter import messagebox
            messagebox.showerror("Error", "Please enter a Parcel Number (PID) first.")
            return "break"
            
        import runsheet_editor
        runsheet_editor.RunsheetEditorWindow(self.root, pid, os.path.dirname(os.path.abspath(__file__)))
        return "break" 
        
    def open_status_tracker(self):
        import status_tracker
        status_tracker.StatusOverviewWindow(self.root, BASE_DIR)

    def open_chat(self):
        # We need the API key to open the chat
        import json
        import os
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
        api_key = ""
        if os.path.exists(config_path):
            try:
                with open(config_path, "r") as f:
                    config = json.load(f)
                    api_key = config.get("GEMINI_API_KEY", "")
            except: pass
            
        if not api_key:
            messagebox.showwarning("Warning", "No Gemini API Key found in config.json. Cannot open Chat.")
            return
            
        SOPChatWindow(self.root, api_key)


import queue
import threading

class KofileStreamingProgressWindow(tk.Toplevel):
    def __init__(self, master, party_name):
        super().__init__(master)
        self.title(f"⚡ Title AI Streaming Engine - {party_name}")
        self.geometry("640x440")
        self.minsize(550, 350)
        self.transient(master)
        
        # Header
        top_frame = ttk.Frame(self, padding=12)
        top_frame.pack(fill=tk.X)
        
        ttk.Label(top_frame, text=f"⚡ Title AI Streaming Engine: {party_name}", font=("Helvetica", 13, "bold")).pack(anchor=tk.W)
        self.status_lbl = ttk.Label(top_frame, text="Starting search on Belmont County Recorder...", font=("Helvetica", 10), foreground="#4a5568")
        self.status_lbl.pack(anchor=tk.W, pady=(2, 4))
        
        self.progress_bar = ttk.Progressbar(top_frame, orient=tk.HORIZONTAL, mode="indeterminate")
        self.progress_bar.pack(fill=tk.X, pady=(4, 8))
        self.progress_bar.start(10)
        
        # Live Activity Log Area
        log_frame = ttk.LabelFrame(self, text="📜 Live Activity & AI Triage Log", padding=8)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 8))
        
        import tkinter.scrolledtext as st
        self.log_text = st.ScrolledText(log_frame, wrap=tk.WORD, font=("Courier", 10))
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # Bottom controls
        btn_frame = ttk.Frame(self, padding=(12, 8))
        btn_frame.pack(fill=tk.X)
        
        self.close_btn = ttk.Button(btn_frame, text="Minimize / Close Window", command=self.destroy)
        self.close_btn.pack(side=tk.RIGHT)

class SOPChatWindow(tk.Toplevel):
    def __init__(self, master, api_key):
        super().__init__(master)
        self.title("Harbinger Land - SOP Chat Agent")
        self.geometry("700x800")
        self.api_key = api_key
        
        self.client = None
        self.chat = None
        self.uploaded_files = []
        self.msg_queue = queue.Queue()
        
        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        
        # Start initialization thread
        self.status_var.set("Initializing Agent and uploading SOPs... Please wait.")
        self.input_entry.config(state="disabled")
        self.send_btn.config(state="disabled")
        
        threading.Thread(target=self._init_chat_session, daemon=True).start()
        self.after(100, self._process_queue)

    def _build_ui(self):
        self.chat_history = tk.Text(self, wrap=tk.WORD, state="disabled", font=("Helvetica", 12))
        self.chat_history.pack(expand=True, fill="both", padx=10, pady=10)
        
        # tags for formatting
        self.chat_history.tag_configure("user", foreground="blue", font=("Helvetica", 12, "bold"))
        self.chat_history.tag_configure("agent", foreground="white")
        self.chat_history.tag_configure("system", foreground="grey", font=("Helvetica", 10, "italic"))
        
        bottom_frame = ttk.Frame(self)
        bottom_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        self.input_entry = ttk.Entry(bottom_frame, font=("Helvetica", 12))
        self.input_entry.pack(side=tk.LEFT, expand=True, fill="x", padx=(0, 5))
        self.input_entry.bind("<Return>", lambda e: self.send_message())
        
        self.send_btn = ttk.Button(bottom_frame, text="Send", command=self.send_message)
        self.send_btn.pack(side=tk.RIGHT)
        
        self.status_var = tk.StringVar()
        self.status_lbl = ttk.Label(self, textvariable=self.status_var, font=("Helvetica", 10))
        self.status_lbl.pack(fill="x", padx=10, pady=(0, 5))
        
    def append_message(self, role, text):
        self.chat_history.config(state="normal")
        if role == "User":
            self.chat_history.insert("end", f"\nYou: {text}\n", "user")
        elif role == "Agent":
            self.chat_history.insert("end", f"\nAgent: {text}\n", "agent")
        else:
            self.chat_history.insert("end", f"\n{text}\n", "system")
        self.chat_history.config(state="disabled")
        self.chat_history.see("end")
        
    def _init_chat_session(self):
        try:
            import ai_parser
            sop_dirs = [
                os.path.join(DRIVE_ROOT, "SOPs"),
                os.path.join(DRIVE_ROOT, "SOPs", "Feedback")
            ]
            
            def progress_cb(status_msg):
                self.msg_queue.put(("status", status_msg))
                
            self.client, self.uploaded_files, self.sys_instruction = ai_parser.initialize_sop_chat(
                self.api_key, 
                sop_dirs, 
                progress_callback=progress_cb
            )
            self.history = []
            self.msg_queue.put(("system", f"Agent ready. Loaded {len(self.uploaded_files)} documents from SOPs and Feedback folders."))
        except Exception as e:
            self.msg_queue.put(("error", f"Initialization failed: {e}"))

    def _process_queue(self):
        try:
            while True:
                msg_type, content = self.msg_queue.get_nowait()
                if msg_type == "status":
                    self.status_var.set(content)
                elif msg_type == "system":
                    self.status_var.set("Ready.")
                    self.input_entry.config(state="normal")
                    self.send_btn.config(state="normal")
                    self.input_entry.focus_set()
                    self.append_message("System", content)
                elif msg_type == "error":
                    self.status_var.set("Error.")
                    self.append_message("System", content)
                elif msg_type == "agent_reply":
                    self.status_var.set("Ready.")
                    self.append_message("Agent", content)
                    self.input_entry.config(state="normal")
                    self.send_btn.config(state="normal")
                    self.input_entry.focus_set()
        except queue.Empty:
            pass
        self.after(100, self._process_queue)
        
    def send_message(self):
        msg = self.input_entry.get().strip()
        if not msg: return
        if not self.client: return
        
        self.append_message("User", msg)
        self.input_entry.delete(0, tk.END)
        self.input_entry.config(state="disabled")
        self.send_btn.config(state="disabled")
        self.status_var.set("Agent is typing...")
        
        threading.Thread(target=self._send_to_gemini, args=(msg,), daemon=True).start()
        
    def _send_to_gemini(self, msg):
        try:
            prompt = ""
            if hasattr(self, 'history') and self.history:
                for turn in self.history:
                    prompt += f"{turn['role']}: {turn['text']}\n"
                prompt += f"\nUser: {msg}"
            else:
                prompt = f"User: {msg}"
                
            contents = self.uploaded_files + [prompt]
            
            response = self.client.models.generate_content(
                model='gemini-3.6-flash',
                contents=contents,
                config={'system_instruction': self.sys_instruction, 'temperature': 0.1}
            )
            
            if not hasattr(self, 'history'):
                self.history = []
            self.history.append({"role": "User", "text": msg})
            self.history.append({"role": "Agent", "text": response.text})
            
            self.msg_queue.put(("agent_reply", response.text))
        except Exception as e:
            self.msg_queue.put(("error", f"Error sending message: {e}"))
            
    def on_close(self):
        # Cleanup files from gemini
        if self.client and self.uploaded_files:
            def cleanup():
                for f in self.uploaded_files:
                    try: self.client.files.delete(name=f.name)
                    except: pass
            threading.Thread(target=cleanup, daemon=True).start()
        self.destroy()

if __name__ == "__main__":
    import tkinterdnd2
    root = tkinterdnd2.TkinterDnD.Tk()
    app = AutomatorApp(root)
    root.mainloop()

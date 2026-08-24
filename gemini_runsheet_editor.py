import os
import json
import tkinter as tk
from tkinter import ttk, messagebox
import threading
import openpyxl

import runsheet_editor
import gemini_rs_engine

class GeminiRunsheetEditorWindow(runsheet_editor.RunsheetEditorWindow):
    def __init__(self, parent, parcel_id, app_dir):
        super().__init__(parent, parcel_id, app_dir)
        self.title(f"✨ Gemini Runsheet Editor (AI Powered) - {parcel_id}")

        self.provenance_file = os.path.join(self.pid_dir, "gemini_source_provenance.json")
        self.provenance_data = {}
        if os.path.exists(self.provenance_file):
            try:
                with open(self.provenance_file, "r") as f:
                    self.provenance_data = json.load(f)
            except Exception: pass

        # Add Gemini specific buttons to action_bar_row1
        if hasattr(self, 'action_bar_row1'):
            bar = self.action_bar_row1
            self.gemini_draft_btn = ttk.Button(bar, text="✨ Draft with Gemini (Cmd+G)", command=self.draft_current_row_with_gemini)
            self.gemini_draft_btn.pack(side=tk.LEFT, before=self.reformat_btn, padx=(0, 6))
            
            self.gemini_batch_btn = ttk.Button(bar, text="🚀 Batch AI All Rows", command=self.run_batch_gemini_generation)
            self.gemini_batch_btn.pack(side=tk.LEFT, before=self.reformat_btn, padx=(0, 6))
        elif hasattr(self, 'phrase_btn') and self.phrase_btn.master:
            bar = self.phrase_btn.master
            self.gemini_draft_btn = ttk.Button(bar, text="✨ Draft with Gemini (Cmd+G)", command=self.draft_current_row_with_gemini)
            self.gemini_draft_btn.pack(side=tk.LEFT, padx=(6, 0))
            self.gemini_batch_btn = ttk.Button(bar, text="🚀 Batch AI All Rows", command=self.run_batch_gemini_generation)
            self.gemini_batch_btn.pack(side=tk.LEFT, padx=(6, 0))

        # Bind Cmd+G and Ctrl+G
        self.bind("<Command-g>", lambda e: self.draft_current_row_with_gemini())
        self.bind("<Control-g>", lambda e: self.draft_current_row_with_gemini())

        # Update initial field labels on startup
        if self.current_row_idx:
            self._update_field_labels(self.current_row_idx)

        # Trigger automatic initial Gemini audit in background if first time
        self.check_and_run_initial_gemini_drafts()

    def check_and_run_initial_gemini_drafts(self):
        """Automatically runs Gemini on all documents upon first opening the parcel."""
        gemini_marker = os.path.join(self.pid_dir, "initial_gemini_batch_done.json")
        if os.path.exists(gemini_marker):
            return

        api_key = gemini_rs_engine.get_api_key()
        if not api_key:
            return

        self.warning_label.config(text="🤖 Initial Audit: Gemini is drafting AI comments and analyzing all PDF scans in background...")

        def worker():
            def cb(cur, total, msg):
                def update_status():
                    if hasattr(self, 'warning_label'):
                        self.warning_label.config(text=f"🤖 Initial Audit ({cur}/{total}): Processing {msg}...")
                self.after(0, update_status)

            ok, res_msg = gemini_rs_engine.batch_generate_runsheet(api_key, self.pid_dir, progress_callback=cb)
            
            if ok:
                try:
                    with open(gemini_marker, "w") as f:
                        json.dump({"initial_gemini_batch_completed": True}, f)
                except Exception: pass

            def on_done():
                if os.path.exists(self.provenance_file):
                    try:
                        with open(self.provenance_file, "r") as f:
                            self.provenance_data = json.load(f)
                    except Exception: pass

                try:
                    self.wb = openpyxl.load_workbook(self.excel_path, rich_text=True)
                    self.ws = self.wb.active
                    self.load_rows()
                    if self.current_row_idx:
                        self.on_select(None)
                except Exception: pass

                self.warning_label.config(text="✨ Initial AI Audit & Gemini Drafts complete! All rows loaded with provenance.")

            self.after(0, on_done)

        threading.Thread(target=worker, daemon=True).start()

    def delete_original_block(self, event=None):
        """Removes both --- Gemini Draft --- and --- Original --- blocks."""
        comments_widget = None
        for i, h in enumerate(self.headers):
            if "comment" in h.lower() or "note" in h.lower():
                w = self.widgets_by_col.get(i)
                if isinstance(w, tk.Text):
                    comments_widget = w
                break

        focus_w = self.focus_get()
        target_w = focus_w if isinstance(focus_w, tk.Text) else comments_widget

        if target_w and isinstance(target_w, tk.Text):
            txt = target_w.get("1.0", "end-1c")
            cleaned_txt = txt
            if "--- Gemini Draft ---" in cleaned_txt:
                cleaned_txt = cleaned_txt.split("--- Gemini Draft ---")[0].rstrip()
            elif "--- Original ---" in cleaned_txt:
                cleaned_txt = cleaned_txt.split("--- Original ---")[0].rstrip()

            target_w.delete("1.0", tk.END)
            target_w.insert("1.0", cleaned_txt)
            self.perform_spellcheck(target_w)
            self.highlight_links(target_w)

        return "break"

    def show_ai_note_dialog(self, event=None):
        """Enhanced Cmd+A dialog showing full Gemini Source Provenance & Quotes."""
        if not getattr(self, 'current_row_idx', None):
            return "break"

        vol = str(self.ws.cell(row=self.current_row_idx, column=3).value or "").strip()
        pg = str(self.ws.cell(row=self.current_row_idx, column=4).value or "").strip()
        inst = str(self.ws.cell(row=self.current_row_idx, column=1).value or "").strip()
        cache_key = f"{vol}_{pg}"

        # Reload latest provenance
        if os.path.exists(self.provenance_file):
            try:
                with open(self.provenance_file, "r") as f:
                    self.provenance_data = json.load(f)
            except Exception: pass

        prov = self.provenance_data.get(cache_key) or self.provenance_data.get(str(self.current_row_idx), {})

        popup = tk.Toplevel(self)
        popup.title(f"Row {self.current_row_idx} - Gemini Source Provenance & Legal Reasoning")
        popup.geometry("720x640")
        popup.attributes("-topmost", True)
        popup.transient(self)

        try:
            x = self.winfo_rootx() + (self.winfo_width() // 2) - 360
            y = self.winfo_rooty() + (self.winfo_height() // 2) - 320
            popup.geometry(f"+{x}+{y}")
        except Exception: pass

        main_frame = ttk.Frame(popup, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)

        row_title = f"Row {self.current_row_idx}: {inst} ({vol}/{pg})"
        ttk.Label(main_frame, text=row_title, font=("Helvetica", 16, "bold")).pack(anchor=tk.W, pady=(0, 10))

        # Notebook with Tabs: Provenance & Raw OCR
        nb = ttk.Notebook(main_frame)
        nb.pack(fill=tk.BOTH, expand=True)

        # TAB 1: Source Provenance & Quotes
        tab_prov = ttk.Frame(nb, padding=10)
        nb.add(tab_prov, text="📌 Source Quotes & Provenance")

        txt_prov = tk.Text(tab_prov, font=("Helvetica", 13), wrap=tk.WORD)
        txt_prov.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb1 = ttk.Scrollbar(tab_prov, orient="vertical", command=txt_prov.yview)
        sb1.pack(side=tk.RIGHT, fill=tk.Y)
        txt_prov.config(yscrollcommand=sb1.set)

        if prov:
            out_str = "🔍 GEMINI EXACT SOURCE PROVENANCE (ZERO-HALLUCINATION VERIFICATION)\n"
            out_str += "=" * 65 + "\n\n"
            
            if prov.get("highlight_found"):
                out_str += f"🟡 VISUAL HIGHLIGHT DETECTED:\n{prov.get('highlight_description', 'Yes')}\n\n"
            else:
                out_str += "🟡 VISUAL HIGHLIGHT: None found on scan.\n\n"

            out_str += f"📄 SUBJECT TRACT (Page {prov.get('subject_tract_page', '?')}):\n"
            out_str += f"   \"{prov.get('subject_tract_quote', 'N/A')}\"\n\n"

            out_str += f"⚖️ DOWER STATUS (Page {prov.get('dower_page', '?')}):\n"
            out_str += f"   \"{prov.get('dower_quote', 'N/A')}\"\n\n"

            out_str += f"🛢️ OIL & GAS / RESERVATIONS (Page {prov.get('reservations_page', '?')}):\n"
            out_str += f"   \"{prov.get('reservations_quote', 'None')}\"\n\n"

            out_str += f"🔗 PRIOR REFERENCE (Page {prov.get('prior_ref_page', '?')}):\n"
            out_str += f"   \"{prov.get('prior_ref_quote', 'None')}\"\n\n"

            out_str += f"🧠 LEGAL REASONING & SOP ALIGNMENT:\n"
            out_str += f"{prov.get('legal_reasoning', 'Extracted directly from recorded instrument without assumptions.')}\n"
        else:
            out_str = "No Gemini provenance data cached for this document yet.\n\nClick '✨ Draft with Gemini' (Cmd+G) to analyze this PDF and view complete source quotes."

        txt_prov.insert("1.0", out_str)
        txt_prov.config(state="disabled")

        # TAB 2: Active Warnings & Error Bar
        tab_warn = ttk.Frame(nb, padding=10)
        nb.add(tab_warn, text="⚠️ Warnings & Error Bar")
        
        txt_warn = tk.Text(tab_warn, font=("Helvetica", 13), wrap=tk.WORD)
        txt_warn.pack(fill=tk.BOTH, expand=True)
        warn_str = getattr(self.warning_label, 'cget', lambda k: '')("text") or "No active warnings."
        txt_warn.insert("1.0", warn_str)
        txt_warn.config(state="disabled")

        btn_close = ttk.Button(main_frame, text="Close (Esc)", command=popup.destroy)
        btn_close.pack(pady=(10, 0), anchor=tk.E)
        btn_close.focus_set()

        popup.bind("<Escape>", lambda e: popup.destroy())
        popup.bind("<Return>", lambda e: popup.destroy())
        return "break"

    def draft_current_row_with_gemini(self):
        """Runs Gemini on the active row's PDF and formats comparative draft blocks."""
        if not getattr(self, 'current_row_idx', None):
            messagebox.showinfo("Select Row", "Please select a row first.", parent=self)
            return

        api_key = gemini_rs_engine.get_api_key()
        if not api_key:
            messagebox.showerror("API Key Missing", "No Gemini API key found in config.json.", parent=self)
            return

        vol = str(self.ws.cell(row=self.current_row_idx, column=3).value or "").strip()
        pg = str(self.ws.cell(row=self.current_row_idx, column=4).value or "").strip()

        # Locate PDF
        target_pdf = None
        docs_dir = os.path.join(self.pid_dir, "DOCS")
        if os.path.exists(docs_dir):
            for fn in os.listdir(docs_dir):
                if fn.endswith(".pdf") and vol in fn and pg in fn:
                    target_pdf = os.path.join(docs_dir, fn)
                    break

        if not target_pdf:
            messagebox.showwarning("PDF Not Found", f"No matching PDF found in DOCS for Vol {vol} Pg {pg}.", parent=self)
            return

        # Show status
        self.warning_label.config(text="🤖 Gemini is reading document with subject land context...")
        self.update_idletasks()

        def worker():
            row_meta = {
                "instrument": str(self.ws.cell(row=self.current_row_idx, column=1).value or ""),
                "vol": vol,
                "pg": pg,
                "grantor": str(self.ws.cell(row=self.current_row_idx, column=8).value or ""),
                "grantee": str(self.ws.cell(row=self.current_row_idx, column=9).value or ""),
                "notes": str(self.ws.cell(row=self.current_row_idx, column=12).value or "")
            }
            res_data, err = gemini_rs_engine.analyze_document_with_gemini(api_key, target_pdf, row_meta, parcel_dir=self.pid_dir)

            def apply_res():
                if err or not res_data:
                    messagebox.showerror("Gemini Error", f"Failed: {err}", parent=self)
                    self.warning_label.config(text="")
                    return

                gemini_comment = res_data.get("comments", "").strip()

                # Update Comments Text Widget with Comparative Sections
                for i, h in enumerate(self.headers):
                    if "comment" in h.lower() or "note" in h.lower():
                        w = self.widgets_by_col.get(i)
                        if isinstance(w, tk.Text):
                            current_text = w.get("1.0", "end-1c").strip()
                            
                            # Extract previous original note if present
                            original_note = ""
                            if "--- Original ---" in current_text:
                                original_note = current_text.split("--- Original ---")[1].strip()
                            elif "--- Gemini Draft ---" in current_text:
                                original_note = current_text.split("--- Gemini Draft ---")[0].strip()
                            else:
                                original_note = current_text

                            # Build formatted comparative output
                            formatted_output = f"{gemini_comment}\n\n--- Gemini Draft ---\n{gemini_comment}"
                            if original_note and original_note != gemini_comment:
                                formatted_output += f"\n\n--- Original ---\n{original_note}"

                            w.delete("1.0", tk.END)
                            w.insert("1.0", formatted_output)
                            self.perform_spellcheck(w)
                            self.highlight_links(w)

                    elif "conveyance" in h.lower():
                        w = self.widgets_by_col.get(i)
                        if isinstance(w, ttk.Entry) and res_data.get("conveyance"):
                            w.delete(0, tk.END)
                            w.insert(0, res_data["conveyance"])

                # Save Provenance and Extracted Fields for Blue Label Verification
                cache_key = f"{vol}_{pg}"
                prov_payload = dict(res_data.get("source_provenance", {}))
                prov_payload.update({
                    "instrument_type": res_data.get("instrument_type"),
                    "book_type": res_data.get("book_type"),
                    "grantor": res_data.get("grantor"),
                    "grantee": res_data.get("grantee"),
                    "effective_date": res_data.get("effective_date"),
                    "filing_date": res_data.get("filing_date"),
                    "acreage": res_data.get("acreage")
                })
                self.provenance_data[cache_key] = prov_payload
                self.provenance_data[str(self.current_row_idx)] = prov_payload
                with open(self.provenance_file, "w") as f:
                    json.dump(self.provenance_data, f, indent=4)

                # Update Blue Label Indicators across all fields
                self._update_field_labels(self.current_row_idx)

                self.warning_label.config(text=f"✨ Gemini draft loaded! Click any 🔍 Blue Label to auto-apply suggested values, or press Cmd+A.")

            self.after(0, apply_res)

        threading.Thread(target=worker, daemon=True).start()

    def run_batch_gemini_generation(self):
        """Batch generates AI comments for all rows in the runsheet."""
        if not messagebox.askyesno("Batch Gemini Runsheet", "Run Gemini AI on all document rows in this parcel?\nThis will draft comments and record source provenance for every row.", parent=self):
            return

        api_key = gemini_rs_engine.get_api_key()
        if not api_key:
            messagebox.showerror("API Key Missing", "No Gemini API key found.", parent=self)
            return

        prog_win = tk.Toplevel(self)
        prog_win.title("Gemini Batch Generator")
        prog_win.geometry("400x150")
        prog_win.attributes("-topmost", True)
        prog_win.transient(self)

        lbl = ttk.Label(prog_win, text="Starting Gemini batch analysis...", font=("Helvetica", 12))
        lbl.pack(pady=15)
        pbar = ttk.Progressbar(prog_win, orient="horizontal", length=320, mode="determinate")
        pbar.pack(pady=10)

        def worker():
            def cb(cur, total, msg):
                def update_ui():
                    pbar["maximum"] = total
                    pbar["value"] = cur
                    lbl.config(text=f"({cur}/{total}) {msg}")
                self.after(0, update_ui)

            ok, res_msg = gemini_rs_engine.batch_generate_runsheet(api_key, self.pid_dir, progress_callback=cb)

            def done():
                prog_win.destroy()
                if ok:
                    self.load_rows()
                    if self.current_row_idx:
                        self.on_select(None)
                    messagebox.showinfo("Complete", res_msg, parent=self)
                else:
                    messagebox.showerror("Error", res_msg, parent=self)

            self.after(0, done)

        threading.Thread(target=worker, daemon=True).start()

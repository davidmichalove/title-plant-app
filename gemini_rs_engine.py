import os
import json
import re
import traceback
import openpyxl
import fitz # PyMuPDF
from google import genai

SOP_DIR = "/Volumes/davidlls/Gemini_SOPs"

def get_api_key():
    config_paths = [
        "/Volumes/davidlls/assignments/app/config.json",
        os.path.expanduser("~/.config/title_plant/config.json"),
        os.path.expanduser("~/.title_plant_config.json")
    ]
    for p in config_paths:
        if os.path.exists(p):
            try:
                with open(p, "r") as f:
                    data = json.load(f)
                    k = data.get("GEMINI_API_KEY", "")
                    if k:
                        return k
            except Exception:
                pass
    return ""

def get_subject_land_context(parcel_dir):
    """
    Extracts the exact target subject land description, parcel number,
    and owner from the TAX/ property cards in the parcel folder.
    """
    tax_dir = os.path.join(parcel_dir, "TAX")
    subject_info = []
    
    if os.path.exists(tax_dir):
        for fn in os.listdir(tax_dir):
            if fn.endswith(".pdf") and ("property card" in fn.lower() or "tax card" in fn.lower()):
                p = os.path.join(tax_dir, fn)
                try:
                    doc = fitz.open(p)
                    txt = ""
                    for page in doc:
                        txt += page.get_text() + "\n"
                    
                    legal_m = re.search(r'Legal\s+Description\s*\n+([^\n]+)', txt, re.IGNORECASE)
                    if legal_m:
                        subject_info.append(f"Legal Description: {legal_m.group(1).strip()}")
                    
                    p_num_m = re.search(r'Parcel\s+Number\s*\n+([^\n]+)', txt, re.IGNORECASE)
                    if p_num_m:
                        subject_info.append(f"Parcel Number: {p_num_m.group(1).strip()}")
                        
                    owner_m = re.search(r'Deeded\s+Name\s*\n+([^\n]+)', txt, re.IGNORECASE)
                    if owner_m:
                        subject_info.append(f"Current Record Owner: {owner_m.group(1).strip()}")
                        
                    dist_m = re.search(r'Tax\s+District\s*\n+([^\n]+)', txt, re.IGNORECASE)
                    if dist_m:
                        subject_info.append(f"Tax District: {dist_m.group(1).strip()}")
                        
                    if subject_info:
                        break
                except Exception:
                    pass

    if not subject_info:
        folder_name = os.path.basename(parcel_dir).replace("PID ", "").strip()
        subject_info.append(f"Target Parcel ID: {folder_name}")

    return "\n".join(subject_info)

def normalize_prior_ref_text(text, parcel_dir=None):
    if not text: return text
    
    def find_book_type(vol, pg):
        vol = str(vol).strip()
        pg = str(pg).strip()
        if parcel_dir and os.path.exists(parcel_dir):
            import glob
            for p in glob.glob(os.path.join(parcel_dir, "**", "*.pdf"), recursive=True):
                fname = os.path.basename(p).upper()
                m = re.search(r'\b(DR|OR|MR|LR|PR|WR|MISC|PL)\s*[-_ ]\s*' + re.escape(vol) + r'\s*[-_ /]\s*' + re.escape(pg) + r'\b', fname)
                if m: return m.group(1).upper()
                m2 = re.search(r'\b(DR|OR|MR|LR|PR|WR|MISC|PL)\s*' + re.escape(vol) + r'\s*[-_ /]\s*' + re.escape(pg) + r'\b', fname)
                if m2: return m2.group(1).upper()
        try:
            return "DR" if int(vol) <= 805 else "OR"
        except:
            return "DR"

    def repl_prior_ref(m):
        full_ref = m.group(1)
        m_std = re.search(r'\b(DR|OR|MR|LR|PR|WR|MISC)\s+(\d+)[/-](\d+)\b', full_ref, re.IGNORECASE)
        if m_std:
            return f"Prior Ref: {m_std.group(1).upper()} {m_std.group(2)}/{m_std.group(3)}"
            
        m_named = re.search(r'\b(Deed\s*(?:Book|Record|Vol)?|Official\s*Records?|Mortgage\s*(?:Book|Record|Vol)?|Lease\s*(?:Book|Record|Vol)?)\s*[:.]?\s*(?:Vol(?:ume)?\.?\s*)?(\d+)[,\s]+(?:Page|Pg|p\.?)\s*(\d+)\b', full_ref, re.IGNORECASE)
        if m_named:
            bname = m_named.group(1).lower()
            vol = m_named.group(2)
            pg = m_named.group(3)
            btype = "DR"
            if "official" in bname: btype = "OR"
            elif "mortgage" in bname: btype = "MR"
            elif "lease" in bname: btype = "LR"
            return f"Prior Ref: {btype} {vol}/{pg}"

        m_vol_pg = re.search(r'\b(?:Vol(?:ume)?\.?|Bk\.?|Book)\s*(\d+)[,\s]+(?:Page|Pg|p\.?)\s*(\d+)\b', full_ref, re.IGNORECASE)
        if m_vol_pg:
            vol = m_vol_pg.group(1)
            pg = m_vol_pg.group(2)
            btype = find_book_type(vol, pg)
            return f"Prior Ref: {btype} {vol}/{pg}"

        m_bare = re.search(r'\b(\d{1,4})[/-](\d{1,4})\b', full_ref)
        if m_bare:
            vol = m_bare.group(1)
            pg = m_bare.group(2)
            btype = find_book_type(vol, pg)
            return f"Prior Ref: {btype} {vol}/{pg}"

        return m.group(0)

    pattern = r'(?:Prior\s*(?:deed\s*)?references?|Prior\s*Ref)\s*[:.]?\s*([^\n\.;]+)'
    return re.sub(pattern, repl_prior_ref, text, flags=re.IGNORECASE)

def analyze_document_with_gemini(api_key, pdf_path, row_meta=None, parcel_dir=None, model="gemini-3.6-flash"):
    """
    Analyzes a single document PDF with Gemini vision.
    Generates an ultra-succinct, pared-down, client-ready Runsheet Comment (2-4 lines max)
    focused strictly on the target subject land, with mandatory Dower on Deeds and Mortgages.
    """
    if not api_key:
        return None, "No Gemini API key provided."
    if not os.path.exists(pdf_path):
        return None, f"PDF file not found: {pdf_path}"

    row_meta = row_meta or {}
    existing_notes = str(row_meta.get('notes', '')).strip()
    meta_info = f"""
DRAFT RUNSHEET METADATA & PRE-EXISTING COMMENTS:
- Preliminary Instrument: {row_meta.get('instrument', '')}
- Book/Vol/Pg: {row_meta.get('vol', '')}/{row_meta.get('pg', '')}
- Grantor: {row_meta.get('grantor', '')}
- Grantee: {row_meta.get('grantee', '')}
- Existing Runsheet Comments: "{existing_notes}"

CRITICAL INSTRUCTION ON EXISTING COMMENTS:
Carefully read and take into account the existing comments above (which originate from the initial runsheet, Phase 1 transformations, or title examiner notes).
- Incorporate and cross-verify all accurate factual details from the existing comment (e.g. probate case numbers, release references, death dates, specific lot/tract designations, or loan terms).
- Refine and distill them into the strict, succinct 2-4 line SOP format.
"""

    subject_land = ""
    if parcel_dir:
        subject_land = get_subject_land_context(parcel_dir)

    system_prompt = f"""
You are an expert Ohio Title Examiner following strict Gulfport Energy & Belmont County Runsheet standards.
Your output comment MUST BE EXTREMELY SUCCINCT, PARED-DOWN, AND BRIEF (2 to 4 lines maximum).

TARGET SUBJECT LAND (TARGET PARCEL):
{subject_land}
(CRITICAL INSTRUCTION: Always focus specifically on the tract/lot matching the target subject property above. If a deed conveys multiple lots or tracts, state the conveyed interest for the target subject property, or note "Undivided X interest" or "ARTI").

REAL RUNSHEET COMMENT EXAMPLES FROM CLIENT SOPS:
- Standard Warranty Deed:
  ARTI
  No dower mentioned
  Prior Ref: DR 372/194

- Warranty Deed with Dower:
  ARTI
  Dower released
  Prior Ref: DR 472/690

- Fractional Deed (e.g. 1/2 or 1/4 interest):
  Undivided 1/2 interest
  Dower released
  Prior Ref: DR 362/222

- Deed with Mineral / Oil & Gas Exception:
  ARTI
  [[BOLD_START]]EXCEPTING all oil and gas rights.[[BOLD_END]]
  No dower mentioned
  Prior Ref: DR 500/100

- Mortgage (with Dower):
  Amount: $50,000.00
  Maturity Date: 05/01/2015
  Dower released
  Release: OR 234/567

- Mortgage (without Dower / Single):
  Amount: $12,946.66
  Maturity Date: Not stated
  No dower mentioned
  Release: MR 107/710

- Release of Mortgage:
  Releases mortgage recorded in MR 988/778

- Oil & Gas Lease / Memo:
  PT 5 yrs
  OTE 5 yrs
  Royalty: Not stated
  Prior Ref: OR 145/891

- Affidavit / Probate:
  Affiant is the daughter of Charles W. McLaughlin (DOD 10/26/2023).
  Terminates life estate reserved at OR 145/891.
  Death Certificate Attached.
  Prior Ref: OR 145/891

STRICT RULES FOR THE "comments" FIELD:
1. EXTREME BREVITY: Do NOT dump legal descriptions, lot numbers lists, or narrative paragraphs. Keep the "comments" field strictly to 2-4 lines using the exact keywords above.
2. SUBJECT LAND FOCUS: Focus exclusively on the target subject land provided above.
3. DOWER IS MANDATORY FOR BOTH DEEDS AND MORTGAGES:
   - State "Dower released" if spouse joins/releases dower/homestead.
   - State "No dower mentioned" if mortgagor/grantor is single, unmarried, or no dower clause is present.
   - State "Dower not stated" if unspecified.
4. MORTGAGES: Always include the 4 standard components:
   Amount: $X
   Maturity Date: MM/DD/YYYY (or "Not stated")
   Dower released (or "No dower mentioned")
   Release: [Book] [Vol]/[Pg] (if released, or omit if unreleased)
5. CONVEYANCE: Use "ARTI" for fee simple conveyance, or state fractional interest (e.g. "Undivided 1/2 interest").
6. OIL & GAS / RESERVATIONS: If O&G, coal, or life estates are reserved/excepted, wrap that single brief line in [[BOLD_START]]...[[BOLD_END]].
7. PRIOR REF: Strictly "Prior Ref: [Book] [Vol]/[Pg]" or case number.
8. SOURCE PROVENANCE (ZERO HALLUCINATIONS): Put all verbatim quotes, exact page numbers, visual highlight descriptions, and legal reasoning inside "source_provenance".

{meta_info}

Return a STRICT JSON object with this exact structure:
{{
  "instrument_type": "Warranty Deed",
  "book_type": "DR",
  "volume": "391",
  "page": "183",
  "effective_date": "MM/DD/YYYY",
  "filing_date": "MM/DD/YYYY",
  "grantor": "...",
  "grantee": "...",
  "conveyance": "Fee Simple",
  "comments": "ARTI\\nDower released\\nPrior Ref: DR 362/222",
  "source_provenance": {{
    "subject_tract_quote": "Verbatim quote of tract from PDF",
    "subject_tract_page": 1,
    "highlight_found": true,
    "highlight_description": "Description of what was highlighted on page X",
    "dower_quote": "Verbatim quote or 'N/A - Grantor executed as single'",
    "dower_page": 2,
    "reservations_quote": "Verbatim quote of reservation or 'None'",
    "reservations_page": 1,
    "prior_ref_quote": "Verbatim quote of prior deed reference or 'None'",
    "prior_ref_page": 1,
    "legal_reasoning": "Step-by-step reasoning for the final comment"
  }}
}}
"""

    client = genai.Client(api_key=api_key)
    sample_file = None
    try:
        sample_file = client.files.upload(file=pdf_path)
        response = client.models.generate_content(
            model=model,
            contents=[sample_file, system_prompt],
            config={"response_mime_type": "application/json"}
        )
        data = json.loads(response.text.strip())
        return data, None
    except Exception as e:
        return None, f"Gemini Error: {str(e)}\n{traceback.format_exc()}"
    finally:
        if sample_file:
            try: client.files.delete(name=sample_file.name)
            except Exception: pass

def batch_generate_runsheet(api_key, parcel_dir, progress_callback=None, model="gemini-3.6-flash"):
    """
    Processes all documents in parcel_dir/DOCS and builds/updates the Gemini Runsheet.
    """
    if not api_key:
        return False, "No API key."

    docs_dir = os.path.join(parcel_dir, "DOCS")
    if not os.path.exists(docs_dir):
        return False, f"DOCS directory not found in {parcel_dir}"

    import glob
    rs_files = glob.glob(os.path.join(parcel_dir, "*RS*.xlsx"))
    if not rs_files:
        return False, "No Runsheet Excel found in parcel directory."

    rs_path = rs_files[0]
    for rf in rs_files:
        if "BLANK" not in rf and "Backup" not in rf and "copy" not in rf:
            rs_path = rf
            break

    provenance_file = os.path.join(parcel_dir, "gemini_source_provenance.json")
    provenance_data = {}
    if os.path.exists(provenance_file):
        try:
            with open(provenance_file, "r") as f:
                provenance_data = json.load(f)
        except Exception: pass

    wb = openpyxl.load_workbook(rs_path)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[2]]

    col_map = {}
    for idx, h in enumerate(headers, 1):
        hl = h.lower()
        if "instrument" in hl and "type" in hl: col_map["inst"] = idx
        elif "book" in hl: col_map["book"] = idx
        elif "vol" in hl: col_map["vol"] = idx
        elif "page" in hl or "pg" in hl: col_map["pg"] = idx
        elif "grantor" in hl: col_map["grantor"] = idx
        elif "grantee" in hl: col_map["grantee"] = idx
        elif "comment" in hl or "note" in hl: col_map["comments"] = idx
        elif "effective" in hl: col_map["eff_date"] = idx
        elif "filing" in hl: col_map["filing_date"] = idx
        elif "conveyance" in hl: col_map["conveyance"] = idx
        elif "acreage" in hl: col_map["acreage"] = idx

    rows_to_process = []
    for r in range(3, ws.max_row + 1):
        vol = str(ws.cell(row=r, column=col_map.get("vol", 3)).value or "").strip()
        pg = str(ws.cell(row=r, column=col_map.get("pg", 4)).value or "").strip()
        if vol or pg:
            rows_to_process.append((r, vol, pg))

    total = len(rows_to_process)
    for i, (r_idx, vol, pg) in enumerate(rows_to_process, 1):
        if progress_callback:
            progress_callback(i, total, f"Row {r_idx} (Vol {vol} Pg {pg})")

        target_pdf = None
        for fn in os.listdir(docs_dir):
            if fn.endswith(".pdf") and vol in fn and pg in fn:
                target_pdf = os.path.join(docs_dir, fn)
                break

        if not target_pdf:
            continue

        row_meta = {
            "instrument": str(ws.cell(row=r_idx, column=col_map.get("inst", 1)).value or ""),
            "vol": vol,
            "pg": pg,
            "grantor": str(ws.cell(row=r_idx, column=col_map.get("grantor", 8)).value or ""),
            "grantee": str(ws.cell(row=r_idx, column=col_map.get("grantee", 9)).value or ""),
            "notes": str(ws.cell(row=r_idx, column=col_map.get("comments", 12)).value or "")
        }

        res_data, err = analyze_document_with_gemini(api_key, target_pdf, row_meta, parcel_dir=parcel_dir, model=model)
        if res_data and isinstance(res_data, dict):
            gemini_comm = normalize_prior_ref_text(res_data.get("comments", "").strip(), parcel_dir=parcel_dir)
            if "comments" in col_map and gemini_comm:
                raw_existing = str(ws.cell(row=r_idx, column=col_map["comments"]).value or "").strip()
                
                original_note = ""
                if "--- Original ---" in raw_existing:
                    original_note = raw_existing.split("--- Original ---")[1].strip()
                elif "--- Gemini Draft ---" in raw_existing:
                    original_note = raw_existing.split("--- Gemini Draft ---")[0].strip()
                else:
                    original_note = raw_existing

                formatted_output = f"{gemini_comm}\n\n--- Gemini Draft ---\n{gemini_comm}"
                if original_note and original_note != gemini_comm:
                    formatted_output += f"\n\n--- Original ---\n{original_note}"

                ws.cell(row=r_idx, column=col_map["comments"]).value = formatted_output

            if "conveyance" in col_map and res_data.get("conveyance"):
                ws.cell(row=r_idx, column=col_map["conveyance"]).value = res_data["conveyance"]
            if "grantor" in col_map and res_data.get("grantor") and not ws.cell(row=r_idx, column=col_map["grantor"]).value:
                ws.cell(row=r_idx, column=col_map["grantor"]).value = res_data["grantor"]
            if "grantee" in col_map and res_data.get("grantee") and not ws.cell(row=r_idx, column=col_map["grantee"]).value:
                ws.cell(row=r_idx, column=col_map["grantee"]).value = res_data["grantee"]

            cache_key = f"{vol}_{pg}"
            prov_payload = dict(res_data.get("source_provenance", {}))
            prov_payload.update({
                "instrument_type": res_data.get("instrument_type"),
                "book_type": res_data.get("book_type"),
                "grantor": res_data.get("grantor"),
                "grantee": res_data.get("grantee"),
                "effective_date": res_data.get("effective_date"),
                "filing_date": res_data.get("filing_date"),
                "acreage": res_data.get("acreage")
            })
            provenance_data[cache_key] = prov_payload
            provenance_data[str(r_idx)] = prov_payload

    wb.save(rs_path)
    with open(provenance_file, "w") as f:
        json.dump(provenance_data, f, indent=4)

    return True, f"Successfully processed {total} rows with Gemini."

import os
import glob
import re
import datetime
import openpyxl

BASE_APP_DIR = os.path.dirname(os.path.abspath(__file__))
SHAPE_FILE_DBF = os.path.join(BASE_APP_DIR, "shape_files", "Belmont_County_Parcels.dbf")

def normalize_parcel_num(p):
    if not p: return ""
    p = p.replace("TEST", "").replace("PID", "").strip()
    m = re.match(r"^(\d+)-(\d+)(\..*)?$", p)
    if m:
        dist = m.group(1)
        num = m.group(2).zfill(5)
        suf = m.group(3) if m.group(3) else ".000"
        return f"{dist}-{num}{suf}"
    return p

def parse_date(date_val):
    if not date_val:
        return None
    if isinstance(date_val, (datetime.datetime, datetime.date)):
        return date_val.strftime("%m/%d/%Y")
    
    s = str(date_val).strip()
    s = re.sub(r'^(DOD|Dated|Effective|Filed|Recorded)\s*', '', s, flags=re.IGNORECASE).strip()
    
    for fmt in ["%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%m/%d/%y", "%m-%d-%y"]:
        try:
            dt = datetime.datetime.strptime(s, fmt)
            if dt.year > 2050:
                dt = dt.replace(year=dt.year - 100)
            return dt.strftime("%m/%d/%Y")
        except ValueError:
            pass
            
    m = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})', s)
    if m:
        mo, day, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yr < 50: yr += 2000
        elif yr < 100: yr += 1900
        try:
            dt = datetime.datetime(yr, mo, day)
            return dt.strftime("%m/%d/%Y")
        except: pass
    return None

def calculate_quarter_section(parcel_num):
    norm_p = normalize_parcel_num(parcel_num)
    try:
        shp_path = os.path.join(BASE_APP_DIR, "shape_files", "Belmont_County_Parcels.shp")
        plss_shp = "/Volumes/davidlls/various_GIS_shapefiles/OH-CADNSDI-v2_SPSNAD83/PLSSFirstDivision.shp"
        if os.path.exists(shp_path):
            import geopandas as gpd
            from shapely.geometry import box
            parcels_gdf = gpd.read_file(shp_path)
            p_matches = parcels_gdf[parcels_gdf["parcel_no"] == norm_p]
            if len(p_matches) == 0:
                p_matches = parcels_gdf[parcels_gdf["parcel_no"].str.contains(norm_p.replace("-",""), na=False)]
                
            if len(p_matches) > 0:
                p_row = p_matches.iloc[0]
                desc_str = str(p_row.get("desc_", "") or "")
                m_q = re.search(r'\b(NW|NE|SW|SE)\b', desc_str, re.IGNORECASE)
                if m_q:
                    return f"{m_q.group(1).upper()}4"
                    
                p_geom = p_row.geometry
                if p_geom is not None and os.path.exists(plss_shp):
                    plss_gdf = gpd.read_file(plss_shp)
                    m_secs = plss_gdf[plss_gdf.intersects(p_geom)]
                    for _, sec_row in m_secs.iterrows():
                        minx, miny, maxx, maxy = sec_row.geometry.bounds
                        midx = (minx + maxx) / 2.0
                        midy = (miny + maxy) / 2.0
                        q_boxes = {
                            "NW": box(minx, midy, midx, maxy),
                            "NE": box(midx, midy, maxx, maxy),
                            "SW": box(minx, miny, midx, midy),
                            "SE": box(midx, miny, maxx, midy),
                        }
                        overlaps = {}
                        for q_code, q_box in q_boxes.items():
                            if p_geom.intersects(q_box):
                                overlaps[q_code] = p_geom.intersection(q_box).area
                        if overlaps:
                            best_q = max(overlaps, key=overlaps.get)
                            return f"{best_q}4"
    except Exception as e:
        print(f"Quarter calculation error: {e}")
    return "SE4"

def get_gis_owner_info(parcel_num):
    if not parcel_num or not os.path.exists(SHAPE_FILE_DBF):
        return {}
        
    norm_p = normalize_parcel_num(parcel_num)
    try:
        import struct
        with open(SHAPE_FILE_DBF, "rb") as f:
            header = f.read(32)
            num_records, header_len, record_len = struct.unpack("<IHH", header[4:12])
            fields = []
            while True:
                field_desc = f.read(32)
                if field_desc[0] == 0x0D:
                    break
                name = field_desc[:11].replace(b"\x00", b"").decode("ascii", errors="ignore")
                length = field_desc[16]
                fields.append((name, length))
            
            f.seek(header_len)
            for _ in range(num_records):
                rec_bytes = f.read(record_len)
                if not rec_bytes: break
                offset = 1
                rec = {}
                for name, length in fields:
                    val = rec_bytes[offset:offset+length].decode("ascii", errors="ignore").strip()
                    rec[name] = val
                    offset += length
                
                p_no = rec.get("parcel_no", "").strip()
                if p_no == norm_p or (len(norm_p) > 5 and norm_p in p_no) or (p_no.replace("-","") == norm_p.replace("-","")):
                    add1 = rec.get("ownadd1", "").strip()
                    add2 = rec.get("ownadd2", "").strip()
                    add3 = rec.get("ownadd3", "").strip()
                    zip_code = rec.get("zip", "").strip()
                    
                    addr_lines = []
                    if add1 and add1.upper() != "NAN": addr_lines.append(add1.title())
                    if add2 and add2.upper() != "NAN": addr_lines.append(add2.title())
                    if add3 and add3.upper() != "NAN": 
                        if zip_code and zip_code not in add3:
                            addr_lines.append(f"{add3.title()} {zip_code}")
                        else:
                            addr_lines.append(add3.title())
                    elif zip_code:
                        addr_lines.append(zip_code)
                        
                    return {
                        "name": rec.get("name", "").strip().title(),
                        "lname": rec.get("lname", "").strip().title(),
                        "fname": rec.get("fname", "").strip().title(),
                        "address_lines": addr_lines,
                        "raw": rec
                    }
    except Exception as e:
        print(f"GIS lookup error: {e}")
    return {}

def format_encumbrance_short(r):
    btype = str(r.get("btype", "")).strip()
    vol = str(r.get("vol", "")).strip()
    pg = str(r.get("pg", "")).strip()
    inst = str(r.get("inst_num", "")).strip()
    if vol and pg and vol.lower() != "na" and pg.lower() != "na":
        return f"{btype} {vol}/{pg}"
    elif vol and vol.lower() != "na":
        return f"{btype} {vol}"
    elif inst and inst.lower() != "na":
        return f"Inst #{inst}"
    else:
        return str(r.get("itype", "Doc")).strip()

def parse_lease_details(r):
    comments = r.get("comments", "")
    eff_dt_str = r.get("eff_dt", "")
    
    term_match = re.search(r"(\d+)\s*(?:yr|year|years)\s*(?:pt|primary\s*term)?", comments, re.IGNORECASE)
    term_years = int(term_match.group(1)) if term_match else 5
    
    option_match = re.search(r"(?:option\s*(?:to\s*renew)?|renewal)[\s:]*(\d+)\s*(?:yr|year|years)", comments, re.IGNORECASE)
    option_years = int(option_match.group(1)) if option_match else 0
    
    term_desc = f"{term_years}yr"
    if option_years:
        term_desc += f" + {option_years}yr Option"
        
    exp_date_str = "HBP"
    if eff_dt_str:
        try:
            dt = datetime.datetime.strptime(eff_dt_str, "%m/%d/%Y")
            total_years = term_years + option_years
            exp_dt = dt.replace(year=dt.year + total_years)
            exp_date_str = exp_dt.strftime("%m/%d/%Y")
        except: pass
        
    royalty_match = re.search(r"royalty[\s:]*([0-9\./%]+|unknown|1/8th?)", comments, re.IGNORECASE)
    royalty_str = royalty_match.group(1).strip() if royalty_match else "20%"
    
    pugh = "None"
    pugh_m = re.search(r"pugh[\s:]*(.*?)(?:\n|$)", comments, re.IGNORECASE)
    if pugh_m and "none" not in pugh_m.group(1).lower():
        pugh = pugh_m.group(1).strip()
        
    depth = "None"
    depth_m = re.search(r"depth[\s:]*(.*?)(?:\n|$)", comments, re.IGNORECASE)
    if depth_m and "none" not in depth_m.group(1).lower():
        depth = depth_m.group(1).strip()
        
    pooling = "Right to Unitize, 640 Pooling, ETC"
    pool_m = re.search(r"pooling[\s:]*(.*?)(?:\n|$)", comments, re.IGNORECASE)
    if pool_m:
        pooling = pool_m.group(1).strip()
        
    btype = r.get("btype", "")
    vol = r.get("vol", "")
    pg = r.get("pg", "")
    bk_pg = f"{btype} {vol}/{pg}"
    is_memo = "memo" in r.get("itype", "").lower() or "memo" in comments.lower()
    
    return {
        "row": r,
        "lessor": r.get("grantor", ""),
        "lessee": r.get("grantee", ""),
        "bk_pg": bk_pg,
        "inst_num": r.get("inst_num", ""),
        "is_memo": is_memo,
        "eff_date": eff_dt_str,
        "term": term_desc,
        "exp_date": exp_date_str,
        "royalty": royalty_str,
        "pugh": pugh,
        "depth": depth,
        "pooling": pooling
    }

class ORCompilerEngine:
    @classmethod
    def compile_data(cls, pid_dir, parcel_num=None, rs_path=None):
        if not parcel_num and pid_dir:
            m = re.search(r'PID\s*([0-9\-\.]+)', os.path.basename(pid_dir), re.IGNORECASE)
            if m: parcel_num = m.group(1).strip()
            
        if not rs_path:
            rs_files = glob.glob(os.path.join(pid_dir, "*RS*.xlsx"))
            valid_rs = [f for f in rs_files if not os.path.basename(f).startswith("~") and not os.path.basename(f).startswith("._") and "backup" not in f.lower() and "blank" not in f.lower()]
            if not valid_rs:
                return None
            rs_path = valid_rs[0]
            
        if not os.path.exists(rs_path):
            return None
            
        wb = openpyxl.load_workbook(rs_path, data_only=True)
        ws = wb.active
        
        # 1. Find Header Row dynamically
        header_row_idx = 1
        col_map = {}
        for r in range(1, min(6, ws.max_row + 1)):
            row_texts = [str(ws.cell(r, c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]
            if any("instrument" in t for t in row_texts) or any("grantor" in t for t in row_texts):
                header_row_idx = r
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(r, c).value or '').strip().lower()
                    if "instrument type" in val or val == "instrument": col_map["itype"] = c
                    elif "book type" in val or "book" in val: col_map["btype"] = c
                    elif "volume" in val or "vol" in val: col_map["vol"] = c
                    elif "page" in val or "pg" in val: col_map["pg"] = c
                    elif "instrument number" in val or "inst" in val: col_map["inst_num"] = c
                    elif "effective date" in val or "effective" in val or "dated" in val: col_map["eff_dt"] = c
                    elif "filing date" in val or "filed" in val or "recording date" in val: col_map["file_dt"] = c
                    elif "grantor" in val: col_map["grantor"] = c
                    elif "grantee" in val: col_map["grantee"] = c
                    elif "comment" in val or "note" in val: col_map["comments"] = c
                break
                
        # Default fallbacks
        if "itype" not in col_map: col_map["itype"] = 1
        if "btype" not in col_map: col_map["btype"] = 2
        if "vol" not in col_map: col_map["vol"] = 3
        if "pg" not in col_map: col_map["pg"] = 4
        if "inst_num" not in col_map: col_map["inst_num"] = 5
        if "eff_dt" not in col_map: col_map["eff_dt"] = 6
        if "file_dt" not in col_map: col_map["file_dt"] = 7
        if "grantor" not in col_map: col_map["grantor"] = 8
        if "grantee" not in col_map: col_map["grantee"] = 9
        if "comments" not in col_map: col_map["comments"] = 12
        
        # 2. Extract Data Rows
        rows = []
        for r in range(header_row_idx + 1, ws.max_row + 1):
            itype = str(ws.cell(r, col_map["itype"]).value or '').strip()
            btype = str(ws.cell(r, col_map["btype"]).value or '').strip()
            vol = str(ws.cell(r, col_map["vol"]).value or '').strip()
            pg = str(ws.cell(r, col_map["pg"]).value or '').strip()
            inst_num = str(ws.cell(r, col_map["inst_num"]).value or '').strip()
            eff_dt = ws.cell(r, col_map["eff_dt"]).value
            file_dt = ws.cell(r, col_map["file_dt"]).value
            grantor = str(ws.cell(r, col_map["grantor"]).value or '').strip()
            grantee = str(ws.cell(r, col_map["grantee"]).value or '').strip()
            comments = str(ws.cell(r, col_map["comments"]).value or '').strip()
            
            if itype.lower() == "none": itype = ""
            if btype.lower() == "none": btype = ""
            if vol.lower() == "none": vol = ""
            if pg.lower() == "none": pg = ""
            if grantor.lower() == "none": grantor = ""
            if grantee.lower() == "none": grantee = ""
            
            if not any([itype, btype, vol, pg, grantor, grantee]):
                continue
                
            parsed_eff = parse_date(eff_dt)
            parsed_file = parse_date(file_dt)
            
            rows.append({
                "row_idx": r,
                "itype": itype,
                "btype": btype,
                "vol": vol,
                "pg": pg,
                "inst_num": inst_num,
                "eff_dt_raw": str(eff_dt or ''),
                "eff_dt": parsed_eff,
                "file_dt": parsed_file,
                "grantor": grantor,
                "grantee": grantee,
                "comments": comments
            })
            
        if not rows:
            return None
            
        # 3. Date Range
        all_dates = []
        for r in rows:
            if r["eff_dt"]:
                try:
                    dt_obj = datetime.datetime.strptime(r["eff_dt"], "%m/%d/%Y")
                    all_dates.append((dt_obj, r["eff_dt"]))
                except: pass
                
        all_dates.sort(key=lambda x: x[0])
        earliest_date = all_dates[0][1] if all_dates else "01/01/1900"
        today_date = datetime.date.today().strftime("%m/%d/%Y")
        
        # 4. Vesting Deed & Surface Owner
        vesting_row = None
        deed_keywords = ["deed", "survivorship", "warranty", "quit claim", "fiduciary", "sheriff", "affidavit for transfer", "certificate of transfer"]
        for r in reversed(rows):
            it_l = r["itype"].lower()
            if any(k in it_l for k in deed_keywords):
                vesting_row = r
                break
                
        gis_info = get_gis_owner_info(parcel_num)
        
        surface_owner_name = ""
        surface_tenancy = ""
        # Default Acquired Year to (2026) / current year
        acquired_year = f"({datetime.date.today().year})"
        
        if vesting_row:
            raw_grantee = vesting_row["grantee"]
            m_ten = re.search(r',\s*(husband and wife.*|for their joint lives.*|as survivorship tenants.*|a single person.*|unmarried.*|widow.*|a corporation.*|an ohio.*|a delaware.*)', raw_grantee, re.IGNORECASE)
            if m_ten:
                surface_tenancy = m_ten.group(1).strip().upper()
                surface_owner_name = raw_grantee[:m_ten.start()].strip().upper()
            else:
                surface_owner_name = raw_grantee.strip().upper()
        elif gis_info.get("name"):
            surface_owner_name = gis_info["name"].upper()
            
        address_lines = gis_info.get("address_lines", [])
        if not address_lines:
            address_lines = ["Belmont County, OH"]
            
        # 5. Quarter Section Calculation
        qtr_val = calculate_quarter_section(parcel_num)
        
        # 6. Encumbrances Scanning (Clean Short Format: "BookType Vol/Pg")
        # A) Easements
        easement_rows = []
        for r in rows:
            it_l = r["itype"].lower()
            if any(k in it_l for k in ["right of way", "easement", "pipeline", "powerline", "electric", "highway", "utility", "telephone", "roadway"]):
                summary = format_encumbrance_short(r)
                easement_rows.append({"row": r, "summary": summary, "included": True})
                
        # B) Oil & Gas Leases
        lease_rows = []
        parsed_leases = []
        for r in rows:
            it_l = r["itype"].lower()
            if any(k in it_l for k in ["lease", "memorandum of lease", "oil and gas lease", "ratification of oil", "addendum to and ratification"]) and not any(k in it_l for k in ["release", "surrender", "assignment"]):
                summary = format_encumbrance_short(r)
                lease_rows.append({"row": r, "summary": summary, "included": True, "status": "Active"})
                parsed_leases.append(parse_lease_details(r))
                
        # C) Mortgages
        mortgage_rows = []
        satisfactions = []
        for r in rows:
            it_l = r["itype"].lower()
            if any(k in it_l for k in ["satisfaction", "release of mortgage", "certificate of satisfaction", "discharge of mortgage"]):
                satisfactions.append(r)
                
        for r in rows:
            it_l = r["itype"].lower()
            if "mortgage" in it_l and not any(k in it_l for k in ["satisfaction", "release", "assignment", "modification"]):
                is_satisfied = False
                ref_target = f"{r['vol']}-{r['pg']}"
                for sat in satisfactions:
                    if (r["vol"] and r["vol"] in sat["comments"]) or (r["pg"] and r["pg"] in sat["comments"]) or (ref_target in sat["comments"]):
                        is_satisfied = True
                        break
                    if sat["grantee"] and r["grantor"] and sat["row_idx"] > r["row_idx"]:
                        b_m = sat["grantee"].split()[0].lower() in r["grantor"].lower()
                        if b_m:
                            is_satisfied = True
                            
                summary = format_encumbrance_short(r)
                mortgage_rows.append({
                    "row": r,
                    "summary": summary,
                    "is_satisfied": is_satisfied,
                    "included": not is_satisfied,
                    "status": "Satisfied" if is_satisfied else "Unreleased"
                })
                
        # Primary Lease for Schedule A (prioritizing base lease / memo over ratifications)
        base_leases = [l for l in parsed_leases if not any(k in l["row"]["itype"].lower() for k in ["amendment", "ratification", "addendum", "assignment"])]
        primary_lease = base_leases[-1] if base_leases else (parsed_leases[-1] if parsed_leases else None)
        
        return {
            "pid_dir": pid_dir,
            "rs_path": rs_path,
            "parcel_num": parcel_num,
            "from_date": earliest_date,
            "to_date": today_date,
            "qtr_val": qtr_val,
            "vesting_deed": vesting_row,
            "surface_owner": {
                "name": surface_owner_name,
                "tenancy": surface_tenancy,
                "address_lines": address_lines,
                "year": acquired_year,
                "interest": "1"
            },
            "mineral_owner": {
                "name": surface_owner_name,
                "tenancy": surface_tenancy,
                "address_lines": address_lines,
                "year": acquired_year,
                "interest": "1"
            },
            "easements": easement_rows,
            "leases": lease_rows,
            "parsed_leases": parsed_leases,
            "primary_lease": primary_lease,
            "mortgages": mortgage_rows,
            "sole_mineral_owner": True,
            "leasehold_mode": "populate" if primary_lease else "open_of_record",
            "delete_notes": True
        }

    @classmethod
    def apply_to_excel(cls, or_path, data):
        if not os.path.exists(or_path):
            raise FileNotFoundError(f"Ownership report file not found: {or_path}")
            
        wb = openpyxl.load_workbook(or_path)
        ws = wb.active
        
        # 1. Update Cell B3 (Caption Paragraph with QTR & Vesting Deed Info)
        b3_val = ws["B3"].value
        if b3_val and isinstance(b3_val, str):
            vd = data.get("vesting_deed")
            qtr = data.get("qtr_val", "SE4")
            
            btype_full = "Deed Records"
            if vd:
                bt = str(vd.get("btype", "")).upper()
                if "OR" in bt or "OFFICIAL" in bt: btype_full = "Official Records"
                elif "MR" in bt or "MORTGAGE" in bt: btype_full = "Mortgage Records"
                else: btype_full = "Deed Records"
                
            new_b3 = b3_val
            new_b3 = new_b3.replace("QUARTER CALL", qtr).replace("<QTR>", qtr).replace("<QUARTER>", qtr).replace("<QTR_CALL>", qtr)
            
            if vd:
                v_itype = str(vd.get("itype", "Deed")).strip()
                v_grantor = str(vd.get("grantor", "")).strip()
                v_grantee = str(vd.get("grantee", "")).strip()
                v_eff_dt = str(vd.get("eff_dt") or vd.get("eff_dt_raw") or "XX/XX/XXXX").strip()
                v_vol = str(vd.get("vol", "XX")).strip()
                v_pg = str(vd.get("pg", "XX")).strip()
                
                new_b3 = new_b3.replace("<INST_TYPE>", v_itype).replace("Instrument Type", v_itype)
                new_b3 = new_b3.replace("<GRANTOR>", v_grantor).replace("from Grantor to Grantee", f"from {v_grantor} to {v_grantee}").replace("Grantor", v_grantor)
                new_b3 = new_b3.replace("<GRANTEE>", v_grantee).replace("Grantee", v_grantee)
                new_b3 = new_b3.replace("effective date XX/XX/XXXX", f"effective date {v_eff_dt}").replace("<EFF_DATE>", v_eff_dt)
                
                new_b3 = new_b3.replace("Volume <VOL>", f"Volume {v_vol}").replace("Vol <VOL>", f"Vol {v_vol}").replace("Volume XX", f"Volume {v_vol}").replace("<VOL>", f"Volume {v_vol}")
                new_b3 = new_b3.replace("Page <PG>", f"Page {v_pg}").replace("Pg <PG>", f"Page {v_pg}").replace("Page XX", f"Page {v_pg}").replace("<PG>", f"Page {v_pg}")
                new_b3 = new_b3.replace("Volume Volume", "Volume").replace("Page Page", "Page")
                
                new_b3 = new_b3.replace("<REC_TYPE>", btype_full).replace("Record Type Records", btype_full).replace("Record Type", btype_full.replace(" Records", ""))
                
            ws["B3"] = new_b3

        # 2. Date Range in cell B62 (or finding "RECORDS EXAMINED FROM AND TO:")
        date_str = f"{data['from_date']} TO {data['to_date']}"
        date_cell_found = False
        for r in range(50, ws.max_row + 1):
            for c in range(1, 4):
                val_str = str(ws.cell(r, c).value or '')
                if "RECORDS EXAMINED FROM AND TO" in val_str:
                    target_c = c + 1 if c + 1 <= ws.max_column else 2
                    ws.cell(r, target_c, date_str)
                    date_cell_found = True
                    break
            if date_cell_found: break
            
        if not date_cell_found:
            ws.cell(62, 2, date_str)
            
        # 3. Prepared By (Row 61: <AGENT> / AGENT NAME -> DAVID MICHALOVE)
        for r in range(50, ws.max_row + 1):
            for c in range(1, 4):
                val_str = str(ws.cell(r, c).value or '')
                if "PREPARED BY" in val_str:
                    target_c = c + 1 if c + 1 <= ws.max_column else 2
                    curr_agent = str(ws.cell(r, target_c).value or '')
                    if not curr_agent or "AGENT" in curr_agent.upper() or "<AGENT" in curr_agent.upper():
                        ws.cell(r, target_c, "DAVID MICHALOVE")
                    break

        # 4. Surface Owner (Rows 15-19)
        so = data.get("surface_owner", {})
        if so.get("name"):
            ws.cell(15, 1, so["name"])
            if so.get("tenancy"):
                ws.cell(16, 1, so["tenancy"])
            addrs = so.get("address_lines", [])
            if len(addrs) > 0: ws.cell(17, 1, addrs[0])
            if len(addrs) > 1: ws.cell(18, 1, addrs[1])
            if so.get("year"): ws.cell(19, 1, so["year"])
            ws.cell(15, 3, so.get("interest", 1))
            
        # 5. Mineral Owner (Rows 23-27)
        mo = data.get("mineral_owner", {})
        if mo.get("name"):
            ws.cell(23, 1, mo["name"])
            if mo.get("tenancy"):
                ws.cell(24, 1, mo["tenancy"])
            addrs = mo.get("address_lines", [])
            if len(addrs) > 0: ws.cell(25, 1, addrs[0])
            if len(addrs) > 1: ws.cell(26, 1, addrs[1])
            if mo.get("year"): ws.cell(27, 1, mo["year"])
            ws.cell(23, 3, mo.get("interest", 1))

        # 6. Sole Mineral Owner -> Clear Jim Doe placeholder (Rows 32 to 40)
        if data.get("sole_mineral_owner", True):
            for r in range(32, 41):
                for c in range(1, 15):
                    ws.cell(r, c).value = None
            ws.cell(23, 3).value = 1.0

        # 7. Leasehold Handling (Populate vs Open of Record)
        l_mode = data.get("leasehold_mode", "open_of_record")
        p_lease = data.get("primary_lease")
        
        if l_mode == "open_of_record" or not p_lease:
            ws.cell(23, 10, "OPEN OF RECORD")
            for r in range(24, 32):
                ws.cell(r, 10).value = None
            if "Leasehold Schedule A" in wb.sheetnames:
                wb.remove(wb["Leasehold Schedule A"])
        else:
            # Auto-populate Leasehold Column J in TR #1 Ownership
            memo_tag = " (Memo)" if p_lease.get("is_memo") else ""
            ws.cell(23, 10, "LEASEHOLD SCHEDULE A")
            ws.cell(24, 10, f"Exp.: {p_lease['exp_date']}, {p_lease['term']}, HBP")
            ws.cell(25, 10, f"Royalty: {p_lease['royalty']}")
            ws.cell(26, 10, f"Book/Page: {p_lease['bk_pg']}{memo_tag}")
            ws.cell(27, 10, f"Instrument Number: {p_lease['inst_num']}")
            ws.cell(28, 10, "Covers Tract #1 Only")
            ws.cell(29, 10, f"Horizontal Pugh Clause: {p_lease['pugh']}")
            ws.cell(30, 10, f"Depth Clause: {p_lease['depth']}")
            ws.cell(31, 10, f"Pooling Clause: {p_lease['pooling']}")
            
            # Auto-populate Leasehold Schedule A tab if present
            if "Leasehold Schedule A" in wb.sheetnames:
                ws_ls = wb["Leasehold Schedule A"]
                ws_ls.cell(11, 1, f"{p_lease['bk_pg']}\n{p_lease['inst_num']}\n(Memo)" if p_lease.get("is_memo") else f"{p_lease['bk_pg']}\n{p_lease['inst_num']}")
                ws_ls.cell(11, 2, p_lease["lessor"])
                ws_ls.cell(11, 3, p_lease["lessee"])
                ws_ls.cell(11, 4, "='TR #1 Ownership'!E23")
                
                # Royalty decimal conversion
                r_num = 0.20
                try:
                    r_clean = p_lease["royalty"].replace("%", "").strip()
                    if "/" in r_clean:
                        n, d = r_clean.split("/")
                        r_num = float(n) / float(d)
                    else:
                        r_num = float(r_clean) / 100.0 if float(r_clean) > 1 else float(r_clean)
                except: pass
                ws_ls.cell(11, 6, r_num)
                ws_ls.cell(11, 8, p_lease["eff_date"])
                ws_ls.cell(11, 9, p_lease["term"])
                ws_ls.cell(11, 10, 1)
                ws_ls.cell(11, 11, p_lease["exp_date"])
                ws_ls.cell(19, 1, p_lease["lessee"])
                ws_ls.cell(8, 1, f"DATED: {datetime.date.today().strftime('%m/%d/%Y')}")
                ws_ls.cell(8, 11, "PREPARED BY: DAVID MICHALOVE")
            
        # 8. Delete Notes Block (Rows 42 to 46)
        if data.get("delete_notes", True):
            note_row = None
            for r in range(40, 52):
                if ws.cell(r, 1).value and "NOTE #" in str(ws.cell(r, 1).value):
                    note_row = r
                    break
            if note_row:
                ws.delete_rows(note_row - 2, 5)

        # 9. Easements
        inc_easements = [e["summary"] for e in data.get("easements", []) if e.get("included")]
        for r in range(40, ws.max_row + 1):
            if ws.cell(r, 1).value and "EASEMENTS & RIGHTS OF WAY" in str(ws.cell(r, 1).value):
                if inc_easements:
                    txt = "\n".join([f"{i+1}) {s}" for i, s in enumerate(inc_easements)])
                    ws.cell(r + 1, 1, txt)
                else:
                    ws.cell(r + 1, 1, "1) None")
                break
                
        # 10. Oil & Gas Leases
        inc_leases = [l["summary"] for l in data.get("leases", []) if l.get("included")]
        for r in range(40, ws.max_row + 1):
            if ws.cell(r, 1).value and "UNRELEASED OIL & GAS LEASES" in str(ws.cell(r, 1).value):
                if inc_leases:
                    txt = "\n".join([f"{i+1}) {s}" for i, s in enumerate(inc_leases)])
                    ws.cell(r + 1, 1, txt)
                else:
                    ws.cell(r + 1, 1, "1) None")
                break
                
        # 11. Unreleased Mortgages
        inc_mortgages = [m["summary"] for m in data.get("mortgages", []) if m.get("included")]
        for r in range(40, ws.max_row + 1):
            if ws.cell(r, 1).value and "UNRELEASED MORTGAGES" in str(ws.cell(r, 1).value):
                if inc_mortgages:
                    txt = "\n".join([f"{i+1}) {s}" for i, s in enumerate(inc_mortgages)])
                    ws.cell(r + 1, 1, txt)
                else:
                    ws.cell(r + 1, 1, "1) None")
                break
                
        wb.save(or_path)
        return True

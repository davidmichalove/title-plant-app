import tkinter as tk
from tkinter import messagebox
from tkinter import ttk, messagebox
import openpyxl
import datetime
import os
import glob
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.rich_text import InlineFont
from spellchecker import SpellChecker
import re

class ToolTip:
    def __init__(self, widget):
        self.widget = widget
        self.tip_window = None

    def show_tip(self, text, x, y):
        if self.tip_window or not text:
            return
        self.tip_window = tk.Toplevel(self.widget)
        self.tip_window.wm_overrideredirect(True)
        self.tip_window.wm_geometry(f"+{x}+{y}")
        
        # Determine width for wrapping if text is long
        max_len = max([len(line) for line in text.split("\n")]) if text else 0
        wrap_length = 500 if max_len > 60 else 0
        
        label = tk.Label(self.tip_window, text=text, background="#ffffe0", foreground="black", relief=tk.SOLID, borderwidth=1, font=("Helvetica", 16), justify=tk.LEFT, wraplength=wrap_length)
        label.pack(ipadx=4, ipady=4)

    def hide_tip(self):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None

class RunsheetEditorWindow(tk.Toplevel):
    def __init__(self, parent, parcel_id, app_dir):
        super().__init__(parent)
        self.title(f"Runsheet Form Editor - {parcel_id}")
        self.geometry("1100x700")
        
        self.parcel_id = parcel_id
        self.spell = SpellChecker()
        
        # Find the runsheet
        self.base_dir = os.path.dirname(app_dir)
        self.pid_dir = os.path.join(self.base_dir, f"PID {parcel_id}")
        matches = glob.glob(os.path.join(self.pid_dir, "*RS*.xlsx"))
        matches = [m for m in matches if "_Backup" not in m and not os.path.basename(m).startswith("~$")]
        if not matches:
            messagebox.showerror("Error", f"Could not find any Runsheet (*RS*.xlsx) for PID {parcel_id} in {self.pid_dir}", parent=self)
            self.destroy()
            return
            
        self.excel_path = matches[0]
        self.status_file = os.path.join(self.pid_dir, "row_statuses.json")
        self.row_statuses = self.load_statuses()
        
        self.dower_file = os.path.join(self.pid_dir, "dower_reviewed.json")
        self.dower_reviewed = {}
        import json
        if os.path.exists(self.dower_file):
            with open(self.dower_file, "r") as f:
                self.dower_reviewed = json.load(f)
                
        self.maturity_file = os.path.join(self.pid_dir, "maturity_reviewed.json")
        self.maturity_reviewed = {}
        if os.path.exists(self.maturity_file):
            with open(self.maturity_file, "r") as f:
                self.maturity_reviewed = json.load(f)
                
        self.warnings_ignored_file = os.path.join(self.pid_dir, "warnings_ignored.json")
        self.warnings_ignored = {}
        if os.path.exists(self.warnings_ignored_file):
            with open(self.warnings_ignored_file, "r") as f:
                self.warnings_ignored = json.load(f)
                
        self.formatted_state_file = os.path.join(self.pid_dir, "initial_formatting_done.json")
        self.trash_file = os.path.join(self.pid_dir, "trash_rows.json")
        self.trash_rows = self.load_trash()
        
        self.ai_cache_file = os.path.join(self.pid_dir, "ai_qc_cache.json")
        self.ai_qc_cache = self.load_ai_cache()
        self.ai_qc_running = False
        
        self.ai_title_cache_file = os.path.join(self.pid_dir, "ai_title_cache.json")
        self.ai_title_cache = {}
        try:
            import json
            if os.path.exists(self.ai_title_cache_file):
                with open(self.ai_title_cache_file, "r") as f:
                    self.ai_title_cache = json.load(f)
        except: pass
        self.db_path = os.path.join(self.pid_dir, "runsheet_backup.db")
        import sqlite3
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''CREATE TABLE IF NOT EXISTS runsheet (
                            row_idx INTEGER PRIMARY KEY,
                            col_1 TEXT, col_2 TEXT, col_3 TEXT, col_4 TEXT, col_5 TEXT,
                            col_6 TEXT, col_7 TEXT, col_8 TEXT, col_9 TEXT, col_10 TEXT,
                            col_11 TEXT, col_12 TEXT, col_13 TEXT, col_14 TEXT, col_15 TEXT,
                            col_16 TEXT, col_17 TEXT, col_18 TEXT, col_19 TEXT, col_20 TEXT,
                            status TEXT, timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
                        )''')
        conn.commit()
        conn.close()
        
        # Create a backup of the original file if one doesn't already exist
        import shutil
        backup_path = self.excel_path.replace(".xlsx", "_Original_Backup.xlsx")
        if not os.path.exists(backup_path):
            try:
                shutil.copy2(self.excel_path, backup_path)
            except Exception as e:
                print(f"Failed to create backup: {e}")
        
        try:
            self.wb = openpyxl.load_workbook(self.excel_path, rich_text=True)
            self.ws = self.wb.active
            self.headers = []
            for cell in self.ws[2]:
                self.headers.append(str(cell.value) if cell.value else f"Col {cell.column}")
            self.cleanup_messy_dates()
            self.initial_format_all_comments()
        except Exception as e:
            messagebox.showerror("Error", f"Could not open Runsheet:\n{e}", parent=self)
            self.destroy()
            return
            
        self.current_row_idx = None
        self.widgets_by_col = {}
        
        # UI Layout
        paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        left_frame = ttk.Frame(paned)
        paned.add(left_frame, weight=0)
        
        ttk.Label(left_frame, text="Rows", font=("Helvetica", 16, "bold")).pack(anchor=tk.W)
        
        list_frame = ttk.Frame(left_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        self.listbox = tk.Listbox(list_frame, width=40, font=("Helvetica", 16), exportselection=False)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox.config(yscrollcommand=scrollbar.set)
        
        self.listbox.bind("<<ListboxSelect>>", self.on_select)
        
        self.listbox_tooltip = ToolTip(self.listbox)
        self.listbox.bind("<Motion>", self.on_listbox_motion)
        self.listbox.bind("<Leave>", lambda e: self.listbox_tooltip.hide_tip())
        
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill=tk.X, pady=(5, 0))
        ttk.Button(btn_frame, text="Move Up", command=self.move_row_up).pack(side=tk.LEFT, expand=True, fill=tk.X)
        ttk.Button(btn_frame, text="Move Down", command=self.move_row_down).pack(side=tk.LEFT, expand=True, fill=tk.X)
        ttk.Button(btn_frame, text="Add New Row", command=self.add_new_row).pack(side=tk.LEFT, expand=True, fill=tk.X)
        
        btn_frame2 = ttk.Frame(left_frame)
        btn_frame2.pack(fill=tk.X, pady=(0, 5))
        ttk.Button(btn_frame2, text="Delete Row", command=self.delete_row).pack(side=tk.LEFT, expand=True, fill=tk.X)
        ttk.Button(btn_frame2, text="View Trash Bin", command=self.view_trash).pack(side=tk.LEFT, expand=True, fill=tk.X)
        
        right_container = ttk.Frame(paned)
        paned.add(right_container, weight=1)
        
        header_area = ttk.Frame(right_container)
        header_area.pack(fill=tk.X, pady=(0, 10))
        
        top_bar = ttk.Frame(header_area)
        top_bar.pack(fill=tk.X)
        
        ttk.Label(top_bar, text=os.path.basename(self.excel_path), font=("Helvetica", 18, "bold")).pack(side=tk.LEFT)
        self.save_btn = ttk.Button(top_bar, text="Save Current Row", command=self.save_row, state="disabled")
        self.save_btn.pack(side=tk.RIGHT)
        self.ogl_btn = ttk.Button(top_bar, text="OGL Form", command=self.open_ogl_form)
        self.ogl_btn.pack(side=tk.RIGHT, padx=5)
        
        self.status_var = tk.StringVar(value="Not Started")
        self.status_combo = ttk.Combobox(top_bar, textvariable=self.status_var, state="readonly", values=["Not Started", "In Progress", "Completed"], width=15)
        self.status_combo.pack(side=tk.RIGHT, padx=10)
        self.status_combo.bind("<<ComboboxSelected>>", self.on_status_change)
        ttk.Label(top_bar, text="Row Status:").pack(side=tk.RIGHT)
        
        # Action Toolbar Row 1 (AI Actions, Sync, Reformat, Shortcuts)
        self.action_bar_row1 = ttk.Frame(header_area)
        self.action_bar_row1.pack(fill=tk.X, pady=(4, 2))
        
        self.reformat_btn = ttk.Button(self.action_bar_row1, text="Reformat Comments", command=self.confirm_and_reformat_all)
        self.reformat_btn.pack(side=tk.LEFT)
        
        self.sync_or_btn = ttk.Button(self.action_bar_row1, text="📊 Sync OR", command=self.open_or_sync_dialog)
        self.sync_or_btn.pack(side=tk.LEFT, padx=(6, 0))
        
        self.retry_qc_btn = ttk.Button(self.action_bar_row1, text="Retry AI Check", command=self.retry_ai_check, state="disabled")
        self.retry_qc_btn.pack(side=tk.LEFT, padx=(6, 0))
        
        self.shortcuts_btn = ttk.Button(self.action_bar_row1, text="⌨️ Shortcuts", command=self.show_shortcuts_dialog)
        self.shortcuts_btn.pack(side=tk.RIGHT)
        
        # Action Toolbar Row 2 (Documents, Phrase Library, RS Excel)
        self.action_bar_row2 = ttk.Frame(header_area)
        self.action_bar_row2.pack(fill=tk.X, pady=(2, 4))
        
        self.open_pdf_btn = ttk.Button(self.action_bar_row2, text="Open Document", command=self.open_pdf_for_row, state="disabled")
        self.open_pdf_btn.pack(side=tk.LEFT)
        
        self.phrase_btn = ttk.Button(self.action_bar_row2, text="Phrase Library", command=self.open_phrase_library)
        self.phrase_btn.pack(side=tk.LEFT, padx=(6, 0))
        
        self.open_excel_btn = ttk.Button(self.action_bar_row2, text="Open RS Excel", command=self.open_excel_runsheet)
        self.open_excel_btn.pack(side=tk.LEFT, padx=(6, 0))
        
        self.canvas = tk.Canvas(right_container)
        v_scroll = ttk.Scrollbar(right_container, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=v_scroll.set)
        
        self.form_frame = ttk.Frame(self.canvas)
        self.canvas_window = self.canvas.create_window((0, 0), window=self.form_frame, anchor="nw")
        
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.form_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>", self.on_canvas_configure)
        
        self.bind("<MouseWheel>", self._on_mousewheel)
        
        # Row Navigation & Actions Shortcuts
        self.bind("<Command-s>", lambda e: self.save_row())
        self.bind("<Control-s>", lambda e: self.save_row())
        self.bind("<Command-p>", self.set_status_in_progress)
        self.bind("<Control-p>", self.set_status_in_progress)
        self.bind("<Command-f>", self.set_status_completed)
        self.bind("<Control-f>", self.set_status_completed)
        self.bind("<Command-a>", self.show_ai_note_dialog)
        self.bind("<Control-a>", self.show_ai_note_dialog)
        self.bind("<Command-Shift-A>", self.show_ai_note_dialog)
        self.bind("<Command-Shift-a>", self.show_ai_note_dialog)
        self.bind("<Control-Shift-A>", self.show_ai_note_dialog)
        self.bind("<Control-Shift-a>", self.show_ai_note_dialog)

        self.bind("<Command-Shift-O>", self.open_or_sync_dialog)
        self.bind("<Command-Shift-o>", self.open_or_sync_dialog)
        self.bind("<Control-Shift-O>", self.open_or_sync_dialog)
        self.bind("<Control-Shift-o>", self.open_or_sync_dialog)
        self.bind("<Command-d>", self.toggle_dower_reviewed)
        self.bind("<Control-d>", self.toggle_dower_reviewed)
        
        # Delete Original Block shortcuts
        self.bind("<Command-Shift-D>", self.delete_original_block)
        self.bind("<Command-Shift-d>", self.delete_original_block)
        self.bind("<Control-Shift-D>", self.delete_original_block)
        self.bind("<Control-Shift-d>", self.delete_original_block)
        self.bind("<Command-Shift-O>", self.delete_original_block)
        self.bind("<Command-Shift-o>", self.delete_original_block)
        self.bind("<Control-Shift-O>", self.delete_original_block)
        self.bind("<Control-Shift-o>", self.delete_original_block)
        self.bind("<Alt-d>", self.delete_original_block)
        self.bind("<Alt-D>", self.delete_original_block)
        self.bind("<Command-o>", lambda e: self.open_pdf_for_row())
        self.bind("<Control-o>", lambda e: self.open_pdf_for_row())
        self.bind("<Command-n>", self.convert_to_normal_case)
        self.bind("<Control-n>", self.convert_to_normal_case)
        self.bind("<Command-l>", lambda e: self.open_phrase_library())
        self.bind("<Control-l>", lambda e: self.open_phrase_library())
        
        # Row Arrow Navigation (Ctrl/Cmd/Alt + Up/Down)
        for seq in ("<Control-Up>", "<Control-Down>", "<Command-Up>", "<Command-Down>", "<Alt-Up>", "<Alt-Down>"):
            handler = self.nav_prev_row if "Up" in seq else self.nav_next_row
            self.bind_all(seq, handler)
            self.listbox.bind(seq, handler)
        
        # Number shortcuts for inserting phrases 1..10 (Ctrl+1 .. Ctrl+9, Ctrl+0)
        for num in range(10):
            self.bind(f"<Control-Key-{num}>", lambda e, n=num: self.insert_phrase_by_num(n))
            self.bind(f"<Command-Key-{num}>", lambda e, n=num: self.insert_phrase_by_num(n))
            
        # Shift+Number & Alt+Number shortcuts for inserting phrases 11..20 (Ctrl+Shift+1..0 / Option+1..0)
        shift_keys = {"exclam": 1, "at": 2, "numbersign": 3, "dollar": 4, "percent": 5, "asciicircum": 6, "ampersand": 7, "asterisk": 8, "parenleft": 9, "parenright": 0}
        for sym, num in shift_keys.items():
            self.bind(f"<Control-{sym}>", lambda e, n=num: self.insert_phrase_by_num(n + 10))
            self.bind(f"<Command-{sym}>", lambda e, n=num: self.insert_phrase_by_num(n + 10))
            self.bind(f"<Control-Shift-Key-{num}>", lambda e, n=num: self.insert_phrase_by_num(n + 10))
            self.bind(f"<Command-Shift-Key-{num}>", lambda e, n=num: self.insert_phrase_by_num(n + 10))

        for num in range(10):
            self.bind(f"<Alt-Key-{num}>", lambda e, n=num: self.insert_phrase_by_num(n + 10))
            self.bind(f"<Option-Key-{num}>", lambda e, n=num: self.insert_phrase_by_num(n + 10))
        
        self.row_indices = []
        self.load_rows()
        self.build_form()
        
        self.auto_run_ai_qc()
        
    def cleanup_messy_dates(self):
        changed = False
        headers = [str(cell.value or "").lower() for cell in self.ws[2]]
        date_cols = [i for i, h in enumerate(headers) if "date" in h]
        
        for row_idx in range(3, self.ws.max_row + 1):
            row = self.ws[row_idx]
            for col_idx in date_cols:
                if col_idx < len(row):
                    cell = row[col_idx]
                    val = str(cell.value or "").strip()
                    if getattr(cell, 'quotePrefix', False):
                        cell.quotePrefix = False
                        changed = True
                        
                    if len(val) > 20 and "GMT" in val:
                        parts = val.split()
                        if len(parts) >= 4 and parts[0] in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]:
                            try:
                                dt = datetime.datetime.strptime(" ".join(parts[:4]), "%a %b %d %Y")
                                cell.value = dt.strftime("%m/%d/%Y")
                                cell.number_format = '@'
                                cell.data_type = 's'
                                changed = True
                            except Exception:
                                pass
                    elif val.startswith("'"):
                        cell.value = val[1:]
                        cell.number_format = '@'
                        cell.data_type = 's'
                        changed = True
        if changed:
            try:
                self.wb.save(self.excel_path)
            except:
                pass

    def _on_mousewheel(self, event):
        if isinstance(event.widget, (tk.Listbox, tk.Text)):
            return
        # macOS scrolling uses event.delta directly
        self.canvas.yview_scroll(int(-1 * event.delta), "units")
        
    def on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_window, width=event.width)
        
    def on_listbox_motion(self, event):
        idx = self.listbox.nearest(event.y)
        if idx < 0:
            self.listbox_tooltip.hide_tip()
            return
            
        bbox = self.listbox.bbox(idx)
        if bbox:
            x, y, w, h = bbox
            if y <= event.y <= y + h:
                item_text = self.listbox.get(idx)
                tip_text = ""
                if "🔵" in item_text: tip_text = "🔵 Within 10 Years"
                elif "🟣" in item_text: tip_text = "🟣 Within 30 Years"
                elif "⚫" in item_text: tip_text = "⚫ Within 40 Years"
                elif "⚪" in item_text: tip_text = "⚪ Over 40 Years"
                
                if tip_text:
                    x_root = self.listbox.winfo_rootx() + event.x + 15
                    y_root = self.listbox.winfo_rooty() + event.y + 15
                    if self.listbox_tooltip.tip_window:
                        self.listbox_tooltip.tip_window.winfo_children()[0].config(text=tip_text)
                        self.listbox_tooltip.tip_window.wm_geometry(f"+{x_root}+{y_root}")
                    else:
                        self.listbox_tooltip.show_tip(tip_text, x_root, y_root)
                else:
                    self.listbox_tooltip.hide_tip()
            else:
                self.listbox_tooltip.hide_tip()
        else:
            self.listbox_tooltip.hide_tip()
        
    def load_rows(self):
        scroll_pos = self.listbox.yview()
        self.listbox.delete(0, tk.END)
        self.row_indices = []
        self.row_warnings = {}
        
        # Pre-scan for volume and page sequence order errors
        eff_date_col = -1
        file_date_col = -1
        for i, h in enumerate(self.headers):
            hl = h.lower()
            if "eff" in hl and "date" in hl: eff_date_col = i
            elif "file" in hl and "date" in hl: file_date_col = i
            elif "date" in hl and eff_date_col == -1: eff_date_col = i

        seq_row_meta = []
        for r_idx in range(3, self.ws.max_row + 1):
            r = self.ws[r_idx]
            if not any(cell.value for cell in r): continue
            
            btype = str(r[1].value or "").strip()
            v_val = str(r[2].value or "").strip()
            p_val = str(r[3].value or "").strip()
            
            eff_dt = str(r[eff_date_col].value or "").strip() if eff_date_col != -1 and eff_date_col < len(r) else ""
            file_dt = str(r[file_date_col].value or "").strip() if file_date_col != -1 and file_date_col < len(r) else ""
            
            p_num = None
            import re
            m_pg = re.search(r'\d+', p_val)
            if m_pg:
                p_num = int(m_pg.group(0))
                
            seq_row_meta.append({
                "row_idx": r_idx,
                "btype": btype,
                "vol": v_val,
                "page": p_val,
                "page_num": p_num,
                "eff_date": eff_dt,
                "file_date": file_dt
            })

        seq_warnings_by_row = {}
        for i in range(len(seq_row_meta)):
            item_a = seq_row_meta[i]
            if not item_a["vol"] or item_a["page_num"] is None: continue
            
            for j in range(i + 1, len(seq_row_meta)):
                item_b = seq_row_meta[j]
                if not item_b["vol"] or item_b["page_num"] is None: continue
                
                # Must be same volume
                if item_a["vol"] == item_b["vol"]:
                    if item_a["btype"] and item_b["btype"] and item_a["btype"].upper() != item_b["btype"].upper():
                        continue
                        
                    # Same effective date or same file date
                    same_eff = item_a["eff_date"] and item_b["eff_date"] and (item_a["eff_date"] == item_b["eff_date"])
                    same_file = item_a["file_date"] and item_b["file_date"] and (item_a["file_date"] == item_b["file_date"])
                    
                    if same_eff or same_file:
                        if item_a["page_num"] > item_b["page_num"]:
                            v = item_a["vol"]
                            p_a = item_a["page"]
                            p_b = item_b["page"]
                            rb_idx = item_b["row_idx"]
                            ra_idx = item_a["row_idx"]
                            d_str = item_a["eff_date"] if same_eff else item_a["file_date"]
                            
                            w_a = f"Page sequence error: Vol {v} Pg {p_a} placed before lower Pg {p_b} (Row {rb_idx}) with same date ({d_str})"
                            w_b = f"Page sequence error: Vol {v} Pg {p_b} placed after higher Pg {p_a} (Row {ra_idx}) with same date ({d_str})"
                            
                            seq_warnings_by_row.setdefault(item_a["row_idx"], []).append(w_a)
                            seq_warnings_by_row.setdefault(item_b["row_idx"], []).append(w_b)

        for row_idx in range(3, self.ws.max_row + 1):
            row = self.ws[row_idx]
            inst_type = str(row[0].value).strip() if row[0].value else ""
            book_type = str(row[1].value).strip() if row[1].value else ""
            vol = str(row[2].value).strip() if row[2].value else ""
            page = str(row[3].value).strip() if row[3].value else ""
            
            # If the row is completely empty, skip it
            if not any(cell.value for cell in row):
                continue
                
            disp = ""
            if book_type: disp += f"{book_type} "
            if vol and page: disp += f"{vol}/{page}"
            elif vol: disp += f"{vol}"
            elif page: disp += f"{page}"
            
            if disp and inst_type: disp += f" - {inst_type}"
            elif inst_type: disp += f"{inst_type}"
            
            if not disp: disp = f"Row {row_idx}"
            else: disp = f"Row {row_idx}: {disp}"
            
            # Date checking for visual indicators
            import datetime
            date_idx = -1
            for h in ["effective date", "filing date", "date"]:
                for idx, col_name in enumerate(self.headers):
                    if h in col_name.lower():
                        date_idx = idx
                        break
                if date_idx != -1: break
                
            prefix = "⚪ "
            date_color = ""
            if date_idx != -1 and date_idx < len(row):
                d_val = row[date_idx].value
                date_obj = None
                if isinstance(d_val, datetime.datetime):
                    date_obj = d_val
                elif isinstance(d_val, str) and d_val.strip():
                    try:
                        clean_date = d_val.strip().lstrip("'")
                        date_obj = datetime.datetime.strptime(clean_date, "%m/%d/%Y")
                    except: pass
                    
                if date_obj:
                    diff_years = (datetime.datetime.now() - date_obj).days / 365.25
                    if diff_years <= 10:
                        prefix = "🔵 "
                    elif diff_years <= 30:
                        prefix = "🟣 "
                    elif diff_years <= 40:
                        prefix = "⚫ "
                    else:
                        prefix = "⚪ "
            
            # QC Notes Validation
            notes_idx = -1
            for idx, col_name in enumerate(self.headers):
                if "note" in col_name.lower() or "comment" in col_name.lower():
                    notes_idx = idx
                    break
                    
            notes_val_raw = ""
            if notes_idx != -1 and notes_idx < len(row):
                notes_val_raw = str(row[notes_idx].value or "")
            notes_val = notes_val_raw.lower()
            
            inst_lower = inst_type.lower()
            needs_warning = False
            warnings = []
            
            # Add sequence warnings if any
            if row_idx in seq_warnings_by_row:
                for w in seq_warnings_by_row[row_idx]:
                    warnings.append(w)
                    needs_warning = True
            
            if "--- Original ---" in notes_val_raw:
                needs_warning = True
                warnings.append("Original text block not deleted.")
            
            if "deed" in inst_lower or "mortgage" in inst_lower:
                if "release" not in inst_lower and "satisfaction" not in inst_lower:
                    if not self.dower_reviewed.get(str(row_idx), False):
                        needs_warning = True
                        warnings.append("Dower unreviewed.")
                        
            if "mortgage" in inst_lower and "release" not in inst_lower and "satisfaction" not in inst_lower:
                if "release" not in notes_val:
                    needs_warning = True
                    warnings.append("Missing 'release' in notes.")
                if "maturity" not in notes_val:
                    needs_warning = True
                    warnings.append("Missing 'maturity' date in notes.")
                    
            if "lease" in inst_lower or "notice" in inst_lower or "memorandum" in inst_lower:
                if "release" not in inst_lower and "satisfaction" not in inst_lower:
                    if "release" not in notes_val:
                        needs_warning = True
                        warnings.append("Missing 'release' in notes.")
            
            # Check for empty fields or 0 acreage for passive warnings
            empty_cols = []
            for col_idx, h in enumerate(self.headers):
                if col_idx >= len(row): break
                h_lower = h.lower()
                if "deed plot" in h_lower or "notes" in h_lower:
                    continue
                val = str(row[col_idx].value or "").strip()
                if not val:
                    empty_cols.append(h)
                if "acreage" in h_lower and val == "0":
                    warnings.append("Acreage is 0")
                    needs_warning = True
                    
            if empty_cols:
                warnings.append(f"Missing {len(empty_cols)} field(s)")
                needs_warning = True

            if self.warnings_ignored.get(str(row_idx), False):
                if "Original text block not deleted." in warnings:
                    warnings = ["Original text block not deleted."]
                    needs_warning = True
                else:
                    needs_warning = False
                    warnings.clear()

            if warnings:
                self.row_warnings[str(row_idx)] = warnings
                        
            status = self.row_statuses.get(str(row_idx), "Not Started")
            
            fg_color = ""
            if needs_warning:
                prefix = "⚠️ " + prefix
                if status != "Completed":
                    fg_color = "red"
            
            disp = prefix + disp
            
            # Append AI Checkmark if cached
            if hasattr(self, 'ai_qc_cache'):
                vol = str(row[2].value).strip() if row[2].value else ""
                pg = str(row[3].value).strip() if row[3].value else ""
                if vol and pg and f"{vol}_{pg}" in self.ai_qc_cache:
                    disp += " ✅"
                    
            if hasattr(self, 'ai_title_cache'):
                vol = str(row[2].value).strip() if row[2].value else ""
                pg = str(row[3].value).strip() if row[3].value else ""
                inst = str(row[0].value).strip().lower() if row[0].value else ""
                if vol and pg and f"{vol}_{pg}" in self.ai_title_cache:
                    cached = self.ai_title_cache[f"{vol}_{pg}"].lower()
                    if cached and inst != cached:
                        disp += " 🔍"

            self.listbox.insert(tk.END, disp)
            
            # Color code listbox item based on status
            if status == "In Progress":
                self.listbox.itemconfig(tk.END, {'bg': '#fff3cd', 'fg': fg_color or 'black'})
            elif status == "Completed":
                self.listbox.itemconfig(tk.END, {'bg': '#d4edda', 'fg': fg_color or 'black'})
            else:
                self.listbox.itemconfig(tk.END, {'bg': '', 'fg': fg_color})
                
            self.row_indices.append(row_idx)
            
        if scroll_pos:
            self.listbox.yview_moveto(scroll_pos[0])
            
    def build_form(self):
        for widget in self.form_frame.winfo_children():
            widget.destroy()
        
        self.widgets_by_col = {}
        self.label_widgets_by_col = {}
        
        qc_frame = ttk.LabelFrame(self.form_frame, text=" QC Reminders (Row Checklist) ")
        qc_frame.pack(fill=tk.X, pady=(5, 5), padx=5)
        
        self.qc_vars = []
        checks = [
            "Dower Mentioned / Released?",
            "Mortgage Released?",
            "Maturity Date Present?",
            "Ignore Warnings"
        ]
        
        def on_dower_toggle():
            if not getattr(self, 'current_row_idx', None): return
            self.dower_reviewed[str(self.current_row_idx)] = self.qc_vars[0].get()
            import json
            with open(self.dower_file, "w") as f:
                json.dump(self.dower_reviewed, f)
                
            # Re-evaluate warnings for this row
            warnings = getattr(self, 'row_warnings', {}).get(str(self.current_row_idx), [])
            
            has_dower = False
            for i, h in enumerate(self.headers):
                if "comment" in h.lower() or "note" in h.lower():
                    widget = self.widgets_by_col.get(i)
                    if isinstance(widget, __import__('tkinter').Text):
                        live_text = widget.get("1.0", "end").lower()
                        has_dower = "dower" in live_text
                    break

            if self.qc_vars[0].get() and has_dower:
                if "Dower unreviewed." in warnings: warnings.remove("Dower unreviewed.")
            else:
                if "Dower unreviewed." not in warnings: warnings.append("Dower unreviewed.")
            
            if warnings:
                self.row_warnings[str(self.current_row_idx)] = warnings
            else:
                if str(self.current_row_idx) in self.row_warnings:
                    del self.row_warnings[str(self.current_row_idx)]
                    
            try:
                idx = self.row_indices.index(self.current_row_idx)
                self._apply_row_color(idx)
            except: pass
            
            # Update warning label safely
            warn_text = ""
            if warnings: warn_text += "⚠️ " + " | ".join(warnings)
            
            # Preserve AI text if it exists
            if getattr(self, 'current_ai_full_text', ""):
                # We need to recreate the ai_res string
                vol = str(self.ws.cell(row=self.current_row_idx, column=3).value or "").strip()
                pg = str(self.ws.cell(row=self.current_row_idx, column=4).value or "").strip()
                cache_key = f"{vol}_{pg}"
                if hasattr(self, 'ai_qc_cache') and cache_key in self.ai_qc_cache:
                    raw_text = self.ai_qc_cache[cache_key]
                    if "SUMMARY:" in raw_text and "FULL TEXT:" in raw_text:
                        parts = raw_text.split("FULL TEXT:")
                        warn_text += "\n🤖 AI Check " + parts[0].replace("SUMMARY:", "").strip()
                    else:
                        warn_text += "\n🤖 AI Check: Result formatted incorrectly. Hover to view."
            
            # Magic Auto-Injector for Releases found by AI QC
            if getattr(self, 'current_ai_full_text', ""):
                vol = str(self.ws.cell(row=self.current_row_idx, column=3).value or "").strip()
                pg = str(self.ws.cell(row=self.current_row_idx, column=4).value or "").strip()
                cache_key = f"{vol}_{pg}"
                if hasattr(self, 'ai_qc_cache') and cache_key in self.ai_qc_cache:
                    raw_text = self.ai_qc_cache[cache_key]
                    # Find patterns like VOL 941 PAGE 285, OR 941/285
                    import re
                    matches = re.finditer(r'(?:VOL(?:UME)?\s*|MR\s*|OR\s*|DR\s*)(\d+)\s*(?:PAGE|PG|/)?\s*(\d+)', raw_text, re.IGNORECASE)
                    for m in matches:
                        v = m.group(1)
                        p = m.group(2)
                        start = max(0, m.start() - 50)
                        end = min(len(raw_text), m.end() + 50)
                        context = raw_text[start:end].lower()
                        if 'release' in context or 'satisfaction' in context or 'see' in context:
                            new_release = f"Release: OR {v}/{p}"
                            # Check if notes_val already has this release
                            notes_val_tmp = ""
                            try:
                                for i, h in enumerate(self.headers):
                                    if "comment" in h.lower() or "note" in h.lower():
                                        widget = self.widgets_by_col.get(i)
                                        if isinstance(widget, __import__('tkinter').Text):
                                            notes_val_tmp = widget.get("1.0", "end").lower()
                                        break
                            except: pass
                            
                            # if no release in notes at all, or if this specific one is missing
                            if "release:" not in notes_val_tmp or (v not in notes_val_tmp and p not in notes_val_tmp):
                                try:
                                    for i, h in enumerate(self.headers):
                                        if "comment" in h.lower() or "note" in h.lower():
                                            widget = self.widgets_by_col.get(i)
                                            if isinstance(widget, __import__('tkinter').Text):
                                                current_text = widget.get("1.0", "end").strip()
                                                new_text = f"{current_text}\n\n{new_release}".strip()
                                                widget.delete("1.0", "end")
                                                widget.insert("1.0", new_text)
                                            break
                                except: pass
                            break # Only inject first found release
            
            self.warning_label.config(text=warn_text.strip())
        self.on_dower_toggle_cb = on_dower_toggle

        def on_ignore_warnings_toggle():
            if not getattr(self, 'current_row_idx', None): return
            self.warnings_ignored[str(self.current_row_idx)] = self.qc_vars[3].get()
            import json
            with open(self.warnings_ignored_file, "w") as f:
                json.dump(self.warnings_ignored, f)
            # Re-evaluate warnings immediately
            self.load_rows(reload_only=True)
            self._update_instrument_label(self.current_row_idx)
            
            # Update warning label
            warn_text = ""
            warnings = getattr(self, 'row_warnings', {}).get(str(self.current_row_idx), [])
            if warnings: warn_text += "⚠️ " + " | ".join(warnings)
            
            if getattr(self, 'current_ai_full_text', ""):
                vol = str(self.ws.cell(row=self.current_row_idx, column=3).value or "").strip()
                pg = str(self.ws.cell(row=self.current_row_idx, column=4).value or "").strip()
                cache_key = f"{vol}_{pg}"
                if hasattr(self, 'ai_qc_cache') and cache_key in self.ai_qc_cache:
                    raw_text = self.ai_qc_cache[cache_key]
                    if "SUMMARY:" in raw_text and "FULL TEXT:" in raw_text:
                        parts = raw_text.split("FULL TEXT:")
                        warn_text += "\n🤖 AI Check " + parts[0].replace("SUMMARY:", "").strip()
                    else:
                        warn_text += "\n🤖 AI Check: Result formatted incorrectly. Hover to view."
            self.warning_label.config(text=warn_text.strip())

        for c in checks:
            var = tk.BooleanVar()
            if c == "Dower Mentioned / Released?":
                chk = ttk.Checkbutton(qc_frame, text=c, variable=var, command=on_dower_toggle)
            elif c == "Ignore Warnings":
                chk = ttk.Checkbutton(qc_frame, text=c, variable=var, command=on_ignore_warnings_toggle)
            else:
                chk = ttk.Checkbutton(qc_frame, text=c, variable=var)
            chk.pack(side=tk.LEFT, padx=10, pady=5)
            self.qc_vars.append(var)
            
            if c == "Mortgage Released?":
                self.release_vol_pg_var = tk.StringVar()
                self.release_entry = ttk.Entry(qc_frame, textvariable=self.release_vol_pg_var, width=15)
                self.release_entry.pack(side=tk.LEFT, padx=(0, 10), pady=5)
                # Add placeholder
                self.release_entry.insert(0, "Vol/Pg")
                def on_focus_in(e):
                    if self.release_entry.get() == 'Vol/Pg':
                        self.release_entry.delete(0, 'end')
                def on_focus_out(e):
                    if not self.release_entry.get():
                        self.release_entry.insert(0, 'Vol/Pg')
                self.release_entry.bind("<FocusIn>", on_focus_in)
                self.release_entry.bind("<FocusOut>", on_focus_out)
            
        self.warning_label = tk.Label(self.form_frame, text="", font=("Helvetica", 16, "bold"), fg="red", bg="#ffeeee", justify=tk.LEFT, anchor="w")
        self.warning_label.pack(fill=tk.X, pady=(0, 10), padx=5)
        
        self.warning_tooltip = ToolTip(self.warning_label)
        def on_warning_enter(event):
            if getattr(self, 'current_ai_full_text', ""):
                x = self.warning_label.winfo_rootx() + event.x + 15
                y = self.warning_label.winfo_rooty() + event.y + 15
                self.warning_tooltip.show_tip(self.current_ai_full_text, x, y)
        self.warning_label.bind("<Enter>", on_warning_enter)
        self.warning_label.bind("<Leave>", lambda e: self.warning_tooltip.hide_tip())
        self.warning_label.bind("<Motion>", lambda e: [self.warning_tooltip.hide_tip(), on_warning_enter(e)] if self.warning_tooltip.tip_window else None)
        
        # Add a tooltip or hint for bolding
        hint_lbl = ttk.Label(self.form_frame, text="Shortcuts: Ctrl+S (Save) | Ctrl+P (In Progress) | Ctrl+N (Title Case) | Ctrl+L (Phrases) | Ctrl+1..9 (Insert Phrase) | Ctrl+↑/↓ (Prev/Next Row)", font=("Helvetica", 11, "italic"))
        hint_lbl.pack(anchor=tk.E, pady=2, padx=5)
        
        for i, header in enumerate(self.headers):
            if i == 13: continue # Hide Column 14 (Notes) from Editor
            
            frame = ttk.Frame(self.form_frame)
            frame.pack(fill=tk.X, pady=5)
            
            lbl = tk.Label(frame, text=header, width=20, anchor=tk.NW, font=("Helvetica", 16, "bold"), fg="systemTextColor", cursor="arrow")
            lbl.pack(side=tk.LEFT, anchor=tk.NW)
            self.label_widgets_by_col[i] = lbl
            
            if i < 4 or header.lower() in ['acreage', 'instrument number', 'filing date', 'effective date']:
                widget = ttk.Entry(frame, font=("Helvetica", 16))
                widget.pack(side=tk.LEFT, fill=tk.X, expand=True)
            else:
                text_container = ttk.Frame(frame)
                text_container.pack(side=tk.LEFT, fill=tk.X, expand=True)
                
                widget = tk.Text(text_container, height=4, font=("Helvetica", 16), wrap=tk.WORD, undo=True, maxundo=-1, autoseparators=True)
                widget.pack(side=tk.TOP, fill=tk.X, expand=True)
                
                # Undo/Redo Bindings
                def _undo(e):
                    try:
                        e.widget.edit_undo()
                    except:
                        pass
                    return "break"
                def _redo(e):
                    try:
                        e.widget.edit_redo()
                    except:
                        pass
                    return "break"
                widget.bind("<Command-z>", _undo)
                widget.bind("<Command-Shift-Z>", _redo)
                widget.bind("<Control-z>", _undo)
                widget.bind("<Control-y>", _redo)
                
                grip = ttk.Label(text_container, text="═", cursor="sb_v_double_arrow", anchor=tk.CENTER, foreground="gray")
                grip.pack(side=tk.TOP, fill=tk.X)
                
                def start_resize(event, w=widget):
                    w.startY = event.y_root
                    w.startH = w.winfo_height()
                    
                def do_resize(event, w=widget):
                    dy = event.y_root - w.startY
                    # Approx 21px per line for Helvetica 14
                    new_lines = max(2, int((w.startH + dy) / 21))
                    w.configure(height=new_lines)
                    
                grip.bind("<Button-1>", start_resize)
                grip.bind("<B1-Motion>", do_resize)
                
                widget.bind("<Tab>", self.focus_next_widget)
                widget.tag_configure("bold", font=("Helvetica", 16, "bold"))
                widget.bind("<Command-b>", self.toggle_bold)
                widget.bind("<Control-b>", self.toggle_bold)
                
                # Spell checker bindings
                widget.tag_configure("misspelled", underline=True, foreground="red")
                widget.bind("<KeyRelease>", self.schedule_spellcheck)
                widget.bind("<Button-2>", self.show_context_menu)
                widget.bind("<Button-3>", self.show_context_menu)
                widget.bind("<Control-Button-1>", self.show_context_menu)
            self.widgets_by_col[i] = widget
            
    def focus_next_widget(self, event):
        event.widget.tk_focusNext().focus()
        return "break"
        
    def toggle_bold(self, event=None):
        if not event or not isinstance(event.widget, tk.Text):
            return "break"
        text_widget = event.widget
        try:
            if "bold" in text_widget.tag_names(tk.SEL_FIRST):
                text_widget.tag_remove("bold", tk.SEL_FIRST, tk.SEL_LAST)
            else:
                text_widget.tag_add("bold", tk.SEL_FIRST, tk.SEL_LAST)
        except tk.TclError:
            pass # No selection
        return "break"
        
    def _update_field_labels(self, row_idx):
        if not getattr(self, 'label_widgets_by_col', None): return
        
        v = str(self.ws.cell(row=row_idx, column=3).value or "").strip()
        p = str(self.ws.cell(row=row_idx, column=4).value or "").strip()
        ck = f"{v}_{p}"
        
        ai_data = {}
        # 1. Load from provenance data if present
        if hasattr(self, 'provenance_data') and self.provenance_data:
            prov = self.provenance_data.get(ck) or self.provenance_data.get(str(row_idx), {})
            if isinstance(prov, dict):
                ai_data.update(prov)
                
        # 2. Check ai_title_cache
        if hasattr(self, 'ai_title_cache') and ck in self.ai_title_cache:
            ai_data['instrument_type'] = self.ai_title_cache[ck]

        for col_idx, lbl in self.label_widgets_by_col.items():
            if col_idx >= len(self.headers): continue
            header = self.headers[col_idx]
            hl = header.lower()
            
            # Exclude conveyance, deed plot, comments, notes
            if "conveyance" in hl or "deed plot" in hl or "comment" in hl or "note" in hl:
                lbl.config(text=header, fg="systemTextColor", cursor="arrow")
                lbl.unbind("<Button-1>")
                continue
                
            base_text = header
            lbl.config(text=base_text, fg="systemTextColor", cursor="arrow")
            lbl.unbind("<Button-1>")
            
            ai_val = None
            if ("instrument" in hl and "type" in hl) or hl == "instrument":
                ai_val = ai_data.get("instrument_type") or ai_data.get("instrument")
            elif "grantor" in hl or "lessor" in hl:
                ai_val = ai_data.get("grantor") or ai_data.get("grantor_lessor")
            elif "grantee" in hl or "lessee" in hl:
                ai_val = ai_data.get("grantee") or ai_data.get("grantee_lessee")
            elif "effective" in hl and "date" in hl:
                ai_val = ai_data.get("effective_date")
            elif "filing" in hl and "date" in hl:
                ai_val = ai_data.get("filing_date")
            elif "acreage" in hl:
                ai_val = ai_data.get("acreage")
            elif "book" in hl and "type" in hl:
                ai_val = ai_data.get("book_type")
            elif "instrument number" in hl:
                ai_val = ai_data.get("instrument_number")
                
            if ai_val:
                ai_val_str = str(ai_val).strip()
                w = self.widgets_by_col.get(col_idx)
                curr_val = ""
                if w:
                    if hasattr(w, 'get') and callable(w.get):
                        curr_val = w.get().strip()
                    elif isinstance(w, tk.Text):
                        curr_val = w.get("1.0", "end-1c").strip()
                        
                def norm(s):
                    import re
                    return re.sub(r'[\s,._/-]+', '', str(s).lower())
                    
                if ai_val_str and norm(curr_val) != norm(ai_val_str):
                    lbl.config(text=f"🔍 {base_text}", fg="systemLinkColor", cursor="hand2")
                    def make_apply(widget=w, target_val=ai_val_str, label=lbl, orig_text=base_text):
                        def apply_val(e):
                            if widget:
                                if hasattr(widget, 'delete') and hasattr(widget, 'insert'):
                                    if isinstance(widget, tk.Text):
                                        widget.delete("1.0", tk.END)
                                        widget.insert("1.0", target_val)
                                    else:
                                        widget.delete(0, tk.END)
                                        widget.insert(0, target_val)
                                    label.config(text=orig_text, fg="systemTextColor", cursor="arrow")
                                    label.unbind("<Button-1>")
                        return apply_val
                    lbl.bind("<Button-1>", make_apply())

    _update_instrument_label = _update_field_labels

    def on_select(self, event):
        selection = self.listbox.curselection()
        if not selection: return
        
        idx = selection[0]
        self.current_row_idx = self.row_indices[idx]
        
        row_cells = self.ws[self.current_row_idx]
        
        for i, widget in self.widgets_by_col.items():
            val = row_cells[i].value if i < len(row_cells) else ""
            header_name = self.headers[i].lower()
            
            if isinstance(widget, ttk.Entry):
                widget.bind("<Button-2>", self.show_context_menu)
                widget.bind("<Button-3>", self.show_context_menu)
                widget.bind("<Control-Button-1>", self.show_context_menu)
                
                if isinstance(val, datetime.datetime):
                    val = val.strftime("%m/%d/%Y")
                elif "acreage" in header_name and isinstance(val, (int, float)):
                    val = f"{val:.6f}"
                else:
                    val = str(val) if val is not None else ""
                
                if "date" in header_name:
                    if isinstance(val, str) and len(val) > 20 and "GMT" in val:
                        parts = val.split()
                        if len(parts) >= 4 and parts[0] in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]:
                            try:
                                dt = datetime.datetime.strptime(" ".join(parts[:4]), "%a %b %d %Y")
                                val = dt.strftime("%m/%d/%Y")
                            except Exception:
                                pass
                    if isinstance(val, str) and val.startswith("'"):
                        val = val[1:]
                        
                widget.delete(0, tk.END)
                widget.insert(0, val)
            elif isinstance(widget, tk.Text):
                widget.delete("1.0", tk.END)
                def insert_parsed_text(w, text, base_bold=False):
                    import re
                    tokens = re.split(r'(\[\[BOLD_START\]\]|\[\[BOLD_END\]\])', text)
                    current_bold = base_bold
                    for token in tokens:
                        if token == '[[BOLD_START]]':
                            current_bold = True
                        elif token == '[[BOLD_END]]':
                            current_bold = base_bold
                        elif token:
                            if current_bold:
                                w.insert(tk.END, token, "bold")
                            else:
                                w.insert(tk.END, token)

                if isinstance(val, CellRichText):
                    for part in val:
                        if isinstance(part, str):
                            insert_parsed_text(widget, part)
                        else:
                            font_is_bold = part.font and part.font.b
                            insert_parsed_text(widget, part.text, base_bold=font_is_bold)
                else:
                    if isinstance(val, datetime.datetime):
                        val = val.strftime("%m/%d/%Y")
                    else:
                        val = str(val) if val is not None else ""
                    insert_parsed_text(widget, val)
                try:
                    widget.edit_reset()
                except:
                    pass
                self.perform_spellcheck(widget)
                self.highlight_links(widget)
                
                # Bind KeyRelease to dynamically highlight links and spellcheck
                widget.bind("<KeyRelease>", lambda e, w=widget: [self.perform_spellcheck(w), self.highlight_links(w)], add="+")
                
        self._update_instrument_label(self.current_row_idx)
        self.save_btn.config(state="normal")
        self.open_pdf_btn.config(state="normal")
        status = self.row_statuses.get(str(self.current_row_idx), "Not Started")
        self.status_var.set(status)
        
        # Display specific warnings
        warnings = getattr(self, 'row_warnings', {}).get(str(self.current_row_idx), [])
        
        ai_res = ""
        self.current_ai_full_text = ""
        vol = str(self.ws.cell(row=self.current_row_idx, column=3).value or "").strip()
        pg = str(self.ws.cell(row=self.current_row_idx, column=4).value or "").strip()
        if vol and pg:
            cache_key = f"{vol}_{pg}"
            if hasattr(self, 'ai_qc_cache') and cache_key in self.ai_qc_cache:
                raw_text = self.ai_qc_cache[cache_key]
                if "SUMMARY:" in raw_text and "FULL TEXT:" in raw_text:
                    parts = raw_text.split("FULL TEXT:")
                    ai_res = "\n🤖 AI Check " + parts[0].replace("SUMMARY:", "").strip()
                    self.current_ai_full_text = parts[1].strip()
                else:
                    ai_res = "\n🤖 AI Check: Result formatted incorrectly. Hover to view."
                    self.current_ai_full_text = raw_text
                
        if hasattr(self, 'retry_qc_btn'):
            if vol and pg: self.retry_qc_btn.config(state="normal")
            else: self.retry_qc_btn.config(state="disabled")

        if warnings or ai_res:
            warn_text = ""
            if warnings: warn_text += "⚠️ " + " | ".join(warnings)
            if ai_res: warn_text += ai_res
            # Magic Auto-Injector for Releases found by AI QC
            if getattr(self, 'current_ai_full_text', ""):
                vol = str(self.ws.cell(row=self.current_row_idx, column=3).value or "").strip()
                pg = str(self.ws.cell(row=self.current_row_idx, column=4).value or "").strip()
                cache_key = f"{vol}_{pg}"
                if hasattr(self, 'ai_qc_cache') and cache_key in self.ai_qc_cache:
                    raw_text = self.ai_qc_cache[cache_key]
                    # Find patterns like VOL 941 PAGE 285, OR 941/285
                    import re
                    matches = re.finditer(r'(?:VOL(?:UME)?\s*|MR\s*|OR\s*|DR\s*)(\d+)\s*(?:PAGE|PG|/)?\s*(\d+)', raw_text, re.IGNORECASE)
                    for m in matches:
                        v = m.group(1)
                        p = m.group(2)
                        start = max(0, m.start() - 50)
                        end = min(len(raw_text), m.end() + 50)
                        context = raw_text[start:end].lower()
                        if 'release' in context or 'satisfaction' in context or 'see' in context:
                            new_release = f"Release: OR {v}/{p}"
                            # Check if notes_val already has this release
                            notes_val_tmp = ""
                            try:
                                for i, h in enumerate(self.headers):
                                    if "comment" in h.lower() or "note" in h.lower():
                                        widget = self.widgets_by_col.get(i)
                                        if isinstance(widget, __import__('tkinter').Text):
                                            notes_val_tmp = widget.get("1.0", "end").lower()
                                        break
                            except: pass
                            
                            # if no release in notes at all, or if this specific one is missing
                            if "release:" not in notes_val_tmp or (v not in notes_val_tmp and p not in notes_val_tmp):
                                try:
                                    for i, h in enumerate(self.headers):
                                        if "comment" in h.lower() or "note" in h.lower():
                                            widget = self.widgets_by_col.get(i)
                                            if isinstance(widget, __import__('tkinter').Text):
                                                current_text = widget.get("1.0", "end").strip()
                                                new_text = f"{current_text}\n\n{new_release}".strip()
                                                widget.delete("1.0", "end")
                                                widget.insert("1.0", new_text)
                                            break
                                except: pass
                            break # Only inject first found release
            
            self.warning_label.config(text=warn_text.strip())
        else:
            self.warning_label.config(text="")
        
        # Reset QC reminders
        if hasattr(self, 'qc_vars'):
            self.qc_vars[0].set(self.dower_reviewed.get(str(self.current_row_idx), False))
            self.qc_vars[1].set(False)
            self.qc_vars[2].set(self.maturity_reviewed.get(str(self.current_row_idx), False))
            if len(self.qc_vars) > 3:
                self.qc_vars[3].set(self.warnings_ignored.get(str(self.current_row_idx), False))
        if hasattr(self, 'release_entry'):
            self.release_entry.delete(0, 'end')
            self.release_entry.insert(0, 'Vol/Pg')
    def move_row_up(self):
        selection = self.listbox.curselection()
        if not selection or selection[0] == 0: return
        
        idx = selection[0]
        curr_row = self.row_indices[idx]
        prev_row = self.row_indices[idx - 1]
        
        for col in range(1, self.ws.max_column + 1):
            temp = self.ws.cell(row=curr_row, column=col).value
            self.ws.cell(row=curr_row, column=col).value = self.ws.cell(row=prev_row, column=col).value
            self.ws.cell(row=prev_row, column=col).value = temp
            
        curr_status = self.row_statuses.get(str(curr_row), "Not Started")
        prev_status = self.row_statuses.get(str(prev_row), "Not Started")
        self.row_statuses[str(curr_row)] = prev_status
        self.row_statuses[str(prev_row)] = curr_status
        self.save_statuses()
        
        self.wb.save(self.excel_path)
        self.load_rows()
        self.listbox.selection_set(idx - 1)
        self.on_select(None)
        
    def move_row_down(self):
        selection = self.listbox.curselection()
        if not selection or selection[0] == len(self.row_indices) - 1: return
        
        idx = selection[0]
        curr_row = self.row_indices[idx]
        next_row = self.row_indices[idx + 1]
        
        for col in range(1, self.ws.max_column + 1):
            temp = self.ws.cell(row=curr_row, column=col).value
            self.ws.cell(row=curr_row, column=col).value = self.ws.cell(row=next_row, column=col).value
            self.ws.cell(row=next_row, column=col).value = temp
            
        curr_status = self.row_statuses.get(str(curr_row), "Not Started")
        next_status = self.row_statuses.get(str(next_row), "Not Started")
        self.row_statuses[str(curr_row)] = next_status
        self.row_statuses[str(next_row)] = curr_status
        self.save_statuses()
        
        self.wb.save(self.excel_path)
        self.load_rows()
        self.listbox.selection_set(idx + 1)
        self.on_select(None)
        
    def add_new_row(self):
        selection = self.listbox.curselection()
        if selection:
            idx = selection[0]
            curr_row = self.row_indices[idx]
            self.ws.insert_rows(curr_row + 1)
            
            # Put a placeholder so it's not skipped as an "empty row"
            self.ws.cell(row=curr_row + 1, column=1, value="New Row")
            
            new_statuses = {}
            for r_str, status in self.row_statuses.items():
                r = int(r_str)
                if r > curr_row:
                    new_statuses[str(r + 1)] = status
                else:
                    new_statuses[str(r)] = status
            self.row_statuses = new_statuses
            self.save_statuses()
            
            self.wb.save(self.excel_path)
            self.load_rows()
            self.listbox.selection_set(idx + 1)
            self.listbox.see(idx + 1)
            self.on_select(None)
        else:
            # Add to the very end
            max_r = self.ws.max_row
            row_data = [""] * self.ws.max_column
            row_data[0] = "New Row"
            self.ws.append(row_data)
            self.wb.save(self.excel_path)
            self.load_rows()
            self.listbox.selection_set(tk.END)
            self.listbox.see(tk.END)
            self.on_select(None)
    def open_ogl_form(self):
        if not self.current_row_idx:
            from tkinter import messagebox
            messagebox.showerror("Error", "Please select a row first.")
            return
            
        dialog = tk.Toplevel(self)
        dialog.title("OGL Commentary Form")
        dialog.geometry("400x320")
        
        # Center dialog
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 400) // 2
        y = self.winfo_y() + (self.winfo_height() - 320) // 2
        dialog.geometry(f"+{x}+{y}")
        dialog.transient(self)
        dialog.grab_set()
        
        # Get default effective date
        eff_date_val = ""
        for i, header in enumerate(self.headers):
            if header.strip().lower() == "effective date":
                if i in self.widgets_by_col:
                    eff_date_val = self.widgets_by_col[i].get()
                break
                
        # Effective Date
        ttk.Label(dialog, text="Effective Date:").grid(row=0, column=0, padx=10, pady=10, sticky="e")
        eff_date_var = tk.StringVar(value=eff_date_val)
        ttk.Entry(dialog, textvariable=eff_date_var, width=15).grid(row=0, column=1, sticky="w")
        
        # Primary Term
        ttk.Label(dialog, text="Primary Term (Years):").grid(row=1, column=0, padx=10, pady=10, sticky="e")
        pt_var = tk.StringVar(value="5")
        ttk.Entry(dialog, textvariable=pt_var, width=10).grid(row=1, column=1, sticky="w")
        
        # Option to Renew
        ttk.Label(dialog, text="Option to Renew:").grid(row=2, column=0, padx=10, pady=10, sticky="e")
        renew_var = tk.StringVar(value="None")
        renew_cb = ttk.Combobox(dialog, textvariable=renew_var, values=["None", "Yes"], state="readonly", width=10)
        renew_cb.grid(row=2, column=1, sticky="w")
        
        # Option Years (only if Yes)
        ttk.Label(dialog, text="Renew Years:").grid(row=3, column=0, padx=10, pady=10, sticky="e")
        renew_years_var = tk.StringVar(value="5")
        renew_years_entry = ttk.Entry(dialog, textvariable=renew_years_var, width=10, state="disabled")
        renew_years_entry.grid(row=3, column=1, sticky="w")
        
        def on_renew_change(e):
            if renew_var.get() == "Yes":
                renew_years_entry.config(state="normal")
            else:
                renew_years_entry.config(state="disabled")
                
        renew_cb.bind("<<ComboboxSelected>>", on_renew_change)
        
        # Royalty
        ttk.Label(dialog, text="Royalty:").grid(row=4, column=0, padx=10, pady=10, sticky="e")
        royalty_var = tk.StringVar(value="Unknown")
        ttk.Entry(dialog, textvariable=royalty_var, width=15).grid(row=4, column=1, sticky="w")
        
        # Release Checkbox & Entry
        release_var = tk.BooleanVar(value=True)
        release_chk = ttk.Checkbutton(dialog, text="No release found of record", variable=release_var)
        release_chk.grid(row=5, column=0, columnspan=2, padx=20, pady=(10, 0), sticky="w")
        
        release_text_var = tk.StringVar(value="Release: ")
        release_entry = ttk.Entry(dialog, textvariable=release_text_var, width=30, state="disabled")
        release_entry.grid(row=6, column=0, columnspan=2, padx=20, pady=(0, 10), sticky="w")
        
        def on_release_toggle(*args):
            if release_var.get():
                release_entry.config(state="disabled")
            else:
                release_entry.config(state="normal")
        release_var.trace_add("write", on_release_toggle)
        
        def on_save():
            eff = eff_date_var.get().strip()
            pt = pt_var.get().strip()
            renew = renew_var.get()
            renew_years = renew_years_var.get().strip()
            royalty = royalty_var.get().strip()
            
            lines = []
            if pt: lines.append(f"{pt} Year PT")
            if eff: lines.append(f"Effective Date: {eff}")
            
            if renew == "Yes":
                lines.append(f"Option to Renew: {renew_years} Year")
            else:
                lines.append("Option to Renew: None")
                
            if royalty: lines.append(f"Royalty: {royalty}")
            
            if release_var.get():
                lines.append("No release found of record")
            else:
                rel_text = release_text_var.get().strip()
                if rel_text:
                    lines.append(rel_text)
            
            summary = "\n".join(lines)
            
            # Find the Notes text widget from widgets_by_col
            # The user's columns might have Notes anywhere, but typically we find it by column name
            notes_widget = None
            for idx, col_name in enumerate(self.headers):
                if "comment" in col_name.strip().lower() or "note" in col_name.strip().lower():
                    if idx in self.widgets_by_col:
                        notes_widget = self.widgets_by_col[idx]
                        break
                        
            if notes_widget and isinstance(notes_widget, tk.Text):
                current = notes_widget.get("1.0", tk.END).strip()
                if current:
                    notes_widget.insert(tk.END, "\n\n" + summary)
                else:
                    notes_widget.insert(tk.END, summary)
            else:
                from tkinter import messagebox
                messagebox.showerror("Error", "Could not find a Notes or Comments column to insert into.")
                
            dialog.destroy()
            
        ttk.Button(dialog, text="Insert", command=on_save).grid(row=7, column=0, columnspan=2, pady=20)
        
    def save_row(self, show_msg=True):
        if not self.current_row_idx: return

        row_cells = self.ws[self.current_row_idx]
        for i, widget in self.widgets_by_col.items():
            cell = self.ws.cell(row=self.current_row_idx, column=i+1)
            
            if isinstance(widget, ttk.Entry):
                val = widget.get().strip()
                if val == "" and cell.value is None:
                    continue
                    
                header_name = self.headers[i].lower()
                        
                # Parse dates
                if "date" in header_name:
                    try:
                        import dateutil.parser
                        dt = dateutil.parser.parse(val).date()
                        val = dt.strftime("%m/%d/%Y")
                    except Exception:
                        pass
                    
                    cell.value = val
                    if val:
                        cell.number_format = '@'
                        cell.data_type = 's'
                else:
                    # Format acreage
                    if "acreage" in header_name:
                        try:
                            val = float(val)
                            cell.number_format = '0.000000'
                        except ValueError:
                            pass
                    cell.value = val
                
            elif isinstance(widget, tk.Text):
                # Dump text and tags
                dump = widget.dump("1.0", tk.END, text=True, tag=True)
                
                # First, build a single full string with bold tags embedded
                full_text = ""
                current_tags = set()
                for item in dump:
                    if item[0] == 'tagon':
                        current_tags.add(item[1])
                        if item[1] == 'bold': full_text += '[[BOLD_START]]'
                    elif item[0] == 'tagoff':
                        if item[1] in current_tags:
                            current_tags.remove(item[1])
                            if item[1] == 'bold': full_text += '[[BOLD_END]]'
                    elif item[0] == 'text':
                        txt = item[1]
                        if item == dump[-1] and txt == '\n':
                            continue # Ignore trailing newline added by tk.Text
                        full_text += txt
                
                # Apply auto-formatting to the FULL text
                header_name = self.headers[i].lower()
                full_text = self.format_cell_text(header_name, full_text)
                
                # Now parse the full string back into parts
                parts = []
                is_bold_context = False
                import re
                token_splits = re.split(r'(\[\[BOLD_START\]\]|\[\[BOLD_END\]\])', full_text)
                for token in token_splits:
                    if token == '[[BOLD_START]]':
                        is_bold_context = True
                    elif token == '[[BOLD_END]]':
                        is_bold_context = False
                    elif token:
                        if is_bold_context:
                            parts.append(TextBlock(InlineFont(b=True), token))
                        else:
                            parts.append(token)
                
                if len(parts) == 0:
                    cell.value = "" if cell.value is not None else None
                elif not any(isinstance(p, TextBlock) for p in parts):
                    val = "".join(parts).strip()
                    if val == "" and cell.value is None:
                        continue
                    cell.value = val
                else:
                    cell.value = CellRichText(*parts)
                    
        # Check for auto-release cross-referencing
        try:
            import re
            comments_col = None
            vol_col = None
            pg_col = None
            book_col = None
            inst_col = None
            for i, h in enumerate(self.headers):
                hl = h.lower()
                if "comments" in hl: comments_col = i + 1
                elif "vol" in hl: vol_col = i + 1
                elif "page" in hl or "pg" in hl: pg_col = i + 1
                elif "book" in hl: book_col = i + 1
                elif "inst" in hl or "type" in hl: inst_col = i + 1
                
            if comments_col and vol_col and pg_col:
                current_comments = str(self.ws.cell(row=self.current_row_idx, column=comments_col).value or "")
                pattern = re.compile(r'Release(?:s)?\s+(?:of\s+)?mortgage\s+recorded\s+in\s+(?:(?:DR|OR|MR|LR|PR|PA|WR|MISC)\s*)?(\d+)[/-](\d+)', re.IGNORECASE)
                
                target_rows_to_update = []
                for m in pattern.finditer(current_comments):
                    t_vol = m.group(1)
                    t_pg = m.group(2)
                    
                    for r_idx in range(3, self.ws.max_row + 1):
                        if r_idx == self.current_row_idx: continue
                        r_vol = str(self.ws.cell(row=r_idx, column=vol_col).value or "").strip()
                        r_pg = str(self.ws.cell(row=r_idx, column=pg_col).value or "").strip()
                        
                        if r_vol == t_vol and r_pg == t_pg:
                            current_book = str(self.ws.cell(row=self.current_row_idx, column=book_col).value or "").strip() if book_col else ""
                            current_v = str(self.ws.cell(row=self.current_row_idx, column=vol_col).value or "").strip()
                            current_p = str(self.ws.cell(row=self.current_row_idx, column=pg_col).value or "").strip()
                            
                            # Do not cross-reference if the release row doesn't have a volume/page yet
                            if not current_v or not current_p:
                                continue
                                
                            release_text = f"Release: {current_book} {current_v}/{current_p}".strip()
                            
                            target_comments = self.ws.cell(row=r_idx, column=comments_col).value
                            tc_str = ""
                            if type(target_comments).__name__ == 'CellRichText':
                                tc_str = "".join(str(p) for p in target_comments)
                            else:
                                tc_str = str(target_comments or "").strip()
                                
                            # Only append if this volume isn't already cross-referenced to prevent infinite loops if user edits it
                            check_str = f"Release: {current_book} {current_v}"
                            if check_str not in tc_str:
                                if tc_str:
                                    tc_str += f"\n{release_text}"
                                else:
                                    tc_str = release_text
                                    
                                self.ws.cell(row=r_idx, column=comments_col).value = tc_str
                                target_rows_to_update.append(r_idx)
                                
        except Exception as e:
            print(f"Error auto-cross-referencing release: {e}")

        # Backup row to SQLite
        try:
            import sqlite3
            conn = sqlite3.connect(self.db_path)
            c = conn.cursor()
            vals = [self.current_row_idx]
            for col in range(1, 21):
                cell_val = self.ws.cell(row=self.current_row_idx, column=col).value
                if type(cell_val).__name__ == 'CellRichText':
                    cell_val = "".join(str(p) for p in cell_val)
                vals.append(str(cell_val) if cell_val is not None else "")
            vals.append(self.row_statuses.get(str(self.current_row_idx), "Not Started"))
            
            # Update listbox emoji
            idx = self.row_indices.index(self.current_row_idx)
            text = self.listbox.get(idx)
            if hasattr(self, 'ai_title_cache') and "🔍" in text:
                vol = str(self.ws.cell(row=self.current_row_idx, column=3).value or "").strip()
                pg = str(self.ws.cell(row=self.current_row_idx, column=4).value or "").strip()
                inst = str(self.ws.cell(row=self.current_row_idx, column=1).value or "").strip().lower()
                ck = f"{vol}_{pg}"
                if ck in self.ai_title_cache:
                    c_title = self.ai_title_cache[ck].lower()
                    if inst == c_title:
                        self.listbox.delete(idx)
                        self.listbox.insert(idx, text.replace(" 🔍", ""))
                        self._apply_row_color(idx)
            
            self._update_instrument_label(self.current_row_idx)
                        
            c.execute('''INSERT INTO runsheet (row_idx, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12, col_13, col_14, col_15, col_16, col_17, col_18, col_19, col_20, status)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                         ON CONFLICT(row_idx) DO UPDATE SET
                         col_1=excluded.col_1, col_2=excluded.col_2, col_3=excluded.col_3, col_4=excluded.col_4, col_5=excluded.col_5,
                         col_6=excluded.col_6, col_7=excluded.col_7, col_8=excluded.col_8, col_9=excluded.col_9, col_10=excluded.col_10,
                         col_11=excluded.col_11, col_12=excluded.col_12, col_13=excluded.col_13, col_14=excluded.col_14, col_15=excluded.col_15,
                         col_16=excluded.col_16, col_17=excluded.col_17, col_18=excluded.col_18, col_19=excluded.col_19, col_20=excluded.col_20,
                         status=excluded.status, timestamp=CURRENT_TIMESTAMP''', vals)
                         
            # Also backup target rows that were auto-updated
            if 'target_rows_to_update' in locals():
                for tr_idx in target_rows_to_update:
                    t_vals = [tr_idx]
                    for col in range(1, 21):
                        cell_val = self.ws.cell(row=tr_idx, column=col).value
                        if type(cell_val).__name__ == 'CellRichText':
                            cell_val = "".join(str(p) for p in cell_val)
                        t_vals.append(str(cell_val) if cell_val is not None else "")
                    t_vals.append(self.row_statuses.get(str(tr_idx), "Not Started"))
                    c.execute('''INSERT INTO runsheet (row_idx, col_1, col_2, col_3, col_4, col_5, col_6, col_7, col_8, col_9, col_10, col_11, col_12, col_13, col_14, col_15, col_16, col_17, col_18, col_19, col_20, status)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                         ON CONFLICT(row_idx) DO UPDATE SET
                         col_1=excluded.col_1, col_2=excluded.col_2, col_3=excluded.col_3, col_4=excluded.col_4, col_5=excluded.col_5,
                         col_6=excluded.col_6, col_7=excluded.col_7, col_8=excluded.col_8, col_9=excluded.col_9, col_10=excluded.col_10,
                         col_11=excluded.col_11, col_12=excluded.col_12, col_13=excluded.col_13, col_14=excluded.col_14, col_15=excluded.col_15,
                         col_16=excluded.col_16, col_17=excluded.col_17, col_18=excluded.col_18, col_19=excluded.col_19, col_20=excluded.col_20,
                         status=excluded.status, timestamp=CURRENT_TIMESTAMP''', t_vals)
                         
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"Failed to backup to DB: {e}")
            
        # If this row is a Mortgage, update the tracking Note automatically!
        try:
            vol, pg, inst, grantor, grantee = "", "", "", "", ""
            for i, h in enumerate(self.headers):
                h_lower = h.lower()
                val = self.ws.cell(row=self.current_row_idx, column=i+1).value
                if val is not None:
                    val_str = str(val)
                    if type(val).__name__ == 'CellRichText':
                        val_str = "".join(str(p) for p in val)
                    if "vol" in h_lower: vol = val_str
                    elif "page" in h_lower or "pg" in h_lower: pg = val_str
                    elif "inst" in h_lower or "type" in h_lower: inst = val_str
                    elif "grantor" in h_lower: grantor = val_str
                    elif "grantee" in h_lower: grantee = val_str

            if "mortgage" in inst.lower() and "release" not in inst.lower() and "satisfaction" not in inst.lower():
                is_released = self.qc_vars[1].get() if len(self.qc_vars) > 1 else False
                release_val = getattr(self, 'release_vol_pg_var', None)
                release_val = release_val.get().strip() if release_val else ""
                if release_val == "Vol/Pg": release_val = ""
                
                import json, datetime
                pid_dir = os.path.dirname(self.excel_path)
                notes_dir = os.path.join(pid_dir, "NOTES")
                os.makedirs(notes_dir, exist_ok=True)
                notes_file = os.path.join(notes_dir, "notes.json")
                
                notes = []
                if os.path.exists(notes_file):
                    with open(notes_file, 'r') as f:
                        try: notes = json.load(f)
                        except: pass
                        
                mortgage_note_idx = None
                for i, note in enumerate(notes):
                    if note.get("subject") == "Mortgage Tracking":
                        mortgage_note_idx = i
                        break
                        
                if mortgage_note_idx is None:
                    notes.append({
                        "subject": "Mortgage Tracking",
                        "content": "List of Mortgages and their Release statuses:\n\n",
                        "date": datetime.datetime.now().strftime("%Y-%m-%d"),
                        "time": datetime.datetime.now().strftime("%I:%M %p")
                    })
                    mortgage_note_idx = len(notes) - 1
                    
                content = notes[mortgage_note_idx].get("content", "")
                lines = content.split('\n')
                
                mortgage_link = f"<link:{vol}-{pg}>"
                entry_prefix = f"• Mortgage {mortgage_link} "
                if is_released:
                    new_status = f"[RELEASED by {release_val}]" if release_val else "[RELEASED]"
                else:
                    new_status = "[UNRELEASED]"
                
                # Use a special tag pattern for hyperlinks
                mortgage_link = f"<link:{vol}-{pg}>"
                release_link = f"<link:{release_val.replace(' ', '').replace('/', '-')}>" if release_val else ""
                
                if release_val:
                    new_line = f"• Mortgage {mortgage_link} - {grantor} to {grantee} -> [RELEASED by {release_link}]"
                else:
                    new_line = f"• Mortgage {mortgage_link} - {grantor} to {grantee} -> {new_status}"
                
                found = False
                for i, line in enumerate(lines):
                    if line.startswith(entry_prefix):
                        lines[i] = new_line
                        found = True
                        break
                        
                if not found:
                    lines.append(new_line)
                    
                notes[mortgage_note_idx]["content"] = '\n'.join(lines)
                notes[mortgage_note_idx]["date"] = datetime.datetime.now().strftime("%Y-%m-%d")
                notes[mortgage_note_idx]["time"] = datetime.datetime.now().strftime("%I:%M %p")
                
                with open(notes_file, 'w') as f:
                    json.dump(notes, f, indent=4)
        except Exception as e:
            print(f"Failed to update mortgage note: {e}")

        try:
            if hasattr(self, 'formatted_state_file') and not os.path.exists(self.formatted_state_file):
                import json
                with open(self.formatted_state_file, "w") as f:
                    json.dump({"initial_formatting_completed": True}, f)
        except Exception: pass

        try:
            self.wb.save(self.excel_path)
            if show_msg:
                self.show_save_success_dialog(self.current_row_idx)
            
            # Refresh listbox display without losing selection
            sel = self.listbox.curselection()
            self.load_rows()
            if sel:
                self.listbox.selection_set(sel[0])
        except PermissionError:
            messagebox.showerror("Error", f"Permission denied. Make sure the file is completely closed in Excel and not read-only.", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"Error saving file:\n{e}", parent=self)

    def schedule_spellcheck(self, event):
        widget = event.widget
        if hasattr(widget, '_spell_timer'):
            widget.after_cancel(widget._spell_timer)
        widget._spell_timer = widget.after(500, lambda: self.perform_spellcheck(widget))
        
    def get_local_docs(self):
        if hasattr(self, '_local_docs_cache'):
            return self._local_docs_cache
            
        import glob
        all_docs = []
        for ext in ("*.pdf", "*.txt", "*.doc", "*.docx", "*.rtf", "*.png", "*.jpg", "*.tif", "*.tiff"):
            all_docs.extend(glob.glob(os.path.join(self.pid_dir, "**", ext), recursive=True))
            all_docs.extend(glob.glob(os.path.join(self.base_dir, ext)))
            
        import re
        valid_pairs = set()
        for doc in all_docs:
            fname = os.path.basename(doc)
            if not fname.startswith("._"):
                m = re.search(r'(\d+)_(\d+)', fname)
                if m: valid_pairs.add(f"{m.group(1)}-{m.group(2)}")
                m = re.search(r'(\d+)-(\d+)', fname)
                if m: valid_pairs.add(f"{m.group(1)}-{m.group(2)}")
                
        self._local_docs_cache = valid_pairs
        return valid_pairs

    def highlight_links(self, widget):
        widget.tag_remove("hyperlink", "1.0", "end")
        text = widget.get("1.0", "end-1c")
        if not text: return
        
        valid_pairs = self.get_local_docs()
        if not valid_pairs: return
        
        widget.tag_configure("hyperlink", foreground="blue", underline=True)
        widget.tag_bind("hyperlink", "<Enter>", lambda e: widget.config(cursor="hand2"))
        widget.tag_bind("hyperlink", "<Leave>", lambda e: widget.config(cursor=""))
        
        def on_click(event):
            index = widget.index(f"@{event.x},{event.y}")
            tags = widget.tag_names(index)
            for t in tags:
                if t.startswith("link_val_"):
                    vol_pg = t.split("link_val_")[1]
                    vol, pg = vol_pg.split('-')
                    
                    vol_col = None
                    pg_col = None
                    for i, h in enumerate(self.headers):
                        if "vol" in h.lower() or "book" in h.lower():
                            vol_col = i + 1
                        elif "page" in h.lower():
                            pg_col = i + 1
                            
                    if not vol_col or not pg_col: return
                    
                    for row_idx in range(3, self.ws.max_row + 1):
                        r_vol = str(self.ws.cell(row=row_idx, column=vol_col).value or "").strip()
                        r_pg = str(self.ws.cell(row=row_idx, column=pg_col).value or "").strip()
                        if r_vol == vol and r_pg == pg:
                            # Implicitly save current row before jumping
                            if self.current_row_idx:
                                try: self.save_row(show_msg=False)
                                except Exception as e: print("SAVE ROW FAILED DURING JUMP:", e)
                            if row_idx not in self.row_indices:
                                self.search_var.set("")
                                self.filter_var.set("All")
                                self.load_rows()
                                
                            if row_idx in self.row_indices:
                                lb_idx = self.row_indices.index(row_idx)
                                self.listbox.selection_clear(0, tk.END)
                                self.listbox.selection_set(lb_idx)
                                self.listbox.see(lb_idx)
                                self.on_select(None)
                            return
        
        widget.tag_bind("hyperlink", "<Button-1>", on_click)
        
        import re
        for vpg in valid_pairs:
            v, p = vpg.split('-')
            
            # Find "Vol X Pg Y"
            pattern1 = re.compile(r'Vol(?:ume|\.)?\s*' + v + r'\b\s*P(?:a)?g(?:e|\.)?\s*' + p + r'\b', re.IGNORECASE)
            for match in pattern1.finditer(text):
                start_idx = f"1.0+{match.start()}c"
                end_idx = f"1.0+{match.end()}c"
                widget.tag_add("hyperlink", start_idx, end_idx)
                widget.tag_add(f"link_val_{vpg}", start_idx, end_idx)
                
            # Find "X-Y"
            pattern2 = re.compile(r'(?<!\d)' + v + r'-' + p + r'(?!\d)')
            for match in pattern2.finditer(text):
                start_idx = f"1.0+{match.start()}c"
                end_idx = f"1.0+{match.end()}c"
                widget.tag_add("hyperlink", start_idx, end_idx)
                widget.tag_add(f"link_val_{vpg}", start_idx, end_idx)
                
            # Find "X/Y"
            pattern3 = re.compile(r'(?<!\d)' + v + r'/' + p + r'(?!\d)')
            for match in pattern3.finditer(text):
                start_idx = f"1.0+{match.start()}c"
                end_idx = f"1.0+{match.end()}c"
                widget.tag_add("hyperlink", start_idx, end_idx)
                widget.tag_add(f"link_val_{vpg}", start_idx, end_idx)
                
    def perform_spellcheck(self, widget):
        widget.tag_remove("misspelled", "1.0", "end")
        text = widget.get("1.0", "end-1c")
        if not text:
            return
            
        words = re.finditer(r'\b[a-zA-Z]+\b', text)
        for match in words:
            word = match.group()
            if word and word.lower() not in self.spell:
                start_idx = f"1.0+{match.start()}c"
                end_idx = f"1.0+{match.end()}c"
                widget.tag_add("misspelled", start_idx, end_idx)
                
    def open_hyperlinked_doc(self, vol_pg):
        vol, pg = vol_pg.split('-')
        import glob, subprocess
        import os
        
        local_docs = []
        for ext in ("*.pdf", "*.txt", "*.doc", "*.docx", "*.rtf", "*.png", "*.jpg", "*.tif", "*.tiff"):
            local_docs.extend(glob.glob(os.path.join(self.pid_dir, "**", ext), recursive=True))
            local_docs.extend(glob.glob(os.path.join(self.base_dir, ext)))
            
        for doc in local_docs:
            fname = os.path.basename(doc)
            if f"{vol}-{pg}" in fname or f"{vol}_{pg}" in fname:
                try:
                    subprocess.run(["open", doc])
                    return
                except: pass
                
        vol_pad = vol.zfill(3) if vol.isdigit() else vol
        vol_pad_4 = vol.zfill(4) if vol.isdigit() else vol
        archives = [
            f"/Volumes/davidlls/drive/DEEDS/DEED {vol_pad}",
            f"/Volumes/davidlls/drive/MTGS/MTG {vol_pad}",
            f"/Volumes/davidlls/drive/extracted/DEEDS/DEED {vol_pad}",
            f"/Volumes/davidlls/drive/extracted/MTGS/MTG {vol_pad}",
            f"/Volumes/davidlls/Belmont_Drive_External/Belmont County Court House/2. Belmont Deeds/DEED {vol_pad}",
            f"/Volumes/davidlls/Belmont_Drive_External/Belmont County Court House/3. Belmont Leases/{vol_pad_4}"
        ]
        for archive_dir in archives:
            if os.path.exists(archive_dir):
                for ext in ("*.pdf", "*.tif", "*.jpg"):
                    for doc in glob.glob(os.path.join(archive_dir, f"*{vol}-{pg}*{ext}")):
                        try:
                            subprocess.run(["open", doc])
                            return
                        except: pass
                    for doc in glob.glob(os.path.join(archive_dir, f"*{vol}_{pg}*{ext}")):
                        try:
                            subprocess.run(["open", doc])
                            return
                        except: pass

    def show_context_menu(self, event):
        widget = event.widget
        menu = tk.Menu(self, tearoff=0)
        
        has_selection = False
        linked_vol_pg = None
        try:
            if isinstance(widget, tk.Text):
                has_selection = widget.tag_ranges(tk.SEL)
                index = widget.index(f"@{event.x},{event.y}")
                tags = widget.tag_names(index)
                for t in tags:
                    if t.startswith("link_val_"):
                        linked_vol_pg = t.split("link_val_")[1]
                        break
            else:
                has_selection = widget.selection_present()
        except: pass
        
        if linked_vol_pg:
            menu.add_command(label="Open Document", command=lambda: self.open_hyperlinked_doc(linked_vol_pg))
            menu.add_separator()

        def to_title_case():
            if not has_selection: return
            try:
                if isinstance(widget, tk.Text):
                    start = widget.index(tk.SEL_FIRST)
                    end = widget.index(tk.SEL_LAST)
                    text = widget.get(start, end)
                else:
                    text = widget.selection_get()
                    start = widget.index(tk.SEL_FIRST)
                    end = widget.index(tk.SEL_LAST)
                    
                # Smart title casing
                minor_words = {'a', 'an', 'and', 'as', 'at', 'but', 'by', 'for', 'in', 'nor', 'of', 'on', 'or', 'so', 'the', 'to', 'up', 'yet', 'et', 'al'}
                words = text.split(' ')
                cased_words = []
                for i, word in enumerate(words):
                    if not word:
                        cased_words.append('')
                        continue
                    
                    lower_word = word.lower()
                    if i > 0 and i < len(words) - 1 and lower_word in minor_words:
                        cased_words.append(lower_word)
                    else:
                        cased_words.append(word.title())
                        
                smart_text = ' '.join(cased_words)
                
                widget.delete(start, end)
                widget.insert(start, smart_text)
            except Exception as e: print(e)

        if has_selection:
            menu.add_command(label="To Title Case (Normal)", command=to_title_case)
            menu.add_separator()
            
        menu.add_command(label="Copy", command=lambda: widget.event_generate("<<Copy>>"))
        menu.add_command(label="Paste", command=lambda: widget.event_generate("<<Paste>>"))
        
        if isinstance(widget, tk.Text):
            try:
                index = widget.index(f"@{event.x},{event.y}")
                tags = widget.tag_names(index)
                if "misspelled" in tags:
                    word_start = widget.index(f"{index} wordstart")
                    word_end = widget.index(f"{index} wordend")
                    word = widget.get(word_start, word_end)
                    if word:
                        suggestions = self.spell.candidates(word)
                        if suggestions:
                            menu.add_separator()
                            for s in list(suggestions)[:5]:
                                def replace_word(s=s, start=word_start, end=word_end, w=widget):
                                    w.delete(start, end)
                                    w.insert(start, s)
                                    self.perform_spellcheck(w)
                                menu.add_command(label=f"Spelling: {s}", command=replace_word)
            except Exception: pass
            
        try:
            vol = str(self.ws.cell(row=self.current_row_idx, column=3).value or "").strip()
            pg = str(self.ws.cell(row=self.current_row_idx, column=4).value or "").strip()
            if vol and pg:
                cache_key = f"{vol}_{pg}"
                if cache_key in self.ai_title_cache:
                    title_suggestion = self.ai_title_cache[cache_key]
                    def apply_title():
                        import tkinter.ttk as ttk
                        if isinstance(widget, tk.Text):
                            widget.delete("1.0", tk.END)
                            widget.insert("1.0", title_suggestion)
                        else:
                            widget.delete(0, tk.END)
                            widget.insert(0, title_suggestion)
                    menu.add_separator()
                    menu.add_command(label=f"✨ Suggest Title: {title_suggestion}", command=apply_title)
                else:
                    def fetch_title():
                        import threading
                        def task():
                            import ai_parser
                            import os
                            import tkinter.ttk as ttk
                            docs_dir = os.path.join(self.pid_dir, "DOCS")
                            target_pdf = None
                            for root, _, files in os.walk(docs_dir):
                                if "Irrelevant" in root: continue
                                for f in files:
                                    if f.lower().endswith(".pdf") and vol in f and pg in f:
                                        target_pdf = os.path.join(root, f)
                                        break
                                if target_pdf: break
                            if not target_pdf: return
                            
                            api_key = ""
                            try:
                                config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
                                import json
                                with open(config_path, "r") as f:
                                    api_key = json.load(f).get("GEMINI_API_KEY", "")
                            except: pass
                            if not api_key: return
                            
                            t = ai_parser.extract_exact_title(api_key, target_pdf)
                            if t:
                                self.ai_title_cache[cache_key] = t
                                self.save_ai_title_cache()
                                def do_update():
                                    if isinstance(widget, tk.Text):
                                        widget.delete("1.0", tk.END)
                                        widget.insert("1.0", t)
                                    else:
                                        widget.delete(0, tk.END)
                                        widget.insert(0, t)
                                self.after(0, do_update)
                        
                        threading.Thread(target=task, daemon=True).start()
                        
                    menu.add_separator()
                    menu.add_command(label="✨ Ask AI for Document Title (Wait 5s)", command=fetch_title)
        except Exception: pass
            
        menu.post(event.x_root, event.y_root)

    def load_statuses(self):
        import json
        if os.path.exists(self.status_file):
            try:
                with open(self.status_file, "r") as f:
                    return json.load(f)
            except Exception:
                return {}
        return {}
        
    def save_statuses(self):
        import json
        try:
            with open(self.status_file, "w") as f:
                json.dump(self.row_statuses, f, indent=4)
        except Exception as e:
            print(f"Error saving statuses: {e}")
            
    def load_trash(self):
        import json
        if os.path.exists(self.trash_file):
            try:
                with open(self.trash_file, "r") as f:
                    return json.load(f)
            except Exception:
                pass
        return []

    def save_trash(self):
        import json
        try:
            with open(self.trash_file, "w") as f:
                json.dump(self.trash_rows, f, indent=4)
        except Exception as e:
            print(f"Error saving trash: {e}")

    def delete_row(self):
        selection = self.listbox.curselection()
        if not selection: return
        
        idx = selection[0]
        curr_row = self.row_indices[idx]
        
        inst = str(self.ws.cell(row=curr_row, column=1).value or "")
        vol = str(self.ws.cell(row=curr_row, column=3).value or "")
        pg = str(self.ws.cell(row=curr_row, column=4).value or "")
        preview = f"{inst} {vol}/{pg}".strip() or f"Row {curr_row}"
        
        if not messagebox.askyesno("Delete Row", f"Are you sure you want to delete this row ({preview})? It will be moved to the Trash Bin.", parent=self):
            return
            
        row_data = {}
        for col in range(1, self.ws.max_column + 1):
            val = self.ws.cell(row=curr_row, column=col).value
            if isinstance(val, datetime.datetime):
                val = val.strftime("%m/%d/%Y")
            row_data[str(col)] = str(val) if val is not None else ""
            
        # Move associated PDF to Irrelevant
        import shutil
        moved_files = []
        if vol and pg:
            docs_dir = os.path.join(self.pid_dir, "DOCS")
            irrel_dir = os.path.join(docs_dir, "Irrelevant")
            if os.path.exists(docs_dir):
                for root, _, files in os.walk(docs_dir):
                    if "Irrelevant" in root: continue
                    for f in files:
                        if not f.lower().endswith(".pdf"): continue
                        if vol in f and pg in f:
                            src = os.path.join(root, f)
                            os.makedirs(irrel_dir, exist_ok=True)
                            dest = os.path.join(irrel_dir, f)
                            try:
                                shutil.move(src, dest)
                                moved_files.append({"filename": f, "original_dir": root})
                            except Exception: pass
                            
        trash_item = {
            "deleted_at": datetime.datetime.now().isoformat(),
            "preview": preview,
            "data": row_data,
            "moved_files": moved_files
        }
        self.trash_rows.append(trash_item)
        self.save_trash()
        
        self.ws.delete_rows(curr_row)
        
        new_statuses = {}
        for r_str, status in self.row_statuses.items():
            r = int(r_str)
            if r > curr_row:
                new_statuses[str(r - 1)] = status
            elif r < curr_row:
                new_statuses[str(r)] = status
        self.row_statuses = new_statuses
        self.save_statuses()
        
        self.wb.save(self.excel_path)
        self.load_rows()
        if idx < self.listbox.size():
            self.listbox.selection_set(idx)
        elif self.listbox.size() > 0:
            self.listbox.selection_set(self.listbox.size() - 1)
        self.on_select(None)
        
    def view_trash(self):
        popup = tk.Toplevel(self)
        popup.title("Trash Bin")
        popup.geometry("600x400")
        
        tree = ttk.Treeview(popup, columns=("Date", "Preview"), show="headings")
        tree.heading("Date", text="Deleted Date")
        tree.heading("Preview", text="Row Preview")
        tree.column("Date", width=150)
        tree.column("Preview", width=400)
        
        for i, item in enumerate(self.trash_rows):
            dt_str = item.get("deleted_at", "")[:19].replace("T", " ")
            tree.insert("", tk.END, iid=str(i), values=(dt_str, item.get("preview", "Row")))
            
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        def restore():
            sel = tree.selection()
            if not sel: return
            idx = int(sel[0])
            item = self.trash_rows.pop(idx)
            self.save_trash()
            
            new_row = self.ws.max_row + 1
            for col_str, val in item.get("data", {}).items():
                self.ws.cell(row=new_row, column=int(col_str)).value = val
                
            # Restore files
            import shutil
            import os
            moved_files = item.get("moved_files", [])
            for f_info in moved_files:
                fname = f_info.get("filename")
                orig_dir = f_info.get("original_dir")
                if fname and orig_dir:
                    irrel_path = os.path.join(self.pid_dir, "DOCS", "Irrelevant", fname)
                    if os.path.exists(irrel_path):
                        try:
                            os.makedirs(orig_dir, exist_ok=True)
                            shutil.move(irrel_path, os.path.join(orig_dir, fname))
                        except Exception: pass
                
            self.wb.save(self.excel_path)
            self.load_rows()
            popup.destroy()
            messagebox.showinfo("Restored", f"Row restored to the bottom (Row {new_row}).", parent=self)
            
        ttk.Button(popup, text="Restore Selected", command=restore).pack(pady=10)
            
    def on_status_change(self, event=None):
        if not self.current_row_idx: return
        new_status = self.status_var.get()
        self.row_statuses[str(self.current_row_idx)] = new_status
        self.save_statuses()
        
        # Auto-save current row edits and excel workbook silently
        try:
            self.save_row(show_msg=False)
        except Exception as e:
            print(f"Auto-save on status change error: {e}")
        
        # update listbox color without full reload
        sel = self.listbox.curselection()
        if sel:
            idx = sel[0]
            r = self.row_indices[idx]
            has_warn = str(r) in getattr(self, 'row_warnings', {})
            fg_color = 'red' if has_warn and new_status != "Completed" else ''
            
            if new_status == "In Progress":
                self.listbox.itemconfig(idx, {'bg': '#fff3cd', 'fg': fg_color or 'black'})
            elif new_status == "Completed":
                self.listbox.itemconfig(idx, {'bg': '#d4edda', 'fg': fg_color or 'black'})
            else:
                self.listbox.itemconfig(idx, {'bg': '', 'fg': fg_color})
                
    def load_ai_cache(self):
        import json
        if os.path.exists(self.ai_cache_file):
            try:
                with open(self.ai_cache_file, "r") as f:
                    return json.load(f)
            except Exception:
                pass
        return {}

    def save_ai_cache(self):
        import json
        try:
            with open(self.ai_cache_file, "w") as f:
                json.dump(self.ai_qc_cache, f, indent=4)
        except Exception as e:
            print(f"Error saving ai cache: {e}")

    def save_ai_title_cache(self):
        import json
        try:
            with open(self.ai_title_cache_file, "w") as f:
                json.dump(self.ai_title_cache, f, indent=4)
        except Exception as e:
            print(f"Error saving title cache: {e}")

    def auto_run_ai_titles(self):
        if getattr(self, "ai_title_running", False): return
        self.ai_title_running = True
        import threading
        threading.Thread(target=self._ai_title_worker, daemon=True).start()

    def _ai_title_worker(self):
        import time, json, os
        import ai_parser
        
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
        api_key = ""
        if os.path.exists(config_path):
            try:
                with open(config_path, "r") as f:
                    config = json.load(f)
                    api_key = config.get("GEMINI_API_KEY", "")
            except: pass
        if not api_key:
            self.ai_title_running = False
            return
            
        docs_dir = os.path.join(self.pid_dir, "DOCS")
        if not os.path.exists(docs_dir):
            self.ai_title_running = False
            return
            
        indices = list(self.row_indices)
        for i, r in enumerate(indices):
            vol = str(self.ws.cell(row=r, column=3).value or "").strip()
            pg = str(self.ws.cell(row=r, column=4).value or "").strip()
            if not vol or not pg: continue
            
            cache_key = f"{vol}_{pg}"
            if cache_key in self.ai_title_cache: continue
            
            target_pdf = None
            for root, _, files in os.walk(docs_dir):
                if "Irrelevant" in root: continue
                for f in files:
                    if f.lower().endswith(".pdf") and vol in f and pg in f:
                        target_pdf = os.path.join(root, f)
                        break
                if target_pdf: break
                
            if target_pdf:
                self.after(0, self._set_loading_icon, r)
                title = ai_parser.extract_exact_title(api_key, target_pdf)
                
                if title:
                    self.ai_title_cache[cache_key] = title
                    self.save_ai_title_cache()
                    
                    def update_ui_for_title(row_idx=r):
                        self._remove_loading_icon(row_idx, add_check=False)
                        if row_idx not in self.row_indices: return
                        idx = self.row_indices.index(row_idx)
                        text = self.listbox.get(idx)
                        if "🔍" not in text:
                            v = str(self.ws.cell(row=row_idx, column=3).value or "").strip()
                            p = str(self.ws.cell(row=row_idx, column=4).value or "").strip()
                            i_val = str(self.ws.cell(row=row_idx, column=1).value or "").strip().lower()
                            if v and p:
                                ck = f"{v}_{p}"
                                if ck in self.ai_title_cache:
                                    c_title = self.ai_title_cache[ck].lower()
                                    if c_title and i_val != c_title:
                                        self.listbox.delete(idx)
                                        self.listbox.insert(idx, text + " 🔍")
                                        self._apply_row_color(idx)
                        if getattr(self, 'current_row_idx', None) == row_idx:
                            self._update_instrument_label(row_idx)
                    self.after(0, update_ui_for_title)
                else:
                    self.after(0, lambda r=r: self._remove_loading_icon(r, add_check=False))
                    
                time.sleep(4.5)
                
        self.ai_title_running = False

    def auto_run_ai_qc(self):
        if getattr(self, "ai_qc_running", False): return
        self.ai_qc_running = True
        import threading
        threading.Thread(target=self._ai_qc_worker, daemon=True).start()

    def _ai_qc_worker(self):
        import time, json, os, datetime
        import ai_parser
        
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
        api_key = ""
        if os.path.exists(config_path):
            try:
                with open(config_path, "r") as f:
                    config = json.load(f)
                    api_key = config.get("GEMINI_API_KEY", "")
            except: pass
        if not api_key:
            self.ai_qc_running = False
            return
            
        docs_dir = os.path.join(self.pid_dir, "DOCS")
        if not os.path.exists(docs_dir):
            self.ai_qc_running = False
            return
            
        # Copy current indices so we aren't affected by live changes immediately
        indices = list(self.row_indices)
        for i, r in enumerate(indices):
            inst = str(self.ws.cell(row=r, column=1).value or "").strip().lower()
            vol = str(self.ws.cell(row=r, column=3).value or "").strip()
            pg = str(self.ws.cell(row=r, column=4).value or "").strip()
            eff_dt = self.ws.cell(row=r, column=8).value
            notes = str(self.ws.cell(row=r, column=14).value or "").lower()
            
            if not vol or not pg: continue
            
            is_deed = "deed" in inst and "release" not in inst and "satisfaction" not in inst
            is_mortgage = "mortgage" in inst and "release" not in inst and "satisfaction" not in inst
            
            if not is_deed and not is_mortgage: continue
            
            needs_qc = False
            if is_deed:
                if "dower" not in notes: needs_qc = True
            if is_mortgage:
                age_days = 0
                if eff_dt and isinstance(eff_dt, datetime.datetime):
                    age_days = (datetime.datetime.now() - eff_dt).days
                elif eff_dt and isinstance(eff_dt, str):
                    try:
                        d = datetime.datetime.strptime(eff_dt, "%m/%d/%Y")
                        age_days = (datetime.datetime.now() - d).days
                    except: pass
                if age_days <= (30*365):
                    if "dower" not in notes or "maturity" not in notes or "release" not in notes:
                        needs_qc = True
                        
            if needs_qc:
                cache_key = f"{vol}_{pg}"
                if cache_key in self.ai_qc_cache: continue
                
                target_pdf = None
                for root, _, files in os.walk(docs_dir):
                    if "Irrelevant" in root: continue
                    for f in files:
                        if f.lower().endswith(".pdf") and vol in f and pg in f:
                            target_pdf = os.path.join(root, f)
                            break
                    if target_pdf: break
                    
                if target_pdf:
                    self.after(0, self._set_loading_icon, r)
                    
                    result = ai_parser.generate_qc_check(api_key, target_pdf, inst)
                    self.ai_qc_cache[cache_key] = result
                    self.save_ai_cache()
                    
                    self.after(0, self._remove_loading_icon, r)
                    
                    # Update currently selected row if it matches
                    if self.current_row_idx == r:
                        self.after(0, lambda: self.on_select(None))
                        
                    time.sleep(4.5)
                    
        self.ai_qc_running = False

    def _set_loading_icon(self, row_idx):
        if row_idx not in self.row_indices: return
        idx = self.row_indices.index(row_idx)
        text = self.listbox.get(idx)
        if "⏳" not in text:
            self.listbox.delete(idx)
            self.listbox.insert(idx, "⏳ " + text)
            self._apply_row_color(idx)
            
    def _remove_loading_icon(self, row_idx, add_check=True):
        if row_idx not in self.row_indices: return
        idx = self.row_indices.index(row_idx)
        text = self.listbox.get(idx)
        if "⏳" in text:
            new_text = text.replace("⏳ ", "")
            if add_check and " ✅" not in new_text:
                new_text += " ✅"
            self.listbox.delete(idx)
            self.listbox.insert(idx, new_text)
            self._apply_row_color(idx)
            
    def retry_ai_check(self):
        if not self.current_row_idx: return
        r = self.current_row_idx
        inst = str(self.ws.cell(row=r, column=1).value or "").strip().lower()
        vol = str(self.ws.cell(row=r, column=3).value or "").strip()
        pg = str(self.ws.cell(row=r, column=4).value or "").strip()
        if not vol or not pg: return
        
        cache_key = f"{vol}_{pg}"
        if cache_key in self.ai_qc_cache:
            del self.ai_qc_cache[cache_key]
            self.save_ai_cache()
            
        if hasattr(self, 'ai_title_cache') and cache_key in self.ai_title_cache:
            del self.ai_title_cache[cache_key]
            self.save_ai_title_cache()
            
        # Remove ✅ and 🔍 if present
        idx = self.row_indices.index(r)
        text = self.listbox.get(idx)
        new_text = text.replace(" ✅", "").replace(" 🔍", "")
        if new_text != text:
            self.listbox.delete(idx)
            self.listbox.insert(idx, new_text)
            self.listbox.selection_set(idx)
        
        self.on_select(None)
        
        import threading
        threading.Thread(target=self._single_ai_qc_worker, args=(r, inst, vol, pg, cache_key), daemon=True).start()
            
    def _single_ai_qc_worker(self, r, inst, vol, pg, cache_key):
        import time, json, os, ai_parser
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
        api_key = ""
        if os.path.exists(config_path):
            try:
                with open(config_path, "r") as f:
                    config = json.load(f)
                    api_key = config.get("GEMINI_API_KEY", "")
            except: pass
        if not api_key: return
        
        import glob
        all_docs = []
        for ext in ("*.pdf",):
            all_docs.extend(glob.glob(os.path.join(self.pid_dir, "**", ext), recursive=True))
            all_docs.extend(glob.glob(os.path.join(self.base_dir, ext)))
            
        target_pdf = None
        for doc in all_docs:
            if "Irrelevant" in doc: continue
            fname = os.path.basename(doc).upper()
            import re
            nums = re.findall(r'\d+', fname)
            # Use strict regex check if we have it, or just substring
            if vol in nums and pg in nums:
                target_pdf = doc
                break
            # Fallback to simple substring
            if vol in fname and pg in fname:
                target_pdf = doc
                break
            
        if target_pdf:
            self.after(0, self._set_loading_icon, r)
            result = ai_parser.generate_qc_check(api_key, target_pdf, inst)
            self.ai_qc_cache[cache_key] = result
            self.save_ai_cache()
            
            title = ai_parser.extract_exact_title(api_key, target_pdf)
            if title:
                self.ai_title_cache[cache_key] = title
                self.save_ai_title_cache()
                
            def finalize():
                self._remove_loading_icon(r, add_check=True)
                if self.current_row_idx == r:
                    self.on_select(None)
                # Apply magnifying glass if needed
                if title:
                    idx = self.row_indices.index(r)
                    text = self.listbox.get(idx)
                    i_val = str(self.ws.cell(row=r, column=1).value or "").strip().lower()
                    if i_val != title.lower() and "🔍" not in text:
                        self.listbox.delete(idx)
                        self.listbox.insert(idx, text + " 🔍")
                        self._apply_row_color(idx)
                if getattr(self, 'current_row_idx', None) == r:
                    self._update_instrument_label(r)
                        
            self.after(0, finalize)
            
    def _apply_row_color(self, idx):
        if idx >= len(self.row_indices): return
        r = self.row_indices[idx]
        status = self.row_statuses.get(str(r), "Not Started")
        has_warn = str(r) in getattr(self, 'row_warnings', {})
        fg_color = 'red' if has_warn and status != "Completed" else ''
        
        if status == "In Progress":
            self.listbox.itemconfig(idx, {'bg': '#fff3cd', 'fg': fg_color or 'black'})
        elif status == "Completed":
            self.listbox.itemconfig(idx, {'bg': '#d4edda', 'fg': fg_color or 'black'})
        else:
            self.listbox.itemconfig(idx, {'bg': '', 'fg': fg_color})
                
    def open_excel_runsheet(self):
        import subprocess
        try:
            subprocess.Popen(["open", self.excel_path])
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel file:\n{e}", parent=self)

    def open_pdf_for_row(self):
        if not self.current_row_idx:
            return
            
        row = self.ws[self.current_row_idx]
        book_type = str(row[1].value).strip() if row[1].value else ""
        vol = str(row[2].value).strip() if row[2].value else ""
        page = str(row[3].value).strip() if row[3].value else ""
        
        if not book_type and not vol and not page:
            messagebox.showinfo("Info", "No Book/Vol/Page data to search for.", parent=self)
            return
            
        import glob
        import subprocess
        # Search for document files recursively in the PID folder, and non-recursively in the root folder
        all_docs = []
        for ext in ("*.pdf", "*.txt", "*.doc", "*.docx", "*.rtf", "*.png", "*.jpg", "*.tif", "*.tiff"):
            all_docs.extend(glob.glob(os.path.join(self.pid_dir, "**", ext), recursive=True))
            # Also include the root (base) directory just in case docs are loose there
            all_docs.extend(glob.glob(os.path.join(self.base_dir, ext)))
        
        matches = []
        for doc in all_docs:
            fname = os.path.basename(doc).lower()
            # Replace underscores and hyphens in filename with spaces for easier matching
            clean_fname = fname.replace("_", " ").replace("-", " ")
            clean_vol = vol.lower()
            clean_page = page.lower()
            clean_book = book_type.lower()
            
            # Match volume, page, and book type
            if clean_vol in clean_fname and clean_page in clean_fname:
                if not clean_book or clean_book in clean_fname:
                    matches.append(doc)
                    
        if matches:
            try:
                subprocess.run(["open", matches[0]])
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open Document:\n{e}", parent=self)
        else:
            messagebox.showwarning("Not Found", f"Could not find a Document for {book_type} {vol}/{page} in {os.path.basename(self.pid_dir)}", parent=self)

    def confirm_and_reformat_all(self):
        confirm = messagebox.askyesno(
            "Reformat Comments",
            "Do you really want to re-run the initial formatting rules on all comments?\n\n"
            "• A timestamped backup copy of your current spreadsheet will be saved first.\n"
            "• This will re-parse ARTI, Amount, Maturity, Dower, and re-generate the Original Notes blocks.\n\n"
            "Do you want to proceed?",
            parent=self
        )
        if not confirm:
            return
            
        # 1. Create timestamped backup
        import datetime, shutil, json
        backups_dir = os.path.join(self.pid_dir, "BACKUPS")
        os.makedirs(backups_dir, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.splitext(os.path.basename(self.excel_path))[0]
        backup_file = os.path.join(backups_dir, f"{base_name}_backup_{ts}.xlsx")
        
        try:
            # Save current state before backing up
            self.wb.save(self.excel_path)
            shutil.copy2(self.excel_path, backup_file)
        except Exception as e:
            messagebox.showerror("Backup Error", f"Failed to create backup prior to reformatting:\n{e}", parent=self)
            return

        # 2. Force reformat all comments
        comments_col = None
        for i, h in enumerate(self.headers):
            if "comment" in str(h).lower() or "note" in str(h).lower():
                comments_col = i + 1
                break
                
        if not comments_col:
            messagebox.showwarning("Warning", "Could not find a Comments/Notes column in this sheet.", parent=self)
            return

        from openpyxl.cell.rich_text import TextBlock, CellRichText
        changed_count = 0
        for row_idx in range(3, self.ws.max_row + 1):
            cell = self.ws.cell(row=row_idx, column=comments_col)
            val = cell.value
            if not val:
                continue
                
            if type(val).__name__ == 'CellRichText':
                txt = "".join(str(p) for p in val)
            else:
                txt = str(val)
                
            # If the user previously had an Original block, extract the original text
            if "--- Original ---" in txt:
                raw_text_to_parse = txt.split("--- Original ---")[1].strip()
            else:
                raw_text_to_parse = txt.replace("\u200B", "").strip()
                
            inst_type = str(self.ws.cell(row=row_idx, column=1).value or "").lower()
            formatted_txt = self.apply_initial_formatting_pipeline(raw_text_to_parse, inst_type)
            
            parts = self._parse_bold_tokens(formatted_txt)
            if not any(isinstance(p, TextBlock) for p in parts):
                cell.value = "".join(parts).strip()
            else:
                cell.value = CellRichText(*parts)
            changed_count += 1

        try:
            self.wb.save(self.excel_path)
            if hasattr(self, 'formatted_state_file'):
                with open(self.formatted_state_file, "w") as f:
                    json.dump({"initial_formatting_completed": True}, f)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save reformatted spreadsheet:\n{e}", parent=self)
            return
            
        # Refresh UI
        self.load_rows()
        if self.current_row_idx and self.current_row_idx in self.row_indices:
            lb_idx = self.row_indices.index(self.current_row_idx)
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(lb_idx)
            self.on_select(None)
            
        messagebox.showinfo(
            "Reformat Complete",
            f"Successfully reformatted {changed_count} row(s)!\n\n"
            f"Prior version backed up to:\n{os.path.basename(backup_file)}",
            parent=self
        )

    def initial_format_all_comments(self):
        # Run one-time initial formatting on raw comments in unformatted rows
        if hasattr(self, 'formatted_state_file') and os.path.exists(self.formatted_state_file):
            return
            
        comments_col = None
        for i, h in enumerate(self.headers):
            if "comment" in str(h).lower() or "note" in str(h).lower():
                comments_col = i + 1
                break
        if not comments_col:
            return
            
        changed = False
        for row_idx in range(3, self.ws.max_row + 1):
            cell = self.ws.cell(row=row_idx, column=comments_col)
            val = cell.value
            if not val:
                continue
                
            if type(val).__name__ == 'CellRichText':
                txt = "".join(str(p) for p in val)
            else:
                txt = str(val)
                
            # If already has the Original block or marker, skip
            if "\u200B" in txt or "--- Original ---" in txt:
                continue
                
            inst_type = str(self.ws.cell(row=row_idx, column=1).value or "").lower()
            formatted_txt = self.apply_initial_formatting_pipeline(txt, inst_type)
            if formatted_txt != txt:
                parts = self._parse_bold_tokens(formatted_txt)
                if not any(isinstance(p, TextBlock) for p in parts):
                    cell.value = "".join(parts).strip()
                else:
                    cell.value = CellRichText(*parts)
                changed = True
                
        if changed:
            try:
                self.wb.save(self.excel_path)
            except Exception as e:
                print(f"Failed to save initially formatted comments: {e}")
                
        # Mark initial formatting as completed for this parcel
        try:
            if hasattr(self, 'formatted_state_file'):
                import json
                with open(self.formatted_state_file, "w") as f:
                    json.dump({"initial_formatting_completed": True}, f)
        except Exception as e:
            print(f"Failed to write formatting state: {e}")

    def _parse_bold_tokens(self, text):
        parts = []
        is_bold_context = False
        import re
        from openpyxl.cell.rich_text import TextBlock, CellRichText
        from openpyxl.cell.text import InlineFont
        token_splits = re.split(r'(\[\[BOLD_START\]\]|\[\[BOLD_END\]\])', text)
        for token in token_splits:
            if token == '[[BOLD_START]]':
                is_bold_context = True
            elif token == '[[BOLD_END]]':
                is_bold_context = False
            elif token:
                if is_bold_context:
                    parts.append(TextBlock(InlineFont(b=True), token))
                else:
                    parts.append(token)
        return parts

    def find_book_type_for_vol_pg(self, vol, pg):
        import glob
        vol = str(vol).strip()
        pg = str(pg).strip()
        
        search_dirs = []
        if hasattr(self, 'excel_path') and self.excel_path:
            pdir = os.path.dirname(self.excel_path)
            search_dirs.extend([pdir, os.path.join(pdir, "DOCS"), os.path.dirname(pdir), os.path.join(os.path.dirname(pdir), "DOCS")])
            
        for sdir in search_dirs:
            if sdir and os.path.exists(sdir):
                for p in glob.glob(os.path.join(sdir, "*.pdf")):
                    fname = os.path.basename(p).upper()
                    m = re.search(r'\b(DR|OR|MR|LR|PR|WR|MISC|PL)\s*[-_ ]\s*' + re.escape(vol) + r'\s*[-_ /]\s*' + re.escape(pg) + r'\b', fname)
                    if m:
                        return m.group(1).upper()
                    m2 = re.search(r'\b(DR|OR|MR|LR|PR|WR|MISC|PL)\s*' + re.escape(vol) + r'\s*[-_ /]\s*' + re.escape(pg) + r'\b', fname)
                    if m2:
                        return m2.group(1).upper()

        if hasattr(self, 'ws') and self.ws:
            for r in range(2, self.ws.max_row + 1):
                r_vol = str(self.ws.cell(row=r, column=3).value or "").strip()
                r_pg = str(self.ws.cell(row=r, column=4).value or "").strip()
                if r_vol == vol and r_pg == pg:
                    r_btype = str(self.ws.cell(row=r, column=2).value or "").strip().upper()
                    if r_btype in ["DR", "OR", "MR", "LR", "PR", "WR", "MISC"]:
                        return r_btype
                    r_itype = str(self.ws.cell(row=r, column=1).value or "").lower()
                    if "deed" in r_itype: return "DR"
                    if "mortgage" in r_itype: return "MR"
                    if "lease" in r_itype: return "LR"

        try:
            v_num = int(vol)
            if v_num <= 805:
                return "DR"
            else:
                return "OR"
        except:
            return "DR"

    def normalize_prior_ref_string(self, prior_ref_raw):
        if not prior_ref_raw:
            return ""
            
        prior_ref_clean = re.sub(r'^(?:Prior\s*(?:deed\s*)?references?|Prior\s*Ref)\s*[:.]?\s*', '', prior_ref_raw, flags=re.IGNORECASE).strip()
        prior_ref_clean = prior_ref_clean.rstrip('.')
        
        # Check standard format: DR 554/912 or DR 554-912
        m_std = re.search(r'\b(DR|OR|MR|LR|PR|WR|MISC)\s+(\d+)[/-](\d+)\b', prior_ref_clean, re.IGNORECASE)
        if m_std:
            btype = m_std.group(1).upper()
            vol = m_std.group(2)
            pg = m_std.group(3)
            return f"Prior Ref: {btype} {vol}/{pg}"

        # Named book: Deed Book 554, Page 912 or Official Records 101, Page 345
        m_named = re.search(r'\b(Deed\s*(?:Book|Record|Vol)?|Official\s*Records?|Mortgage\s*(?:Book|Record|Vol)?|Lease\s*(?:Book|Record|Vol)?)\s*[:.]?\s*(?:Vol(?:ume)?\.?\s*)?(\d+)[,\s]+(?:Page|Pg|p\.?)\s*(\d+)\b', prior_ref_clean, re.IGNORECASE)
        if m_named:
            bname = m_named.group(1).lower()
            vol = m_named.group(2)
            pg = m_named.group(3)
            btype = "DR"
            if "official" in bname: btype = "OR"
            elif "mortgage" in bname: btype = "MR"
            elif "lease" in bname: btype = "LR"
            return f"Prior Ref: {btype} {vol}/{pg}"

        # Generic Vol 554, Page 912 or Volume 4, Page 3
        m_vol_pg = re.search(r'\b(?:Vol(?:ume)?\.?|Bk\.?|Book)\s*(\d+)[,\s]+(?:Page|Pg|p\.?)\s*(\d+)\b', prior_ref_clean, re.IGNORECASE)
        if m_vol_pg:
            vol = m_vol_pg.group(1)
            pg = m_vol_pg.group(2)
            btype = self.find_book_type_for_vol_pg(vol, pg)
            return f"Prior Ref: {btype} {vol}/{pg}"

        # Bare numbers: 554/912 or 554-912
        m_bare = re.search(r'\b(\d{1,4})[/-](\d{1,4})\b', prior_ref_clean)
        if m_bare:
            vol = m_bare.group(1)
            pg = m_bare.group(2)
            btype = self.find_book_type_for_vol_pg(vol, pg)
            return f"Prior Ref: {btype} {vol}/{pg}"

        return f"Prior Ref: {prior_ref_clean}"

    def apply_initial_formatting_pipeline(self, txt, inst_type=""):
        import re
        raw_original = txt
        prior_ref_str = ""
        txt = re.sub(r'Prior\s*deed\s*references?:?', 'Prior Ref:', txt, flags=re.IGNORECASE)
        txt = re.sub(r'Prior\s*references?:?', 'Prior Ref:', txt, flags=re.IGNORECASE)
        
        # Remove duplicate adjacent book identifiers (e.g. 'DR DR' -> 'DR')
        dup_book_pattern = r'\b(DR|OR|MR|LR|PR|PA|WR|MISC|DB|MB|PB)\s+(?=(?:DR|OR|MR|LR|PR|PA|WR|MISC|DB|MB|PB)\b)'
        while re.search(dup_book_pattern, txt, flags=re.IGNORECASE):
            txt = re.sub(dup_book_pattern, '', txt, flags=re.IGNORECASE)

        # Extract and normalize Prior Ref
        prior_ref_match = re.search(r'(Prior Ref:[^\n\.]*(?:\.|$))', txt, re.IGNORECASE)
        if prior_ref_match:
            prior_ref_str = self.normalize_prior_ref_string(prior_ref_match.group(1).strip())
            txt = txt.replace(prior_ref_match.group(0), "").strip()

        # EXCEPTING / RESERVING
        txt = re.sub(r'(?i)\bexcepting\b', 'EXCEPTING', txt)
        def format_reserving_oag(m):
            sentence = m.group(0).strip()
            sentence = re.sub(r'(?i)\breserving\b', 'RESERVING', sentence)
            if "[[BOLD_START]]" not in sentence:
                sentence = f"[[BOLD_START]]{sentence}[[BOLD_END]]"
            return f"\n\n{sentence}\n\n"
        txt = re.sub(r'(?i)[^.\n]*\breserving\b[^.\n]*oil\s+and\s+gas[^.\n]*(?:\.|$)', format_reserving_oag, txt)

        def format_own_line(m):
            s = m.group(1).strip()
            return f"\n\n{s}\n\n"
        txt = re.sub(r'(?:^|(?<=\.))\s*((?:EXCEPTING|EXCEPTIONS?|RESERVES?|RESERVATIONS?|RESERVING)\b[^.\n]*(?:\.|$))', format_own_line, txt, flags=re.IGNORECASE)

        # RELEASES / SATISFACTIONS
        if re.search(r'(?:release|satisfaction)\s+(?:of\s+)?mortgage|releases\s+mortgage', txt, re.IGNORECASE) or "release" in inst_type.lower() or "satisfaction" in inst_type.lower():
            def repl_rel(m):
                btype = m.group(1) or ""
                vol = m.group(2)
                pg = m.group(3)
                if not btype or btype.upper() in ["BOOK", "VOL", "VOLUME", "RECORD", "."]:
                    btype = self.find_book_type_for_vol_pg(vol, pg)
                    if btype == "DR": btype = "MR"
                else:
                    btype = btype.upper()
                return f"Releases mortgage recorded in {btype} {vol}/{pg}\nFull satisfaction. Clears lien from the property title."

            rel_pattern = r'(?:Releases?\s*(?:of\s*)?(?:mortgage\s*)?(?:recorded\s*)?(?:in\s*)?(?:by\s*)?:?\s*(?:SEE\s*)?)(?:(?:Book|Vol(?:ume)?\.?|Record)\s*)?(DR|OR|MR|LR|PR|PA|WR|MISC|\.)?\s*(\d+)[-/\s,]+(?:PAGE\s*|PG\s*|p\.?\s*)?(\d+)(?:\.?\s*(?:Full\s+satisfaction\.?\s*)?(?:Clears\s+lien\s+from\s+the\s+property\s+title\.?)?)?'
            txt = re.sub(rel_pattern, repl_rel, txt, flags=re.IGNORECASE)

        # ARTI
        if "ARTI\n" not in txt:
            arti_pattern = re.compile(r'(?i)(?:[^.]*?\bconvey(?:s|ed)?\b\s+)?[^.]*?all,?\s+(?:of\s+)?(?:its\s+|his\s+|her\s+|their\s+)?rights?,?\s*title,?\s*(?:and|&)\s*interest[^.]*(?:\.|$)' )
            match = arti_pattern.search(txt)
            if match:
                extracted = match.group(0).strip()
                extracted = re.sub(r'(?i)\bconvey(?:s|ed)?\s+(?:of\s+)?(?:its\s+|his\s+|her\s+|their\s+)?all,?\s+(?:of\s+)?(?:its\s+|his\s+|her\s+|their\s+)?rights?,?\s*title,?\s*(?:and|&)\s*interest\b', 'Conveys all right, title, and interest', extracted)
                extracted = re.sub(r'(?i)\ball,?\s+(?:of\s+)?(?:its\s+|his\s+|her\s+|their\s+)?rights?,?\s*title,?\s*(?:and|&)\s*interest\b', 'all right, title, and interest', extracted)
                txt = txt.replace(match.group(0).strip(), "").strip()
                txt = re.sub(r'^[,\.]\s*', '', txt)
                txt = re.sub(r'\s+([,\.])', r'\1', txt)
                txt = re.sub(r'[,;\s]+\.', '.', txt)
                txt = re.sub(r'\.+', '.', txt)
                txt = txt.strip()
                if txt: txt = txt[0].upper() + txt[1:]
                txt = f"ARTI\n{extracted}\n{txt}" if txt else f"ARTI\n{extracted}"

        # AMOUNT & MATURITY
        if "Amount: " not in txt and "Maturity Date: " not in txt:
            amount_str = ""
            maturity_str = ""
            amount_match = re.search(r'(?<!Amount: )(\$\d{1,3}(?:,\d{3})*(?:\.\d{2})?)', txt)
            if amount_match:
                amount = amount_match.group(1)
                amount_str = f"Amount: {amount}"
                txt = txt.replace(amount, "").strip()
                if txt.lower().startswith("loan"):
                    txt = txt[4:].strip()

                import dateutil.parser
                date_match = re.search(r'(?:due\s+|payable\s+on\s+|on\s+or\s+before\s+)?\b((?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\s+\d{1,2},?\s+\d{4}|\d{1,2}/\d{1,2}/\d{2,4})\b', txt, re.IGNORECASE)
                if date_match:
                    try:
                        dt = dateutil.parser.parse(date_match.group(1))
                        maturity_str = f"Maturity Date: {dt.strftime('%m/%d/%Y')}"
                        txt = txt.replace(date_match.group(0), "").strip()
                    except:
                        maturity_str = "Maturity Date: Not stated."
                else:
                    if "Maturity Date:" not in txt:
                        maturity_str = "Maturity Date: Not stated."
                    
            if amount_str or maturity_str:
                txt = re.sub(r'^[,\.]\s*', '', txt)
                txt = re.sub(r'\s+([,\.])', r'\1', txt)
                txt = re.sub(r'[,;\s]+\.', '.', txt)
                txt = re.sub(r'\.+', '.', txt)
                txt = txt.strip()
                if txt: txt = txt[0].upper() + txt[1:]
            
                if txt.startswith("ARTI\n"):
                    arti_lines = txt.split('\n')
                    arti_part = arti_lines[0] + '\n' + arti_lines[1]
                    rest = '\n'.join(arti_lines[2:]).strip()
                
                    final_parts = [arti_part, ""]
                    if amount_str: final_parts.append(amount_str)
                    if maturity_str: final_parts.append(maturity_str)
                    if amount_str or maturity_str: final_parts.append("")
                    if rest: final_parts.append(rest)
                    txt = '\n'.join(final_parts).strip()
                else:
                    final_parts = []
                    if amount_str: final_parts.append(amount_str)
                    if maturity_str: final_parts.append(maturity_str)
                    if amount_str or maturity_str: final_parts.append("")
                    if txt: final_parts.append(txt)
                    txt = '\n'.join(final_parts).strip()

        # Dower
        is_dower_applicable = ("deed" in inst_type or "mortgage" in inst_type) and "release" not in inst_type and "satisfaction" not in inst_type
        dower_str = ""
        if re.search(r'(?i)dower\s+(?:rights\s+)?(?:is\s+)?released', txt):
            txt = re.sub(r'(?i)dower\s+(?:rights\s+)?(?:is\s+)?released\.?', '', txt).strip()
            txt = re.sub(r'(?i)dower\s+mentioned\s*that\s*is\s*released\.?', '', txt).strip()
            if is_dower_applicable:
                dower_str = "Dower released."
        elif re.search(r'(?i)no\s+dower|dower\s+(?:is\s+)?not\s+stated', txt):
            txt = re.sub(r'(?i)(?:no\s+dower(?:\s+mentioned)?|dower\s+(?:is\s+)?not\s+stated)\.?', '', txt).strip()
            if is_dower_applicable:
                dower_str = "No dower mentioned."
        else:
            if is_dower_applicable:
                dower_str = "No dower mentioned."

        if dower_str:
            txt = f"{txt}\n{dower_str}".strip()

        if prior_ref_str:
            txt = f"{txt}\n{prior_ref_str}".strip()

        if txt != raw_original:
            txt = f"{txt}\n\n--- Original ---\n{raw_original}"

        txt = "\u200B" + txt
        return txt

    def format_cell_text(self, header_name, txt):
        if not isinstance(txt, str):
            return txt
            
        import re
        # Collapse multiple spaces and tabs into a single space
        txt = re.sub(r'[ \t]{2,}', ' ', txt)
        
        # Remove duplicate adjacent book identifiers (e.g. 'DR DR' -> 'DR', always deleting the front duplicate)
        dup_book_pattern = r'\b(DR|OR|MR|LR|PR|PA|WR|MISC|DB|MB|PB)\s+(?=(?:DR|OR|MR|LR|PR|PA|WR|MISC|DB|MB|PB)\b)'
        while re.search(dup_book_pattern, txt, flags=re.IGNORECASE):
            txt = re.sub(dup_book_pattern, '', txt, flags=re.IGNORECASE)
        
        if "grantor" in header_name or "grantee" in header_name:
            txt = re.sub(r'(?i)\bhusband\s+and\s+wife\b', 'husband and wife', txt)
            txt = re.sub(r'(?i)\bhusband\s*&\s*wife\b', 'husband and wife', txt)
            
        if "comments" in header_name:
            import re
            
            # Dynamic Dower Check (Runs on save to catch Instrument type changes)
            inst_type = ""
            try:
                for i, h in enumerate(self.headers):
                    if h.lower() == 'instrument':
                        widget = self.widgets_by_col.get(i)
                        if widget and hasattr(widget, 'get'):
                            inst_type = widget.get().lower()
                        break
            except: pass
            
            is_dower_applicable = ("deed" in inst_type or "mortgage" in inst_type) and "release" not in inst_type and "satisfaction" not in inst_type
            text_without_original = txt.split("--- Original ---")[0]
            if is_dower_applicable and not re.search(r'(?i)dower', text_without_original):
                if "--- Original ---" in txt:
                    txt = txt.replace("--- Original ---", "No dower mentioned.\n\n--- Original ---")
                else:
                    txt = f"{txt}\nNo dower mentioned.".strip()

            def format_released(m):
                book = m.group(1) or ''
                vol = m.group(2)
                pg = m.group(3)
                book = book.strip().upper()
                if not book or book in ["BOOK", "VOL", "VOLUME", "RECORD", "."]:
                    book = self.find_book_type_for_vol_pg(vol, pg)
                    if book == "DR": book = "MR"
                
                if "release" in inst_type or "satisfaction" in inst_type or "releases mortgage" in txt.lower():
                    return f'Releases mortgage recorded in {book} {vol}/{pg}\nFull satisfaction. Clears lien from the property title.'
                else:
                    return f'Release: {book} {vol}/{pg}'
                    
            rel_pattern = r'(?:Release(?:s|d)?\s*(?:of\s*)?(?:mortgage\s*)?(?:recorded\s*)?(?:in\s*)?(?:by\s*)?:?\s*(?:SEE\s*)?)(?:(?:Book|Vol(?:ume)?\.?|Record)\s*)?(DR|OR|MR|LR|PR|PA|WR|MISC|\.)?\s*(\d+)[-/\s,]+(?:PAGE\s*|PG\s*|p\.?\s*)?(\d+)(?:\.?\s*(?:Full\s+satisfaction\.?\s*)?(?:Clears\s+lien\s+from\s+the\s+property\s+title\.?)?)?'
            txt = re.sub(rel_pattern, format_released, txt, flags=re.IGNORECASE)
            txt = re.sub(r'([^\n\u200B])\s*(Releases mortgage recorded in)', r'\1\n\2', txt)
            txt = re.sub(r'([^\n\u200B])\s*(Release:)', r'\1\n\2', txt)
            txt = re.sub(r'(Release:\s*(?:[A-Z]+\s*)?\d+[-/]\d+)[.,;]?\s+(?=[A-Za-z0-9])', r'\1\n', txt)
            txt = re.sub(r'(Release:\s*(?:[A-Z]+\s*)?\d+[-/]\d+)\.$', r'\1', txt)
            
            # Prior Ref normalization in comments
            def repl_prior_ref_match(m):
                return "\n" + self.normalize_prior_ref_string(m.group(0))
            txt = re.sub(r'(?:Prior\s*(?:deed\s*)?references?|Prior\s*Ref)\s*[:.]?\s*[^\n\.;]+', repl_prior_ref_match, txt, flags=re.IGNORECASE)
            txt = re.sub(r'\n{3,}', '\n\n', txt)

            if "\u200B" not in txt:
                txt = "\u200B" + txt
                
            return txt
        return txt

    def nav_prev_row(self, event=None):
        if not self.row_indices:
            return "break"
        if self.current_row_idx and self.current_row_idx in self.row_indices:
            curr_pos = self.row_indices.index(self.current_row_idx)
        else:
            sel = self.listbox.curselection()
            curr_pos = sel[0] if sel else 0
            
        if curr_pos > 0:
            if self.current_row_idx:
                try: self.save_row(show_msg=False)
                except Exception: pass
            new_pos = curr_pos - 1
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(new_pos)
            self.listbox.see(new_pos)
            self.on_select(None)
        return "break"

    def nav_next_row(self, event=None):
        if not self.row_indices:
            return "break"
        if self.current_row_idx and self.current_row_idx in self.row_indices:
            curr_pos = self.row_indices.index(self.current_row_idx)
        else:
            sel = self.listbox.curselection()
            curr_pos = sel[0] if sel else 0
            
        if curr_pos < len(self.row_indices) - 1:
            if self.current_row_idx:
                try: self.save_row(show_msg=False)
                except Exception: pass
            new_pos = curr_pos + 1
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(new_pos)
            self.listbox.see(new_pos)
            self.on_select(None)
        return "break"

    def set_status_in_progress(self, event=None):
        self.status_var.set("In Progress")
        self.on_status_change(None)
        return "break"

    def set_status_completed(self, event=None):
        self.status_var.set("Completed")
        self.on_status_change(None)
        return "break"

    def delete_original_block(self, event=None):
        comments_widget = None
        for i, h in enumerate(self.headers):
            if "comment" in h.lower() or "note" in h.lower():
                w = self.widgets_by_col.get(i)
                if isinstance(w, tk.Text):
                    comments_widget = w
                break
        
        focus_w = self.focus_get()
        target_w = focus_w if isinstance(focus_w, tk.Text) else comments_widget
        
        if target_w and isinstance(target_w, tk.Text):
            txt = target_w.get("1.0", "end-1c")
            if "--- Original ---" in txt:
                new_txt = txt.split("--- Original ---")[0].rstrip()
                target_w.delete("1.0", tk.END)
                target_w.insert("1.0", new_txt)
                
                self.perform_spellcheck(target_w)
                self.highlight_links(target_w)
                
                # Re-evaluate warnings immediately
                if getattr(self, 'current_row_idx', None):
                    warnings = getattr(self, 'row_warnings', {}).get(str(self.current_row_idx), [])
                    if "Original text block not deleted." in warnings:
                        warnings.remove("Original text block not deleted.")
                        if warnings:
                            self.row_warnings[str(self.current_row_idx)] = warnings
                        else:
                            if str(self.current_row_idx) in self.row_warnings:
                                del self.row_warnings[str(self.current_row_idx)]
                        
                        try:
                            idx = self.row_indices.index(self.current_row_idx)
                            self._apply_row_color(idx)
                        except: pass
                        
                        warn_text = ""
                        if warnings: warn_text += "⚠️ " + " | ".join(warnings)
                        if getattr(self, 'current_ai_full_text', ""):
                            vol = str(self.ws.cell(row=self.current_row_idx, column=3).value or "").strip()
                            pg = str(self.ws.cell(row=self.current_row_idx, column=4).value or "").strip()
                            cache_key = f"{vol}_{pg}"
                            if hasattr(self, 'ai_qc_cache') and cache_key in self.ai_qc_cache:
                                raw_text = self.ai_qc_cache[cache_key]
                                if "SUMMARY:" in raw_text and "FULL TEXT:" in raw_text:
                                    parts = raw_text.split("FULL TEXT:")
                                    warn_text += "\n🤖 AI Check " + parts[0].replace("SUMMARY:", "").strip()
                        self.warning_label.config(text=warn_text.strip())
        return "break"

    def toggle_dower_reviewed(self, event=None):
        if getattr(self, 'qc_vars', None) and len(self.qc_vars) > 0:
            current_val = self.qc_vars[0].get()
            self.qc_vars[0].set(not current_val)
            if hasattr(self, 'on_dower_toggle_cb') and callable(self.on_dower_toggle_cb):
                self.on_dower_toggle_cb()
        return "break"

    def convert_to_normal_case(self, event=None):
        widget = self.focus_get()
        if isinstance(widget, tk.Text):
            try:
                if widget.tag_ranges(tk.SEL):
                    sel_text = widget.get(tk.SEL_FIRST, tk.SEL_LAST)
                    new_text = sel_text.title()
                    first_idx = widget.index(tk.SEL_FIRST)
                    widget.delete(tk.SEL_FIRST, tk.SEL_LAST)
                    widget.insert(first_idx, new_text)
                    widget.tag_add(tk.SEL, first_idx, f"{first_idx}+{len(new_text)}c")
                else:
                    txt = widget.get("1.0", "end-1c")
                    widget.delete("1.0", tk.END)
                    widget.insert("1.0", txt.title())
            except Exception: pass
        elif isinstance(widget, (ttk.Entry, tk.Entry)):
            try:
                if widget.selection_present():
                    start = widget.index(tk.SEL_FIRST)
                    end = widget.index(tk.SEL_LAST)
                    sel_text = widget.get()[start:end]
                    new_text = sel_text.title()
                    widget.delete(start, end)
                    widget.insert(start, new_text)
                    widget.selection_range(start, start + len(new_text))
                else:
                    txt = widget.get()
                    widget.delete(0, tk.END)
                    widget.insert(0, txt.title())
            except Exception: pass
        return "break"

    def insert_phrase_by_num(self, num, event=None):
        import json
        app_dir = os.path.dirname(os.path.abspath(__file__))
        phrases_file = os.path.join(app_dir, "phrases.json")
        default_phrases = [
            "Dower released.",
            "No dower mentioned.",
            "Expired by term.",
            "Releases mortgage recorded in",
            "for their joint lives, remainder to the survivor of them",
            "Maturity Date: Not stated.",
            "No release found of record.",
            "Prior Ref:",
            "Additional lands included, not part of subject.",
            "ARTI"
        ]
        phrases = default_phrases[:]
        if os.path.exists(phrases_file):
            try:
                with open(phrases_file, "r") as f:
                    phrases = json.load(f)
            except: pass

        if num == 0:
            idx = 9
        elif num == 20 or num == 10:
            idx = 19
        else:
            idx = num - 1
        if 0 <= idx < len(phrases):
            phrase = phrases[idx]
            widget = self.focus_get()
            if isinstance(widget, tk.Text):
                try:
                    widget.insert(tk.INSERT, phrase)
                    self.perform_spellcheck(widget)
                    self.highlight_links(widget)
                except Exception: pass
            elif isinstance(widget, (ttk.Entry, tk.Entry)):
                try:
                    widget.insert(tk.INSERT, phrase)
                except Exception: pass
        return "break"

    def show_save_success_dialog(self, row_idx):
        popup = tk.Toplevel(self)
        popup.title("Success")
        popup.geometry("320x150")
        popup.attributes("-topmost", True)
        popup.resizable(False, False)
        popup.transient(self)
        
        # Center over editor window
        try:
            x = self.winfo_rootx() + (self.winfo_width() // 2) - 160
            y = self.winfo_rooty() + (self.winfo_height() // 2) - 75
            popup.geometry(f"+{x}+{y}")
        except Exception: pass
        
        popup.grab_set()
        
        frame = ttk.Frame(popup, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text=f"✅  Row {row_idx} saved!", font=("Helvetica", 15, "bold")).pack(pady=(0, 15))
        
        def close_dialog(event=None):
            try:
                popup.grab_release()
                popup.destroy()
            except Exception: pass
            return "break"
            
        ok_btn = ttk.Button(frame, text="OK", command=close_dialog)
        ok_btn.pack()
        ok_btn.focus_set()
        
        ok_btn.bind("<Return>", close_dialog)
        ok_btn.bind("<KP_Enter>", close_dialog)
        ok_btn.bind("<space>", close_dialog)
        popup.bind("<Return>", close_dialog)
        popup.bind("<KP_Enter>", close_dialog)
        popup.bind("<space>", close_dialog)
        popup.bind("<Escape>", close_dialog)
        
        self.wait_window(popup)

    def show_ai_note_dialog(self, event=None):
        if not getattr(self, 'current_row_idx', None):
            return "break"
            
        popup = tk.Toplevel(self)
        popup.title(f"Row {self.current_row_idx} - AI Note & Error Bar Details")
        popup.geometry("640x520")
        popup.attributes("-topmost", True)
        popup.transient(self)
        
        try:
            x = self.winfo_rootx() + (self.winfo_width() // 2) - 320
            y = self.winfo_rooty() + (self.winfo_height() // 2) - 260
            popup.geometry(f"+{x}+{y}")
        except Exception: pass
        
        main_frame = ttk.Frame(popup, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        row_title = f"Row {self.current_row_idx} Details"
        vol = str(self.ws.cell(row=self.current_row_idx, column=3).value or "").strip()
        pg = str(self.ws.cell(row=self.current_row_idx, column=4).value or "").strip()
        inst = str(self.ws.cell(row=self.current_row_idx, column=1).value or "").strip()
        if vol or pg or inst:
            row_title += f" ({inst} {vol}/{pg})".strip()
            
        ttk.Label(main_frame, text=row_title, font=("Helvetica", 16, "bold")).pack(anchor=tk.W, pady=(0, 10))
        
        warn_frame = ttk.LabelFrame(main_frame, text=" ⚠️ Active Warnings / Error Bar ", padding=10)
        warn_frame.pack(fill=tk.X, pady=(0, 10))
        
        warnings = getattr(self, 'row_warnings', {}).get(str(self.current_row_idx), [])
        warn_text = getattr(self, 'warning_label', None)
        active_error_str = warn_text.cget("text") if warn_text else ""
        if not active_error_str:
            if warnings:
                active_error_str = "⚠️ " + " | ".join(warnings)
            else:
                active_error_str = "✅ No active errors or warnings for this row."
                
        lbl_err = tk.Label(warn_frame, text=active_error_str, font=("Helvetica", 14), fg="#b30000" if "⚠️" in active_error_str else "#006600", justify=tk.LEFT, anchor="w", wraplength=580)
        lbl_err.pack(fill=tk.X)
        
        ai_frame = ttk.LabelFrame(main_frame, text=" 🤖 Gemini AI QC Analysis & Note ", padding=10)
        ai_frame.pack(fill=tk.BOTH, expand=True)
        
        txt_box = tk.Text(ai_frame, font=("Helvetica", 14), wrap=tk.WORD)
        txt_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        sb = ttk.Scrollbar(ai_frame, orient="vertical", command=txt_box.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        txt_box.config(yscrollcommand=sb.set)
        
        full_ai_content = getattr(self, 'current_ai_full_text', "").strip()
        if not full_ai_content:
            cache_key = f"{vol}_{pg}"
            if hasattr(self, 'ai_qc_cache') and cache_key in self.ai_qc_cache:
                full_ai_content = self.ai_qc_cache[cache_key].strip()
                
        if not full_ai_content:
            full_ai_content = "No Gemini AI QC note cached for this document.\n(Click 'Retry AI Check' on the toolbar if you'd like Gemini to re-analyze this document)."
            
        txt_box.insert("1.0", full_ai_content)
        txt_box.config(state="disabled")
        
        btn_close = ttk.Button(main_frame, text="Close (Esc)", command=popup.destroy)
        btn_close.pack(pady=(10, 0), anchor=tk.E)
        btn_close.focus_set()
        
        popup.bind("<Escape>", lambda e: popup.destroy())
        popup.bind("<Return>", lambda e: popup.destroy())
        
        return "break"

    def open_or_sync_dialog(self, event=None):
        import or_sync_dialog
        p_id = getattr(self, 'parcel_id', getattr(self, 'pid', ''))
        rs_file = getattr(self, 'excel_path', None)
        or_sync_dialog.ORSyncDialog(self, self.pid_dir, parcel_num=p_id, rs_path=rs_file)
        return "break"

    def show_shortcuts_dialog(self):
        popup = tk.Toplevel(self)
        popup.title("Keyboard Shortcuts")
        popup.geometry("560x450")
        popup.attributes("-topmost", True)
        popup.resizable(False, False)
        
        main_frame = ttk.Frame(popup, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(main_frame, text="⌨️ Keyboard Shortcuts Cheat Sheet", font=("Helvetica", 16, "bold")).pack(anchor=tk.W, pady=(0, 10))
        
        shortcuts = [
            ("Ctrl + S / Cmd + S", "Save current row"),
            ("Ctrl + P / Cmd + P", "Set status to 'In Progress'"),
            ("Ctrl + F / Cmd + F", "Set status to 'Completed'"),
            ("Ctrl + A / Cmd + A", "View Gemini AI Note & Error Bar Details"),
            ("Ctrl + D / Cmd + D", "Toggle 'Dower Reviewed' checkbox"),
            ("Cmd+Shift+O / Ctrl+Shift+O", "📊 Sync data to Ownership Report (*OR*.xlsx)"),
            ("Cmd+Shift+D", "Delete '--- Original ---' notes block"),
            ("Ctrl + O / Cmd + O", "Open Document (PDF for current row)"),
            ("Ctrl + N / Cmd + N", "Convert selection or field to Title Case"),
            ("Ctrl + L / Cmd + L", "Open Phrase Library (Type number or search + Enter)"),
            ("Ctrl + 1 .. 9, 0", "Insert Phrase #1 through #10 at cursor"),
            ("Cmd/Ctrl + Shift + 1..0", "Insert Phrase #11 through #20 at cursor"),
            ("Option / Alt + 1..0", "Insert Phrase #11 through #20 at cursor"),
            ("Ctrl + ↑ / Ctrl + ↓", "Save & jump to Previous / Next Row"),
            ("Alt + ↑ / Alt + ↓", "Save & jump to Previous / Next Row"),
            ("↑ / ↓ (in listbox)", "Quick select & load row into editor"),
            ("Ctrl + B / Cmd + B", "Toggle bold on selected text"),
            ("Left-Click Link", "Save current row & jump to referenced row"),
            ("Right-Click Link", "Open referenced PDF Document")
        ]
        
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        tree = ttk.Treeview(tree_frame, columns=("shortcut", "action"), show="headings", height=11)
        tree.heading("shortcut", text="Shortcut Key")
        tree.heading("action", text="Action / Description")
        tree.column("shortcut", width=180, anchor=tk.W)
        tree.column("action", width=340, anchor=tk.W)
        
        for sc, act in shortcuts:
            tree.insert("", tk.END, values=(sc, act))
            
        tree.pack(fill=tk.BOTH, expand=True)
        
        ttk.Button(main_frame, text="Close", command=popup.destroy).pack(pady=(10, 0), anchor=tk.E)

    def open_phrase_library(self):
        import json
        app_dir = os.path.dirname(os.path.abspath(__file__))
        phrases_file = os.path.join(app_dir, "phrases.json")
        
        default_phrases = [
            "Dower released.",
            "No dower mentioned.",
            "Expired by term.",
            "Releases mortgage recorded in",
            "for their joint lives, remainder to the survivor of them",
            "Maturity Date: Not stated.",
            "No release found of record.",
            "Prior Ref:",
            "Additional lands included, not part of subject.",
            "ARTI"
        ]
        
        if not os.path.exists(phrases_file):
            phrases = default_phrases[:]
            with open(phrases_file, "w") as f:
                json.dump(phrases, f, indent=4)
        else:
            try:
                with open(phrases_file, "r") as f:
                    phrases = json.load(f)
            except:
                phrases = default_phrases[:]
                
        popup = tk.Toplevel(self)
        popup.title("Phrase Library (Type number or search, press Enter)")
        popup.geometry("480x520")
        popup.attributes("-topmost", True)
        
        # Search / Number Jump bar
        search_frame = ttk.Frame(popup)
        search_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        ttk.Label(search_frame, text="🔍 Jump to # / Filter:").pack(side=tk.LEFT)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, font=("Helvetica", 14))
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        
        list_frame = ttk.Frame(popup)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        listbox = tk.Listbox(list_frame, font=("Helvetica", 15), exportselection=False)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        listbox.config(yscrollcommand=scrollbar.set)
        
        filtered_indices = list(range(len(phrases)))
        
        def refresh_list(*args):
            nonlocal filtered_indices
            query = search_var.get().strip().lower()
            listbox.delete(0, tk.END)
            filtered_indices = []
            
            for idx, p in enumerate(phrases):
                display_p = p.replace("\n", " ↵ ")
                if idx < 9:
                    num_prefix = f"[{idx + 1}] "
                elif idx == 9:
                    num_prefix = "[0] "
                else:
                    num_prefix = f"[{idx + 1}] "
                
                full_item_str = f"{num_prefix}{display_p}"
                
                # Check match against query or exact index
                if not query:
                    listbox.insert(tk.END, full_item_str)
                    filtered_indices.append(idx)
                elif query.isdigit():
                    target_num = int(query)
                    if (target_num == 0 and idx == 9) or (target_num == idx + 1) or query in str(idx + 1):
                        listbox.insert(tk.END, full_item_str)
                        filtered_indices.append(idx)
                elif query in full_item_str.lower():
                    listbox.insert(tk.END, full_item_str)
                    filtered_indices.append(idx)
                    
            if listbox.size() > 0:
                listbox.selection_clear(0, tk.END)
                listbox.selection_set(0)
                listbox.activate(0)

        search_var.trace_add("write", refresh_list)
        refresh_list()
        search_entry.focus_set()
        
        btn_frame = ttk.Frame(popup)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)
        
        def insert_at_cursor(event=None):
            selection = listbox.curselection()
            if not selection and listbox.size() > 0:
                selection = (0,)
            if not selection:
                return
            orig_idx = filtered_indices[selection[0]]
            phrase = phrases[orig_idx]
            
            focus_widget = self.focus_lastfor()
            
            if isinstance(focus_widget, tk.Text):
                try:
                    focus_widget.insert(tk.INSERT, phrase)
                    self.perform_spellcheck(focus_widget)
                    self.highlight_links(focus_widget)
                except tk.TclError:
                    pass
            elif isinstance(focus_widget, (ttk.Entry, tk.Entry)):
                try:
                    focus_widget.insert(tk.INSERT, phrase)
                except tk.TclError:
                    pass
            else:
                popup.clipboard_clear()
                popup.clipboard_append(phrase)
                
            popup.destroy()
            return "break"
                
        def copy_to_clipboard():
            selection = listbox.curselection()
            if not selection and listbox.size() > 0:
                selection = (0,)
            if not selection:
                return
            orig_idx = filtered_indices[selection[0]]
            phrase = phrases[orig_idx]
            popup.clipboard_clear()
            popup.clipboard_append(phrase)
            
        ttk.Button(btn_frame, text="Insert at Cursor (Enter)", command=insert_at_cursor).pack(side=tk.LEFT, expand=True, fill=tk.X)
        ttk.Button(btn_frame, text="Copy to Clipboard", command=copy_to_clipboard).pack(side=tk.LEFT, expand=True, fill=tk.X)
        
        # Keybindings inside Phrase Library popup
        search_entry.bind("<Return>", insert_at_cursor)
        listbox.bind("<Return>", insert_at_cursor)
        listbox.bind("<Double-Button-1>", insert_at_cursor)
        popup.bind("<Escape>", lambda e: popup.destroy())
        
        def on_down_from_search(event):
            listbox.focus_set()
            if listbox.size() > 0 and not listbox.curselection():
                listbox.selection_set(0)
            return "break"
            
        search_entry.bind("<Down>", on_down_from_search)
        
        add_frame = ttk.Frame(popup)
        add_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        new_entry = tk.Text(add_frame, font=("Helvetica", 14), height=3, wrap=tk.WORD, undo=True, maxundo=-1, autoseparators=True)
        new_entry.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        
        def add_phrase():
            val = new_entry.get("1.0", tk.END).strip()
            if val:
                phrases.append(val)
                with open(phrases_file, "w") as f:
                    json.dump(phrases, f, indent=4)
                new_entry.delete("1.0", tk.END)
                refresh_list()
                
        ttk.Button(add_frame, text="Save Phrase", command=add_phrase).pack(side=tk.TOP, pady=(5, 0), anchor=tk.E)
        
        def delete_phrase():
            selection = listbox.curselection()
            if not selection:
                return
            orig_idx = filtered_indices[selection[0]]
            del phrases[orig_idx]
            with open(phrases_file, "w") as f:
                json.dump(phrases, f, indent=4)
            refresh_list()
            
        ttk.Button(btn_frame, text="Delete", command=delete_phrase).pack(side=tk.RIGHT, expand=False)
